"""Building the downloadable "who ordered what" spreadsheet handed to a
manufacturer/printer to place a bulk merchandise order. Kept separate from
bulk_import.py: that module is member-import-specific by name and
docstring, this is shop-specific and has nothing to do with importing
anything.
"""

import re
import uuid
from io import BytesIO

import openpyxl
from django.core.cache import cache
from django.utils.translation import gettext_lazy as _

TEMPLATE_COLUMNS = [_("Order"), _("Last name"), _("First name"), _("Option"), _("Number"), _("Name"), _("Quantity")]

#: Characters Excel refuses in a sheet title, plus the 31-char length limit.
_INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")

#: How long a just-built export stays downloadable after the mutating POST
#: redirects for it -- the auto-download script (order_list.html) fires
#: within moments of the page loading; this is just headroom for a slow
#: connection or a tab left idle mid-request.
STASHED_EXPORT_TTL_SECONDS = 300


def _sheet_title(name, taken):
    title = _INVALID_SHEET_TITLE_CHARS.sub("", name)[:31] or "Sheet"
    if title not in taken:
        return title
    for suffix in range(2, 100):
        candidate = f"{title[: 31 - len(str(suffix)) - 1]} {suffix}"
        if candidate not in taken:
            return candidate
    raise ValueError("Could not find a free sheet title.")  # pragma: no cover -- 98 same-named products is not a real case


def build_production_export(lines):
    """One sheet per product, one row per OrderLine -- ``lines`` is expected
    pre-filtered to what should actually go out (see shop.services.
    production.pending_production_lines), this just lays it out. Order
    within a product's sheet follows ``lines`` as given."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    by_product = {}
    for line in lines:
        by_product.setdefault(line.product, []).append(line)

    taken_titles = set()
    for product, product_lines in by_product.items():
        title = _sheet_title(product.name, taken_titles)
        taken_titles.add(title)
        sheet = workbook.create_sheet(title=title)

        sheet.append([str(column) for column in TEMPLATE_COLUMNS])
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        sheet.freeze_panes = "A2"

        for line in product_lines:
            person = line.beneficiary or line.order.purchaser
            sheet.append([line.order.number, person.last_name, person.first_name, line.variant.name if line.variant_id else "", line.personalization_number, line.personalization_name, line.quantity])

        for column_index, column_name in enumerate(TEMPLATE_COLUMNS, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = max(12, len(str(column_name)) + 2)

    return workbook


def stash_production_export(club, workbook, filename):
    """Saves a built workbook under a one-time token, keyed to ``club`` so a
    guessed/leaked token can't pull another club's file (see
    pop_production_export). Exists because a mutating export (management.
    views.OrderProductionExportView) redirects back to a reloaded order_list
    -- so its Production column actually shows the marking that just
    happened -- rather than returning the file directly: a browser doesn't
    navigate away from a page when a form POST's response is a file
    download, so the page would otherwise look unchanged even though the
    marking succeeded server-side. The redirect target auto-triggers the
    real download via this token instead."""
    buffer = BytesIO()
    workbook.save(buffer)
    token = uuid.uuid4().hex
    cache.set(f"production-export:{token}", (club.pk, filename, buffer.getvalue()), timeout=STASHED_EXPORT_TTL_SECONDS)
    return token


def pop_production_export(club, token):
    """Retrieves and deletes a stash -- one-time use, and only for the club
    it was stashed under. ``None`` if the token is missing, expired, or
    belongs to a different club."""
    key = f"production-export:{token}"
    stashed = cache.get(key)
    cache.delete(key)
    if stashed is None:
        return None
    club_id, filename, content = stashed
    if club_id != club.pk:
        return None
    return filename, content
