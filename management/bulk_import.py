"""Mass-uploading members from an Excel template: extracting a workbook into plain
row data, and validating that data into what will be created.

Kept as two separate functions so binary parsing happens exactly once (at upload
time) while validation -- the part that must behave identically whether it's
building the preview or actually creating records -- runs against plain data both
times (see MemberImportView / MemberImportConfirmView in management/views.py).
"""

from datetime import date, datetime

import openpyxl
from django.db.models import Q
from django.utils.translation import gettext_lazy as _
from openpyxl.worksheet.datavalidation import DataValidation

from club.models import ClubMembership
from members.models import FamilyMembership, Member

from .forms import MemberForm

TEMPLATE_COLUMNS = ["first_name", "last_name", "date_of_birth", "email", "phone", "emergency_phone", "license", "status", "fee_status", "family_group", "family_role", "membership_kind"]
REQUIRED_HEADER_COLUMNS = {"first_name", "last_name"}

TEMPLATE_EXAMPLE_ROWS = [
    # A standalone member -- no family link.
    ["Alex", "Morgan", date(2012, 5, 14), "alex.morgan@example.com", "+32470123456", "+32470654321", "", ClubMembership.StatusChoices.ACTIVE, ClubMembership.FeeStatus.UNPAID, "", "", ClubMembership.Kind.MEMBER],
    # A parent and child linked together: same family_group value, one row each.
    # The child has no email of its own -- it gets a login only if given one via
    # the "Grant login" action later, same as adding a family by hand.
    #
    # membership_kind sits next to family_role because it answers the question
    # family_role raises: this parent is a `guardian`, so they hold the login and
    # can be contacted but don't count as a member and owe no fee. Put `member`
    # there instead for a parent who also plays -- the two are independent, which
    # is why it is its own column rather than inferred from family_role.
    ["Taylor", "Doe", "", "taylor.doe@example.com", "+32470654322", "", "", ClubMembership.StatusChoices.ACTIVE, "", "Doe family", FamilyMembership.FamilyRole.PARENT, ClubMembership.Kind.GUARDIAN],
    ["Jamie", "Doe", date(2014, 3, 2), "", "", "+32470654322", "", ClubMembership.StatusChoices.ACTIVE, ClubMembership.FeeStatus.UNPAID, "Doe family", FamilyMembership.FamilyRole.CHILD, ClubMembership.Kind.MEMBER],
]


def build_member_import_template():
    """The downloadable .xlsx: header row, one example row, and a dropdown on the
    status/fee_status columns so a cell can't be typo'd into an invalid value."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Members"

    sheet.append(TEMPLATE_COLUMNS)
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    sheet.freeze_panes = "A2"

    for example_row in TEMPLATE_EXAMPLE_ROWS:
        sheet.append(example_row)

    for column_name, choices in (("membership_kind", ClubMembership.Kind), ("status", ClubMembership.StatusChoices), ("fee_status", ClubMembership.FeeStatus), ("family_role", FamilyMembership.FamilyRole)):
        column_index = TEMPLATE_COLUMNS.index(column_name) + 1
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        options = ",".join(choices.values)
        validation = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}1000")

    for column_index, column_name in enumerate(TEMPLATE_COLUMNS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = max(12, len(column_name) + 2)

    return workbook


def read_member_import_workbook(file):
    """Uploaded .xlsx -> list[dict], one dict per row keyed by TEMPLATE_COLUMNS.

    Normalizes every cell to a plain str/None here: openpyxl returns typed cells
    (a date-formatted cell comes back as a datetime.date, a phone number typed as
    digits-only can come back as a number), so this is the one place that has to
    deal with that -- everything downstream, including the session, only ever
    sees plain strings.
    """
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook.active

    header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if not REQUIRED_HEADER_COLUMNS.issubset(header):
        missing = REQUIRED_HEADER_COLUMNS - set(header)
        raise ValueError(_("This doesn't look like the template — missing column(s): %(columns)s.") % {"columns": ", ".join(sorted(missing))})

    rows = []
    for excel_row in sheet.iter_rows(min_row=2):
        values = {}
        for column_name, cell in zip(header, excel_row, strict=False):
            if column_name not in TEMPLATE_COLUMNS:
                continue
            values[column_name] = _cell_to_str(cell.value)

        if not any(values.values()):
            continue  # a fully blank row (trailing spreadsheet padding) isn't a row to import
        rows.append(values)

    return rows


def _cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_member_import_rows(rows, club):
    """list[dict] (as returned by read_member_import_workbook) -> one result per
    row: {line_number, raw, member, membership_kwargs, errors}. `member` is an
    unsaved Member instance (ready to .save()) when the row is valid, else None."""
    results = []
    seen_emails = set()

    for line_number, raw in enumerate(rows, start=2):
        errors = []
        member_fields = {key: raw.get(key, "") for key in ("first_name", "last_name", "date_of_birth", "email", "phone", "emergency_phone")}
        form = MemberForm(data=member_fields)

        member = None
        if form.is_valid():
            member = form.save(commit=False)
        else:
            for field_errors in form.errors.values():
                errors.extend(field_errors)

        email = member_fields["email"].strip()
        if email:
            if email.lower() in seen_emails:
                errors.append(_("Duplicate email in this file."))
            seen_emails.add(email.lower())
            already_in_club = Member.objects.filter(member_of__club=club).filter(Q(email__iexact=email) | Q(user__email__iexact=email)).exists()
            if already_in_club:
                errors.append(_("Already a member of this club."))

        membership_kwargs, status_fee_errors = _parse_membership_fields(raw)
        errors.extend(status_fee_errors)

        family_group, family_role, family_errors = _parse_family_fields(raw)
        errors.extend(family_errors)

        # The two columns are otherwise independent -- a parent may or may not also
        # be a member -- but a child is the member the guardian is attached *to*,
        # so that particular combination is always a mistake.
        if family_role == FamilyMembership.FamilyRole.CHILD and membership_kwargs["kind"] == ClubMembership.Kind.GUARDIAN:
            errors.append(_("A child is always a member, so membership_kind cannot be 'guardian'."))

        results.append(
            {
                "line_number": line_number,
                "raw": raw,
                "member": member if not errors else None,
                "membership_kwargs": membership_kwargs,
                "family_group": family_group,
                "family_role": family_role,
                "errors": errors,
            }
        )

    return results


def _parse_membership_fields(raw):
    errors = []
    license_number = raw.get("license", "").strip()

    kind = raw.get("membership_kind", "").strip()
    if kind:
        kind_value = _match_choice(kind, ClubMembership.Kind)
        if kind_value is None:
            errors.append(_("Invalid membership_kind '%(value)s'.") % {"value": kind})
    else:
        # Blank means member: the overwhelmingly common row, and the value every
        # file written before this column existed effectively carried.
        kind_value = ClubMembership.Kind.MEMBER

    status = raw.get("status", "").strip()
    if status:
        status_value = _match_choice(status, ClubMembership.StatusChoices)
        if status_value is None:
            errors.append(_("Invalid status '%(value)s'.") % {"value": status})
    else:
        status_value = ClubMembership.StatusChoices.ACTIVE

    fee_status = raw.get("fee_status", "").strip()
    if fee_status:
        fee_status_value = _match_choice(fee_status, ClubMembership.FeeStatus)
        if fee_status_value is None:
            errors.append(_("Invalid fee status '%(value)s'.") % {"value": fee_status})
    else:
        fee_status_value = ClubMembership.FeeStatus.UNPAID

    return {"kind": kind_value, "license": license_number, "status": status_value, "fee_status": fee_status_value}, errors


def _parse_family_fields(raw):
    """family_group is a freeform label -- rows sharing the same non-blank value
    (within this file only; it isn't matched against families already in the
    club) are linked into one family. family_role decides whether the row gets a
    login: PARENT/GUARDIAN do (via their email), CHILD/OTHER don't -- same as
    registering a family by hand (see members.services.family)."""
    errors = []
    family_group = raw.get("family_group", "").strip()
    family_role_raw = raw.get("family_role", "").strip()

    if not family_group:
        # family_role=child on its own is the migration case this exists for: a
        # child the club holds with no parent on file yet. It gets a family of
        # its own so there is something for a parent to join later -- see
        # members.services.claims. Any other lone role is still a mistake.
        if family_role_raw:
            family_role = _match_choice(family_role_raw, FamilyMembership.FamilyRole)
            if family_role == FamilyMembership.FamilyRole.CHILD:
                return "", family_role, errors
            errors.append(_("family_role given without a family_group. Only 'child' is allowed on its own, for a child whose parent will register later."))
        return "", None, errors

    if not family_role_raw:
        errors.append(_("family_role is required when family_group is set."))
        return family_group, None, errors

    family_role = _match_choice(family_role_raw, FamilyMembership.FamilyRole)
    if family_role is None:
        errors.append(_("Invalid family_role '%(value)s'.") % {"value": family_role_raw})

    return family_group, family_role, errors


def _match_choice(value, choices):
    for choice_value in choices.values:
        if choice_value.lower() == value.lower():
            return choice_value
    return None
