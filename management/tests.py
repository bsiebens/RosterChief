import datetime
import json
import os
import sys
from decimal import Decimal
from io import BytesIO
from unittest import mock
from urllib.parse import parse_qs, urlparse

import openpyxl
from allauth.mfa.models import Authenticator
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.cache import cache
from django.core.files.uploadedfile import SimpleUploadedFile
from django.template.loader import render_to_string
from django.test import TestCase, override_settings
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from waffle import get_waffle_flag_model

from billing.models import Plan, PlanPrice
from billing.services.dues import record_payment, subscribe
from club.models import Club, ClubMembership, ClubRole, DuesInvoice, EvaluationManager, FeePayment, MemberRequirementStatus, OnboardingRequirement, Season, ShopManager, Sponsor
from club.services.invoicing import DuesInvoicePDFError, create_or_resend_invoice
from club.services.onboarding import mark_complete
from evaluations.models import EvaluationSettings, PlayerEvaluation
from evaluations.services import current_rubric_form
from events.models import Attendance, Competition, Event, EventReferee, EventSeries, EventTask, EventTaskClaim, Location, Opponent, RefereeSignup
from events.services.calendar import week_bounds
from events.services.notifications import notify_new_event
from events.services.rbihf_import import RBIHFImportError
from events.services.recurrence import detach_occurrence, generate_occurrences
from events.services.referees import add_external_referee
from formbuilder.models import Answer, FormSend, Submission
from formbuilder.models import Field as FormBuilderField
from formbuilder.models import Form as FormBuilderForm
from formbuilder.services.notifications import notify_form_send
from management.bulk_import import TEMPLATE_COLUMNS
from management.email_previews import EMAIL_PREVIEWS
from management.pdf import PDFExportError, _tint_with_white, referee_form_colors, render_pdf
from management.pdf_previews import PDF_PREVIEWS
from management.recurrence_ui import build_rrule, describe_rrule, parse_rrule
from management.shop_export import build_production_export, stash_production_export
from members.models import Family, FamilyMembership, Group, GroupMembership, Member, ParentClaim
from members.services.claims import children_awaiting_a_parent
from news.models import News, NewsPhoto
from news.services import _send_and_mark_notified
from notifications.models import Notification
from registration.models import RegistrationBatch, RegistrationDetails
from registration.services import EntryInput, submit_registration
from registration.services.invoicing import active_batch_entries, batch_totals
from registration.services.notifications import confirm_and_send_invoice
from shop.models import Discount, DiscountType, Invoice, Order, OrderLine, Payment, Product, ProductCategory, ProductionStatus, ProductRegistrantDiscountTier, ProductVariant, Voucher, VoucherConsumption
from shop.services.invoices import ShopInvoicePDFError, create_invoice_for_order
from teams.models import NumberPool, NumberReservation, Position, RefereeLevel, RefereeProfile, StaffAssignment, Team, TeamMembership, TeamPhoto
from teams.services import eligible_roster_members

User = get_user_model()

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def make_import_workbook(rows):
    """An in-memory .xlsx upload -- header matching the real template, plus
    whatever data rows a test wants to exercise."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(TEMPLATE_COLUMNS)
    for row in rows:
        sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile("members.xlsx", buffer.read(), content_type=XLSX_CONTENT_TYPE)


def enrol_mfa(user):
    return Authenticator.objects.create(user=user, type=Authenticator.Type.TOTP, data={"secret": "JBSWY3DPEHPK3PXP"})


def make_season(club):
    # Must genuinely cover *today*: current_season()/teams_staffed_by() key off
    # Season.covering(club, timezone.localdate()), not just any season row.
    today = timezone.localdate()
    return Season.objects.create(club=club, start_date=today - datetime.timedelta(days=30), end_date=today + datetime.timedelta(days=300))


@override_settings(
    ROSTERCHIEF_BASE_DOMAIN="rosterchief.app",
    ALLOWED_HOSTS=["rosterchief.app", "ajax-united.rosterchief.app", "rival-fc.rosterchief.app", "testserver"],
)
class ManagementTestBase(TestCase):
    # setUpTestData, not setUp: the club/season/admin fixture is read-only for
    # almost every test, so building it once per class (inside the class-wide
    # transaction Django rolls back) instead of once per test saves the whole
    # suite a lot of inserts and password hashes. Django hands each test its own
    # deepcopy of these attributes, so a test that mutates one stays isolated.
    @classmethod
    def setUpTestData(cls):
        cls.club = Club.objects.create(name="Ajax United", slug="ajax-united")
        cls.season = make_season(cls.club)

        cls.admin_user = User.objects.create_user(email="admin@example.com", password="pw-secret-123")
        cls.admin_member = Member.objects.create(user=cls.admin_user, first_name="Ada", last_name="Admin")
        ClubMembership.objects.create(club=cls.club, member=cls.admin_member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=cls.club, member=cls.admin_member).update(role=ClubRole.Roles.ADMIN)
        enrol_mfa(cls.admin_user)

    def club_get(self, name, *args, params=None):
        return self.client.get(reverse(f"management:{name}", args=args), data=params, HTTP_HOST="ajax-united.rosterchief.app")

    def club_post(self, name, data, *args):
        return self.client.post(reverse(f"management:{name}", args=args), data, HTTP_HOST="ajax-united.rosterchief.app")


class AccessTests(ManagementTestBase):
    def test_anonymous_is_sent_to_login(self):
        response = self.club_get("home")

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("account_login"), response.url)

    def test_a_club_member_with_no_role_or_staff_assignment_gets_403(self):
        plain_user = User.objects.create_user(email="plain@example.com", password="pw-secret-123")
        self.client.force_login(plain_user)

        self.assertEqual(self.club_get("home").status_code, 403)

    def test_a_plain_active_club_member_gets_403(self):
        # An active ClubMembership auto-grants a MEMBER ClubRole (club/signals.py) --
        # every signed-up player has one. That alone must not be enough to get in, or
        # players would reach the staff UI they're explicitly excluded from.
        player_user = User.objects.create_user(email="player@example.com", password="pw-secret-123")
        player_member = Member.objects.create(user=player_user, first_name="Paul", last_name="Player")
        ClubMembership.objects.create(club=self.club, member=player_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(player_user)

        self.assertEqual(self.club_get("home").status_code, 403)

    def test_a_club_admin_can_reach_it(self):
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("home").status_code, 200)

    def test_it_does_not_exist_on_the_base_domain(self):
        # The management UI manages one club; the mirror image of controlpanel
        # refusing to exist on a club subdomain.
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("management:home"), HTTP_HOST="rosterchief.app")

        self.assertEqual(response.status_code, 404)

    def test_staff_with_only_a_staff_assignment_can_reach_staff_pages(self):
        # No ClubRole at all -- authority comes purely from a current-season
        # StaffAssignment, per club.services.access.
        coach_user = User.objects.create_user(email="coach@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U12", short_name="U12")
        position = Position.objects.create(club=self.club, name="Coach", short_name="C", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        self.assertEqual(self.club_get("member_list").status_code, 200)

    def test_staff_without_admin_role_cannot_reach_admin_only_pages(self):
        coach_user = User.objects.create_user(email="coach2@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cody", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U13", short_name="U13")
        position = Position.objects.create(club=self.club, name="Coach2", short_name="C2", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        # position_list itself is open to any staff (see TeamAndPositionAccessTests)
        # -- creating and editing positions stays admin-only.
        self.assertEqual(self.club_get("position_create").status_code, 403)
        self.assertEqual(self.club_post("member_create", {"first_name": "X", "last_name": "Y"}).status_code, 403)


class NavLinkTests(ManagementTestBase):
    """The global navbar's "Management" link, next to Django admin -- see
    management.context_processors.management_link and templates/_base.html."""

    def test_a_club_admin_sees_the_link_on_the_club_subdomain(self):
        self.client.force_login(self.admin_user)

        self.assertContains(self.club_get("home"), reverse("management:home"))

    def test_a_plain_active_club_member_does_not_see_the_link(self):
        player_user = User.objects.create_user(email="player2@example.com", password="pw-secret-123")
        player_member = Member.objects.create(user=player_user, first_name="Pia", last_name="Player")
        ClubMembership.objects.create(club=self.club, member=player_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.login(email="player2@example.com", password="pw-secret-123")

        # 403 on /manage/ itself, but the link must not appear on pages this user *can*
        # reach either -- assert against a page outside the gate: the account view.
        response = self.client.get(reverse("mfa_index"), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertNotContains(response, reverse("management:home"))

    def test_the_link_is_absent_on_the_base_domain_even_for_a_club_admin(self):
        # has_management_access requires a resolved club; the control panel/base domain
        # has none, so the link -- which points at a single club's management app --
        # correctly never appears there regardless of who's signed in.
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("mfa_index"), HTTP_HOST="rosterchief.app")

        self.assertNotContains(response, reverse("management:home"))


class ActiveNavHighlightTests(ManagementTestBase):
    """The sidebar/mobile nav highlights whichever section the current page
    belongs to -- see management.context_processors.active_nav_section."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_the_members_list_page_highlights_members(self):
        response = self.club_get("member_list")

        self.assertContains(response, f'class="nav-item active" href="{reverse("management:member_list")}"')
        self.assertNotContains(response, f'class="nav-item active" href="{reverse("management:home")}"')

    def test_a_member_detail_sub_page_still_highlights_members(self):
        # member_detail has no nav entry of its own -- it belongs to the Members
        # section, same as member_list, member_update, family_detail, etc.
        member = Member.objects.create(first_name="Sub", last_name="Page")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("member_detail", member.pk)

        self.assertContains(response, f'class="nav-item active" href="{reverse("management:member_list")}"')

    def test_the_dashboard_highlights_dashboard_only(self):
        response = self.club_get("home")

        self.assertContains(response, f'class="nav-item active" href="{reverse("management:home")}"')
        self.assertNotContains(response, f'class="nav-item active" href="{reverse("management:member_list")}"')

    def test_roles_positions_and_referee_levels_pages_all_highlight_settings(self):
        # Roles/Positions/Referee levels moved out of Members/Teams and into Settings --
        # club-wide setup, not day-to-day people/roster work.
        for name in ("role_list", "position_list", "referee_level_list"):
            with self.subTest(name=name):
                response = self.club_get(name)

                self.assertContains(response, f'class="nav-item active" href="{reverse("management:club_settings")}"')

    def test_referee_management_page_highlights_calendar(self):
        # Referee management is the operational "who's covering this game" page, so it
        # lives under Calendar with Events/Locations/Opponents, not under Settings.
        response = self.club_get("referee_management")

        self.assertContains(response, f'class="nav-item active" href="{reverse("management:event_list")}"')

    def test_sponsor_list_page_highlights_finance(self):
        # Sponsors sit under Finance, underneath Dues & billing.
        response = self.club_get("sponsor_list")

        self.assertContains(response, f'class="nav-item active" href="{reverse("management:registration_invoice_queue")}"')

    def test_registration_invoice_detail_page_highlights_invoices(self):
        # Reached from Invoices, not the Registrations queue -- "back" from
        # here naturally goes there.
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="nav-reg@example.com", invoice_sent_at=timezone.now())

        response = self.club_get("registration_invoice_detail", batch.pk)

        self.assertContains(response, f'class="nav-item active" href="{reverse("management:registration_invoice_queue")}"')
        self.assertContains(response, f'class="nav-subitem active" href="{reverse("management:invoice_list")}"')

    def test_adding_or_editing_a_product_still_highlights_finance(self):
        # product_create/product_update had no entry in _NAV_SECTIONS at all,
        # so the whole sidebar (and the Products sub-item) went dark the
        # moment you left product_list -- confusing on a page reachable only
        # via "New product"/"Edit" from that same list.
        cache.clear()
        self.addCleanup(cache.clear)
        get_waffle_flag_model().objects.create(name="shop").clubs.add(self.club)
        product = Product.objects.create(club=self.club, name="Nav Check Jersey", price=Decimal("10.00"))
        for name, args in (("product_create", ()), ("product_update", (product.pk,))):
            with self.subTest(name=name):
                response = self.club_get(name, *args)

                self.assertContains(response, f'class="nav-item active" href="{reverse("management:registration_invoice_queue")}"')
                self.assertContains(response, f'class="nav-subitem active" href="{reverse("management:product_list")}"')

    def test_adding_or_editing_a_discount_still_highlights_finance(self):
        cache.clear()
        self.addCleanup(cache.clear)
        get_waffle_flag_model().objects.create(name="shop").clubs.add(self.club)
        discount = Discount.objects.create(club=self.club, name="Nav Check Discount", code="NAVCHECK")
        for name, args in (("discount_create", ()), ("discount_update", (discount.pk,))):
            with self.subTest(name=name):
                response = self.club_get(name, *args)

                self.assertContains(response, f'class="nav-item active" href="{reverse("management:registration_invoice_queue")}"')
                self.assertContains(response, f'class="nav-subitem active" href="{reverse("management:discount_list")}"')


class SidebarThemingTests(ManagementTestBase):
    """The sidebar's own background/foreground follow Club.primary_color (base.html
    emits --tenant-sidebar-bg/fg; assets/management.css's --color-sidebar-* tokens
    fall back to the fixed dark-ink look when unset). Both primary_content_color and
    secondary_content_color are the WCAG-luminance black-or-white choice already
    computed on Club itself (club.tests.ClubBrandingModelTests covers that math) --
    this just checks the value actually reaches the page."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_a_set_primary_colour_feeds_the_sidebar_background(self):
        self.club.primary_color = "#0f766e"
        self.club.save(update_fields=["primary_color"])

        response = self.club_get("home")

        self.assertContains(response, "--tenant-sidebar-bg: #0f766e;")

    def test_no_primary_colour_means_no_sidebar_override(self):
        response = self.club_get("home")

        self.assertNotContains(response, "--tenant-sidebar-bg")

    def test_a_pale_primary_colour_switches_the_sidebar_text_to_black(self):
        self.club.primary_color = "#fef08a"
        self.club.save(update_fields=["primary_color"])

        response = self.club_get("home")

        self.assertContains(response, "--tenant-sidebar-fg: #000000;")

    def test_a_dark_primary_colour_keeps_the_sidebar_text_white(self):
        self.club.primary_color = "#1e40af"
        self.club.save(update_fields=["primary_color"])

        response = self.club_get("home")

        self.assertContains(response, "--tenant-sidebar-fg: #ffffff;")

    def test_a_set_secondary_colour_also_feeds_the_active_nav_items_text_colour(self):
        self.club.secondary_color = "#fef08a"
        self.club.save(update_fields=["secondary_color"])

        response = self.club_get("home")

        self.assertContains(response, "--tenant-club-content: #000000;")

    def test_no_secondary_colour_means_no_active_nav_text_override(self):
        response = self.club_get("home")

        self.assertNotContains(response, "--tenant-club-content")


class MemberManagementTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_member_list_is_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_season = make_season(other_club)
        other_member = Member.objects.create(first_name="Other", last_name="Person")
        ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("member_list")

        self.assertNotContains(response, "Other Person")

    def test_creating_a_member_also_signs_them_up_for_the_current_season(self):
        response = self.club_post("member_create", {"first_name": "New", "last_name": "Player", "email": "new@example.com"})

        member = Member.objects.get(first_name="New", last_name="Player")
        self.assertRedirects(response, reverse("management:member_detail", args=[member.pk]))
        self.assertTrue(ClubMembership.objects.filter(club=self.club, member=member, season=self.season).exists())

    def test_updating_a_member(self):
        member = Member.objects.create(first_name="Old", last_name="Name")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        # This member has a current-season membership, so its section renders too --
        # one combined submit, so its (required) fields must come along.
        self.club_post(
            "member_update",
            {"first_name": "New", "last_name": "Name", "kind": ClubMembership.Kind.MEMBER, "status": ClubMembership.StatusChoices.ACTIVE, "fee_status": ClubMembership.FeeStatus.UNPAID},
            member.pk,
        )

        member.refresh_from_db()
        self.assertEqual(member.first_name, "New")


class MemberAttendanceSparklineTests(ManagementTestBase):
    """The D7-alike attendance card on member_detail.html --
    events.services.attendance.member_attendance_sparkline/_counts."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.position = Position.objects.create(club=cls.club, name="Forward", short_name="FW")
        cls.player = Member.objects.create(first_name="Peter", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        TeamMembership.objects.create(team=cls.team, member=cls.player, season=cls.season, position=cls.position)

    def setUp(self):
        self.client.force_login(self.admin_user)

    def make_attendance(self, *, status, start, showed_up=None):
        event = Event.objects.create(club=self.club, title="Training", kind=Event.EventKind.TRAINING, start=start, season=self.season)
        # A future event's roster is auto-synced on event.teams.add() (see
        # events/signals.py), which already creates this row -- a past event's
        # isn't (history is never rewritten), so get_or_create covers both.
        event.teams.add(self.team)
        attendance, created = Attendance.objects.get_or_create(event=event, member=self.player, defaults={"status": status, "showed_up": showed_up})
        if not created:
            attendance.status, attendance.showed_up = status, showed_up
            attendance.save(update_fields=["status", "showed_up"])
        return attendance

    def test_the_card_is_hidden_for_a_member_not_rostered_this_season(self):
        unrostered = Member.objects.create(first_name="Not", last_name="Rostered")
        ClubMembership.objects.create(club=self.club, member=unrostered, season=self.season, status=ClubMembership.StatusChoices.PENDING)

        response = self.club_get("member_detail", unrostered.pk)

        self.assertNotContains(response, "attendance-sparkline")

    def test_the_card_is_hidden_for_a_guardian(self):
        guardian = Member.objects.create(first_name="Gale", last_name="Guardian")
        ClubMembership.objects.create(club=self.club, member=guardian, season=self.season, kind=ClubMembership.Kind.GUARDIAN, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("member_detail", guardian.pk)

        self.assertNotContains(response, "attendance-sparkline")

    def test_a_past_present_event_counts_as_present(self):
        self.make_attendance(status=Attendance.AttendanceStatus.PRESENT, start=timezone.now() - datetime.timedelta(days=2))

        response = self.club_get("member_detail", self.player.pk)

        self.assertEqual(response.context["attendance_counts"]["present"], 1)
        self.assertEqual(response.context["attendance_sparkline"][0]["state"], "present")

    def test_a_past_absent_event_counts_as_absent(self):
        self.make_attendance(status=Attendance.AttendanceStatus.ABSENT, start=timezone.now() - datetime.timedelta(days=2))

        response = self.club_get("member_detail", self.player.pk)

        self.assertEqual(response.context["attendance_counts"]["absent"], 1)
        self.assertEqual(response.context["attendance_sparkline"][0]["state"], "absent")

    def test_a_future_event_is_upcoming_regardless_of_rsvp_status(self):
        self.make_attendance(status=Attendance.AttendanceStatus.SELECTED, start=timezone.now() + datetime.timedelta(days=2))

        response = self.club_get("member_detail", self.player.pk)

        self.assertEqual(response.context["attendance_sparkline"][0]["state"], "upcoming")
        # Only past events feed the Present/Absent/No-reply totals.
        self.assertEqual(response.context["attendance_counts"]["present"], 0)
        self.assertEqual(response.context["attendance_counts"]["absent"], 0)
        self.assertEqual(response.context["attendance_counts"]["no_reply"], 0)

    def test_a_past_unanswered_event_counts_as_no_reply(self):
        self.make_attendance(status=Attendance.AttendanceStatus.NO_RESPONSE, start=timezone.now() - datetime.timedelta(days=2))

        response = self.club_get("member_detail", self.player.pk)

        self.assertEqual(response.context["attendance_counts"]["no_reply"], 1)

    def test_a_present_rsvp_checked_in_as_absent_counts_as_a_no_show(self):
        self.make_attendance(status=Attendance.AttendanceStatus.PRESENT, start=timezone.now() - datetime.timedelta(days=2), showed_up=False)

        response = self.club_get("member_detail", self.player.pk)

        self.assertEqual(response.context["attendance_counts"]["no_shows"], 1)
        self.assertContains(response, "No-shows")

    def test_an_unchecked_present_rsvp_is_not_a_no_show(self):
        self.make_attendance(status=Attendance.AttendanceStatus.PRESENT, start=timezone.now() - datetime.timedelta(days=2))

        response = self.club_get("member_detail", self.player.pk)

        self.assertEqual(response.context["attendance_counts"]["no_shows"], 0)

    def test_the_sparkline_is_capped_at_twelve_bars_oldest_first(self):
        for day in range(15, 0, -1):
            self.make_attendance(status=Attendance.AttendanceStatus.PRESENT, start=timezone.now() - datetime.timedelta(days=day))

        response = self.club_get("member_detail", self.player.pk)

        sparkline = response.context["attendance_sparkline"]
        self.assertEqual(len(sparkline), 12)
        starts = [bar["event"].start for bar in sparkline]
        self.assertEqual(starts, sorted(starts))


class TeamManagementTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_team_list_is_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        Team.objects.create(club=other_club, name="Rival Team", short_name="RT")

        response = self.club_get("team_list")

        self.assertNotContains(response, "Rival Team")

    def test_creating_a_team(self):
        response = self.club_post("team_create", {"name": "U15", "short_name": "U15", "referee_management": "club"})

        team = Team.objects.get(club=self.club, name="U15")
        self.assertRedirects(response, reverse("management:team_detail", args=[team.pk]))

    def test_creating_a_federation_managed_team(self):
        self.club_post("team_create", {"name": "U15", "short_name": "U15", "referee_management": "federation"})

        team = Team.objects.get(club=self.club, name="U15")
        self.assertEqual(team.referee_management, Team.RefereeManagement.FEDERATION)

    def test_creating_a_team_with_a_pool(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)

        response = self.club_post("team_create", {"name": "U15", "short_name": "U15", "referee_management": "club", "pool": str(pool.pk)})

        team = Team.objects.get(club=self.club, name="U15")
        self.assertRedirects(response, reverse("management:team_detail", args=[team.pk]))
        self.assertEqual(team.pool, pool)

    def test_the_pool_dropdown_only_offers_this_clubs_pools(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        NumberPool.objects.create(club=other_club, name="Rival Pool", min_number=1, max_number=99)

        response = self.club_get("team_create")

        self.assertNotContains(response, "Rival Pool")

    def test_deleting_a_team(self):
        team = Team.objects.create(club=self.club, name="U16", short_name="U16")

        response = self.club_post("team_delete", {}, team.pk)

        self.assertRedirects(response, reverse("management:team_list"))
        self.assertFalse(Team.objects.filter(pk=team.pk).exists())

    def test_deleting_a_team_cascades_its_roster_and_staff(self):
        team = Team.objects.create(club=self.club, name="U17", short_name="U17")
        position = Position.objects.create(club=self.club, name="Coach17", short_name="C17", staff_position=True)
        member = Member.objects.create(first_name="Sam", last_name="Staffer")
        StaffAssignment.objects.create(team=team, member=member, season=self.season, position=position)

        self.club_post("team_delete", {}, team.pk)

        self.assertFalse(StaffAssignment.objects.filter(team=team).exists())

    def test_a_non_admin_cannot_delete_a_team(self):
        team = Team.objects.create(club=self.club, name="U18", short_name="U18")
        coach_user = User.objects.create_user(email="coach-team-delete@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="CoachTeamDelete", short_name="CTD", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("team_delete", {}, team.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Team.objects.filter(pk=team.pk).exists())


class TeamRosterStaffTests(ManagementTestBase):
    """Roster/staff management folded into the team page -- see
    management.views.TeamDetailView and the TeamRoster*/TeamStaff* views."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.player_position = Position.objects.create(club=cls.club, name="Forward", short_name="FW", staff_position=False)
        cls.coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)

        cls.player = Member.objects.create(first_name="Peter", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

        # A coach for each team: "this team's coach may, that team's coach may not"
        # is the whole point of this class, so both actors are standing fixtures.
        cls.team_coach = cls.make_team_coach(cls.team, "coach-roster@example.com")
        cls.other_team_coach = cls.make_team_coach(cls.other_team, "coach-roster-other@example.com")

    @classmethod
    def make_team_coach(cls, team, email):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=team, member=coach_member, season=cls.season, position=cls.coach_position)
        return coach_user

    def test_nav_no_longer_lists_roster_or_staff(self):
        # Not a bare "Roster"/"Staff" substring check -- "RosterChief" branding is
        # on every page regardless. The nav's old icons are a safe, specific proxy.
        self.client.force_login(self.admin_user)

        response = self.club_get("team_list")

        self.assertNotContains(response, "clipboard-list")
        self.assertNotContains(response, "hard-hat")

    def test_roster_and_staff_urls_no_longer_resolve(self):
        with self.assertRaises(NoReverseMatch):
            reverse("management:roster_list")
        with self.assertRaises(NoReverseMatch):
            reverse("management:staff_list")

    def test_season_switcher_defaults_to_the_current_season(self):
        self.client.force_login(self.admin_user)
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Peter Player")

    def test_a_pending_members_roster_row_shows_a_pending_badge(self):
        self.client.force_login(self.admin_user)
        pending_member = Member.objects.create(first_name="Penny", last_name="Pending")
        ClubMembership.objects.create(club=self.club, member=pending_member, season=self.season, status=ClubMembership.StatusChoices.PENDING)
        TeamMembership.objects.create(team=self.team, season=self.season, member=pending_member)
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)

        response = self.club_get("team_detail", self.team.pk)

        rows = {membership.member_id: membership.club_status_pending for membership in response.context["roster"]}
        self.assertTrue(rows[pending_member.pk])
        self.assertFalse(rows[self.player.pk])
        self.assertContains(response, "Pending")

    def test_season_switcher_honours_the_query_param(self):
        other_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 1, 1), end_date=datetime.date(2020, 12, 31))
        TeamMembership.objects.create(team=self.team, season=other_season, member=self.player, position=self.player_position)
        self.client.force_login(self.admin_user)

        default_response = self.club_get("team_detail", self.team.pk)
        other_response = self.client.get(f"{reverse('management:team_detail', args=[self.team.pk])}?season={other_season.pk}", HTTP_HOST="ajax-united.rosterchief.app")

        # Not assertNotContains("Peter Player") on the default response -- he's
        # still a valid pick in the "Add player" combobox even when he isn't on
        # *this* season's roster, so his name legitimately appears there too.
        self.assertContains(default_response, "No one on the roster for this season yet.")
        self.assertContains(other_response, "Peter Player")

    def test_a_teams_own_coach_can_add_a_player(self):
        self.client.force_login(self.team_coach)

        response = self.club_post("team_roster_add", {"member": str(self.player.pk), "position": str(self.player_position.pk), "jersey_number": "9"}, self.team.pk, self.season.pk)

        self.assertRedirects(response, f"{reverse('management:team_detail', args=[self.team.pk])}?season={self.season.pk}")
        membership = TeamMembership.objects.get(team=self.team, season=self.season, member=self.player)
        self.assertEqual(membership.jersey_number, 9)

    def test_a_different_teams_coach_cannot_add_a_player(self):
        self.client.force_login(self.other_team_coach)

        response = self.club_post("team_roster_add", {"member": str(self.player.pk), "position": str(self.player_position.pk)}, self.team.pk, self.season.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season).exists())

    def test_a_different_teams_coach_can_still_view_the_team(self):
        self.client.force_login(self.other_team_coach)

        response = self.club_get("team_detail", self.team.pk)

        self.assertEqual(response.status_code, 200)

    def test_admin_can_add_a_player_to_any_team(self):
        self.client.force_login(self.admin_user)

        self.club_post("team_roster_add", {"member": str(self.player.pk), "position": str(self.player_position.pk)}, self.team.pk, self.season.pk)

        self.assertTrue(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.player).exists())

    def test_the_add_player_dropdown_excludes_a_member_with_no_active_membership(self):
        lapsed_player = Member.objects.create(first_name="Lex", last_name="Lapsed")
        ClubMembership.objects.create(club=self.club, member=lapsed_player, season=self.season, status=ClubMembership.StatusChoices.LAPSED)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertNotContains(response, "Lex Lapsed")

    def test_the_add_player_dropdown_includes_a_member_active_only_next_season(self):
        next_season = Season.objects.create(club=self.club, start_date=self.season.end_date + datetime.timedelta(days=1), end_date=self.season.end_date + datetime.timedelta(days=300))
        upcoming_player = Member.objects.create(first_name="Uma", last_name="Upcoming")
        ClubMembership.objects.create(club=self.club, member=upcoming_player, season=next_season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Uma Upcoming")

    def test_the_add_player_dropdown_excludes_a_member_active_only_in_a_past_season(self):
        past_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 1, 1), end_date=datetime.date(2020, 12, 31))
        past_player = Member.objects.create(first_name="Pip", last_name="Past")
        ClubMembership.objects.create(club=self.club, member=past_player, season=past_season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertNotContains(response, "Pip Past")

    def test_adding_the_same_member_twice_fails_with_a_form_error_not_a_500(self):
        self.client.force_login(self.admin_user)
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)

        response = self.club_post("team_roster_add", {"member": str(self.player.pk), "position": str(self.player_position.pk)}, self.team.pk, self.season.pk)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 1)

    def test_a_duplicate_jersey_number_fails_with_a_form_error_not_a_500(self):
        # team/season aren't TeamMembershipForm fields, so Django's own
        # validate_unique() can't see unique_jersey_number_per_team_per_season --
        # this constraint only gets checked because the form does it by hand.
        other_player = Member.objects.create(first_name="Olly", last_name="Other")
        ClubMembership.objects.create(club=self.club, member=other_player, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.club_post("team_roster_add", {"member": str(other_player.pk), "position": str(self.player_position.pk), "jersey_number": "7"}, self.team.pk, self.season.pk)

        self.assertEqual(response.status_code, 302)
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=other_player).exists())

    def test_editing_a_roster_entry_to_a_clashing_jersey_number_fails_gracefully(self):
        other_player = Member.objects.create(first_name="Olly", last_name="Other")
        ClubMembership.objects.create(club=self.club, member=other_player, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        other_membership = TeamMembership.objects.create(team=self.team, season=self.season, member=other_player, position=self.player_position, jersey_number=8)
        self.client.force_login(self.admin_user)

        response = self.club_post(
            "team_roster_update",
            {"member": str(other_player.pk), "position": str(self.player_position.pk), "jersey_number": "7"},
            self.team.pk,
            other_membership.pk,
        )

        self.assertEqual(response.status_code, 302)
        other_membership.refresh_from_db()
        self.assertEqual(other_membership.jersey_number, 8)

    def test_a_jersey_number_taken_on_a_teammate_team_in_the_same_pool_is_rejected(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        self.team.pool = pool
        self.team.save()
        self.other_team.pool = pool
        self.other_team.save()
        other_player = Member.objects.create(first_name="Olly", last_name="Other")
        ClubMembership.objects.create(club=self.club, member=other_player, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        TeamMembership.objects.create(team=self.other_team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.club_post("team_roster_add", {"member": str(other_player.pk), "position": str(self.player_position.pk), "jersey_number": "7"}, self.team.pk, self.season.pk)

        self.assertEqual(response.status_code, 302)
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=other_player).exists())

    def test_the_roster_edit_form_is_prefilled_from_a_pending_requested_number(self):
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com")
        membership = ClubMembership.objects.get(club=self.club, member=self.player, season=self.season)
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=self.team, requested_jersey_number=9)
        roster_entry = TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertEqual(response.context["roster"][0].pk, roster_entry.pk)
        self.assertEqual(response.context["roster"][0].edit_form.initial.get("jersey_number"), 9)

    def test_the_roster_edit_form_does_not_override_an_already_set_number(self):
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com")
        membership = ClubMembership.objects.get(club=self.club, member=self.player, season=self.season)
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=self.team, requested_jersey_number=9)
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=3)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        # initial=None here (jersey_number was already set), so this is
        # ModelForm's own model_to_dict(instance) -- the already-saved 3,
        # never overwritten by the requested 9.
        self.assertEqual(response.context["roster"][0].edit_form.initial.get("jersey_number"), 3)

    def test_a_pool_scoped_team_still_lets_a_member_keep_their_own_number_on_edit(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        self.team.pool = pool
        self.team.save()
        membership = TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.club_post(
            "team_roster_update",
            {"member": str(self.player.pk), "position": str(self.player_position.pk), "jersey_number": "7", "is_captain": "on"},
            self.team.pk,
            membership.pk,
        )

        self.assertEqual(response.status_code, 302)
        membership.refresh_from_db()
        self.assertTrue(membership.is_captain)

    def test_editing_a_roster_entry_updates_it(self):
        membership = TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=9)
        self.client.force_login(self.admin_user)

        self.club_post(
            "team_roster_update",
            {"member": str(self.player.pk), "position": str(self.player_position.pk), "jersey_number": "10", "is_captain": "on"},
            self.team.pk,
            membership.pk,
        )

        membership.refresh_from_db()
        self.assertEqual(membership.jersey_number, 10)
        self.assertTrue(membership.is_captain)

    def test_removing_a_roster_entry_deletes_it(self):
        membership = TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)
        self.client.force_login(self.admin_user)

        self.club_post("team_roster_remove", {}, self.team.pk, membership.pk)

        self.assertFalse(TeamMembership.objects.filter(pk=membership.pk).exists())

    def test_a_teams_own_coach_can_assign_staff(self):
        physio_position = Position.objects.create(club=self.club, name="Physio", short_name="PH", staff_position=True)
        physio = Member.objects.create(first_name="Pat", last_name="Physio")
        ClubMembership.objects.create(club=self.club, member=physio, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.team_coach)

        self.club_post("team_staff_add", {"member": str(physio.pk), "position": str(physio_position.pk)}, self.team.pk, self.season.pk)

        self.assertTrue(StaffAssignment.objects.filter(team=self.team, season=self.season, member=physio).exists())

    def test_the_assign_staff_dropdown_excludes_a_member_active_only_in_a_past_season(self):
        # Same eligibility rule as the "Add player" dropdown -- see
        # test_the_add_player_dropdown_excludes_a_member_active_only_in_a_past_season.
        past_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 1, 1), end_date=datetime.date(2020, 12, 31))
        past_member = Member.objects.create(first_name="Sam", last_name="Stale")
        ClubMembership.objects.create(club=self.club, member=past_member, season=past_season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertNotContains(response, "Sam Stale")

    def test_a_different_teams_coach_cannot_assign_staff(self):
        physio_position = Position.objects.create(club=self.club, name="Physio", short_name="PH", staff_position=True)
        physio = Member.objects.create(first_name="Pat", last_name="Physio")
        self.client.force_login(self.other_team_coach)

        response = self.club_post("team_staff_add", {"member": str(physio.pk), "position": str(physio_position.pk)}, self.team.pk, self.season.pk)

        self.assertEqual(response.status_code, 403)

    def test_removing_a_staff_assignment_deletes_it(self):
        assignment = StaffAssignment.objects.create(team=self.team, season=self.season, member=self.player, position=self.coach_position)
        self.client.force_login(self.admin_user)

        self.club_post("team_staff_remove", {}, self.team.pk, assignment.pk)

        self.assertFalse(StaffAssignment.objects.filter(pk=assignment.pk).exists())


class TeamBulkAddTests(ManagementTestBase):
    """Adding many people to a team's roster/staff in one submit -- see
    management.views.TeamBulkAddView. One formset row per assignment; the
    position picked decides whether the row means a roster entry or a staff
    one, and one bad row rejects the whole submit rather than half-saving."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.player_position = Position.objects.create(club=cls.club, name="Forward", short_name="FW", staff_position=False)
        cls.coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)

        cls.player = Member.objects.create(first_name="Peter", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        cls.other_player = Member.objects.create(first_name="Olly", last_name="Other")
        ClubMembership.objects.create(club=cls.club, member=cls.other_player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        # A third eligible player, so a submit can pair one perfectly valid row with
        # one bad one -- see test_a_good_row_alongside_a_bad_one_is_not_saved_either.
        cls.third_player = Member.objects.create(first_name="Tara", last_name="Third")
        ClubMembership.objects.create(club=cls.club, member=cls.third_player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

        coach_user = User.objects.create_user(email="coach-bulk@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=cls.other_team, member=coach_member, season=cls.season, position=cls.coach_position)
        cls.other_team_coach = coach_user

    def bulk_data(self, *rows):
        """A formset POST body: one dict per row, plus the management form."""
        data = {"form-TOTAL_FORMS": str(len(rows)), "form-INITIAL_FORMS": "0", "form-MIN_NUM_FORMS": "0", "form-MAX_NUM_FORMS": "1000"}
        for index, row in enumerate(rows):
            for name, value in row.items():
                data[f"form-{index}-{name}"] = value
        return data

    def row(self, member, position, jersey_number="", is_captain=False, is_alternate_captain=False):
        row = {"member": str(member.pk), "position": str(position.pk), "jersey_number": str(jersey_number)}
        # Unchecked boxes aren't submitted at all, so only add the key when set --
        # sending "off" would still read as True to a Django BooleanField.
        if is_captain:
            row["is_captain"] = "on"
        if is_alternate_captain:
            row["is_alternate_captain"] = "on"
        return row

    def bulk_add(self, *rows):
        return self.club_post("team_bulk_add", self.bulk_data(*rows), self.team.pk, self.season.pk)

    def test_the_picker_offers_eligible_members_and_marks_who_is_already_on(self):
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_bulk_add", self.team.pk, self.season.pk)

        self.assertContains(response, "Peter Player — on roster")
        self.assertContains(response, "Olly Other")

    def test_the_position_picker_marks_staff_positions_for_the_jersey_toggle(self):
        # PositionSelect stamps data-staff on staff options; bulk-add-rows.js reads
        # it to grey out the jersey input, which a StaffAssignment has no field for.
        self.client.force_login(self.admin_user)

        response = self.club_get("team_bulk_add", self.team.pk, self.season.pk)

        self.assertContains(response, "data-staff")

    def test_a_lapsed_member_is_not_offered(self):
        lapsed = Member.objects.create(first_name="Lex", last_name="Lapsed")
        ClubMembership.objects.create(club=self.club, member=lapsed, season=self.season, status=ClubMembership.StatusChoices.LAPSED)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_bulk_add", self.team.pk, self.season.pk)

        self.assertNotContains(response, "Lex Lapsed")

    def test_admin_can_add_two_players_in_one_submit(self):
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.player_position, 9), self.row(self.other_player, self.player_position))

        self.assertRedirects(response, f"{reverse('management:team_detail', args=[self.team.pk])}?season={self.season.pk}")
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 2)
        self.assertEqual(TeamMembership.objects.get(team=self.team, member=self.player).jersey_number, 9)

    def test_a_staff_position_creates_a_staff_assignment_not_a_roster_entry(self):
        # The position is what picks the role -- there is no separate control.
        self.client.force_login(self.admin_user)

        self.bulk_add(self.row(self.player, self.coach_position))

        self.assertTrue(StaffAssignment.objects.filter(team=self.team, season=self.season, member=self.player).exists())
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.player).exists())

    def test_a_member_can_be_added_as_both_player_and_staff_via_two_rows(self):
        self.client.force_login(self.admin_user)

        self.bulk_add(self.row(self.player, self.player_position), self.row(self.player, self.coach_position))

        self.assertTrue(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.player).exists())
        self.assertTrue(StaffAssignment.objects.filter(team=self.team, season=self.season, member=self.player).exists())

    def test_a_jersey_number_on_a_staff_row_is_rejected(self):
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.coach_position, 9))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(StaffAssignment.objects.filter(team=self.team, season=self.season).exists())

    def test_a_jersey_clashing_with_an_existing_player_rejects_the_submit(self):
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.other_player, self.player_position, 7))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already taken")
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.other_player).exists())
        # All-or-nothing: nothing was saved, so the roster is untouched apart from
        # the entry that was already there before the submit.
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 1)

    def test_a_jersey_clashing_across_teams_in_the_same_pool_rejects_the_submit(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        self.team.pool = pool
        self.team.save()
        self.other_team.pool = pool
        self.other_team.save()
        TeamMembership.objects.create(team=self.other_team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.other_player, self.player_position, 7))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.other_player).exists())

    def test_a_good_row_alongside_a_bad_one_is_not_saved_either(self):
        # The all-or-nothing guarantee proper: the other tests here submit rows that
        # are either all bad or bad in pairs, so none of them would notice a valid
        # row slipping through on its own. If it did, the re-rendered form would
        # re-add it on the next submit.
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.other_player, self.player_position, 7), self.row(self.third_player, self.player_position, 12))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.third_player).exists())
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 1)

    def test_captaincy_is_saved(self):
        self.client.force_login(self.admin_user)

        self.bulk_add(
            self.row(self.player, self.player_position, 9, is_captain=True),
            self.row(self.other_player, self.player_position, 10, is_alternate_captain=True),
        )

        self.assertTrue(TeamMembership.objects.get(team=self.team, member=self.player).is_captain)
        self.assertTrue(TeamMembership.objects.get(team=self.team, member=self.other_player).is_alternate_captain)

    def test_a_row_left_alone_is_neither_captain_nor_alternate(self):
        self.client.force_login(self.admin_user)

        self.bulk_add(self.row(self.player, self.player_position))

        membership = TeamMembership.objects.get(team=self.team, member=self.player)
        self.assertFalse(membership.is_captain)
        self.assertFalse(membership.is_alternate_captain)

    def test_captain_and_alternate_on_one_row_is_rejected(self):
        # Contradictory rather than a club policy -- how many captains a team may
        # have is deliberately left unconstrained, see TeamBulkAddRowForm.clean.
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.player_position, is_captain=True, is_alternate_captain=True))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "not both")
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season).exists())

    def test_captaincy_on_a_staff_row_is_rejected(self):
        # Captaincy lives on TeamMembership; a staff row becomes a StaffAssignment,
        # which has no such field to put it in.
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.coach_position, is_captain=True))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "doesn&#x27;t apply to a staff position")
        self.assertFalse(StaffAssignment.objects.filter(team=self.team, season=self.season, member=self.player).exists())

    def test_two_captains_in_one_submit_are_allowed(self):
        # Pins the deliberate absence of a rule: neither the model nor the
        # single-add form limits a team to one captain, so bulk-add doesn't invent
        # that limit either -- it would be bypassable by adding players one at a time.
        self.client.force_login(self.admin_user)

        self.bulk_add(
            self.row(self.player, self.player_position, 9, is_captain=True),
            self.row(self.other_player, self.player_position, 10, is_captain=True),
        )

        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season, is_captain=True).count(), 2)

    def test_two_rows_claiming_the_same_jersey_are_rejected(self):
        # Neither row clashes with anything already saved -- only with each other,
        # which no single row can see. See BaseTeamBulkAddFormSet.clean.
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.player_position, 7), self.row(self.other_player, self.player_position, 7))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "more than one row")
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 0)

    def test_the_same_member_twice_in_the_same_role_is_rejected(self):
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.player_position), self.row(self.player, self.player_position))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "listed twice")
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 0)

    def test_an_invalid_submit_re_renders_the_rows_that_were_typed(self):
        # The whole point of all-or-nothing: nothing the user typed is lost.
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position, jersey_number=7)
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.other_player, self.player_position, 7))

        self.assertContains(response, f'value="{self.other_player.pk}" selected')

    def test_a_member_already_on_the_roster_is_rejected(self):
        TeamMembership.objects.create(team=self.team, season=self.season, member=self.player, position=self.player_position)
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(self.player, self.player_position))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already on this team&#x27;s roster")
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season, member=self.player).count(), 1)

    def test_a_lapsed_member_cannot_be_added_via_a_crafted_post(self):
        # Eligibility is recomputed from eligible_roster_members server-side, so a
        # member id that was never offered fails the field's own queryset lookup.
        lapsed = Member.objects.create(first_name="Lex", last_name="Lapsed")
        ClubMembership.objects.create(club=self.club, member=lapsed, season=self.season, status=ClubMembership.StatusChoices.LAPSED)
        self.client.force_login(self.admin_user)

        response = self.bulk_add(self.row(lapsed, self.player_position))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(TeamMembership.objects.filter(team=self.team, season=self.season, member=lapsed).exists())

    def test_submitting_nothing_says_so_instead_of_adding(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("team_bulk_add", self.bulk_data({}, {}), self.team.pk, self.season.pk)

        self.assertRedirects(response, f"{reverse('management:team_detail', args=[self.team.pk])}?season={self.season.pk}")
        self.assertEqual(TeamMembership.objects.filter(team=self.team, season=self.season).count(), 0)

    def test_a_different_teams_coach_cannot_bulk_add(self):
        self.client.force_login(self.other_team_coach)

        response = self.bulk_add(self.row(self.player, self.player_position))

        self.assertEqual(response.status_code, 403)


class GroupManagementTests(ManagementTestBase):
    """Generic named collections of members -- see management.views.Group* and
    management.forms.GroupForm. Deliberately has no team/referee knowledge at
    all; see MemberRefereeEligibilityTests for that (teams.RefereeProfile)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        # One team only -- groups are deliberately team-agnostic; the team exists
        # solely so make_non_admin_coach has somewhere to be staffed.
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.member = Member.objects.create(first_name="Peter", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def make_non_admin_coach(self, email="coach-groups@example.com"):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=position)
        return coach_user

    def test_list_is_admin_only(self):
        self.client.force_login(self.make_non_admin_coach())
        self.assertEqual(self.club_get("group_list").status_code, 403)

    def test_the_lists_edit_button_goes_to_the_group_page_not_straight_to_the_form(self):
        # Same as Teams: Edit lands on the overview, which is where the members
        # live, and offers its own Edit for the rename form. The list pages that
        # do jump straight to a form (Locations, Opponents, Sponsors, Positions,
        # Referee levels) have no detail page to land on at all.
        group = Group.objects.create(club=self.club, name="Referees")
        self.client.force_login(self.admin_user)

        response = self.club_get("group_list")

        self.assertContains(response, reverse("management:group_detail", args=[group.pk]))
        self.assertNotContains(response, reverse("management:group_update", args=[group.pk]))

    def test_admin_can_create_a_group(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("group_create", {"name": "Referees"})

        group = Group.objects.get(club=self.club, name="Referees")
        self.assertRedirects(response, reverse("management:group_detail", args=[group.pk]))

    def test_editing_a_group_renames_it(self):
        self.client.force_login(self.admin_user)
        group = Group.objects.create(club=self.club, name="Old name")

        self.club_post("group_update", {"name": "New name"}, group.pk)

        group.refresh_from_db()
        self.assertEqual(group.name, "New name")

    def test_deleting_a_group_removes_it(self):
        self.client.force_login(self.admin_user)
        group = Group.objects.create(club=self.club, name="Doomed")

        self.club_post("group_delete", {}, group.pk)

        self.assertFalse(Group.objects.filter(pk=group.pk).exists())

    def bulk_data(self, *member_pks):
        """A formset POST body: one row per member, plus the management form."""
        data = {"form-TOTAL_FORMS": str(len(member_pks)), "form-INITIAL_FORMS": "0", "form-MIN_NUM_FORMS": "0", "form-MAX_NUM_FORMS": "1000"}
        for index, member_pk in enumerate(member_pks):
            data[f"form-{index}-member"] = str(member_pk)
        return data

    def test_bulk_add_offers_members_and_marks_the_ones_already_in(self):
        group = Group.objects.create(club=self.club, name="Referees")
        GroupMembership.objects.create(group=group, member=self.member)
        other_member = Member.objects.create(first_name="Olly", last_name="Other")
        ClubMembership.objects.create(club=self.club, member=other_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_get("group_bulk_add", group.pk)

        self.assertContains(response, "already in this group")
        self.assertContains(response, "Olly Other")

    def test_bulk_add_adds_the_members_in_the_rows(self):
        group = Group.objects.create(club=self.club, name="Referees")
        other_member = Member.objects.create(first_name="Olly", last_name="Other")
        ClubMembership.objects.create(club=self.club, member=other_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_post("group_bulk_add", self.bulk_data(self.member.pk, other_member.pk), group.pk)

        self.assertRedirects(response, reverse("management:group_detail", args=[group.pk]))
        self.assertEqual(GroupMembership.objects.filter(group=group).count(), 2)

    def test_bulk_add_cannot_re_add_an_existing_member_via_a_crafted_post(self):
        group = Group.objects.create(club=self.club, name="Referees")
        GroupMembership.objects.create(group=group, member=self.member)
        self.client.force_login(self.admin_user)

        response = self.club_post("group_bulk_add", self.bulk_data(self.member.pk), group.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already in this group")
        self.assertEqual(GroupMembership.objects.filter(group=group, member=self.member).count(), 1)

    def test_bulk_add_rejects_the_same_member_listed_twice(self):
        group = Group.objects.create(club=self.club, name="Referees")
        self.client.force_login(self.admin_user)

        response = self.club_post("group_bulk_add", self.bulk_data(self.member.pk, self.member.pk), group.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "listed twice")
        self.assertEqual(GroupMembership.objects.filter(group=group).count(), 0)

    def test_bulk_add_with_no_rows_filled_in_adds_nothing(self):
        group = Group.objects.create(club=self.club, name="Referees")
        self.client.force_login(self.admin_user)

        response = self.club_post("group_bulk_add", self.bulk_data("", ""), group.pk)

        self.assertRedirects(response, reverse("management:group_detail", args=[group.pk]))
        self.assertEqual(GroupMembership.objects.filter(group=group).count(), 0)

    def test_removing_a_member_deletes_the_membership(self):
        group = Group.objects.create(club=self.club, name="Referees")
        membership = GroupMembership.objects.create(group=group, member=self.member)
        self.client.force_login(self.admin_user)

        self.club_post("group_member_remove", {}, group.pk, membership.pk)

        self.assertFalse(GroupMembership.objects.filter(pk=membership.pk).exists())

    def test_groups_are_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_group = Group.objects.create(club=other_club, name="Rival Referees")
        self.client.force_login(self.admin_user)

        response = self.club_get("group_list")

        self.assertNotContains(response, "Rival Referees")
        self.assertEqual(self.club_get("group_detail", other_group.pk).status_code, 404)


ONE_PIXEL_PNG = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"


def make_image_file(name="photo.png"):
    return SimpleUploadedFile(name, ONE_PIXEL_PNG, content_type="image/png")


class TeamPhotoTests(ManagementTestBase):
    """One photo per (team, season), uploaded from the team page -- see
    management.views.TeamPhotoSetView/TeamPhotoDeleteView."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)

        cls.team_coach = cls.make_team_coach(cls.team, "coach-photo@example.com")
        cls.other_team_coach = cls.make_team_coach(cls.other_team, "coach-photo-other@example.com")

        # A staff member on this team with a *non*-management position: allowed
        # into the management app, but never offered the upload controls.
        staff_user = User.objects.create_user(email="physio-photo@example.com", password="pw-secret-123")
        staff_member = Member.objects.create(user=staff_user, first_name="Pat", last_name="Physio")
        position = Position.objects.create(club=cls.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=cls.team, member=staff_member, season=cls.season, position=position)
        cls.plain_staff = staff_user

    @classmethod
    def make_team_coach(cls, team, email):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=team, member=coach_member, season=cls.season, position=cls.coach_position)
        return coach_user

    def test_uploading_creates_a_photo(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("team_photo_set", {"image": make_image_file()}, self.team.pk, self.season.pk)

        self.assertRedirects(response, f"{reverse('management:team_detail', args=[self.team.pk])}?season={self.season.pk}")
        self.assertEqual(TeamPhoto.objects.filter(team=self.team, season=self.season).count(), 1)

    def test_uploading_again_replaces_it_in_place(self):
        self.client.force_login(self.admin_user)
        self.club_post("team_photo_set", {"image": make_image_file("first.png")}, self.team.pk, self.season.pk)
        first = TeamPhoto.objects.get(team=self.team, season=self.season)

        self.club_post("team_photo_set", {"image": make_image_file("second.png")}, self.team.pk, self.season.pk)

        self.assertEqual(TeamPhoto.objects.filter(team=self.team, season=self.season).count(), 1)
        second = TeamPhoto.objects.get(team=self.team, season=self.season)
        self.assertEqual(first.pk, second.pk)
        self.assertIn("second", second.image.name)

    def test_a_different_teams_coach_cannot_upload(self):
        self.client.force_login(self.other_team_coach)

        response = self.club_post("team_photo_set", {"image": make_image_file()}, self.team.pk, self.season.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(TeamPhoto.objects.filter(team=self.team).exists())

    def test_this_teams_coach_can_upload(self):
        self.client.force_login(self.team_coach)

        response = self.club_post("team_photo_set", {"image": make_image_file()}, self.team.pk, self.season.pk)

        self.assertRedirects(response, f"{reverse('management:team_detail', args=[self.team.pk])}?season={self.season.pk}")
        self.assertTrue(TeamPhoto.objects.filter(team=self.team, season=self.season).exists())

    def test_deleting_removes_the_photo(self):
        TeamPhoto.objects.create(team=self.team, season=self.season, image=make_image_file())
        self.client.force_login(self.admin_user)

        response = self.club_post("team_photo_delete", {}, self.team.pk, self.season.pk)

        self.assertRedirects(response, f"{reverse('management:team_detail', args=[self.team.pk])}?season={self.season.pk}")
        self.assertFalse(TeamPhoto.objects.filter(team=self.team, season=self.season).exists())

    def test_the_team_page_shows_the_photo_when_set(self):
        TeamPhoto.objects.create(team=self.team, season=self.season, image=make_image_file())
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Replace photo")
        self.assertNotContains(response, "No photo uploaded")

    def test_the_team_page_shows_a_placeholder_when_not_set(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "No photo uploaded")
        self.assertContains(response, "Upload photo")

    def test_upload_ui_is_hidden_from_a_plain_staff_member(self):
        TeamPhoto.objects.create(team=self.team, season=self.season, image=make_image_file())
        self.client.force_login(self.plain_staff)

        response = self.club_get("team_detail", self.team.pk)

        self.assertNotContains(response, "Replace photo")
        self.assertNotContains(response, "Upload photo")


class PositionManagementTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_position_list_is_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        Position.objects.create(club=other_club, name="Rival Coach", short_name="RC", staff_position=True)

        response = self.club_get("position_list")

        self.assertNotContains(response, "Rival Coach")

    def test_creating_a_position(self):
        response = self.club_post("position_create", {"name": "Physio", "short_name": "PH", "ordering": 0, "staff_position": "on", "management_position": ""})

        position = Position.objects.get(club=self.club, name="Physio")
        self.assertRedirects(response, reverse("management:position_list"))
        self.assertTrue(position.staff_position)
        self.assertFalse(position.management_position)

    def test_updating_a_position(self):
        position = Position.objects.create(club=self.club, name="Old name", short_name="ON")

        self.club_post("position_update", {"name": "New name", "short_name": "NN", "ordering": 0}, position.pk)

        position.refresh_from_db()
        self.assertEqual(position.name, "New name")

    def test_a_management_position_must_also_be_a_staff_position(self):
        response = self.club_post("position_create", {"name": "Bad", "short_name": "B", "ordering": 0, "management_position": "on"})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Position.objects.filter(club=self.club, name="Bad").exists())
        self.assertFormError(response.context["form"], "management_position", "A management position must also be a staff position.")

    def test_deleting_an_unused_position(self):
        position = Position.objects.create(club=self.club, name="Physio", short_name="PH")

        response = self.club_post("position_delete", {}, position.pk)

        self.assertRedirects(response, reverse("management:position_list"))
        self.assertFalse(Position.objects.filter(pk=position.pk).exists())

    def test_deleting_a_position_still_on_a_roster_is_refused(self):
        team = Team.objects.create(club=self.club, name="First Team", short_name="1st")
        member = Member.objects.create(first_name="Roster", last_name="Player")
        position = Position.objects.create(club=self.club, name="Goalkeeper", short_name="GK")
        TeamMembership.objects.create(team=team, member=member, season=self.season, position=position)

        response = self.club_post("position_delete", {}, position.pk)

        self.assertRedirects(response, reverse("management:position_list"))
        self.assertTrue(Position.objects.filter(pk=position.pk).exists())

    def test_deleting_a_position_is_admin_only(self):
        position = Position.objects.create(club=self.club, name="Physio", short_name="PH")
        coach_user = User.objects.create_user(email="coach-position-delete@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="Delete Test Team", short_name="DTT")
        coach_position = Position.objects.create(club=self.club, name="Coach", short_name="C", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=coach_position)
        self.client.force_login(coach_user)

        response = self.club_post("position_delete", {}, position.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Position.objects.filter(pk=position.pk).exists())


class NumberPoolManagementTests(ManagementTestBase):
    """Admin-managed jersey-number pools -- see management.views.NumberPool*
    and teams.NumberPool. Assigning a pool to a team happens on the team's
    own edit form (TeamManagementTests), not here."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_number_pool_list_is_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        NumberPool.objects.create(club=other_club, name="Rival Pool", min_number=1, max_number=99)

        response = self.club_get("number_pool_list")

        self.assertNotContains(response, "Rival Pool")

    def test_creating_a_number_pool(self):
        response = self.club_post("number_pool_create", {"name": "Youth", "min_number": "1", "max_number": "99"})

        pool = NumberPool.objects.get(club=self.club, name="Youth")
        self.assertRedirects(response, reverse("management:number_pool_list"))
        self.assertEqual(pool.min_number, 1)
        self.assertEqual(pool.max_number, 99)

    def test_min_number_must_not_exceed_max_number(self):
        response = self.club_post("number_pool_create", {"name": "Bad", "min_number": "99", "max_number": "1"})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(NumberPool.objects.filter(club=self.club, name="Bad").exists())
        self.assertFormError(response.context["form"], "max_number", "Must be greater than or equal to the minimum number.")

    def test_updating_a_number_pool(self):
        pool = NumberPool.objects.create(club=self.club, name="Old name", min_number=1, max_number=50)

        self.club_post("number_pool_update", {"name": "New name", "min_number": "1", "max_number": "60"}, pool.pk)

        pool.refresh_from_db()
        self.assertEqual(pool.name, "New name")
        self.assertEqual(pool.max_number, 60)

    def test_the_list_page_shows_assigned_teams(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        Team.objects.create(club=self.club, name="U10 Boys", short_name="U10B", pool=pool)

        response = self.club_get("number_pool_list")

        self.assertContains(response, "U10B")

    def test_deleting_a_number_pool(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)

        response = self.club_post("number_pool_delete", {}, pool.pk)

        self.assertRedirects(response, reverse("management:number_pool_list"))
        self.assertFalse(NumberPool.objects.filter(pk=pool.pk).exists())

    def test_deleting_a_pool_clears_it_on_assigned_teams(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        team = Team.objects.create(club=self.club, name="U10 Boys", short_name="U10B", pool=pool)

        self.club_post("number_pool_delete", {}, pool.pk)

        team.refresh_from_db()
        self.assertIsNone(team.pool)

    def test_deleting_a_pool_is_admin_only(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        coach_user = User.objects.create_user(email="coach-pool-delete@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="Delete Test Team", short_name="DTT")
        coach_position = Position.objects.create(club=self.club, name="Coach-pool", short_name="CP", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=coach_position)
        self.client.force_login(coach_user)

        response = self.club_post("number_pool_delete", {}, pool.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(NumberPool.objects.filter(pk=pool.pk).exists())

    def test_creating_a_pool_is_admin_only(self):
        coach_user = User.objects.create_user(email="coach-pool-create@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="Create Test Team", short_name="CTT")
        coach_position = Position.objects.create(club=self.club, name="Coach-pool-create", short_name="CPC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=coach_position)
        self.client.force_login(coach_user)

        response = self.club_post("number_pool_create", {"name": "Youth", "min_number": "1", "max_number": "99"})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(NumberPool.objects.filter(club=self.club, name="Youth").exists())

    def test_the_list_page_is_visible_to_any_staff(self):
        coach_user = User.objects.create_user(email="coach-pool-view@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="View Test Team", short_name="VTT")
        coach_position = Position.objects.create(club=self.club, name="Coach-pool-view", short_name="CPV", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=coach_position)
        self.client.force_login(coach_user)

        self.assertEqual(self.club_get("number_pool_list").status_code, 200)


class RefereeLevelManagementTests(ManagementTestBase):
    """Admin-managed referee qualification tiers -- see
    management.views.RefereeLevel* and teams.RefereeLevel."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")

    def make_non_admin_coach(self, email="coach-levels@example.com"):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=position)
        return coach_user

    def test_list_is_visible_to_any_staff(self):
        RefereeLevel.objects.create(club=self.club, name="Regional")
        self.client.force_login(self.make_non_admin_coach())

        response = self.club_get("referee_level_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Regional")

    def test_create_is_admin_only(self):
        self.client.force_login(self.make_non_admin_coach())

        response = self.club_post("referee_level_create", {"name": "Regional", "teams": []})

        self.assertEqual(response.status_code, 403)

    def test_admin_can_create_a_level_with_teams(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("referee_level_create", {"name": "Regional", "teams": [str(self.team.pk), str(self.other_team.pk)]})

        level = RefereeLevel.objects.get(club=self.club, name="Regional")
        self.assertRedirects(response, reverse("management:referee_level_list"))
        self.assertEqual(set(level.teams.all()), {self.team, self.other_team})

    def test_admin_can_update_a_levels_teams(self):
        level = RefereeLevel.objects.create(club=self.club, name="Regional")
        level.teams.add(self.team)
        self.client.force_login(self.admin_user)

        self.club_post("referee_level_update", {"name": "Regional", "teams": [str(self.other_team.pk)]}, level.pk)

        self.assertEqual(set(level.teams.all()), {self.other_team})

    def test_list_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        RefereeLevel.objects.create(club=other_club, name="Rival Level")
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_level_list")

        self.assertNotContains(response, "Rival Level")

    def test_admin_can_link_a_level_to_inherit_from_another(self):
        regional = RefereeLevel.objects.create(club=self.club, name="Regional")
        regional.teams.add(self.team)
        self.client.force_login(self.admin_user)

        response = self.club_post("referee_level_create", {"name": "National", "teams": [str(self.other_team.pk)], "inherits_from": str(regional.pk)})

        national = RefereeLevel.objects.get(club=self.club, name="National")
        self.assertRedirects(response, reverse("management:referee_level_list"))
        self.assertEqual(national.inherits_from, regional)
        self.assertEqual(national.eligible_team_ids(), {self.team.pk, self.other_team.pk})

    def test_a_level_cannot_be_set_to_inherit_from_itself(self):
        level = RefereeLevel.objects.create(club=self.club, name="Regional")
        self.client.force_login(self.admin_user)

        self.club_post("referee_level_update", {"name": "Regional", "teams": [], "inherits_from": str(level.pk)}, level.pk)

        level.refresh_from_db()
        self.assertIsNone(level.inherits_from)

    def test_the_inherits_from_dropdown_only_offers_this_clubs_levels(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        RefereeLevel.objects.create(club=other_club, name="Rival Level")
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_level_create")

        self.assertNotContains(response, "Rival Level")

    def test_the_list_shows_what_a_level_inherits(self):
        regional = RefereeLevel.objects.create(club=self.club, name="Regional")
        RefereeLevel.objects.create(club=self.club, name="National", inherits_from=regional)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_level_list")

        self.assertContains(response, "Regional")
        self.assertContains(response, "everything")

    def test_deleting_an_unused_level(self):
        level = RefereeLevel.objects.create(club=self.club, name="Regional")
        self.client.force_login(self.admin_user)

        response = self.club_post("referee_level_delete", {}, level.pk)

        self.assertRedirects(response, reverse("management:referee_level_list"))
        self.assertFalse(RefereeLevel.objects.filter(pk=level.pk).exists())

    def test_deleting_a_level_held_by_a_referee_is_refused(self):
        level = RefereeLevel.objects.create(club=self.club, name="Regional")
        member = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        RefereeProfile.objects.create(member=member, level=level, valid_until=timezone.localdate() + datetime.timedelta(days=30))
        self.client.force_login(self.admin_user)

        response = self.club_post("referee_level_delete", {}, level.pk)

        self.assertRedirects(response, reverse("management:referee_level_list"))
        self.assertTrue(RefereeLevel.objects.filter(pk=level.pk).exists())

    def test_deleting_a_level_inherited_by_another_is_refused(self):
        regional = RefereeLevel.objects.create(club=self.club, name="Regional")
        RefereeLevel.objects.create(club=self.club, name="National", inherits_from=regional)
        self.client.force_login(self.admin_user)

        response = self.club_post("referee_level_delete", {}, regional.pk)

        self.assertRedirects(response, reverse("management:referee_level_list"))
        self.assertTrue(RefereeLevel.objects.filter(pk=regional.pk).exists())

    def test_deleting_a_level_is_admin_only(self):
        level = RefereeLevel.objects.create(club=self.club, name="Regional")
        self.client.force_login(self.make_non_admin_coach())

        response = self.club_post("referee_level_delete", {}, level.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(RefereeLevel.objects.filter(pk=level.pk).exists())


class RefereeListViewTests(ManagementTestBase):
    """The club-wide referee overview -- see management.views.RefereeListView."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.level = RefereeLevel.objects.create(club=cls.club, name="Regional")
        cls.level.teams.add(cls.team)
        cls.member = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_lists_a_valid_referee_with_level_and_teams(self):
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=timezone.localdate() + datetime.timedelta(days=30))

        response = self.club_get("referee_list")

        self.assertContains(response, "Ref Eree")
        self.assertContains(response, "Regional")
        self.assertContains(response, "First Team")
        self.assertContains(response, "Valid")

    def test_shows_expired_status(self):
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=timezone.localdate() - datetime.timedelta(days=1))

        response = self.club_get("referee_list")

        self.assertContains(response, "Expired")

    def test_shows_no_level_status(self):
        RefereeProfile.objects.create(member=self.member, valid_until=timezone.localdate() + datetime.timedelta(days=30))

        response = self.club_get("referee_list")

        self.assertContains(response, "No level")

    def test_shows_no_validity_set_status(self):
        RefereeProfile.objects.create(member=self.member, level=self.level)

        response = self.club_get("referee_list")

        self.assertContains(response, "No validity set")

    def test_a_member_with_no_referee_profile_is_not_listed(self):
        other_member = Member.objects.create(first_name="Not", last_name="Referee")
        ClubMembership.objects.create(club=self.club, member=other_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("referee_list")

        self.assertNotContains(response, "Not Referee")

    def test_visible_to_any_staff(self):
        coach_user = User.objects.create_user(email="coach-reflist@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=coach_position)
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=timezone.localdate() + datetime.timedelta(days=30))
        # Give the coach visibility into self.member too (members_visible_to
        # scopes a non-admin to teams they're staffed on).
        player_position = Position.objects.create(club=self.club, name="Forward", short_name="FW")
        TeamMembership.objects.create(team=self.team, member=self.member, season=self.season, position=player_position)
        self.client.force_login(coach_user)

        response = self.club_get("referee_list")

        self.assertEqual(response.status_code, 200)


class NumberListViewTests(ManagementTestBase):
    """The Numbers overview page -- see management.views.NumberListView and
    teams.services.numbers, which the tile states are built on top of."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.pool = NumberPool.objects.create(club=cls.club, name="Youth", min_number=1, max_number=5)
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st", pool=cls.pool)
        cls.member = Member.objects.create(first_name="Jane", last_name="Doe")

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_visible_to_any_staff(self):
        coach_user = User.objects.create_user(email="coach-numlist@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=coach_position)
        self.client.force_login(coach_user)

        self.assertEqual(self.club_get("number_list").status_code, 200)

    def test_not_visible_without_any_management_access(self):
        plain_user = User.objects.create_user(email="plain-numlist@example.com", password="pw-secret-123")
        ClubMembership.objects.create(club=self.club, member=Member.objects.create(user=plain_user, first_name="Pat", last_name="Plain"), season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(plain_user)

        self.assertEqual(self.club_get("number_list").status_code, 403)

    def test_no_pools_shows_a_message(self):
        self.pool.delete()

        response = self.club_get("number_list")

        self.assertContains(response, "No number pools yet")

    def test_defaults_to_the_first_pool(self):
        response = self.club_get("number_list")

        self.assertEqual(response.context["pool"], self.pool)

    def test_can_switch_pools_via_query_param(self):
        other_pool = NumberPool.objects.create(club=self.club, name="Senior", min_number=1, max_number=5)

        response = self.club_get("number_list", params={"pool": str(other_pool.pk)})

        self.assertEqual(response.context["pool"], other_pool)

    def test_an_unclaimed_number_is_available(self):
        response = self.club_get("number_list")

        tiles = {tile["number"]: tile for tile in response.context["tiles"]}
        self.assertEqual(tiles[1]["state"], "available")

    def test_a_number_placed_this_season_is_taken(self):
        TeamMembership.objects.create(team=self.team, member=self.member, season=self.season, jersey_number=3)

        response = self.club_get("number_list")

        tiles = {tile["number"]: tile for tile in response.context["tiles"]}
        self.assertEqual(tiles[3]["state"], "taken")
        self.assertEqual(tiles[3]["holders"], [f"{self.member} ({self.team.short_name})"])

    def test_a_number_placed_last_season_only_is_previous(self):
        previous_season = Season.objects.create(club=self.club, start_date=self.season.start_date - datetime.timedelta(days=365), end_date=self.season.start_date - datetime.timedelta(days=1))
        TeamMembership.objects.create(team=self.team, member=self.member, season=previous_season, jersey_number=4)

        response = self.club_get("number_list")

        tiles = {tile["number"]: tile for tile in response.context["tiles"]}
        self.assertEqual(tiles[4]["state"], "previous")

    def test_a_pending_registration_request_is_pending(self):
        membership = ClubMembership.objects.create(club=self.club, member=self.member, season=self.season)
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Jane", contact_last_name="Doe", contact_email="jane@example.com")
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=self.team, requested_jersey_number=5)

        response = self.club_get("number_list")

        tiles = {tile["number"]: tile for tile in response.context["tiles"]}
        self.assertEqual(tiles[5]["state"], "pending")
        self.assertEqual(tiles[5]["holders"], [f"{self.member} ({self.team.short_name})"])

    def test_held_by_shows_the_holders_team(self):
        TeamMembership.objects.create(team=self.team, member=self.member, season=self.season, jersey_number=3)

        response = self.club_get("number_list")

        self.assertContains(response, f'data-holders="{self.member} ({self.team.short_name})"')

    def test_a_manually_reserved_number_is_reserved(self):
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=2, note="Retired -- #2")

        response = self.club_get("number_list")

        tiles = {tile["number"]: tile for tile in response.context["tiles"]}
        self.assertEqual(tiles[2]["state"], "reserved")
        self.assertEqual(tiles[2]["reservation"], reservation)
        self.assertContains(response, "Retired -- #2")

    def test_the_reserving_user_is_shown_on_the_tile(self):
        NumberReservation.objects.create(club=self.club, pool=self.pool, number=2, note="Retired -- #2", reserved_by=self.admin_user)

        response = self.club_get("number_list")

        self.assertContains(response, f'data-reserved-by="{self.admin_user}"')

    def test_reserving_an_available_number(self):
        response = self.club_post("number_reservation_create", {"pool": str(self.pool.pk), "number": "1", "note": "Keep aside"})

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        reservation = NumberReservation.objects.get(pool=self.pool, number=1)
        self.assertEqual(reservation.note, "Keep aside")
        self.assertEqual(reservation.reserved_by, self.admin_user)

    def test_reserving_a_number_outside_the_pool_range_is_rejected(self):
        self.club_post("number_reservation_create", {"pool": str(self.pool.pk), "number": "99", "note": ""})

        self.assertFalse(NumberReservation.objects.filter(pool=self.pool, number=99).exists())

    def test_reserving_an_already_reserved_number_is_rejected(self):
        NumberReservation.objects.create(club=self.club, pool=self.pool, number=1)

        self.club_post("number_reservation_create", {"pool": str(self.pool.pk), "number": "1", "note": "duplicate"})

        self.assertEqual(NumberReservation.objects.filter(pool=self.pool, number=1).count(), 1)

    def test_releasing_a_reservation(self):
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=1)

        response = self.club_post("number_reservation_release", {}, reservation.pk)

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        self.assertFalse(NumberReservation.objects.filter(pk=reservation.pk).exists())

    def _make_coach(self, email):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name=f"Coach {email}", short_name="C", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=position)
        return coach_user

    def test_the_reserving_staff_member_can_release_their_own_reservation(self):
        coach_user = self._make_coach("coach-release-own@example.com")
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=1, reserved_by=coach_user)
        self.client.force_login(coach_user)

        response = self.club_post("number_reservation_release", {}, reservation.pk)

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        self.assertFalse(NumberReservation.objects.filter(pk=reservation.pk).exists())

    def test_another_staff_member_cannot_release_someone_elses_reservation(self):
        reserving_user = self._make_coach("coach-release-a@example.com")
        other_coach = self._make_coach("coach-release-b@example.com")
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=1, reserved_by=reserving_user)
        self.client.force_login(other_coach)

        response = self.club_post("number_reservation_release", {}, reservation.pk)

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        self.assertTrue(NumberReservation.objects.filter(pk=reservation.pk).exists())

    def test_an_admin_can_release_someone_elses_reservation(self):
        coach_user = self._make_coach("coach-release-c@example.com")
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=1, reserved_by=coach_user)
        self.client.force_login(self.admin_user)

        response = self.club_post("number_reservation_release", {}, reservation.pk)

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        self.assertFalse(NumberReservation.objects.filter(pk=reservation.pk).exists())

    def test_a_member_admin_can_release_someone_elses_reservation(self):
        coach_user = self._make_coach("coach-release-d@example.com")
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=1, reserved_by=coach_user)
        member_admin_user = User.objects.create_user(email="member-admin-release@example.com", password="pw-secret-123")
        member_admin_member = Member.objects.create(user=member_admin_user, first_name="Mia", last_name="MemberAdmin")
        ClubMembership.objects.create(club=self.club, member=member_admin_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=self.club, member=member_admin_member).update(role=ClubRole.Roles.MEMBER_ADMIN)
        enrol_mfa(member_admin_user)
        self.client.force_login(member_admin_user)

        response = self.club_post("number_reservation_release", {}, reservation.pk)

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        self.assertFalse(NumberReservation.objects.filter(pk=reservation.pk).exists())

    def test_a_reservation_with_no_reserving_user_cannot_be_released_by_a_plain_staff_member(self):
        coach_user = self._make_coach("coach-release-e@example.com")
        reservation = NumberReservation.objects.create(club=self.club, pool=self.pool, number=1, reserved_by=None)
        self.client.force_login(coach_user)

        response = self.club_post("number_reservation_release", {}, reservation.pk)

        self.assertRedirects(response, f"{reverse('management:number_list')}?pool={self.pool.pk}")
        self.assertTrue(NumberReservation.objects.filter(pk=reservation.pk).exists())


class ClubRoleManagementTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Future", last_name="Editor")
        # An active membership already grants an implicit MEMBER role (club/signals.py) --
        # granting EDITOR must promote that row, not insert a second one.
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_granting_a_role_promotes_the_existing_membership_role(self):
        self.club_post("role_create", {"member": str(self.member.pk), "role": ClubRole.Roles.EDITOR})

        role = ClubRole.objects.get(club=self.club, member=self.member)
        self.assertEqual(role.role, ClubRole.Roles.EDITOR)
        self.assertEqual(ClubRole.objects.filter(club=self.club, member=self.member).count(), 1)

    def test_revoking_a_role(self):
        role = ClubRole.objects.get(club=self.club, member=self.member)

        self.club_post("role_revoke", {}, role.pk)

        self.assertFalse(ClubRole.objects.filter(pk=role.pk).exists())

    def test_the_plain_member_role_never_appears_on_the_list(self):
        # self.admin_member and self.member both hold an implicit MEMBER role --
        # noise this page must never show. (self.member's name still legitimately
        # appears once, in the "Grant role" modal's member picker.)
        response = self.club_get("role_list")

        self.assertContains(response, "No one has the Editor role yet.")

    def test_a_granted_role_appears_under_its_own_section(self):
        self.club_post("role_create", {"member": str(self.member.pk), "role": ClubRole.Roles.EDITOR})

        response = self.club_get("role_list")

        self.assertContains(response, "Future Editor")

    def test_grant_role_is_a_modal_on_the_list_page(self):
        response = self.club_get("role_list")

        self.assertContains(response, 'id="grant_role_modal"')

    def test_each_section_explains_what_the_role_grants(self):
        response = self.club_get("role_list")

        self.assertContains(response, "Full control over the club")
        self.assertContains(response, "Can create and edit events")

    def test_an_invalid_submission_redirects_back_to_the_list_instead_of_a_page(self):
        response = self.club_post("role_create", {"member": "", "role": ClubRole.Roles.EDITOR})

        self.assertRedirects(response, reverse("management:role_list"))


class FamilyManagementTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_registering_a_family_creates_a_login_parent_and_a_login_less_child(self):
        response = self.club_post(
            "family_create",
            {
                "parent_first_name": "Pat",
                "parent_last_name": "Parent",
                "parent_email": "pat.parent@example.com",
                "child_first_name": "Cody",
                "child_last_name": "Child",
                "child_date_of_birth": "2015-04-01",
            },
        )

        family = Family.objects.get(memberships__member__first_name="Cody")
        self.assertRedirects(response, reverse("management:family_detail", args=[family.pk]))

        parent = Member.objects.get(first_name="Pat", last_name="Parent")
        child = Member.objects.get(first_name="Cody", last_name="Child")
        self.assertEqual(parent.user.email, "pat.parent@example.com")
        self.assertFalse(parent.user.has_usable_password(), "should set a password via the reset link, not be given one")
        self.assertIsNone(child.user)

        self.assertEqual(FamilyMembership.objects.get(family=family, member=parent).role, FamilyMembership.FamilyRole.PARENT)
        self.assertEqual(FamilyMembership.objects.get(family=family, member=child).role, FamilyMembership.FamilyRole.CHILD)

        # Both get signed up for the current season, same as a plain MemberCreateView.
        self.assertTrue(ClubMembership.objects.filter(club=self.club, member=parent, season=self.season).exists())
        self.assertTrue(ClubMembership.objects.filter(club=self.club, member=child, season=self.season).exists())

    def test_reusing_an_existing_login_by_email(self):
        # A parent who's already a Member elsewhere (an existing login) must be
        # reused, not duplicated, when registered onto a second family.
        existing_user = User.objects.create_user(email="existing@example.com", password="pw-secret-123")
        existing_member = Member.objects.create(user=existing_user, first_name="Existing", last_name="Parent")

        self.club_post(
            "family_create",
            {
                "parent_first_name": "Ignored",
                "parent_last_name": "Ignored",
                "parent_email": "existing@example.com",
                "child_first_name": "New",
                "child_last_name": "Kid",
            },
        )

        self.assertEqual(Member.objects.filter(user=existing_user).count(), 1)
        family = Family.objects.get(memberships__member__first_name="New")
        self.assertIn(existing_member, family.guardians)

    def make_existing_family(self):
        # A family only counts as "of this club" once at least one of its members
        # has actually signed up (families_of_club, management/views.py) -- exactly
        # what registering the first child through this app already does.
        family = Family.objects.create()
        first_kid = Member.objects.create(first_name="First", last_name="Kid")
        FamilyMembership.objects.create(family=family, member=first_kid, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=self.club, member=first_kid, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        return family

    def test_adding_a_child_to_an_existing_family(self):
        family = self.make_existing_family()

        response = self.club_post("family_add_child", {"first_name": "Second", "last_name": "Kid", "date_of_birth": "2018-01-01"}, family.pk)

        self.assertRedirects(response, reverse("management:family_detail", args=[family.pk]))
        self.assertEqual(family.children.count(), 2)
        self.assertTrue(ClubMembership.objects.filter(club=self.club, member__first_name="Second", season=self.season).exists())

    def test_adding_a_parent_to_an_existing_family_is_idempotent(self):
        family = self.make_existing_family()

        self.club_post("family_add_parent", {"email": "new.parent@example.com", "first_name": "New", "last_name": "Parent"}, family.pk)
        self.club_post("family_add_parent", {"email": "new.parent@example.com", "first_name": "", "last_name": ""}, family.pk)

        self.assertEqual(family.guardians.count(), 1)


class ParentClaimViewTests(ManagementTestBase):
    """The public claim form and the admin review queue -- see
    members.services.claims; the service-level guarantees live in
    members.tests.ParentClaimTests."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.child = Member.objects.create(first_name="Jamie", last_name="Doe", date_of_birth=datetime.date(2014, 3, 2))
        ClubMembership.objects.create(club=cls.club, member=cls.child, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=cls.child, role=FamilyMembership.FamilyRole.CHILD)

    def claim_payload(self, **overrides):
        payload = {
            "parent_first_name": "Taylor",
            "parent_last_name": "Doe",
            "parent_email": "taylor.doe@example.com",
            "child_first_name": "Jamie",
            "child_last_name": "Doe",
            "child_date_of_birth": "2014-03-02",
        }
        payload.update(overrides)
        return payload

    def submit(self, *, follow=False, **overrides):
        return self.client.post(reverse("members:parent_claim"), self.claim_payload(**overrides), HTTP_HOST="ajax-united.rosterchief.app", follow=follow)

    def test_the_claim_form_is_reachable_without_signing_in(self):
        response = self.client.get(reverse("members:parent_claim"), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 200)

    def test_submitting_records_a_pending_claim_and_redirects_back_with_a_flash(self):
        response = self.submit()

        # fetch_redirect_response=False: assertRedirects' own probe GET would
        # otherwise consume the one-shot flash before the assertion below gets to see it.
        self.assertRedirects(response, reverse("members:parent_claim"), fetch_redirect_response=False)
        self.assertEqual(ParentClaim.objects.filter(club=self.club, status=ParentClaim.Status.PENDING).count(), 1)
        page = self.client.get(reverse("members:parent_claim"), HTTP_HOST="ajax-united.rosterchief.app")
        self.assertContains(page, "Request received")

    def test_an_unmatched_claim_flashes_the_same_message_as_a_matched_one(self):
        # The page must not tell an anonymous submitter which children exist.
        matched = self.submit(follow=True)
        unmatched = self.submit(child_first_name="Nobody", child_last_name="Here", parent_email="other@example.com", follow=True)

        self.assertEqual([str(m) for m in matched.context["messages"]], [str(m) for m in unmatched.context["messages"]])

    def test_the_flash_mentions_the_clubs_contact_email_when_set(self):
        self.club.contact_email = "info@ajax-united.example.com"
        self.club.save(update_fields=["contact_email"])

        response = self.submit(follow=True)

        self.assertContains(response, "info@ajax-united.example.com")

    def test_submitting_creates_no_account(self):
        # A public form that made a User per submission would be a spam magnet;
        # the account is created on approval, once a human has vouched for it.
        self.submit()

        self.assertFalse(User.objects.filter(email="taylor.doe@example.com").exists())

    def test_open_signup_is_closed(self):
        response = self.client.get(reverse("account_signup"), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 403)

    def test_the_queue_is_admin_only(self):
        coach_user = User.objects.create_user(email="coach-claims@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="First Team", short_name="1st")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        self.assertEqual(self.club_get("parent_claim_list").status_code, 403)

    def test_the_queue_lists_a_pending_claim_and_the_unclaimed_child(self):
        self.submit()
        self.client.force_login(self.admin_user)

        response = self.club_get("parent_claim_list")

        self.assertContains(response, "taylor.doe@example.com")
        self.assertContains(response, "Jamie Doe")

    def test_the_child_dropdown_matches_the_height_of_the_buttons_beside_it(self):
        # Regression: ClaimReviewForm.child used to hardcode its own class=
        # attrs, which rendered a second class="..." on the <select> alongside
        # the one templatetags/field.html builds (including the size modifier)
        # -- the two never merge, so select-sm silently never reached the page
        # and the dropdown stood taller than the btn-sm/input-sm around it.
        self.submit()
        self.client.force_login(self.admin_user)

        response = self.club_get("parent_claim_list")

        self.assertContains(response, "select-sm")
        html = response.content.decode()
        select_tag = html[html.index("<select") : html.index(">", html.index("<select"))]
        self.assertEqual(select_tag.count('class="'), 1)

    def test_approving_links_the_parent_as_a_guardian(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        parent = Member.objects.get(user__email="taylor.doe@example.com")
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=parent).kind, ClubMembership.Kind.GUARDIAN)
        claim.refresh_from_db()
        self.assertEqual(claim.status, ParentClaim.Status.APPROVED)

    def test_approving_emails_the_parent_a_working_set_password_link(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        self.assertEqual(len(mail.outbox), 1)
        sent = mail.outbox[0]
        self.assertEqual(sent.to, ["taylor.doe@example.com"])
        self.assertIn("Jamie", sent.body)

        [reset_path] = [line for line in sent.body.splitlines() if "/accounts/password/reset/key/" in line]
        parent = Member.objects.get(user__email="taylor.doe@example.com")
        self.assertFalse(parent.user.has_usable_password())
        response = self.client.get(reset_path.strip(), HTTP_HOST="ajax-united.rosterchief.app", follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "password1")

    def test_the_email_carries_an_html_alternative_alongside_the_plain_text(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        sent = mail.outbox[0]
        [(html_body, mimetype)] = sent.alternatives
        self.assertEqual(mimetype, "text/html")
        self.assertIn("Jamie", html_body)
        self.assertIn(self.club.name, html_body)
        [reset_path] = [line for line in sent.body.splitlines() if "/accounts/password/reset/key/" in line]
        self.assertIn(reset_path.strip(), html_body)

    def test_the_initials_badge_text_contrasts_against_the_fallback_colour(self):
        # self.club has no secondary_color set, so the badge falls back to
        # #ec4899 -- white text on that (the old hardcoded default) reads
        # far worse than black (contrast_color("#ec4899") == "#000000"), see
        # club/templatetags/club_email.py::contrast_color.
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        [(html_body, _mimetype)] = mail.outbox[0].alternatives
        self.assertIn("background-color:#ec4899", html_body)
        self.assertIn("color:#000000", html_body)

    def test_the_email_mentions_the_clubs_contact_email_when_set(self):
        self.club.contact_email = "info@ajax-united.example.com"
        self.club.save(update_fields=["contact_email"])
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        self.assertIn("info@ajax-united.example.com", mail.outbox[0].body)

    def test_a_send_failure_still_leaves_the_claim_linked(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        with mock.patch("members.services.claims.EmailMultiAlternatives.send", side_effect=OSError("smtp down")):
            response = self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        self.assertRedirects(response, reverse("management:parent_claim_list"))
        claim.refresh_from_db()
        self.assertEqual(claim.status, ParentClaim.Status.APPROVED)
        self.assertTrue(Member.objects.filter(user__email="taylor.doe@example.com").exists())

    def test_approving_without_choosing_a_child_changes_nothing(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {}, claim.pk)

        claim.refresh_from_db()
        self.assertTrue(claim.is_pending)
        self.assertFalse(User.objects.filter(email="taylor.doe@example.com").exists())

    def test_a_child_who_already_has_a_parent_cannot_be_chosen(self):
        # The shortlist is only ever children with nobody on file, so approving
        # can never quietly re-parent a child who already has one.
        other_child = Member.objects.create(first_name="Sam", last_name="Roe", date_of_birth=datetime.date(2013, 1, 1))
        ClubMembership.objects.create(club=self.club, member=other_child, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=other_child, role=FamilyMembership.FamilyRole.CHILD)
        existing_parent = Member.objects.create(first_name="Existing", last_name="Roe")
        FamilyMembership.objects.create(family=family, member=existing_parent, role=FamilyMembership.FamilyRole.PARENT)
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_approve", {"child": str(other_child.pk)}, claim.pk)

        claim.refresh_from_db()
        self.assertTrue(claim.is_pending)

    def test_rejecting_records_the_reason_and_links_nobody(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        self.club_post("parent_claim_reject", {"note": "Not on our records."}, claim.pk)

        claim.refresh_from_db()
        self.assertEqual(claim.status, ParentClaim.Status.REJECTED)
        self.assertFalse(User.objects.filter(email="taylor.doe@example.com").exists())

    def test_the_parent_sees_their_child_after_approval(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)
        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        self.client.force_login(User.objects.get(email="taylor.doe@example.com"))
        response = self.client.get(reverse("members:my_family"), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "Jamie Doe")

    def test_the_child_dropdown_is_searchable_and_the_top_match_is_preselected(self):
        self.submit()
        self.client.force_login(self.admin_user)

        response = self.club_get("parent_claim_list")

        self.assertContains(response, 'data-searchable="true"')
        html = response.content.decode()
        select_tag = html[html.index("<select") : html.index("</select>", html.index("<select"))]
        self.assertIn(f'value="{self.child.pk}" selected', select_tag)

    def test_a_reject_modal_exists_for_each_pending_claim(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)

        response = self.club_get("parent_claim_list")

        self.assertContains(response, f'id="claim_reject_modal_{claim.pk}"')
        self.assertContains(response, f"document.getElementById('claim_reject_modal_{claim.pk}').showModal()")

    def test_the_history_section_shows_a_claim_reviewed_this_season(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)
        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)

        response = self.club_get("parent_claim_list")

        self.assertContains(response, "Already dealt with")
        self.assertContains(response, "taylor.doe@example.com")

    def test_the_history_section_hides_a_claim_reviewed_last_season(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)
        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)
        claim.refresh_from_db()
        claim.reviewed_at = timezone.make_aware(datetime.datetime.combine(self.season.start_date - datetime.timedelta(days=1), datetime.time()))
        claim.save(update_fields=["reviewed_at"])

        response = self.club_get("parent_claim_list")

        self.assertNotContains(response, "Already dealt with")

    def test_the_history_section_is_empty_without_a_current_season(self):
        self.submit()
        claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)
        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, claim.pk)
        # Push the season into the future rather than deleting it -- ClubMembership.season
        # is PROTECT, and the club being between seasons is the scenario under test,
        # not the absence of any Season row at all.
        future_start = timezone.localdate() + datetime.timedelta(days=100)
        self.season.start_date = future_start
        self.season.end_date = future_start + datetime.timedelta(days=300)
        self.season.save(update_fields=["start_date", "end_date"])

        response = self.club_get("parent_claim_list")

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Already dealt with")

    def test_a_signed_in_parent_sees_locked_read_only_fields_instead_of_blank_inputs(self):
        # Re-typing name/email risks a typo forking off a second account -- the
        # fields are shown read-only rather than editable-but-pre-filled.
        user = User.objects.create_user(email="already.parent@example.com", password="pw-secret-123")
        Member.objects.create(user=user, first_name="Already", last_name="Parent")
        self.client.force_login(user)

        response = self.client.get(reverse("members:parent_claim"), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "Submitting as")
        self.assertContains(response, "Already Parent")
        self.assertContains(response, "already.parent@example.com")
        self.assertNotContains(response, 'name="parent_first_name"')
        self.assertNotContains(response, 'name="parent_last_name"')
        self.assertNotContains(response, 'name="parent_email"')

    def test_a_signed_in_parents_claim_reuses_their_account_and_ignores_posted_parent_fields(self):
        user = User.objects.create_user(email="already.parent@example.com", password="pw-secret-123")
        Member.objects.create(user=user, first_name="Already", last_name="Parent")
        self.client.force_login(user)

        response = self.client.post(
            reverse("members:parent_claim"),
            {
                # Even if a tampered request smuggled these in, the fields don't
                # exist on the locked form and the view never reads POST for them.
                "parent_first_name": "Someone",
                "parent_last_name": "Else",
                "parent_email": "not-me@example.com",
                "child_first_name": "Jamie",
                "child_last_name": "Doe",
                "child_date_of_birth": "2014-03-02",
            },
            HTTP_HOST="ajax-united.rosterchief.app",
        )

        self.assertEqual(response.status_code, 302)
        claim = ParentClaim.objects.get(club=self.club, child_first_name="Jamie")
        self.assertEqual(claim.submitted_by_user, user)
        self.assertEqual(claim.parent_first_name, "Already")
        self.assertEqual(claim.parent_last_name, "Parent")
        self.assertEqual(claim.parent_email, "already.parent@example.com")
        self.assertFalse(User.objects.filter(email="not-me@example.com").exists())

    def test_an_anonymous_submission_has_no_linked_user(self):
        self.submit()

        claim = ParentClaim.objects.get(club=self.club)

        self.assertIsNone(claim.submitted_by_user)

    def test_a_second_claim_from_a_signed_in_parent_merges_the_new_child_into_their_existing_family(self):
        self.submit()
        first_claim = ParentClaim.objects.get(club=self.club)
        self.client.force_login(self.admin_user)
        self.club_post("parent_claim_approve", {"child": str(self.child.pk)}, first_claim.pk)

        parent_user = User.objects.get(email="taylor.doe@example.com")
        second_child = Member.objects.create(first_name="Robin", last_name="Doe", date_of_birth=datetime.date(2016, 5, 1))
        ClubMembership.objects.create(club=self.club, member=second_child, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        second_family = Family.objects.create()
        FamilyMembership.objects.create(family=second_family, member=second_child, role=FamilyMembership.FamilyRole.CHILD)

        self.client.force_login(parent_user)
        self.client.post(
            reverse("members:parent_claim"),
            {"child_first_name": "Robin", "child_last_name": "Doe", "child_date_of_birth": "2016-05-01"},
            HTTP_HOST="ajax-united.rosterchief.app",
        )
        second_claim = ParentClaim.objects.get(club=self.club, child_first_name="Robin")

        self.client.force_login(self.admin_user)
        self.club_post("parent_claim_approve", {"child": str(second_child.pk)}, second_claim.pk)

        # One account, one household with both children -- not a parent split
        # across two Family rows, and not a duplicate User/Member.
        self.assertEqual(User.objects.filter(email="taylor.doe@example.com").count(), 1)
        parent = Member.objects.get(user=parent_user)
        family = FamilyMembership.objects.get(member=parent).family
        self.assertCountEqual(family.children, [self.child, second_child])
        self.assertFalse(Family.objects.filter(pk=second_family.pk).exists())


class GuardianViewTests(ManagementTestBase):
    """How a guardian -- a parent attached to the club only through their child --
    behaves across the management UI. See club.models.ClubMembership.Kind; the
    model-level guarantees live in club.tests.GuardianMembershipTests."""

    def family_payload(self, **overrides):
        payload = {
            "parent_first_name": "Pat",
            "parent_last_name": "Parent",
            "parent_email": "pat.parent@example.com",
            "child_first_name": "Cody",
            "child_last_name": "Child",
            "child_date_of_birth": "2015-04-01",
        }
        payload.update(overrides)
        return payload

    def test_registering_a_family_makes_the_parent_a_guardian_and_the_child_a_member(self):
        self.client.force_login(self.admin_user)

        self.club_post("family_create", self.family_payload())

        parent = Member.objects.get(first_name="Pat", last_name="Parent")
        child = Member.objects.get(first_name="Cody", last_name="Child")
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=parent).kind, ClubMembership.Kind.GUARDIAN)
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=child).kind, ClubMembership.Kind.MEMBER)

    def test_ticking_also_a_member_enrols_the_parent_as_one(self):
        self.client.force_login(self.admin_user)

        self.club_post("family_create", self.family_payload(parent_is_member="on"))

        parent = Member.objects.get(first_name="Pat", last_name="Parent")
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=parent).kind, ClubMembership.Kind.MEMBER)

    def test_a_guardian_is_absent_from_the_member_list(self):
        self.client.force_login(self.admin_user)
        self.club_post("family_create", self.family_payload())

        response = self.club_get("member_list")

        self.assertContains(response, "Cody Child")
        self.assertNotContains(response, "Pat Parent")

    def test_a_guardian_is_not_counted_in_the_membership_kpis(self):
        # Measured as a delta, not an absolute: the base fixture's admin is a
        # member too, and this is about what registering a family *adds*.
        self.client.force_login(self.admin_user)
        before = self.club_get("membership_list").context["kpi_total"]

        self.club_post("family_create", self.family_payload())

        after = self.club_get("membership_list").context["kpi_total"]
        # The child only. A guardian owes nothing, so counting them would
        # overstate the roll and everything derived from it.
        self.assertEqual(after - before, 1)

    def test_a_guardians_own_page_is_still_reachable(self):
        # Excluded from the member *list*, not from the club: an admin has to be
        # able to open them, edit them, and switch them to a member.
        self.client.force_login(self.admin_user)
        self.club_post("family_create", self.family_payload())
        parent = Member.objects.get(first_name="Pat", last_name="Parent")

        self.assertEqual(self.club_get("member_detail", parent.pk).status_code, 200)
        self.assertEqual(self.club_get("member_update", parent.pk).status_code, 200)

    def test_a_guardian_can_still_be_put_in_a_group(self):
        # The stated exception: a parent may well sit on a committee.
        self.client.force_login(self.admin_user)
        self.club_post("family_create", self.family_payload())
        parent = Member.objects.get(first_name="Pat", last_name="Parent")
        group = Group.objects.create(club=self.club, name="Committee")

        response = self.club_get("group_bulk_add", group.pk)

        self.assertContains(response, str(parent.pk))

    def test_a_guardian_is_not_offered_for_a_team_roster(self):
        self.client.force_login(self.admin_user)
        self.club_post("family_create", self.family_payload())
        parent = Member.objects.get(first_name="Pat", last_name="Parent")
        team = Team.objects.create(club=self.club, name="First Team", short_name="1st")

        response = self.club_get("team_bulk_add", team.pk, self.season.pk)

        self.assertNotContains(response, str(parent.pk))


class MemberListFamilyColumnTests(ManagementTestBase):
    """The member list is one flat table -- family is a column (each member's family/
    role attached in Python, management.views.MemberListView), not a grouping."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_family(self, parent_name, child_name):
        family = Family.objects.create()
        parent = Member.objects.create(first_name=parent_name, last_name="Guardian")
        child = Member.objects.create(first_name=child_name, last_name="Kid")
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=family, member=child, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=self.club, member=parent, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubMembership.objects.create(club=self.club, member=child, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        return family, parent, child

    def test_a_family_renders_as_one_group_on_the_member_list(self):
        family, parent, child = self.make_family("Pat", "Cody")

        response = self.club_get("member_list")

        self.assertContains(response, str(family))
        self.assertContains(response, str(parent))
        self.assertContains(response, str(child))

    def test_a_member_with_no_family_has_an_empty_family_column(self):
        loner = Member.objects.create(first_name="Lone", last_name="Member")
        ClubMembership.objects.create(club=self.club, member=loner, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("member_list")

        members_by_pk = {member.pk: member for member in response.context["members"]}
        self.assertEqual(members_by_pk[loner.pk].family_memberships_display, [])

    def test_a_family_members_column_shows_the_family_and_role(self):
        family, parent, child = self.make_family("Pat", "Cody")

        response = self.club_get("member_list")

        members_by_pk = {member.pk: member for member in response.context["members"]}
        parent_fms = members_by_pk[parent.pk].family_memberships_display
        child_fms = members_by_pk[child.pk].family_memberships_display
        self.assertEqual([fm.family for fm in parent_fms], [family])
        self.assertEqual(parent_fms[0].role, "parent")
        self.assertEqual(child_fms[0].role, "child")

    def test_a_member_in_two_families_shows_both_on_the_member_list(self):
        family, parent, _child = self.make_family("Pat", "Cody")
        other_family = Family.objects.create()
        FamilyMembership.objects.create(family=other_family, member=parent, role=FamilyMembership.FamilyRole.OTHER)

        response = self.club_get("member_list")

        members_by_pk = {member.pk: member for member in response.context["members"]}
        parent_fms = members_by_pk[parent.pk].family_memberships_display
        self.assertEqual({fm.family for fm in parent_fms}, {family, other_family})

    def test_families_of_other_clubs_do_not_appear(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_season = make_season(other_club)
        other_member = Member.objects.create(first_name="Other", last_name="Kid")
        ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.ACTIVE)
        other_family = Family.objects.create()
        FamilyMembership.objects.create(family=other_family, member=other_member, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_get("member_list")

        self.assertNotContains(response, str(other_family))
        self.assertNotContains(response, "Other Kid")

    def test_a_non_admin_never_sees_a_family_mate_outside_their_own_visibility(self):
        # group_by_family must bucket the already-scoped members_visible_to() result,
        # never Family.guardians/.children directly -- those ignore scoping entirely
        # and would leak a family-mate a non-admin has no other reason to see.
        _family, parent, child = self.make_family("Pat", "Cody")

        coach_user = User.objects.create_user(email="coach3@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U14", short_name="U14")
        position = Position.objects.create(club=self.club, name="Coach3", short_name="C3", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("member_list")

        self.assertNotContains(response, str(parent))
        self.assertNotContains(response, str(child))


class MemberListStatusColumnTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_current_season_status_is_shown(self):
        member = Member.objects.create(first_name="Sam", last_name="Pending")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.PENDING)

        response = self.club_get("member_list")

        members_by_pk = {m.pk: m for m in response.context["members"]}
        self.assertEqual(members_by_pk[member.pk].current_membership.status, ClubMembership.StatusChoices.PENDING)
        self.assertContains(response, "Pending")

    def test_a_member_with_no_current_season_membership_shows_a_dash(self):
        ClubMembership.objects.filter(club=self.club, member=self.admin_member).delete()

        response = self.club_get("member_list")

        members_by_pk = {m.pk: m for m in response.context["members"]}
        self.assertIsNone(members_by_pk[self.admin_member.pk].current_membership)

    def test_a_lapsed_season_membership_does_not_count_as_current(self):
        lapsed_season = Season.objects.create(club=self.club, start_date=self.season.start_date - datetime.timedelta(days=400), end_date=self.season.start_date - datetime.timedelta(days=40))
        member = Member.objects.create(first_name="Old", last_name="Season")
        ClubMembership.objects.create(club=self.club, member=member, season=lapsed_season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("member_list")

        members_by_pk = {m.pk: m for m in response.context["members"]}
        self.assertIsNone(members_by_pk[member.pk].current_membership)


class MemberGrantLoginTests(ManagementTestBase):
    """A login-less child getting their own account -- see
    members.services.family.grant_login and management.forms.GrantLoginForm."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.family = Family.objects.create()
        cls.child = Member.objects.create(first_name="Cody", last_name="Kid")
        FamilyMembership.objects.create(family=cls.family, member=cls.child, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=cls.club, member=cls.child, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_family_page_offers_the_button_for_a_login_less_child(self):
        response = self.club_get("family_detail", self.family.pk)

        self.assertContains(response, reverse("management:member_grant_login", args=[self.child.pk]))

    def test_no_button_once_the_child_already_has_a_login(self):
        self.child.user = User.objects.create_user(email="already@example.com", password="pw-secret-123")
        self.child.save()

        response = self.club_get("family_detail", self.family.pk)

        self.assertNotContains(response, reverse("management:member_grant_login", args=[self.child.pk]))

    def test_no_button_for_a_parent(self):
        parent = Member.objects.create(first_name="Pat", last_name="Parent")
        FamilyMembership.objects.create(family=self.family, member=parent, role=FamilyMembership.FamilyRole.PARENT)
        ClubMembership.objects.create(club=self.club, member=parent, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("family_detail", self.family.pk)

        self.assertNotContains(response, reverse("management:member_grant_login", args=[parent.pk]))

    def test_the_form_is_prefilled_with_the_childs_contact_email_if_set(self):
        self.child.email = "cody.kid@example.com"
        self.child.save()

        response = self.club_get("family_detail", self.family.pk)

        self.assertContains(response, "cody.kid@example.com")

    def test_granting_a_login_creates_a_usable_account(self):
        response = self.club_post("member_grant_login", {"email": "cody.kid@example.com"}, self.child.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.child.pk]))
        self.child.refresh_from_db()
        self.assertIsNotNone(self.child.user)
        self.assertEqual(self.child.user.email, "cody.kid@example.com")
        self.assertFalse(self.child.user.has_usable_password())

    def test_an_email_already_in_use_is_rejected(self):
        User.objects.create_user(email="taken@example.com", password="pw-secret-123")

        response = self.club_post("member_grant_login", {"email": "taken@example.com"}, self.child.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.child.pk]))
        self.child.refresh_from_db()
        self.assertIsNone(self.child.user)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-grant@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U18", short_name="U18")
        position = Position.objects.create(club=self.club, name="Coach7", short_name="C7", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("member_grant_login", {"email": "cody.kid@example.com"}, self.child.pk)

        self.assertEqual(response.status_code, 403)
        self.child.refresh_from_db()
        self.assertIsNone(self.child.user)


class MemberRefereeEligibilityTests(ManagementTestBase):
    """A member's referee level and validity, set from their own page -- see
    management.views.MemberRefereeEligibilityUpdateView and
    teams.RefereeProfile. Eligible teams are derived from the level
    (teams.RefereeLevel), not picked here. Deliberately unrelated to
    members.Group."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.level = RefereeLevel.objects.create(club=cls.club, name="Regional")
        cls.level.teams.add(cls.team, cls.other_team)
        cls.member = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        cls.future_date = timezone.localdate() + datetime.timedelta(days=30)

    def test_member_page_shows_not_eligible_for_any_team_by_default(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("member_detail", self.member.pk)

        self.assertContains(response, "Not eligible to referee for any team.")

    def test_admin_can_set_level_and_validity_for_a_member_with_no_profile_yet(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("member_referee_eligibility_update", {"level": str(self.level.pk), "valid_until": self.future_date.isoformat()}, self.member.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.member.pk]))
        profile = RefereeProfile.objects.get(member=self.member)
        self.assertEqual(profile.level, self.level)
        self.assertEqual(profile.valid_until, self.future_date)
        self.assertEqual(set(profile.eligible_teams), {self.team, self.other_team})

    def test_admin_can_update_an_existing_profile(self):
        other_level = RefereeLevel.objects.create(club=self.club, name="National")
        profile = RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=self.future_date)
        self.client.force_login(self.admin_user)

        self.club_post("member_referee_eligibility_update", {"level": str(other_level.pk), "valid_until": self.future_date.isoformat()}, self.member.pk)

        profile.refresh_from_db()
        self.assertEqual(profile.level, other_level)

    def test_admin_can_clear_the_level_to_make_someone_ineligible(self):
        profile = RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=self.future_date)
        self.client.force_login(self.admin_user)

        self.club_post("member_referee_eligibility_update", {"level": "", "valid_until": self.future_date.isoformat()}, self.member.pk)

        profile.refresh_from_db()
        self.assertIsNone(profile.level)
        self.assertFalse(profile.is_eligible)

    def test_member_page_shows_eligible_teams_when_valid(self):
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=self.future_date)
        self.client.force_login(self.admin_user)

        response = self.club_get("member_detail", self.member.pk)

        self.assertContains(response, "First Team")

    def test_member_page_shows_a_warning_once_expired(self):
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=timezone.localdate() - datetime.timedelta(days=1))
        self.client.force_login(self.admin_user)

        response = self.club_get("member_detail", self.member.pk)

        self.assertContains(response, "Not currently eligible to referee")
        self.assertNotContains(response, "First Team")

    def test_team_page_lists_eligible_referees(self):
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=self.future_date)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Ref Eree")

    def test_team_page_lists_a_referee_eligible_via_inheritance(self):
        # self.level ("Regional") already qualifies for self.team; a higher
        # level that inherits from it (with no teams of its own) should still
        # show up here -- see RefereeLevel.eligible_team_ids.
        national = RefereeLevel.objects.create(club=self.club, name="National", inherits_from=self.level)
        national_ref = Member.objects.create(first_name="Nat", last_name="Ional")
        RefereeProfile.objects.create(member=national_ref, level=national, valid_until=self.future_date)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Nat Ional")

    def test_team_page_excludes_an_expired_referee(self):
        # "Ref Eree" alone also matches the (unrelated) add-player/assign-staff
        # dropdowns, which list every active club member regardless of referee
        # status -- assert on the eligible-referees panel's own empty state.
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=timezone.localdate() - datetime.timedelta(days=1))
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "No one yet.")

    def test_team_page_shows_a_federation_note_instead_of_eligible_referees(self):
        RefereeProfile.objects.create(member=self.member, level=self.level, valid_until=self.future_date)
        self.team.referee_management = Team.RefereeManagement.FEDERATION
        self.team.save(update_fields=["referee_management"])
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "managed by the federation")
        self.assertNotContains(response, "No one yet.")

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-referee-elig@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("member_referee_eligibility_update", {"level": str(self.level.pk), "valid_until": self.future_date.isoformat()}, self.member.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(RefereeProfile.objects.filter(member=self.member).exists())

    def test_non_admin_does_not_see_the_edit_button(self):
        coach_user = User.objects.create_user(email="coach-referee-elig2@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=coach_position)
        # Puts self.member within the coach's visibility (members_visible_to) so
        # the response is a real 200 -- otherwise "not contains" would trivially
        # pass on a 404 for the wrong reason.
        player_position = Position.objects.create(club=self.club, name="Forward", short_name="FW")
        TeamMembership.objects.create(team=self.team, member=self.member, season=self.season, position=player_position)
        self.client.force_login(coach_user)

        response = self.club_get("member_detail", self.member.pk)

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, reverse("management:member_referee_eligibility_update", args=[self.member.pk]))


class MemberFamilyAttachDetachTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.standalone = Member.objects.create(first_name="Stan", last_name="Alone")
        ClubMembership.objects.create(club=cls.club, member=cls.standalone, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_attaching_to_a_new_family(self):
        response = self.club_post("member_attach_family", {"role": FamilyMembership.FamilyRole.PARENT, "family": ""}, self.standalone.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.standalone.pk]))
        family = Family.objects.get(memberships__member=self.standalone)
        self.assertIn(self.standalone, family.guardians)

    def test_attaching_to_an_existing_family(self):
        family = Family.objects.create()
        kid = Member.objects.create(first_name="Existing", last_name="Kid")
        ClubMembership.objects.create(club=self.club, member=kid, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        FamilyMembership.objects.create(family=family, member=kid, role=FamilyMembership.FamilyRole.CHILD)

        self.club_post("member_attach_family", {"role": FamilyMembership.FamilyRole.PARENT, "family": str(family.pk)}, self.standalone.pk)

        self.assertIn(self.standalone, family.guardians)
        self.assertEqual(Family.objects.filter(memberships__member=self.standalone).count(), 1)

    def test_detaching_from_family_removes_it_when_left_empty(self):
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_post("member_detach_family", {}, self.standalone.pk, family.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.standalone.pk]))
        self.assertFalse(FamilyMembership.objects.filter(member=self.standalone).exists())
        self.assertFalse(Family.objects.filter(pk=family.pk).exists())

    def test_detaching_from_family_keeps_it_when_others_remain(self):
        family = Family.objects.create()
        sibling = Member.objects.create(first_name="Sibling", last_name="Kid")
        ClubMembership.objects.create(club=self.club, member=sibling, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=sibling, role=FamilyMembership.FamilyRole.CHILD)

        self.club_post("member_detach_family", {}, self.standalone.pk, family.pk)

        self.assertTrue(Family.objects.filter(pk=family.pk).exists())
        self.assertIn(sibling, family.children)

    def test_detaching_from_one_family_keeps_membership_in_another(self):
        family = Family.objects.create()
        other_family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=other_family, member=self.standalone, role=FamilyMembership.FamilyRole.OTHER)

        self.club_post("member_detach_family", {}, self.standalone.pk, family.pk)

        self.assertFalse(FamilyMembership.objects.filter(member=self.standalone, family=family).exists())
        self.assertTrue(FamilyMembership.objects.filter(member=self.standalone, family=other_family).exists())

    def test_a_member_can_join_a_second_family(self):
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_post("member_attach_family", {"role": FamilyMembership.FamilyRole.OTHER, "family": ""}, self.standalone.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.standalone.pk]))
        self.assertEqual(FamilyMembership.objects.filter(member=self.standalone).count(), 2)
        self.assertTrue(FamilyMembership.objects.filter(member=self.standalone, family=family).exists())

    def test_member_detail_shows_a_card_per_family(self):
        family = Family.objects.create()
        other_family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=other_family, member=self.standalone, role=FamilyMembership.FamilyRole.OTHER)

        response = self.club_get("member_detail", self.standalone.pk)

        self.assertEqual(len(response.context["family_groups"]), 2)
        self.assertContains(response, str(family))
        self.assertContains(response, str(other_family))
        # "Add to family" stays available even though the member is already in two.
        self.assertContains(response, "Add to family")

    def test_attach_form_excludes_families_already_joined(self):
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_get("member_detail", self.standalone.pk)

        self.assertNotIn(family, response.context["attach_to_family_form"].fields["family"].queryset)

    def test_a_childs_page_shows_their_guardians_phone_numbers(self):
        family = Family.objects.create()
        parent = Member.objects.create(first_name="Pat", last_name="Parent", phone="+32470000001", emergency_phone="+32470000002")
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)

        response = self.club_get("member_detail", self.standalone.pk)

        self.assertIn(parent, response.context["guardians"])
        self.assertContains(response, "Pat Parent")
        self.assertContains(response, "tel:+32 470 00 00 01")
        self.assertContains(response, "tel:+32 470 00 00 02")

    def test_a_parents_page_does_not_show_guardian_numbers(self):
        # self.standalone has no CHILD membership anywhere -- the section must not appear.
        response = self.club_get("member_detail", self.standalone.pk)

        self.assertEqual(len(response.context["guardians"]), 0)
        self.assertNotContains(response, "Parent/guardian contact")

    def test_a_guardian_with_no_phone_numbers_gets_no_dial_buttons(self):
        family = Family.objects.create()
        parent = Member.objects.create(first_name="Pat", last_name="Parent")
        FamilyMembership.objects.create(family=family, member=self.standalone, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)

        response = self.club_get("member_detail", self.standalone.pk)

        self.assertContains(response, "Pat Parent")
        self.assertNotContains(response, "tel:")


class MemberDeleteTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Doomed", last_name="Member")
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_deleting_a_member(self):
        response = self.club_post("member_delete", {}, self.member.pk)

        self.assertRedirects(response, reverse("management:member_list"))
        self.assertFalse(Member.objects.filter(pk=self.member.pk).exists())

    def test_deleting_the_last_member_of_a_family_cleans_it_up(self):
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=self.member, role=FamilyMembership.FamilyRole.CHILD)

        self.club_post("member_delete", {}, self.member.pk)

        self.assertFalse(Family.objects.filter(pk=family.pk).exists())

    def test_a_member_referenced_by_an_order_cannot_be_deleted(self):
        # shop.Order.purchaser is PROTECT -- deleting must fail gracefully, not 500.
        Order.objects.create(club=self.club, purchaser=self.member, total=Decimal("10.00"))

        response = self.club_post("member_delete", {}, self.member.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.member.pk]))
        self.assertTrue(Member.objects.filter(pk=self.member.pk).exists())


class MemberClubMembershipFormTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Fee", last_name="Payer")
        cls.membership = ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_editing_a_member_also_updates_their_current_membership(self):
        response = self.club_post(
            "member_update",
            {
                "first_name": "Fee",
                "last_name": "Payer",
                "kind": ClubMembership.Kind.MEMBER,
                "license": "BE-9999",
                "status": ClubMembership.StatusChoices.ACTIVE,
                "fee_status": ClubMembership.FeeStatus.PAID,
            },
            self.member.pk,
        )

        self.assertRedirects(response, reverse("management:member_detail", args=[self.member.pk]))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.license, "BE-9999")
        self.assertEqual(self.membership.fee_status, ClubMembership.FeeStatus.PAID)

    def test_membership_section_is_fillable_when_not_yet_rostered_this_season(self):
        member = Member.objects.create(first_name="Unrostered", last_name="Member")
        # Visible (has *a* membership in this club), but not for the current season --
        # this is exactly the case that should now offer a fillable section to sign
        # them up, rather than hiding it because nothing exists yet.
        old_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 8, 1), end_date=datetime.date(2021, 5, 31))
        ClubMembership.objects.create(club=self.club, member=member, season=old_season, status=ClubMembership.StatusChoices.LAPSED)

        response = self.club_get("member_update", member.pk)

        self.assertContains(response, "This season")
        self.assertFalse(ClubMembership.objects.filter(club=self.club, member=member, season=self.season).exists())

    def test_saving_the_membership_section_signs_up_a_previously_unrostered_member(self):
        member = Member.objects.create(first_name="Unrostered", last_name="Member")
        # A ClubRole, not a ClubMembership -- gives the admin visibility into this
        # member without a season-bound row already existing, which is exactly the
        # "not signed up for anything yet" case this test means to exercise.
        ClubRole.objects.create(club=self.club, member=member, role=ClubRole.Roles.MEMBER)

        response = self.club_post(
            "member_update",
            {"first_name": "Unrostered", "last_name": "Member", "kind": ClubMembership.Kind.MEMBER, "status": ClubMembership.StatusChoices.ACTIVE, "fee_status": ClubMembership.FeeStatus.PAID},
            member.pk,
        )

        self.assertRedirects(response, reverse("management:member_detail", args=[member.pk]))
        membership = ClubMembership.objects.get(club=self.club, member=member, season=self.season)
        self.assertEqual(membership.status, ClubMembership.StatusChoices.ACTIVE)
        self.assertEqual(membership.fee_status, ClubMembership.FeeStatus.PAID)

    def test_membership_section_absent_with_no_season_at_all(self):
        # A fresh club, never given a season -- there's nothing sensible to sign up
        # for, so unlike the "not rostered yet" case above, this stays hidden.
        empty_club = Club.objects.create(name="No Season FC", slug="no-season-fc")
        admin_member = Member.objects.create(first_name="Empty", last_name="Admin")
        ClubRole.objects.create(club=empty_club, member=admin_member, role=ClubRole.Roles.ADMIN)
        admin_user = User.objects.create_user(email="noseasonadmin@example.com", password="pw-secret-123")
        admin_member.user = admin_user
        admin_member.save()
        enrol_mfa(admin_user)
        self.client.force_login(admin_user)
        member = Member.objects.create(first_name="No", last_name="Season")
        ClubRole.objects.create(club=empty_club, member=member, role=ClubRole.Roles.MEMBER)

        with override_settings(ALLOWED_HOSTS=["rosterchief.app", "ajax-united.rosterchief.app", "no-season-fc.rosterchief.app", "testserver"]):
            response = self.client.get(reverse("management:member_update", args=[member.pk]), HTTP_HOST="no-season-fc.rosterchief.app")

        self.assertNotContains(response, "This season")

    def test_detail_page_shows_season_history(self):
        old_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 8, 1), end_date=datetime.date(2021, 5, 31))
        ClubMembership.objects.create(club=self.club, member=self.member, season=old_season, status=ClubMembership.StatusChoices.LAPSED, fee_status=ClubMembership.FeeStatus.UNPAID)

        response = self.club_get("member_detail", self.member.pk)

        self.assertContains(response, f"{self.season.start_date:%Y} - {self.season.end_date:%Y}")
        self.assertContains(response, f"{old_season.start_date:%Y} - {old_season.end_date:%Y}")
        self.assertContains(response, "lapsed")

    def test_formatted_phone_number_is_shown(self):
        self.member.phone = "+32476123456"
        self.member.save()

        response = self.club_get("member_detail", self.member.pk)

        self.assertContains(response, "+32 476 12 34 56")


class FamilyDetailViewTests(ManagementTestBase):
    """The family overview page, reached by clicking a family's name -- there's no
    standalone "Families" nav entry or list."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.family = Family.objects.create()
        cls.parent = Member.objects.create(first_name="Pat", last_name="Guardian")
        cls.child = Member.objects.create(first_name="Cody", last_name="Kid")
        FamilyMembership.objects.create(family=cls.family, member=cls.parent, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=cls.family, member=cls.child, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=cls.club, member=cls.parent, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubMembership.objects.create(club=cls.club, member=cls.child, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_shows_every_member_of_the_family(self):
        response = self.club_get("family_detail", self.family.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Guardian")
        self.assertContains(response, "Kid")

    def test_a_family_from_another_club_404s(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_family = Family.objects.create()
        other_member = Member.objects.create(first_name="Other", last_name="Kid")
        other_season = make_season(other_club)
        ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.ACTIVE)
        FamilyMembership.objects.create(family=other_family, member=other_member, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_get("family_detail", other_family.pk)

        self.assertEqual(response.status_code, 404)

    def test_only_admins_see_edit_and_delete_actions(self):
        # Give the coach visibility into this family's child (rostered on their team),
        # so the assertion below actually exercises the admin-only button gating --
        # not just "the coach can't see this row at all".
        coach_user = User.objects.create_user(email="coach4@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U15", short_name="U15")
        staff_position = Position.objects.create(club=self.club, name="Coach4", short_name="C4", staff_position=True, management_position=True)
        player_position = Position.objects.create(club=self.club, name="Player4", short_name="P4", staff_position=False)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=staff_position)
        TeamMembership.objects.create(team=team, member=self.child, season=self.season, position=player_position)
        self.client.force_login(coach_user)

        response = self.club_get("family_detail", self.family.pk)

        self.assertContains(response, "Kid")
        # The edit icon links straight to member_detail, same as the row's own name
        # link -- not admin-exclusive, so member_delete is the actual gating signal.
        self.assertNotContains(response, reverse("management:member_delete", args=[self.child.pk]))

    def test_admin_sees_a_remove_from_family_action_per_row(self):
        response = self.club_get("family_detail", self.family.pk)

        self.assertContains(response, reverse("management:member_detach_family", args=[self.parent.pk, self.family.pk]))
        self.assertContains(response, reverse("management:member_detach_family", args=[self.child.pk, self.family.pk]))

    def test_non_admin_does_not_see_the_remove_from_family_action(self):
        coach_user = User.objects.create_user(email="coach6@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U17", short_name="U17")
        staff_position = Position.objects.create(club=self.club, name="Coach6", short_name="C6", staff_position=True, management_position=True)
        player_position = Position.objects.create(club=self.club, name="Player6", short_name="P6", staff_position=False)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=staff_position)
        TeamMembership.objects.create(team=team, member=self.child, season=self.season, position=player_position)
        self.client.force_login(coach_user)

        response = self.club_get("family_detail", self.family.pk)

        self.assertNotContains(response, reverse("management:member_detach_family", args=[self.child.pk, self.family.pk]))

    def test_removing_a_family_member_from_the_row_button_detaches_only_them(self):
        response = self.club_post("member_detach_family", {}, self.child.pk, self.family.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.child.pk]))
        self.assertFalse(FamilyMembership.objects.filter(family=self.family, member=self.child).exists())
        self.assertTrue(FamilyMembership.objects.filter(family=self.family, member=self.parent).exists())


class FamilyMembershipRoleUpdateTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.family = Family.objects.create()
        cls.member = Member.objects.create(first_name="Cody", last_name="Kid")
        cls.membership = FamilyMembership.objects.create(family=cls.family, member=cls.member, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_admin_can_change_the_role(self):
        response = self.club_post("family_membership_role_update", {"role": FamilyMembership.FamilyRole.GUARDIAN}, self.family.pk, self.member.pk)

        self.assertRedirects(response, reverse("management:family_detail", args=[self.family.pk]))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.role, FamilyMembership.FamilyRole.GUARDIAN)

    def test_only_changes_the_role_in_the_targeted_family(self):
        other_family = Family.objects.create()
        other_membership = FamilyMembership.objects.create(family=other_family, member=self.member, role=FamilyMembership.FamilyRole.OTHER)

        self.club_post("family_membership_role_update", {"role": FamilyMembership.FamilyRole.PARENT}, self.family.pk, self.member.pk)

        self.membership.refresh_from_db()
        other_membership.refresh_from_db()
        self.assertEqual(self.membership.role, FamilyMembership.FamilyRole.PARENT)
        self.assertEqual(other_membership.role, FamilyMembership.FamilyRole.OTHER)

    def test_an_invalid_role_is_rejected(self):
        response = self.club_post("family_membership_role_update", {"role": "not-a-role"}, self.family.pk, self.member.pk)

        self.assertRedirects(response, reverse("management:family_detail", args=[self.family.pk]))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.role, FamilyMembership.FamilyRole.CHILD)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-role@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U16", short_name="U16")
        position = Position.objects.create(club=self.club, name="Coach5", short_name="C5", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("family_membership_role_update", {"role": FamilyMembership.FamilyRole.PARENT}, self.family.pk, self.member.pk)

        self.assertEqual(response.status_code, 403)
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.role, FamilyMembership.FamilyRole.CHILD)

    def test_a_family_from_another_club_404s(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_season = make_season(other_club)
        other_family = Family.objects.create()
        other_member = Member.objects.create(first_name="Other", last_name="Kid")
        ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.ACTIVE)
        FamilyMembership.objects.create(family=other_family, member=other_member, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_post("family_membership_role_update", {"role": FamilyMembership.FamilyRole.PARENT}, other_family.pk, other_member.pk)

        self.assertEqual(response.status_code, 404)

    def test_the_dropdown_reflects_the_role_specific_to_that_family(self):
        # Same member, a different role in a second family -- the "others" bucket
        # used to read person.family_memberships.first(), which could silently show
        # a role from the wrong family. This is exactly that scenario.
        other_family = Family.objects.create()
        FamilyMembership.objects.create(family=other_family, member=self.member, role=FamilyMembership.FamilyRole.OTHER)

        this_family_response = self.club_get("family_detail", self.family.pk)
        other_family_response = self.club_get("family_detail", other_family.pk)

        self.assertContains(this_family_response, 'value="child" selected')
        self.assertContains(other_family_response, 'value="other" selected')

    def test_redirects_to_next_when_changed_from_the_member_page(self):
        next_url = reverse("management:member_detail", args=[self.member.pk])

        response = self.club_post("family_membership_role_update", {"role": FamilyMembership.FamilyRole.GUARDIAN, "next": next_url}, self.family.pk, self.member.pk)

        self.assertRedirects(response, next_url)

    def test_ignores_an_unsafe_next_url(self):
        response = self.club_post("family_membership_role_update", {"role": FamilyMembership.FamilyRole.GUARDIAN, "next": "https://evil.example.com/steal"}, self.family.pk, self.member.pk)

        self.assertRedirects(response, reverse("management:family_detail", args=[self.family.pk]))

    def test_member_detail_page_sends_its_own_url_as_next(self):
        response = self.club_get("member_detail", self.member.pk)

        self.assertContains(response, f'name="next" value="{reverse("management:member_detail", args=[self.member.pk])}"')

    def test_family_detail_page_sends_no_next(self):
        response = self.club_get("family_detail", self.family.pk)

        self.assertNotContains(response, 'name="next"')


class MembershipListViewTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        # The base class's own admin ClubMembership (fee_status defaults to UNPAID)
        # would otherwise pollute every count below -- ClubRole (not ClubMembership)
        # is what actually makes them an admin, so this is safe to drop.
        ClubMembership.objects.filter(club=cls.club, member=cls.admin_member).delete()

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_membership(self, first_name, last_name, *, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.UNPAID, season=None, license=""):
        member = Member.objects.create(first_name=first_name, last_name=last_name)
        return ClubMembership.objects.create(club=self.club, member=member, season=season or self.season, status=status, fee_status=fee_status, license=license)

    def test_kpi_counts_for_the_current_season(self):
        self.make_membership("Paid", "One", fee_status=ClubMembership.FeeStatus.PAID)
        self.make_membership("Partial", "One", fee_status=ClubMembership.FeeStatus.PARTIALLY_PAID)
        self.make_membership("Unpaid", "One", fee_status=ClubMembership.FeeStatus.UNPAID)
        self.make_membership("Unpaid", "Two", fee_status=ClubMembership.FeeStatus.UNPAID)
        self.make_membership("Waived", "One", fee_status=ClubMembership.FeeStatus.WAIVED)

        response = self.club_get("membership_list")

        self.assertEqual(response.context["kpi_total"], 5)
        self.assertEqual(response.context["kpi_paid"], 1)
        self.assertEqual(response.context["kpi_partial"], 1)
        self.assertEqual(response.context["kpi_unpaid"], 2)
        self.assertEqual(response.context["kpi_waived"], 1)
        self.assertEqual(response.context["kpi_paid_rate"], 20)

    def test_a_cancelled_memberships_stale_fee_status_does_not_count_toward_the_kpis(self):
        # cancel_membership never touches fee_status -- a cancelled membership
        # can still read UNPAID from before it was cancelled, but nothing is
        # genuinely owed (or unsettled) on it any more.
        self.make_membership("Unpaid", "One", fee_status=ClubMembership.FeeStatus.UNPAID)
        self.make_membership("Cancelled", "One", status=ClubMembership.StatusChoices.CANCELLED, fee_status=ClubMembership.FeeStatus.UNPAID)

        response = self.club_get("membership_list")

        self.assertEqual(response.context["kpi_total"], 1)
        self.assertEqual(response.context["kpi_unpaid"], 1)

    def test_a_cancelled_unpaid_membership_is_excluded_from_the_default_list(self):
        cancelled = self.make_membership("Cancelled", "Two", status=ClubMembership.StatusChoices.CANCELLED, fee_status=ClubMembership.FeeStatus.UNPAID)

        response = self.club_get("membership_list")

        self.assertNotIn(cancelled, response.context["memberships"])

    def test_explicitly_filtering_by_cancelled_status_still_shows_it(self):
        cancelled = self.make_membership("Cancelled", "Three", status=ClubMembership.StatusChoices.CANCELLED, fee_status=ClubMembership.FeeStatus.UNPAID)

        response = self.club_get("membership_list", params={"status": "cancelled", "fee_status": "all"})

        self.assertIn(cancelled, response.context["memberships"])

    def test_a_cancelled_membership_has_no_record_payment_or_mark_paid_actions(self):
        cancelled = self.make_membership("Cancelled", "Four", status=ClubMembership.StatusChoices.CANCELLED, fee_status=ClubMembership.FeeStatus.UNPAID)

        response = self.club_get("membership_list", params={"status": "cancelled", "fee_status": "all"})

        found = next(row for row in response.context["memberships"] if row.pk == cancelled.pk)
        self.assertIsNone(found.record_payment_form)

    def test_default_list_excludes_paid_but_includes_waived(self):
        paid = self.make_membership("Paid", "One", fee_status=ClubMembership.FeeStatus.PAID)
        unpaid = self.make_membership("Unpaid", "One", fee_status=ClubMembership.FeeStatus.UNPAID)
        waived = self.make_membership("Waived", "One", fee_status=ClubMembership.FeeStatus.WAIVED)

        response = self.club_get("membership_list")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertNotIn(paid.pk, ids)
        self.assertIn(unpaid.pk, ids)
        self.assertIn(waived.pk, ids)

    def test_fee_status_filter_narrows_to_one_value(self):
        unpaid = self.make_membership("Unpaid", "One", fee_status=ClubMembership.FeeStatus.UNPAID)
        waived = self.make_membership("Waived", "One", fee_status=ClubMembership.FeeStatus.WAIVED)

        response = self.client.get(reverse("management:membership_list") + "?fee_status=waived", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertNotIn(unpaid.pk, ids)
        self.assertIn(waived.pk, ids)

    def test_fee_status_all_includes_paid(self):
        paid = self.make_membership("Paid", "One", fee_status=ClubMembership.FeeStatus.PAID)

        response = self.client.get(reverse("management:membership_list") + "?fee_status=all", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertIn(paid.pk, ids)

    def test_status_filter(self):
        active = self.make_membership("Active", "One", status=ClubMembership.StatusChoices.ACTIVE)
        pending = self.make_membership("Pending", "One", status=ClubMembership.StatusChoices.PENDING)

        response = self.client.get(reverse("management:membership_list") + "?status=pending", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertNotIn(active.pk, ids)
        self.assertIn(pending.pk, ids)

    def test_team_filter(self):
        on_team = self.make_membership("OnTeam", "Kid")
        off_team = self.make_membership("OffTeam", "Kid")
        team = Team.objects.create(club=self.club, name="U10", short_name="U10")
        position = Position.objects.create(club=self.club, name="Player", short_name="P")
        TeamMembership.objects.create(team=team, member=on_team.member, season=self.season, position=position)

        response = self.client.get(reverse("management:membership_list") + f"?team={team.pk}", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertIn(on_team.pk, ids)
        self.assertNotIn(off_team.pk, ids)

    def test_search_filter(self):
        match = self.make_membership("Findme", "Person")
        other = self.make_membership("Other", "Person")

        response = self.client.get(reverse("management:membership_list") + "?q=Findme", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertIn(match.pk, ids)
        self.assertNotIn(other.pk, ids)

    def test_search_matches_by_family_surname(self):
        # Searching "Smith" should find a family member even when their own name
        # isn't Smith -- e.g. a parent with a different surname than their kid.
        family = Family.objects.create()
        smith_kid = self.make_membership("Junior", "Smith")
        FamilyMembership.objects.create(family=family, member=smith_kid.member, role=FamilyMembership.FamilyRole.CHILD)
        other_parent = self.make_membership("Alex", "Jones")
        FamilyMembership.objects.create(family=family, member=other_parent.member, role=FamilyMembership.FamilyRole.PARENT)
        unrelated = self.make_membership("Nobody", "Related")

        response = self.client.get(reverse("management:membership_list") + "?q=Smith", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertIn(smith_kid.pk, ids)
        self.assertIn(other_parent.pk, ids)
        self.assertNotIn(unrelated.pk, ids)

    def test_search_matches_an_explicit_family_name(self):
        family = Family.objects.create(name="The Andersons")
        membership = self.make_membership("Pat", "Vandermeer")
        FamilyMembership.objects.create(family=family, member=membership.member, role=FamilyMembership.FamilyRole.PARENT)
        unrelated = self.make_membership("Nobody", "Related")

        response = self.client.get(reverse("management:membership_list") + "?q=Anderson", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertIn(membership.pk, ids)
        self.assertNotIn(unrelated.pk, ids)

    def test_family_column_shows_every_family_a_member_belongs_to(self):
        family = Family.objects.create()
        other_family = Family.objects.create()
        membership = self.make_membership("Multi", "Family")
        FamilyMembership.objects.create(family=family, member=membership.member, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=other_family, member=membership.member, role=FamilyMembership.FamilyRole.OTHER)

        response = self.club_get("membership_list")

        by_pk = {m.pk: m for m in response.context["memberships"]}
        family_ids = {fm.family_id for fm in by_pk[membership.pk].member.family_memberships_display}
        self.assertEqual(family_ids, {family.pk, other_family.pk})

    def test_season_filter_switches_away_from_current(self):
        old_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 8, 1), end_date=datetime.date(2021, 5, 31))
        current_membership = self.make_membership("Current", "Season")
        old_membership = self.make_membership("Old", "Season", season=old_season, fee_status=ClubMembership.FeeStatus.UNPAID)

        response = self.client.get(reverse("management:membership_list") + f"?season={old_season.pk}", HTTP_HOST="ajax-united.rosterchief.app")

        ids = {m.pk for m in response.context["memberships"]}
        self.assertIn(old_membership.pk, ids)
        self.assertNotIn(current_membership.pk, ids)
        # KPIs stay pinned to the *current* season regardless of the list's own filter.
        self.assertEqual(response.context["current_season"], self.season)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-membership@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U19", short_name="U19")
        position = Position.objects.create(club=self.club, name="Coach8", short_name="C8", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("membership_list")

        self.assertEqual(response.status_code, 403)

    def test_nav_entry_is_admin_only(self):
        admin_response = self.club_get("home")
        self.assertContains(admin_response, reverse("management:registration_invoice_queue"))

    def test_a_members_own_email_is_shown_with_no_role_label(self):
        member = Member.objects.create(first_name="Ella", last_name="Emailed", email="ella@example.com")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season)

        response = self.club_get("membership_list")

        self.assertContains(response, "ella@example.com")
        self.assertNotContains(response, "(parent)")
        self.assertNotContains(response, "(guardian)")

    def test_a_parents_email_is_shown_and_labelled_parent(self):
        child = Member.objects.create(first_name="Cam", last_name="Childless")
        parent = Member.objects.create(first_name="Pia", last_name="Parent", email="pia@example.com")
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=child, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)
        ClubMembership.objects.create(club=self.club, member=child, season=self.season)

        response = self.club_get("membership_list")

        self.assertContains(response, "pia@example.com")
        self.assertContains(response, "(parent)")

    def test_a_guardians_email_is_shown_and_labelled_guardian(self):
        child = Member.objects.create(first_name="Cam", last_name="Childless")
        guardian = Member.objects.create(first_name="Gia", last_name="Guardian", email="gia@example.com")
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=child, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=guardian, role=FamilyMembership.FamilyRole.GUARDIAN)
        ClubMembership.objects.create(club=self.club, member=child, season=self.season)

        response = self.club_get("membership_list")

        self.assertContains(response, "gia@example.com")
        self.assertContains(response, "(guardian)")

    def test_shows_both_parents_emails(self):
        child = Member.objects.create(first_name="Cam", last_name="Childless")
        parent_one = Member.objects.create(first_name="Bernard", last_name="One", email="bernard@example.com")
        parent_two = Member.objects.create(first_name="Charlotte", last_name="Two", email="charlotte@example.com")
        family = Family.objects.create()
        FamilyMembership.objects.create(family=family, member=child, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=parent_one, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=family, member=parent_two, role=FamilyMembership.FamilyRole.PARENT)
        ClubMembership.objects.create(club=self.club, member=child, season=self.season)

        response = self.club_get("membership_list")

        self.assertContains(response, "bernard@example.com")
        self.assertContains(response, "charlotte@example.com")

    def test_shows_a_dash_when_nobody_has_an_email(self):
        member = Member.objects.create(first_name="No", last_name="Email")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season)

        response = self.club_get("membership_list")

        self.assertContains(response, ">-<")

    def test_a_confirmed_registration_invoice_shows_as_sent(self):
        # Before this, the Invoice column only ever read DuesInvoice -- a
        # membership billed through a confirmed registration instead kept
        # showing "Not sent" even though a real invoice went out.
        member = Member.objects.create(first_name="Reg", last_name="Istered")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, fee_amount=Decimal("50.00"))
        due_date = timezone.now().date() + datetime.timedelta(days=14)
        batch = RegistrationBatch.objects.create(
            club=self.club, season=self.season, contact_first_name="Reg", contact_last_name="Istered", contact_email="reg@example.com", invoice_number="REG-2026-00001", invoice_sent_at=timezone.now(), invoice_due_date=due_date
        )
        RegistrationDetails.objects.create(membership=membership, batch=batch, price=Decimal("50.00"))

        response = self.club_get("membership_list")

        self.assertNotContains(response, "Not sent")
        self.assertContains(response, reverse("management:registration_invoice_pdf", args=[batch.pk]))

    def test_an_unconfirmed_registration_still_shows_not_sent(self):
        member = Member.objects.create(first_name="Reg", last_name="Pending")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season)
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Reg", contact_last_name="Pending", contact_email="reg-pending@example.com")
        RegistrationDetails.objects.create(membership=membership, batch=batch, price=Decimal("50.00"))

        response = self.club_get("membership_list")

        self.assertContains(response, "Not sent")

    def test_an_explicit_dues_invoice_takes_priority_over_a_registration_one(self):
        member = Member.objects.create(first_name="Both", last_name="Invoiced", email="both@example.com")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, fee_amount=Decimal("50.00"))
        due_date = timezone.now().date() + datetime.timedelta(days=14)
        batch = RegistrationBatch.objects.create(
            club=self.club, season=self.season, contact_first_name="Both", contact_last_name="Invoiced", contact_email="both@example.com", invoice_number="REG-2026-00002", invoice_sent_at=timezone.now(), invoice_due_date=due_date
        )
        RegistrationDetails.objects.create(membership=membership, batch=batch, price=Decimal("50.00"))
        dues_invoice = DuesInvoice.objects.create(club=self.club, membership=membership, amount=Decimal("50.00"), due_date=timezone.now().date(), sent_at=timezone.now())

        response = self.club_get("membership_list")

        self.assertContains(response, reverse("management:membership_invoice_detail", args=[membership.pk]))
        self.assertNotContains(response, reverse("management:registration_invoice_pdf", args=[batch.pk]))
        self.assertIsNotNone(dues_invoice)

    def make_registered_membership(self, first_name, last_name, *, invoice_number, contact_first_name="Reg", contact_last_name="Parent", contact_email=None):
        member = Member.objects.create(first_name=first_name, last_name=last_name)
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, fee_amount=Decimal("50.00"))
        batch = RegistrationBatch.objects.create(
            club=self.club,
            season=self.season,
            contact_first_name=contact_first_name,
            contact_last_name=contact_last_name,
            contact_email=contact_email or f"{invoice_number}@example.com",
            invoice_number=invoice_number,
            invoice_sent_at=timezone.now(),
        )
        RegistrationDetails.objects.create(membership=membership, batch=batch, price=Decimal("50.00"))
        return member, membership, batch

    def test_search_matches_a_registrations_invoice_number(self):
        member, membership, _batch = self.make_registered_membership("Kid", "One", invoice_number="REG-2026-00099")
        other, _other_membership, _other_batch = self.make_registered_membership("Other", "Two", invoice_number="REG-2026-00100")

        response = self.club_get("membership_list", params={"q": "REG-2026-00099"})

        self.assertContains(response, member.first_name)
        self.assertNotContains(response, other.first_name)
        self.assertEqual(list(response.context["memberships"]), [membership])

    def test_search_matches_the_registrations_contact_name(self):
        # None of the kids in the batch are themselves named "Parent" --
        # only the search-by-registration join finds them.
        member, _membership, _batch = self.make_registered_membership("Kid", "One", invoice_number="REG-2026-00101", contact_first_name="Uniquely", contact_last_name="Named")

        response = self.club_get("membership_list", params={"q": "Uniquely"})

        self.assertContains(response, member.first_name)

    def test_group_by_registration_clusters_siblings_under_one_heading(self):
        # Two kids from the same submission (RegistrationBatch) -- the whole
        # point of grouping is seeing them clustered together instead of
        # scattered alphabetically among everyone else.
        batch = RegistrationBatch.objects.create(
            club=self.club, season=self.season, contact_first_name="Shared", contact_last_name="Parent", contact_email="shared@example.com", invoice_number="REG-2026-00102", invoice_sent_at=timezone.now()
        )
        first = Member.objects.create(first_name="Alice", last_name="Sibling")
        first_membership = ClubMembership.objects.create(club=self.club, member=first, season=self.season, fee_amount=Decimal("50.00"))
        RegistrationDetails.objects.create(membership=first_membership, batch=batch, price=Decimal("50.00"))
        second = Member.objects.create(first_name="Bob", last_name="Sibling")
        second_membership = ClubMembership.objects.create(club=self.club, member=second, season=self.season, fee_amount=Decimal("50.00"))
        RegistrationDetails.objects.create(membership=second_membership, batch=batch, price=Decimal("50.00"))

        response = self.club_get("membership_list", params={"group": "registration"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "REG-2026-00102")
        self.assertContains(response, first.first_name)
        self.assertContains(response, second.first_name)
        self.assertContains(response, reverse("management:registration_invoice_detail", args=[batch.pk]))

    def test_grouping_is_on_by_default_on_a_bare_page_load(self):
        self.make_registered_membership("Solo", "Kid", invoice_number="REG-2026-00103")

        response = self.club_get("membership_list")

        self.assertTrue(response.context["group_by_registration"])
        self.assertContains(response, "REG-2026-00103 &middot;")

    def test_unchecking_the_toggle_turns_grouping_off(self):
        # An unchecked checkbox submits nothing at all -- but once the filter
        # form has genuinely been submitted (its other fields always post a
        # value), a missing "group" means "unchecked", not "never asked".
        self.make_registered_membership("Solo", "Kid", invoice_number="REG-2026-00104")

        response = self.club_get("membership_list", params={"fee_status": "all"})

        self.assertFalse(response.context["group_by_registration"])
        self.assertNotContains(response, "REG-2026-00104 &middot;")


class MembershipMarkPaidTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Owed", last_name="Fee")
        cls.membership = ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.PENDING, fee_status=ClubMembership.FeeStatus.UNPAID)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_marking_paid_settles_the_fee_but_leaves_status_pending(self):
        # Settling the fee, even through this explicit admin action, never activates
        # on its own -- club.services.onboarding.approve_one/approve_all_clean (the
        # Sign-up page) is the only path to ACTIVE, since it's the one place that
        # also checks the onboarding checklist.
        response = self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)]})

        self.assertRedirects(response, reverse("management:membership_list"))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.PENDING)
        self.assertEqual(self.membership.fee_status, ClubMembership.FeeStatus.PAID)
        self.assertIsNone(self.membership.activated_at)

    def test_marking_paid_does_not_grant_the_member_role(self):
        # MEMBER is granted off status ACTIVE (club/signals.py's post_save signal) --
        # paying the fee alone doesn't reach that, so the role must not appear either.
        self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)]})

        self.assertFalse(ClubRole.objects.filter(club=self.club, member=self.member, role=ClubRole.Roles.MEMBER).exists())

    def test_does_not_overwrite_an_existing_activated_at(self):
        earlier = datetime.date(2026, 1, 1)
        self.membership.activated_at = earlier
        self.membership.save()

        self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)]})

        self.membership.refresh_from_db()
        self.assertEqual(self.membership.activated_at, earlier)

    def test_does_not_touch_unselected_rows(self):
        other = ClubMembership.objects.create(club=self.club, member=Member.objects.create(first_name="Not", last_name="Selected"), season=self.season, status=ClubMembership.StatusChoices.PENDING, fee_status=ClubMembership.FeeStatus.UNPAID)

        self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)]})

        other.refresh_from_db()
        self.assertEqual(other.status, ClubMembership.StatusChoices.PENDING)
        self.assertEqual(other.fee_status, ClubMembership.FeeStatus.UNPAID)

    def test_a_membership_from_another_club_is_ignored(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_season = make_season(other_club)
        other_member = Member.objects.create(first_name="Other", last_name="Club")
        other_membership = ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.PENDING, fee_status=ClubMembership.FeeStatus.UNPAID)

        self.club_post("membership_mark_paid", {"membership_ids": [str(other_membership.pk)]})

        other_membership.refresh_from_db()
        self.assertEqual(other_membership.status, ClubMembership.StatusChoices.PENDING)

    def test_redirects_to_a_safe_next(self):
        next_url = reverse("management:membership_list") + "?status=pending"

        response = self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)], "next": next_url})

        self.assertRedirects(response, next_url)

    def test_ignores_an_unsafe_next(self):
        response = self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)], "next": "https://evil.example.com/"})

        self.assertRedirects(response, reverse("management:membership_list"))

    def test_nothing_selected_is_a_harmless_no_op(self):
        response = self.club_post("membership_mark_paid", {})

        self.assertRedirects(response, reverse("management:membership_list"))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.PENDING)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-markpaid@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U20", short_name="U20")
        position = Position.objects.create(club=self.club, name="Coach9", short_name="C9", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("membership_mark_paid", {"membership_ids": [str(self.membership.pk)]})

        self.assertEqual(response.status_code, 403)
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.PENDING)


class MembershipRecordPaymentTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Owed", last_name="Fee")
        cls.membership = ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.PENDING, fee_status=ClubMembership.FeeStatus.UNPAID, fee_amount=Decimal("150.00"))

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_recording_a_partial_payment(self):
        response = self.club_post("membership_record_payment", {"amount": "50.00", "method": FeePayment.Method.CASH, "reference": "R1"}, self.membership.pk)

        self.assertRedirects(response, reverse("management:membership_list"))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.amount_paid, Decimal("50.00"))
        self.assertEqual(self.membership.fee_status, ClubMembership.FeeStatus.PARTIALLY_PAID)
        payment = FeePayment.objects.get(membership=self.membership)
        self.assertEqual(payment.reference, "R1")
        self.assertEqual(payment.recorded_by, self.admin_user)

    def test_recording_the_full_remaining_amount_settles_the_fee_but_leaves_status_pending(self):
        self.club_post("membership_record_payment", {"amount": "150.00", "method": FeePayment.Method.BANK_TRANSFER}, self.membership.pk)

        self.membership.refresh_from_db()
        self.assertEqual(self.membership.fee_status, ClubMembership.FeeStatus.PAID)
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.PENDING)

    def test_an_invalid_amount_is_rejected_without_recording_anything(self):
        response = self.club_post("membership_record_payment", {"amount": "0", "method": FeePayment.Method.CASH}, self.membership.pk)

        self.assertRedirects(response, reverse("management:membership_list"))
        self.assertFalse(FeePayment.objects.filter(membership=self.membership).exists())
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.amount_paid, Decimal("0.00"))

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-recordpay@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U22", short_name="U22")
        position = Position.objects.create(club=self.club, name="Coach11", short_name="C11", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("membership_record_payment", {"amount": "50.00", "method": FeePayment.Method.CASH}, self.membership.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(FeePayment.objects.filter(membership=self.membership).exists())


class MembershipMarkFullyPaidTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Owed", last_name="Fee")
        cls.membership = ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.PENDING, fee_status=ClubMembership.FeeStatus.UNPAID, fee_amount=Decimal("150.00"))

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_settles_the_remaining_balance_in_one_click(self):
        response = self.club_post("membership_mark_fully_paid", {}, self.membership.pk)

        self.assertRedirects(response, reverse("management:membership_list"))
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.fee_status, ClubMembership.FeeStatus.PAID)
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.PENDING)
        self.assertEqual(self.membership.amount_paid, Decimal("150.00"))

    def test_a_membership_from_another_club_404s(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_season = make_season(other_club)
        other_member = Member.objects.create(first_name="Other", last_name="Club")
        other_membership = ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, fee_amount=Decimal("100.00"))

        response = self.club_post("membership_mark_fully_paid", {}, other_membership.pk)

        self.assertEqual(response.status_code, 404)

    def test_redirects_to_a_safe_next(self):
        next_url = reverse("management:membership_list") + "?status=pending"

        response = self.club_post("membership_mark_fully_paid", {"next": next_url}, self.membership.pk)

        self.assertRedirects(response, next_url)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-fullypaid@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U23", short_name="U23")
        position = Position.objects.create(club=self.club, name="Coach12", short_name="C12", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("membership_mark_fully_paid", {}, self.membership.pk)

        self.assertEqual(response.status_code, 403)
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.PENDING)


class MembershipListButtonVisibilityTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_membership(self, fee_status):
        member = Member.objects.create(first_name=fee_status, last_name="Row")
        return ClubMembership.objects.create(club=self.club, member=member, season=self.season, fee_status=fee_status, fee_amount=Decimal("100.00"))

    def test_buttons_hidden_for_a_paid_row(self):
        membership = self.make_membership(ClubMembership.FeeStatus.PAID)

        response = self.client.get(reverse("management:membership_list") + "?fee_status=all", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertNotContains(response, reverse("management:membership_mark_fully_paid", args=[membership.pk]))

    def test_buttons_hidden_for_a_waived_row(self):
        membership = self.make_membership(ClubMembership.FeeStatus.WAIVED)

        response = self.client.get(reverse("management:membership_list") + "?fee_status=all", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertNotContains(response, reverse("management:membership_mark_fully_paid", args=[membership.pk]))

    def test_buttons_shown_for_an_unpaid_row(self):
        membership = self.make_membership(ClubMembership.FeeStatus.UNPAID)

        response = self.client.get(reverse("management:membership_list") + "?fee_status=all", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, reverse("management:membership_mark_fully_paid", args=[membership.pk]))


class RenderPdfTests(TestCase):
    """management.pdf.render_pdf itself -- see billing/tests.py's equivalent
    coverage of billing.services.invoices.render_pdf, same lazy-import shape."""

    def test_the_pdf_library_is_only_needed_when_a_pdf_is_asked_for(self):
        with mock.patch.dict(sys.modules, {"weasyprint": mock.MagicMock()}):
            sys.modules["weasyprint"].HTML.return_value.write_pdf.return_value = b"%PDF-1.7"

            self.assertEqual(render_pdf("<p>hi</p>"), b"%PDF-1.7")

    def test_a_missing_pdf_library_says_what_is_missing(self):
        with mock.patch.dict(sys.modules, {"weasyprint": None}), self.assertRaises(PDFExportError):
            render_pdf("<p>hi</p>")


class MembershipExportPdfTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Print", last_name="Me")
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.UNPAID)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_downloads_as_a_pdf(self):
        with mock.patch("management.views.membership_list_pdf", return_value=b"%PDF-fake") as renderer:
            response = self.club_get("membership_export_pdf")

        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(".pdf", response["Content-Disposition"])
        self.assertEqual(response.content, b"%PDF-fake")
        renderer.assert_called_once()

    def test_export_uses_the_same_filters_as_the_page(self):
        # Same fixture as the "on-team" filter test for the page itself -- the PDF
        # must reflect whatever's filtered, not the whole club.
        other = Member.objects.create(first_name="Other", last_name="Person")
        ClubMembership.objects.create(club=self.club, member=other, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.UNPAID)

        with mock.patch("management.views.membership_list_pdf", return_value=b"%PDF-fake") as renderer:
            self.client.get(reverse("management:membership_export_pdf") + "?q=Print", HTTP_HOST="ajax-united.rosterchief.app")

        context = renderer.call_args[0][0]
        names = {m.member.last_name for m in context["memberships"]}
        self.assertEqual(names, {"Me"})

    def test_a_missing_pdf_library_is_reported_rather_than_a_500(self):
        # WeasyPrint needs native libs. Without them the button must explain itself.
        with mock.patch("management.views.membership_list_pdf", side_effect=PDFExportError("PDF rendering needs the native pango/cairo libraries.")):
            response = self.club_get("membership_export_pdf")
        response = self.client.get(response.url, HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "pango")

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-export@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U21", short_name="U21")
        position = Position.objects.create(club=self.club, name="Coach10", short_name="C10", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("membership_export_pdf")

        self.assertEqual(response.status_code, 403)


class MembershipSendInvoicesTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)
        self.member = Member.objects.create(first_name="Jane", last_name="Doe", email="jane@example.com")
        self.membership = ClubMembership.objects.create(club=self.club, member=self.member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_amount=Decimal("150.00"))

    def test_sending_creates_and_emails_an_invoice(self):
        self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "14"})

        invoice = DuesInvoice.objects.get(membership=self.membership)
        self.assertIsNotNone(invoice.sent_at)
        self.assertEqual(invoice.sent_to_email, "jane@example.com")
        self.assertEqual(invoice.amount, Decimal("150.00"))
        self.assertEqual(invoice.due_date, timezone.now().date() + datetime.timedelta(days=14))
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["jane@example.com"])

    def test_the_email_carries_an_html_alternative(self):
        self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "14"})

        [(html_body, mimetype)] = mail.outbox[0].alternatives
        self.assertEqual(mimetype, "text/html")
        self.assertIn(self.club.name, html_body)

    def test_falls_back_to_a_guardians_email(self):
        self.member.email = ""
        self.member.save(update_fields=["email"])
        family = Family.objects.create()
        parent = Member.objects.create(first_name="Pat", last_name="Doe", email="pat@example.com")
        FamilyMembership.objects.create(family=family, member=self.member, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)

        self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "14"})

        invoice = DuesInvoice.objects.get(membership=self.membership)
        self.assertEqual(invoice.sent_to_email, "pat@example.com")
        self.assertTrue(invoice.sent_to_guardian)

    def test_a_member_with_no_reachable_email_is_skipped(self):
        self.member.email = ""
        self.member.save(update_fields=["email"])

        response = self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "14"})

        self.assertFalse(DuesInvoice.objects.filter(membership=self.membership).exists())
        self.assertEqual(len(mail.outbox), 0)
        response = self.client.get(response.url, HTTP_HOST="ajax-united.rosterchief.app")
        self.assertContains(response, "no email on file")

    def test_resending_updates_the_same_invoice(self):
        self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "14"})
        first_number = DuesInvoice.objects.get(membership=self.membership).number

        self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "30"})

        self.assertEqual(DuesInvoice.objects.filter(membership=self.membership).count(), 1)
        invoice = DuesInvoice.objects.get(membership=self.membership)
        self.assertEqual(invoice.number, first_number)
        self.assertEqual(invoice.due_date, timezone.now().date() + datetime.timedelta(days=30))
        self.assertEqual(len(mail.outbox), 2)

    def test_no_selection_shows_an_error(self):
        response = self.club_post("membership_send_invoices", {"due_in_days": "14"})

        self.assertFalse(DuesInvoice.objects.exists())
        response = self.client.get(response.url, HTTP_HOST="ajax-united.rosterchief.app")
        self.assertContains(response, "Select at least one member")

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-invoice@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U15", short_name="U15")
        position = Position.objects.create(club=self.club, name="Coach11", short_name="C11", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("membership_send_invoices", {"membership_ids": [str(self.membership.pk)], "due_in_days": "14"})

        self.assertEqual(response.status_code, 403)


class MembershipSendInvoiceRemindersTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_invoice(self, *, due_date, fee_status=ClubMembership.FeeStatus.UNPAID, email="jane@example.com"):
        member = Member.objects.create(first_name="Jane", last_name="Doe", email=email)
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_amount=Decimal("100.00"), fee_status=fee_status)
        return DuesInvoice.objects.create(club=self.club, membership=membership, number="DUE-2026-00001", amount=Decimal("100.00"), due_date=due_date, sent_at=timezone.now(), sent_to_email=email)

    def test_reminds_an_overdue_unpaid_invoice(self):
        invoice = self.make_invoice(due_date=timezone.now().date() - datetime.timedelta(days=1))

        self.club_post("membership_send_invoice_reminders", {})

        invoice.refresh_from_db()
        self.assertEqual(invoice.reminder_count, 1)
        self.assertIsNotNone(invoice.last_reminder_sent_at)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["jane@example.com"])

    def test_does_not_remind_one_not_yet_due(self):
        invoice = self.make_invoice(due_date=timezone.now().date() + datetime.timedelta(days=5))

        self.club_post("membership_send_invoice_reminders", {})

        invoice.refresh_from_db()
        self.assertEqual(invoice.reminder_count, 0)
        self.assertEqual(len(mail.outbox), 0)

    def test_does_not_remind_a_paid_invoice(self):
        invoice = self.make_invoice(due_date=timezone.now().date() - datetime.timedelta(days=1), fee_status=ClubMembership.FeeStatus.PAID)

        self.club_post("membership_send_invoice_reminders", {})

        invoice.refresh_from_db()
        self.assertEqual(invoice.reminder_count, 0)

    def test_nothing_to_remind_notifies_gracefully(self):
        response = self.club_post("membership_send_invoice_reminders", {})

        response = self.client.get(response.url, HTTP_HOST="ajax-united.rosterchief.app")
        self.assertContains(response, "Nothing to remind")

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-reminder@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U16", short_name="U16")
        position = Position.objects.create(club=self.club, name="Coach12", short_name="C12", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_post("membership_send_invoice_reminders", {})

        self.assertEqual(response.status_code, 403)


class DuesInvoiceDetailViewTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)
        self.member = Member.objects.create(first_name="Jane", last_name="Doe", email="jane@example.com")
        self.membership = ClubMembership.objects.create(club=self.club, member=self.member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_amount=Decimal("150.00"))
        self.invoice = DuesInvoice.objects.create(club=self.club, membership=self.membership, number="DUE-2026-00001", amount=Decimal("150.00"), due_date=timezone.now().date(), sent_at=timezone.now(), sent_to_email="jane@example.com")

    def test_shows_the_invoice(self):
        response = self.club_get("membership_invoice_detail", self.membership.pk)

        self.assertContains(response, "DUE-2026-00001")
        self.assertContains(response, "jane@example.com")

    def test_404_when_the_membership_has_no_invoice(self):
        other_member = Member.objects.create(first_name="No", last_name="Invoice")
        other_membership = ClubMembership.objects.create(club=self.club, member=other_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("membership_invoice_detail", other_membership.pk)

        self.assertEqual(response.status_code, 404)

    def test_downloads_as_a_pdf(self):
        with mock.patch("management.views.invoice_pdf", return_value=b"%PDF-fake") as renderer:
            response = self.club_get("membership_invoice_pdf", self.membership.pk)

        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertEqual(response.content, b"%PDF-fake")
        renderer.assert_called_once()

    def test_a_missing_pdf_library_is_reported_rather_than_a_500(self):
        with mock.patch("management.views.invoice_pdf", side_effect=DuesInvoicePDFError("PDF rendering needs the native pango/cairo libraries.")):
            response = self.club_get("membership_invoice_pdf", self.membership.pk)
        response = self.client.get(response.url, HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "pango")

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-invoice-detail@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U17", short_name="U17")
        position = Position.objects.create(club=self.club, name="Coach13", short_name="C13", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("membership_invoice_detail", self.membership.pk)

        self.assertEqual(response.status_code, 403)


class MemberListRowActionsTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.member = Member.objects.create(first_name="Row", last_name="Actions")
        ClubMembership.objects.create(club=cls.club, member=cls.member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def test_admin_sees_edit_and_delete_buttons(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("member_list")

        self.assertContains(response, reverse("management:member_delete", args=[self.member.pk]))

    def test_non_admin_does_not_see_edit_and_delete_buttons(self):
        # Roster self.member on the coach's team so this actually tests the
        # admin-only button gating, not just "the coach can't see this row at all".
        coach_user = User.objects.create_user(email="coach5@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U16", short_name="U16")
        staff_position = Position.objects.create(club=self.club, name="Coach5", short_name="C5", staff_position=True, management_position=True)
        player_position = Position.objects.create(club=self.club, name="Player5", short_name="P5", staff_position=False)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=staff_position)
        TeamMembership.objects.create(team=team, member=self.member, season=self.season, position=player_position)
        self.client.force_login(coach_user)

        response = self.club_get("member_list")

        self.assertContains(response, "Row")
        self.assertNotContains(response, reverse("management:member_delete", args=[self.member.pk]))


class HomeViewTests(ManagementTestBase):
    """The dashboard reuses controlpanel.services.statistics' club_attention/
    club_charts/club_statistics -- already club-scoped, so directly usable for this
    club's own staff. Financial pieces (fee chart, Shop stat group) are admin-only,
    same line the nav already draws around the Shop section."""

    def make_coach(self, email):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name=f"Team-{email}", short_name="T")
        position = Position.objects.create(club=self.club, name=f"Coach-{email}", short_name="C", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        return coach_user

    def test_admin_sees_the_financial_sections(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="fees-chart"')
        self.assertContains(response, 'id="signups-chart"')
        self.assertContains(response, "Renewal rate")
        self.assertContains(response, "Orders unpaid")
        self.assertContains(response, "md:grid-cols-4")

    def test_non_admin_staff_does_not_see_the_financial_sections(self):
        self.client.force_login(self.make_coach("coach6@example.com"))

        response = self.club_get("home")

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'id="fees-chart"')
        self.assertNotContains(response, 'id="signups-chart"')
        self.assertNotContains(response, "Renewal rate")
        self.assertNotContains(response, "Orders unpaid")
        self.assertContains(response, "md:grid-cols-3")

    def test_the_season_pill_shows_the_season_years_not_the_short_name(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertContains(response, f"Season {self.season.start_date.year}-{self.season.end_date.year}")
        self.assertNotContains(response, self.season.name)

    def test_total_members_kpi_counts_current_season_members(self):
        # self.admin_user already has one ACTIVE ClubMembership for self.season
        # (ManagementTestBase.setUpTestData) -- a guardian-kind row must not count.
        member = Member.objects.create(first_name="Extra", last_name="Player")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, kind=ClubMembership.Kind.MEMBER, status=ClubMembership.StatusChoices.ACTIVE)
        guardian = Member.objects.create(first_name="Guardian", last_name="Only")
        ClubMembership.objects.create(club=self.club, member=guardian, season=self.season, kind=ClubMembership.Kind.GUARDIAN, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["member_count"], 2)

    def test_total_members_trend_compares_to_the_previous_season(self):
        previous_season = Season.objects.create(club=self.club, start_date=self.season.start_date - datetime.timedelta(days=365), end_date=self.season.start_date - datetime.timedelta(days=1))
        for i in range(3):
            member = Member.objects.create(first_name=f"Last{i}", last_name="Season")
            ClubMembership.objects.create(club=self.club, member=member, season=previous_season, kind=ClubMembership.Kind.MEMBER, status=ClubMembership.StatusChoices.ACTIVE)
        # self.admin_user is the only current-season member -- 1 now vs 3 last season.
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["member_count_change"], -2)
        self.assertContains(response, "-2")

    def test_no_previous_season_shows_no_comparison_available(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertIsNone(response.context["member_count_change"])
        self.assertContains(response, "No previous season to compare")

    def test_missing_documentation_kpi_hidden_without_any_requirements_configured(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertFalse(response.context["requirements_configured"])
        self.assertNotContains(response, "Missing documentation")

    def test_missing_documentation_kpi_counts_open_checklist_items_and_appears_in_needs_attention(self):
        OnboardingRequirement.objects.create(club=self.club, name="Medical certificate")
        member = Member.objects.create(first_name="Needs", last_name="Docs")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, kind=ClubMembership.Kind.MEMBER, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        # self.admin_user also has an open item for the same requirement -- 2 total.
        self.assertTrue(response.context["requirements_configured"])
        self.assertEqual(response.context["missing_documentation_count"], 2)
        self.assertContains(response, "Missing documentation")
        self.assertContains(response, "members have an open checklist item")

    def test_missing_documentation_does_not_appear_once_everyone_is_complete(self):
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Medical certificate")
        membership = ClubMembership.objects.get(club=self.club, member=self.admin_member, season=self.season)
        MemberRequirementStatus.objects.create(membership=membership, requirement=requirement, is_complete=True)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["missing_documentation_count"], 0)
        self.assertNotContains(response, "have an open checklist item")

    def test_the_user_menu_links_to_password_mfa_and_logout(self):
        # base.html's own chrome, exercised through any management page.
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertContains(response, f'href="{reverse("account_change_password")}"')
        self.assertContains(response, f'href="{reverse("mfa_index")}"')
        self.assertContains(response, f'href="{reverse("account_logout")}"')

    def test_published_news_is_shown_to_everyone(self):
        item = News.objects.create(club=self.club, title="Season kickoff", body="Body.")
        item.publish()
        self.client.force_login(self.make_coach("coach-news-home@example.com"))

        response = self.club_get("home")

        self.assertContains(response, "Season kickoff")

    def test_a_draft_or_scheduled_news_item_is_not_shown_on_the_home_page(self):
        draft = News.objects.create(club=self.club, title="Still a draft", body="Body.")
        scheduled = News.objects.create(club=self.club, title="Scheduled for later", body="Body.")
        scheduled.publish(at=timezone.now() + datetime.timedelta(days=7))
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertNotContains(response, draft.title)
        self.assertNotContains(response, scheduled.title)

    def test_upcoming_events_are_listed_in_order_and_future_only(self):
        now = timezone.now()
        past = Event.objects.create(club=self.club, kind=Event.EventKind.TRAINING, title="Past training", start=now - datetime.timedelta(days=1))
        soon = Event.objects.create(club=self.club, kind=Event.EventKind.TRAINING, title="Sooner training", start=now + datetime.timedelta(days=1))
        later = Event.objects.create(club=self.club, kind=Event.EventKind.GAME, title="Later game", start=now + datetime.timedelta(days=5))
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        body = response.content.decode()
        self.assertNotIn(past.title, body)
        self.assertLess(body.index(soon.title), body.index(later.title))

    def test_loads_fine_with_no_season_and_no_events(self):
        # A fresh club, never given a season -- ClubMembership.season is PROTECT, so
        # this is a new club rather than deleting self.season out from under setUp's
        # own ClubMembership.
        empty_club = Club.objects.create(name="Empty FC", slug="empty-fc")
        admin_member = Member.objects.create(first_name="Empty", last_name="Admin")
        ClubRole.objects.create(club=empty_club, member=admin_member, role=ClubRole.Roles.ADMIN)
        admin_user = User.objects.create_user(email="emptyadmin@example.com", password="pw-secret-123")
        admin_member.user = admin_user
        admin_member.save()
        enrol_mfa(admin_user)
        self.client.force_login(admin_user)

        with override_settings(ALLOWED_HOSTS=["rosterchief.app", "ajax-united.rosterchief.app", "empty-fc.rosterchief.app", "testserver"]):
            response = self.client.get(reverse("management:home"), HTTP_HOST="empty-fc.rosterchief.app")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "cannot take a signup")
        self.assertContains(response, "Nothing scheduled.")


class MemberBulkImportTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_non_admin_staff(self):
        coach_user = User.objects.create_user(email="coach-import@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U14", short_name="U14")
        position = Position.objects.create(club=self.club, name="CoachImport", short_name="CI", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        return coach_user

    def test_template_download_has_expected_headers_and_dropdowns(self):
        response = self.club_get("member_import_template")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_CONTENT_TYPE)

        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]
        self.assertEqual(header, TEMPLATE_COLUMNS)
        self.assertTrue(sheet.data_validations.dataValidation)

    def test_non_admin_can_download_template_but_not_upload(self):
        coach_user = self.make_non_admin_staff()
        self.client.force_login(coach_user)

        template_response = self.club_get("member_import_template")
        upload_response = self.club_post("member_import", {"file": make_import_workbook([["Jamie", "Kid", "", "jamie@example.com", "", "", "", "", ""]])})

        self.assertEqual(template_response.status_code, 200)
        self.assertEqual(upload_response.status_code, 403)

    def test_anonymous_is_redirected_to_login_for_both(self):
        self.client.logout()

        self.assertEqual(self.club_get("member_import_template").status_code, 302)
        self.assertEqual(self.club_get("member_import").status_code, 302)

    def test_a_clean_row_previews_as_will_create(self):
        upload = make_import_workbook([["Jamie", "Kid", "2010-01-01", "jamie.kid@example.com", "+32470111111", "", "", "", ""]])

        response = self.club_post("member_import", {"file": upload})

        self.assertEqual(response.context["valid_count"], 1)
        self.assertEqual(response.context["skipped_count"], 0)
        result = response.context["results"][0]
        self.assertIsNotNone(result["member"])
        self.assertEqual(result["membership_kwargs"]["status"], ClubMembership.StatusChoices.ACTIVE)
        self.assertEqual(result["membership_kwargs"]["fee_status"], ClubMembership.FeeStatus.UNPAID)

    def test_a_row_missing_a_required_name_is_skipped(self):
        upload = make_import_workbook([["", "Noname", "", "", "", "", "", "", ""]])

        response = self.club_post("member_import", {"file": upload})

        self.assertEqual(response.context["valid_count"], 0)
        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(result["errors"])

    def test_an_invalid_email_is_skipped(self):
        upload = make_import_workbook([["Bad", "Email", "", "not-an-email", "", "", "", "", ""]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])

    def test_an_invalid_status_value_is_skipped(self):
        upload = make_import_workbook([["Bad", "Status", "", "bad.status@example.com", "", "", "", "not-a-status", ""]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(any("status" in error.lower() for error in result["errors"]))

    def test_a_duplicate_email_within_the_file_is_flagged_on_the_second_row(self):
        upload = make_import_workbook(
            [
                ["First", "Dup", "", "dup@example.com", "", "", "", "", ""],
                ["Second", "Dup", "", "dup@example.com", "", "", "", "", ""],
            ]
        )

        response = self.club_post("member_import", {"file": upload})

        results = response.context["results"]
        self.assertIsNotNone(results[0]["member"])
        self.assertIsNone(results[1]["member"])
        self.assertIn("Duplicate email in this file.", results[1]["errors"])

    def test_an_email_already_in_the_club_is_skipped(self):
        upload = make_import_workbook([["Ada", "Admin", "", self.admin_user.email, "", "", "", "", ""]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertIn("Already a member of this club.", result["errors"])

    def test_confirm_creates_only_the_valid_rows_with_defaults(self):
        upload = make_import_workbook(
            [
                ["Jamie", "Kid", "2010-01-01", "jamie.kid@example.com", "", "", "LIC-1", "", ""],
                ["", "Noname", "", "", "", "", "", "", ""],
            ]
        )
        self.club_post("member_import", {"file": upload})

        response = self.club_post("member_import_confirm", {})

        self.assertRedirects(response, reverse("management:member_list"))
        member = Member.objects.get(email="jamie.kid@example.com")
        membership = ClubMembership.objects.get(club=self.club, member=member)
        self.assertEqual(membership.status, ClubMembership.StatusChoices.ACTIVE)
        self.assertEqual(membership.fee_status, ClubMembership.FeeStatus.UNPAID)
        self.assertEqual(membership.license, "LIC-1")
        self.assertFalse(Member.objects.filter(last_name="Noname").exists())

    def test_confirm_without_a_prior_upload_creates_nothing(self):
        response = self.club_post("member_import_confirm", {})

        self.assertRedirects(response, reverse("management:member_import"))
        self.assertEqual(Member.objects.filter(last_name="Kid").count(), 0)

    def test_confirm_is_admin_only(self):
        coach_user = self.make_non_admin_staff()
        upload = make_import_workbook([["Jamie", "Kid", "", "jamie.kid2@example.com", "", "", "", "", ""]])
        self.club_post("member_import", {"file": upload})

        self.client.force_login(coach_user)
        response = self.club_post("member_import_confirm", {})

        self.assertEqual(response.status_code, 403)

    def test_a_family_group_links_a_parent_and_child_and_grants_the_parent_a_login(self):
        upload = make_import_workbook(
            [
                ["Taylor", "Doe", "", "taylor.doe@example.com", "", "", "", "", "", "Doe family", "parent"],
                ["Jamie", "Doe", "2014-03-02", "", "", "", "", "", "", "Doe family", "child"],
            ]
        )
        self.club_post("member_import", {"file": upload})

        self.club_post("member_import_confirm", {})

        parent = Member.objects.get(email="taylor.doe@example.com")
        child = Member.objects.get(first_name="Jamie", last_name="Doe")
        self.assertIsNotNone(parent.user_id)
        self.assertTrue(User.objects.filter(email="taylor.doe@example.com").exists())
        self.assertIsNone(child.user_id)
        family = Family.objects.get(memberships__member=parent)
        self.assertEqual(family, Family.objects.get(memberships__member=child))
        self.assertEqual(FamilyMembership.objects.get(family=family, member=parent).role, FamilyMembership.FamilyRole.PARENT)
        self.assertEqual(FamilyMembership.objects.get(family=family, member=child).role, FamilyMembership.FamilyRole.CHILD)

    def test_membership_kind_guardian_creates_a_guardian_not_a_member(self):
        # Columns are positional; membership_kind is the last one.
        upload = make_import_workbook(
            [
                ["Taylor", "Doe", "", "taylor.guardian@example.com", "", "", "", "", "", "Doe family", "parent", "guardian"],
                ["Jamie", "Doe", "2014-03-02", "", "", "", "", "", "", "Doe family", "child", "member"],
            ]
        )
        self.club_post("member_import", {"file": upload})

        self.club_post("member_import_confirm", {})

        parent = Member.objects.get(email="taylor.guardian@example.com")
        child = Member.objects.get(first_name="Jamie", last_name="Doe")
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=parent).kind, ClubMembership.Kind.GUARDIAN)
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=child).kind, ClubMembership.Kind.MEMBER)

    def test_a_blank_membership_kind_still_means_member(self):
        # Every file written before the column existed carried this implicitly.
        upload = make_import_workbook([["Solo", "Blankkind", "", "solo.blank@example.com", "", "", "", "", "", "", "", ""]])
        self.club_post("member_import", {"file": upload})

        self.club_post("member_import_confirm", {})

        member = Member.objects.get(email="solo.blank@example.com")
        self.assertEqual(ClubMembership.objects.get(club=self.club, member=member).kind, ClubMembership.Kind.MEMBER)

    def test_a_lone_child_row_gets_a_family_of_its_own(self):
        # The migration case: children arrive with no parents on file. A family of
        # one is what makes "nobody responsible for this child" visible -- see
        # members.services.claims.families_awaiting_a_parent.
        upload = make_import_workbook([["Jamie", "Lonechild", "2014-03-02", "", "", "", "", "", "", "", "child", ""]])
        self.club_post("member_import", {"file": upload})

        self.club_post("member_import_confirm", {})

        child = Member.objects.get(first_name="Jamie", last_name="Lonechild")
        self.assertIn(child, children_awaiting_a_parent(self.club))

    def test_a_lone_parent_row_is_still_an_error(self):
        # Only `child` is meaningful without a family_group; a parent with nobody
        # to be a parent *of* is a mistake in the file.
        upload = make_import_workbook([["Odd", "Loneparent", "", "odd.lone@example.com", "", "", "", "", "", "", "parent", ""]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(any("family_group" in error for error in result["errors"]))

    def test_a_child_marked_as_a_guardian_is_an_error(self):
        # A child is the member the guardian is attached *to*.
        upload = make_import_workbook([["Jamie", "Doe", "2014-03-02", "", "", "", "", "", "", "Doe family", "child", "guardian"]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(any("child is always a member" in error.lower() for error in result["errors"]))

    def test_an_invalid_membership_kind_is_reported(self):
        upload = make_import_workbook([["Odd", "Kind", "", "odd.kind@example.com", "", "", "", "", "", "", "", "sponsor"]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(any("membership_kind" in error for error in result["errors"]))

    def test_family_role_without_a_group_is_an_error(self):
        upload = make_import_workbook([["Odd", "Row", "", "odd.row@example.com", "", "", "", "", "", "", "parent"]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(any("family_group" in error.lower() for error in result["errors"]))

    def test_family_group_without_a_role_is_an_error(self):
        upload = make_import_workbook([["Odd", "Row", "", "odd.row2@example.com", "", "", "", "", "", "Odd family", ""]])

        response = self.club_post("member_import", {"file": upload})

        result = response.context["results"][0]
        self.assertIsNone(result["member"])
        self.assertTrue(any("family_role" in error.lower() for error in result["errors"]))

    def test_a_standalone_row_is_not_linked_to_any_family(self):
        upload = make_import_workbook([["Solo", "Standalone", "", "solo@example.com", "", "", "", "", "", "", ""]])
        self.club_post("member_import", {"file": upload})

        self.club_post("member_import_confirm", {})

        member = Member.objects.get(email="solo@example.com")
        self.assertFalse(FamilyMembership.objects.filter(member=member).exists())


class NewsManagementTests(ManagementTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")

        # The three actors this class draws its lines between: a coach with a
        # management position (may draft, may not publish), a plain staff member
        # (may not touch news at all), and an EDITOR (may publish and may keep
        # editing afterwards).
        cls.coach_manager = User.objects.create_user(email="coach-news@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=cls.coach_manager, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=cls.team, member=coach_member, season=cls.season, position=coach_position)

        cls.plain_staff = User.objects.create_user(email="physio-news@example.com", password="pw-secret-123")
        staff_member = Member.objects.create(user=cls.plain_staff, first_name="Pat", last_name="Physio")
        staff_position = Position.objects.create(club=cls.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=cls.team, member=staff_member, season=cls.season, position=staff_position)

        cls.editor = User.objects.create_user(email="editor-news@example.com", password="pw-secret-123")
        editor_member = Member.objects.create(user=cls.editor, first_name="Eve", last_name="Editor")
        ClubMembership.objects.create(club=cls.club, member=editor_member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=cls.club, member=editor_member).update(role=ClubRole.Roles.EDITOR)
        enrol_mfa(cls.editor)  # ClubRole ADMIN/EDITOR requires a second factor; StaffAssignment-only doesn't.

    def test_list_is_scoped_to_the_club(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        News.objects.create(club=other_club, title="Rival news", body="Body.")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_list")

        self.assertNotContains(response, "Rival news")

    def test_the_status_filter_chips_count_each_bucket(self):
        News.objects.create(club=self.club, title="Draft one", body="Body.")
        News.objects.create(club=self.club, title="Draft two", body="Body.")
        News.objects.create(club=self.club, title="Live", body="Body.", status=News.Status.PUBLISHED, published_at=timezone.now())
        News.objects.create(club=self.club, title="Upcoming", body="Body.", status=News.Status.PUBLISHED, published_at=timezone.now() + datetime.timedelta(days=3))
        self.client.force_login(self.admin_user)

        response = self.club_get("news_list")

        self.assertEqual(response.context["counts"], {"all": 4, "draft": 2, "scheduled": 1, "published": 1})

    def test_the_draft_filter_chip_narrows_the_list_to_drafts(self):
        News.objects.create(club=self.club, title="A draft", body="Body.")
        News.objects.create(club=self.club, title="Already live", body="Body.", status=News.Status.PUBLISHED, published_at=timezone.now())
        self.client.force_login(self.admin_user)

        response = self.club_get("news_list", params={"status": "draft"})

        self.assertContains(response, "A draft")
        self.assertNotContains(response, "Already live")

    def test_selecting_an_item_previews_it_on_the_right(self):
        News.objects.create(club=self.club, title="First post", body="First body.")
        second = News.objects.create(club=self.club, title="Second post", body="Second body.")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_list", params={"selected": str(second.pk)})

        self.assertEqual(response.context["news_item"], second)
        self.assertContains(response, "Second body.")
        self.assertNotContains(response, "First body.")

    def test_the_preview_renders_markdown_as_html_not_raw_source(self):
        News.objects.create(club=self.club, title="Formatted post", body="**Bold** and a [link](https://example.com).")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_list")

        self.assertContains(response, "<strong>Bold</strong>")
        self.assertContains(response, 'href="https://example.com"')
        self.assertNotContains(response, "**Bold**")

    def test_the_detail_page_also_renders_markdown_as_html(self):
        item = News.objects.create(club=self.club, title="Formatted post", body="**Bold** text.")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_detail", item.pk)

        self.assertContains(response, "<strong>Bold</strong>")

    def test_the_first_item_previews_by_default_when_none_is_selected(self):
        # News.Meta.ordering is "-created", so the most recently created row
        # (second) sorts first and is what should preview with no ?selected=.
        News.objects.create(club=self.club, title="First post", body="First body.")
        second = News.objects.create(club=self.club, title="Second post", body="Second body.")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_list")

        self.assertEqual(response.context["news_item"], second)

    def test_a_coach_manager_can_create_a_draft(self):
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_create", {"title": "Season Kickoff", "body": "Big news.", "visibility": News.Visibility.INTERNAL, "teams": [str(self.team.pk)]})

        item = News.objects.get(club=self.club, title="Season Kickoff")
        self.assertRedirects(response, reverse("management:news_detail", args=[item.pk]))
        self.assertEqual(item.status, News.Status.DRAFT)

    def test_an_english_translation_can_be_added_alongside_the_original(self):
        self.client.force_login(self.coach_manager)

        self.club_post(
            "news_create",
            {"title": "Seizoensstart", "title_en": "Season kickoff", "body": "We beginnen het seizoen.", "body_en": "We're starting the season.", "visibility": News.Visibility.INTERNAL, "teams": [str(self.team.pk)]},
        )

        item = News.objects.get(club=self.club, title="Seizoensstart")
        self.assertEqual(item.title_en, "Season kickoff")
        self.assertEqual(item.body_en, "We're starting the season.")

    def test_the_english_translation_is_optional(self):
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_create", {"title": "Seizoensstart", "body": "We beginnen het seizoen.", "visibility": News.Visibility.INTERNAL, "teams": [str(self.team.pk)]})

        item = News.objects.get(club=self.club, title="Seizoensstart")
        self.assertRedirects(response, reverse("management:news_detail", args=[item.pk]))
        self.assertEqual(item.title_en, "")

    def test_detail_page_shows_the_english_translation_when_set(self):
        item = News.objects.create(club=self.club, title="Seizoensstart", body="We beginnen het seizoen.", title_en="Season kickoff", body_en="We're starting the season.")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_detail", item.pk)

        self.assertContains(response, "Season kickoff")
        # Rendered through news.services.render_body_html (Markdown -> sanitised
        # HTML), not Django's own auto-escaping -- a plain apostrophe, not &#x27;.
        self.assertContains(response, "We're starting the season.")

    def test_detail_page_notes_a_missing_english_translation(self):
        # The NL/EN toggle is always present now (not conditional on a
        # translation existing) -- untranslated items instead get a fallback
        # note, and the EN pane falls back to showing the Dutch content
        # (News.effective_title_en/effective_body_en) rather than sitting empty.
        item = News.objects.create(club=self.club, title="Seizoensstart", body="We beginnen het seizoen.")
        self.client.force_login(self.admin_user)

        response = self.club_get("news_detail", item.pk)

        self.assertContains(response, "No English translation yet")
        self.assertContains(response, 'data-lang-content="en"')

    def test_plain_staff_cannot_create_news(self):
        self.client.force_login(self.plain_staff)

        response = self.club_post("news_create", {"title": "Not allowed", "body": "Body.", "visibility": News.Visibility.INTERNAL})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(News.objects.filter(club=self.club, title="Not allowed").exists())

    def test_coach_manager_cannot_publish(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_publish", {"published_at": "2026-08-10T10:00"}, item.pk)

        self.assertEqual(response.status_code, 403)
        item.refresh_from_db()
        self.assertEqual(item.status, News.Status.DRAFT)

    def test_editor_can_publish(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.editor)

        self.club_post("news_publish", {"published_at": "2026-08-10T10:00"}, item.pk)

        item.refresh_from_db()
        self.assertEqual(item.status, News.Status.PUBLISHED)

    def test_publishing_with_a_future_date_leaves_it_scheduled(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.editor)
        future = timezone.now() + datetime.timedelta(days=7)

        self.club_post("news_publish", {"published_at": future.strftime("%Y-%m-%dT%H:%M")}, item.pk)

        item.refresh_from_db()
        self.assertTrue(item.is_scheduled)

    def test_publishing_with_now_makes_it_live_immediately(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.editor)

        self.club_post("news_publish", {"published_at": timezone.now().strftime("%Y-%m-%dT%H:%M")}, item.pk)

        item.refresh_from_db()
        self.assertFalse(item.is_scheduled)

    def test_notify_members_checkbox_emails_the_audience(self):
        member = Member.objects.create(first_name="Jamie", last_name="Doe", email="jamie@example.com")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        User.objects.create_user(email="jamie@example.com", password="pw-secret-123")
        member.user = User.objects.get(email="jamie@example.com")
        member.save(update_fields=["user"])
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.editor)

        # dispatch_send_publish_notification fires on a real background thread now
        # (see news/services.py) -- run it inline here instead, via side_effect, so
        # the email is guaranteed sent by the time this assertion runs rather than
        # racing a thread that may not have finished yet.
        with mock.patch("management.views.dispatch_send_publish_notification", side_effect=_send_and_mark_notified):
            self.club_post("news_publish", {"published_at": timezone.now().strftime("%Y-%m-%dT%H:%M"), "notify_members": "on"}, item.pk)

        # Not asserting an exact outbox size: the base fixture's own staff
        # members (self.editor etc.) may also be active MEMBER-kind club
        # members in this same season, so they legitimately get one too.
        sent_to = [address for message in mail.outbox for address in message.to]
        self.assertIn("jamie@example.com", sent_to)

    def test_leaving_notify_members_unchecked_sends_nothing(self):
        member = Member.objects.create(first_name="Jamie", last_name="Doe", email="jamie@example.com")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.editor)

        self.club_post("news_publish", {"published_at": timezone.now().strftime("%Y-%m-%dT%H:%M")}, item.pk)

        self.assertEqual(len(mail.outbox), 0)

    def test_notify_members_checkbox_creates_an_in_app_notification_and_pushes_it(self):
        # End-to-end: the publish-time toggle all the way through to the mobile
        # app's own inbox and its push channel -- not just the email side
        # test_notify_members_checkbox_emails_the_audience above already covers.
        from mobile.models import PushSubscription

        member_user = User.objects.create_user(email="jamie@example.com", password="pw-secret-123")
        member = Member.objects.create(first_name="Jamie", last_name="Doe", email="jamie@example.com", user=member_user)
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        PushSubscription.objects.create(club=self.club, member=member, endpoint="https://push.example.com/jamie", p256dh="key1", auth="key2")
        item = News.objects.create(club=self.club, title="Signed: New Player", body="Body.")
        self.client.force_login(self.editor)

        # See test_notify_members_checkbox_emails_the_audience above for why this patches
        # dispatch_send_publish_notification with a synchronous side_effect.
        with (
            override_settings(VAPID_PRIVATE_KEY="test-private-key", VAPID_ADMIN_EMAIL="admin@example.com"),
            mock.patch("mobile.services.push.webpush") as webpush_mock,
            mock.patch("management.views.dispatch_send_publish_notification", side_effect=_send_and_mark_notified),
        ):
            self.club_post("news_publish", {"published_at": timezone.now().strftime("%Y-%m-%dT%H:%M"), "notify_members": "on"}, item.pk)

        notification = Notification.objects.get(member=member, title="Signed: New Player")
        self.assertIsNotNone(notification.sent_at)
        webpush_mock.assert_called_once()
        self.assertEqual(json.loads(webpush_mock.call_args.kwargs["data"])["title"], "Signed: New Player")

        # And it actually shows up in the member's own mobile inbox, not just the DB row.
        self.client.logout()
        self.client.force_login(member_user)
        response = self.client.get(reverse("mobile:notifications"), HTTP_HOST="ajax-united.rosterchief.app")
        self.assertContains(response, "Signed: New Player")
        self.assertEqual(response.context["unread_notification_count"], 1)

    def test_a_coach_manager_can_submit_a_draft_for_review(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.coach_manager)

        self.club_post("news_submit_for_review", {}, item.pk)

        item.refresh_from_db()
        self.assertEqual(item.status, News.Status.PENDING_REVIEW)

    def test_submitting_for_review_notifies_editors_in_app_only(self):
        editor_member = Member.objects.get(user=self.editor)
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.coach_manager)

        self.club_post("news_submit_for_review", {}, item.pk)

        self.assertTrue(Notification.objects.filter(member=editor_member, title__contains="Draft item").exists())
        self.assertEqual(len(mail.outbox), 0)

    def test_plain_staff_cannot_submit_for_review(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.plain_staff)

        response = self.club_post("news_submit_for_review", {}, item.pk)

        self.assertEqual(response.status_code, 403)
        item.refresh_from_db()
        self.assertEqual(item.status, News.Status.DRAFT)

    def test_an_editor_still_publishes_directly_without_a_review_step(self):
        # The button _news_preview.html shows for can_publish is Publish, never
        # Send for review -- editors skip the queue entirely.
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.editor)

        response = self.club_get("news_detail", item.pk)

        self.assertContains(response, reverse("management:news_publish", args=[item.pk]))
        self.assertNotContains(response, reverse("management:news_submit_for_review", args=[item.pk]))

    def test_unpublishing_reverts_to_draft(self):
        item = News.objects.create(club=self.club, title="Live item", body="Body.")
        item.publish()
        self.client.force_login(self.editor)

        self.club_post("news_unpublish", {}, item.pk)

        item.refresh_from_db()
        self.assertEqual(item.status, News.Status.DRAFT)
        self.assertIsNone(item.published_at)

    def test_a_coach_manager_can_edit_someone_elses_draft(self):
        item = News.objects.create(club=self.club, title="Old title", body="Body.")
        self.client.force_login(self.coach_manager)

        self.club_post("news_update", {"title": "New title", "body": "Body.", "visibility": News.Visibility.INTERNAL}, item.pk)

        item.refresh_from_db()
        self.assertEqual(item.title, "New title")

    def test_a_coach_manager_cannot_edit_once_published(self):
        item = News.objects.create(club=self.club, title="Old title", body="Body.")
        item.publish()
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_update", {"title": "New title", "body": "Body.", "visibility": News.Visibility.INTERNAL}, item.pk)

        self.assertEqual(response.status_code, 403)

    def test_an_editor_can_still_edit_once_published(self):
        item = News.objects.create(club=self.club, title="Old title", body="Body.")
        item.publish()
        self.client.force_login(self.editor)

        self.club_post("news_update", {"title": "New title", "body": "Body.", "visibility": News.Visibility.INTERNAL}, item.pk)

        item.refresh_from_db()
        self.assertEqual(item.title, "New title")

    def test_uploading_multiple_photos_creates_one_per_file_and_marks_the_first_main(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        self.client.force_login(self.coach_manager)
        images = [
            SimpleUploadedFile("one.jpg", b"fake-bytes-one", content_type="image/jpeg"),
            SimpleUploadedFile("two.jpg", b"fake-bytes-two", content_type="image/jpeg"),
        ]

        self.club_post("news_photo_upload", {"images": images}, item.pk)

        self.assertEqual(item.photos.count(), 2)
        self.assertEqual(item.photos.filter(is_main=True).count(), 1)

    def test_set_main_moves_the_main_flag(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        first = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"), is_main=True)
        second = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("two.jpg", b"two", content_type="image/jpeg"), is_main=False)
        self.client.force_login(self.coach_manager)

        self.club_post("news_photo_set_main", {}, item.pk, second.pk)

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertFalse(first.is_main)
        self.assertTrue(second.is_main)

    def test_setting_the_focal_point_stores_it(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        photo = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"))
        self.client.force_login(self.coach_manager)

        self.club_post("news_photo_set_focal_point", {"focal_x": "20", "focal_y": "80"}, item.pk, photo.pk)

        photo.refresh_from_db()
        self.assertEqual(photo.focal_x, 20)
        self.assertEqual(photo.focal_y, 80)

    def test_the_focal_point_is_clamped_to_0_100(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        photo = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"))
        self.client.force_login(self.coach_manager)

        self.club_post("news_photo_set_focal_point", {"focal_x": "-5", "focal_y": "500"}, item.pk, photo.pk)

        photo.refresh_from_db()
        self.assertEqual(photo.focal_x, 0)
        self.assertEqual(photo.focal_y, 100)

    def test_a_non_numeric_focal_point_is_rejected(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        photo = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"))
        self.client.force_login(self.coach_manager)

        self.club_post("news_photo_set_focal_point", {"focal_x": "nope", "focal_y": "50"}, item.pk, photo.pk)

        photo.refresh_from_db()
        self.assertEqual(photo.focal_x, 50)  # default, untouched

    def test_deleting_a_photo_removes_it(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        photo = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"))
        photo_path = photo.image.path
        self.client.force_login(self.coach_manager)

        self.club_post("news_photo_delete", {}, item.pk, photo.pk)

        self.assertFalse(NewsPhoto.objects.filter(pk=photo.pk).exists())
        self.assertFalse(os.path.exists(photo_path))

    def test_deleting_the_main_photo_promotes_another_one(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        main = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"), is_main=True)
        other = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("two.jpg", b"two", content_type="image/jpeg"), is_main=False)
        self.client.force_login(self.coach_manager)

        self.club_post("news_photo_delete", {}, item.pk, main.pk)

        other.refresh_from_db()
        self.assertTrue(other.is_main)

    def test_deleting_the_only_photo_leaves_nothing_to_promote(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        photo = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"), is_main=True)
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_photo_delete", {}, item.pk, photo.pk)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(item.photos.count(), 0)

    def test_a_coach_manager_can_delete_a_draft(self):
        item = News.objects.create(club=self.club, title="Draft item", body="Body.")
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_delete", {}, item.pk)

        self.assertRedirects(response, reverse("management:news_list"))
        self.assertFalse(News.objects.filter(pk=item.pk).exists())

    def test_a_coach_manager_cannot_delete_once_published(self):
        item = News.objects.create(club=self.club, title="Live item", body="Body.")
        item.publish()
        self.client.force_login(self.coach_manager)

        response = self.club_post("news_delete", {}, item.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(News.objects.filter(pk=item.pk).exists())

    def test_an_editor_can_delete_once_published(self):
        item = News.objects.create(club=self.club, title="Live item", body="Body.")
        item.publish()
        self.client.force_login(self.editor)

        self.club_post("news_delete", {}, item.pk)

        self.assertFalse(News.objects.filter(pk=item.pk).exists())

    def test_deleting_a_news_item_removes_its_photos(self):
        item = News.objects.create(club=self.club, title="Match report", body="Body.")
        photo = NewsPhoto.objects.create(news_item=item, image=SimpleUploadedFile("one.jpg", b"one", content_type="image/jpeg"))
        photo_path = photo.image.path
        self.client.force_login(self.coach_manager)

        self.club_post("news_delete", {}, item.pk)

        self.assertFalse(NewsPhoto.objects.filter(pk=photo.pk).exists())
        self.assertFalse(os.path.exists(photo_path))

    def test_the_edit_and_delete_buttons_are_hidden_once_published_for_a_coach_manager(self):
        item = News.objects.create(club=self.club, title="Live item", body="Body.")
        item.publish()
        self.client.force_login(self.coach_manager)

        response = self.club_get("news_list")

        self.assertNotContains(response, reverse("management:news_update", args=[item.pk]))


class TeamAttendancePanelTests(ManagementTestBase):
    """The attendance KPI panel on the team page -- see
    management.views.TeamDetailView and events.services.attendance."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.position = Position.objects.create(club=cls.club, name="Forward", short_name="FW", staff_position=False)
        cls.player = Member.objects.create(first_name="Peter", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        TeamMembership.objects.create(team=cls.team, season=cls.season, member=cls.player, position=cls.position)

    def make_past_training(self, days_ago=1):
        event = Event.objects.create(club=self.club, title="Practice", kind=Event.EventKind.TRAINING, season=self.season, start=timezone.now() - datetime.timedelta(days=days_ago))
        event.teams.add(self.team)
        return event

    def test_attendance_panel_shows_the_rate_and_rankings(self):
        event = self.make_past_training()
        Attendance.objects.create(event=event, member=self.player, status=Attendance.AttendanceStatus.PRESENT)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Attendance rate")
        self.assertContains(response, "Peter Player")

    def test_a_present_rsvp_without_a_check_in_is_never_a_no_show(self):
        event = self.make_past_training()
        Attendance.objects.create(event=event, member=self.player, status=Attendance.AttendanceStatus.PRESENT)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "None recorded.")

    def test_a_checked_in_no_show_appears_in_the_panel(self):
        event = self.make_past_training()
        attendance = Attendance.objects.create(event=event, member=self.player, status=Attendance.AttendanceStatus.PRESENT, showed_up=False)
        self.client.force_login(self.admin_user)

        response = self.club_get("team_detail", self.team.pk)

        self.assertContains(response, "Peter Player")
        self.assertContains(response, attendance.event.title)
        self.assertNotContains(response, "None recorded.")


class TeamAndPositionAccessTests(ManagementTestBase):
    """Non-admin coaches/managers: scoped to their own teams, read-only on
    positions -- see club.mixins.TeamManagerRequiredMixin and
    management.views.TeamListView/PositionListView."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.own_team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.manager_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        cls.coach_user = User.objects.create_user(email="coach3@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=cls.coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=cls.own_team, member=coach_member, season=cls.season, position=cls.manager_position)

    def test_a_coach_only_sees_their_own_team_in_the_list(self):
        self.client.force_login(self.coach_user)

        response = self.club_get("team_list")

        self.assertContains(response, "First Team")
        self.assertNotContains(response, "Second Team")

    def test_an_admin_sees_every_team_in_the_list(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("team_list")

        self.assertContains(response, "First Team")
        self.assertContains(response, "Second Team")

    def test_a_coach_does_not_see_the_new_team_button(self):
        self.client.force_login(self.coach_user)

        response = self.club_get("team_list")

        self.assertNotContains(response, reverse("management:team_create"))

    def test_a_coach_does_not_see_the_edit_button_on_their_team_page(self):
        self.client.force_login(self.coach_user)

        response = self.club_get("team_detail", self.own_team.pk)

        self.assertNotContains(response, reverse("management:team_update", args=[self.own_team.pk]))

    def test_a_coach_can_view_positions_but_not_edit_them(self):
        self.client.force_login(self.coach_user)

        response = self.club_get("position_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Head Coach")
        self.assertNotContains(response, reverse("management:position_create"))
        self.assertNotContains(response, reverse("management:position_update", args=[self.manager_position.pk]))

    def test_a_coach_cannot_create_or_edit_a_position(self):
        self.client.force_login(self.coach_user)

        self.assertEqual(self.club_get("position_create").status_code, 403)
        self.assertEqual(self.club_get("position_update", self.manager_position.pk).status_code, 403)


class TeamListCountsTests(ManagementTestBase):
    """Player/staff counts on the team list -- see TeamListView.get_queryset."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.player_position = Position.objects.create(club=cls.club, name="Forward", short_name="FW", staff_position=False)
        cls.coach_position = Position.objects.create(club=cls.club, name="Coach", short_name="C", staff_position=True, management_position=True)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_counts_reflect_the_current_seasons_roster_and_staff(self):
        peter = Member.objects.create(first_name="Peter", last_name="Player")
        paula = Member.objects.create(first_name="Paula", last_name="Player")
        cara = Member.objects.create(first_name="Cara", last_name="Coach")
        TeamMembership.objects.create(team=self.team, season=self.season, member=peter, position=self.player_position)
        TeamMembership.objects.create(team=self.team, season=self.season, member=paula, position=self.player_position)
        StaffAssignment.objects.create(team=self.team, season=self.season, member=cara, position=self.coach_position)

        response = self.club_get("team_list")

        # Not .get(pk=...): the team list is now paginated, and a Page's
        # object_list is a sliced queryset -- Django refuses to .filter()/.get()
        # a queryset once it's been sliced.
        team = next(t for t in response.context["teams"] if t.pk == self.team.pk)
        self.assertEqual(team.player_count, 2)
        self.assertEqual(team.staff_count, 1)

    def test_counts_exclude_a_different_season(self):
        other_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 1, 1), end_date=datetime.date(2020, 12, 31))
        peter = Member.objects.create(first_name="Peter", last_name="Player")
        TeamMembership.objects.create(team=self.team, season=other_season, member=peter, position=self.player_position)

        response = self.club_get("team_list")

        team = next(t for t in response.context["teams"] if t.pk == self.team.pk)
        self.assertEqual(team.player_count, 0)


class LocationOpponentManagementTests(ManagementTestBase):
    """Full CRUD for Location/Opponent -- restricted to ADMIN and anyone with a
    current-season management position, see club.mixins.ManagementPositionRequiredMixin
    and management.views.LocationListView/OpponentListView (and their Create/Update/Delete
    siblings)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")

        # The two actors either side of the line this class draws: a management
        # position (allowed) and a plain staff position (refused).
        cls.coach_manager = User.objects.create_user(email="coach-loc@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=cls.coach_manager, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=cls.team, member=coach_member, season=cls.season, position=coach_position)

        cls.plain_staff = User.objects.create_user(email="physio-loc@example.com", password="pw-secret-123")
        staff_member = Member.objects.create(user=cls.plain_staff, first_name="Pat", last_name="Physio")
        staff_position = Position.objects.create(club=cls.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=cls.team, member=staff_member, season=cls.season, position=staff_position)

    # --- Locations ---------------------------------------------------------

    def test_a_management_position_can_view_the_location_list(self):
        Location.objects.create(club=self.club, name="Main Field", address="1 St", city="Town", zip_code="1000", country="BE")
        self.client.force_login(self.coach_manager)

        response = self.club_get("location_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Main Field")

    def test_the_location_form_renders_country_as_a_dropdown(self):
        # Regression: CountryField's widget reports widget_type "lazyselect", which
        # the form_field templatetag didn't recognise -- it fell through to the
        # "input" case and rendered a plain <input type="lazyselect"> (i.e. a
        # broken text box), not a <select>.
        self.client.force_login(self.coach_manager)

        response = self.club_get("location_create")

        self.assertNotContains(response, 'type="lazyselect"')
        self.assertContains(response, "Belgium")
        self.assertContains(response, "<select")

    def test_plain_staff_cannot_view_the_location_list(self):
        self.client.force_login(self.plain_staff)

        self.assertEqual(self.club_get("location_list").status_code, 403)

    def test_a_management_position_can_create_a_location(self):
        self.client.force_login(self.coach_manager)

        response = self.club_post("location_create", {"name": "New Field", "address": "2 St", "city": "Town", "zip_code": "1000", "country": "BE"})

        self.assertRedirects(response, reverse("management:location_list"))
        self.assertTrue(Location.objects.filter(club=self.club, name="New Field").exists())

    def test_plain_staff_cannot_create_a_location(self):
        self.client.force_login(self.plain_staff)

        response = self.club_post("location_create", {"name": "New Field", "address": "2 St", "city": "Town", "zip_code": "1000", "country": "BE"})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Location.objects.filter(club=self.club, name="New Field").exists())

    def test_a_management_position_can_edit_a_location(self):
        location = Location.objects.create(club=self.club, name="Old name", address="1 St", city="Town", zip_code="1000", country="BE")
        self.client.force_login(self.coach_manager)

        self.club_post("location_update", {"name": "New name", "address": "1 St", "city": "Town", "zip_code": "1000", "country": "BE"}, location.pk)

        location.refresh_from_db()
        self.assertEqual(location.name, "New name")

    def test_a_management_position_can_delete_a_location(self):
        location = Location.objects.create(club=self.club, name="Doomed", address="1 St", city="Town", zip_code="1000", country="BE")
        self.client.force_login(self.coach_manager)

        response = self.club_post("location_delete", {}, location.pk)

        self.assertRedirects(response, reverse("management:location_list"))
        self.assertFalse(Location.objects.filter(pk=location.pk).exists())

    def test_deleting_a_location_nulls_it_on_events_instead_of_erroring(self):
        location = Location.objects.create(club=self.club, name="Doomed", address="1 St", city="Town", zip_code="1000", country="BE")
        event = Event.objects.create(club=self.club, title="Match", start=timezone.now() + datetime.timedelta(days=1), location=location)
        self.client.force_login(self.admin_user)

        self.club_post("location_delete", {}, location.pk)

        event.refresh_from_db()
        self.assertIsNone(event.location)

    def test_an_admin_has_full_rights_without_any_staff_assignment(self):
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("location_list").status_code, 200)
        response = self.club_post("location_create", {"name": "Admin Field", "address": "3 St", "city": "Town", "zip_code": "1000", "country": "BE"})
        self.assertRedirects(response, reverse("management:location_list"))

    # --- Opponents -----------------------------------------------------------

    def test_a_management_position_can_view_the_opponent_list(self):
        Opponent.objects.create(club=self.club, name="Rivals FC")
        self.client.force_login(self.coach_manager)

        response = self.club_get("opponent_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Rivals FC")

    def test_plain_staff_cannot_view_the_opponent_list(self):
        self.client.force_login(self.plain_staff)

        self.assertEqual(self.club_get("opponent_list").status_code, 403)

    def test_a_management_position_can_create_an_opponent_with_a_logo(self):
        self.client.force_login(self.coach_manager)
        # Opponent.logo is a real ImageField (unlike NewsPhoto.image, set outside any
        # ModelForm) -- Django's ImageField.clean() runs it through Pillow, so this
        # needs to actually decode as an image, not just carry an image/png header.
        one_pixel_png = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82"
        logo = SimpleUploadedFile("logo.png", one_pixel_png, content_type="image/png")

        response = self.club_post("opponent_create", {"name": "Rivals FC", "logo": logo})

        self.assertRedirects(response, reverse("management:opponent_list"))
        opponent = Opponent.objects.get(club=self.club, name="Rivals FC")
        self.assertTrue(opponent.logo)

    def test_a_management_position_can_delete_an_opponent(self):
        opponent = Opponent.objects.create(club=self.club, name="Doomed FC")
        self.client.force_login(self.coach_manager)

        response = self.club_post("opponent_delete", {}, opponent.pk)

        self.assertRedirects(response, reverse("management:opponent_list"))
        self.assertFalse(Opponent.objects.filter(pk=opponent.pk).exists())

    def test_plain_staff_cannot_delete_an_opponent(self):
        opponent = Opponent.objects.create(club=self.club, name="Safe FC")
        self.client.force_login(self.plain_staff)

        response = self.club_post("opponent_delete", {}, opponent.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Opponent.objects.filter(pk=opponent.pk).exists())


class SponsorManagementTests(ManagementTestBase):
    """Full CRUD for Sponsor -- admin-only, unlike Location/Opponent which any
    management position can maintain (see club.mixins.ClubAdminRequiredMixin
    and management.views.Sponsor*View)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")

        # Neither of these may touch sponsors -- a management position is enough
        # for Location/Opponent but not here, and plain staff never qualifies.
        cls.coach_manager = User.objects.create_user(email="coach-sponsor@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=cls.coach_manager, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=cls.team, member=coach_member, season=cls.season, position=coach_position)

        cls.plain_staff = User.objects.create_user(email="physio-sponsor@example.com", password="pw-secret-123")
        staff_member = Member.objects.create(user=cls.plain_staff, first_name="Pat", last_name="Physio")
        staff_position = Position.objects.create(club=cls.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=cls.team, member=staff_member, season=cls.season, position=staff_position)

    def sponsor_data(self, **overrides):
        data = {"name": "Acme Corp", "url": "https://acme.example.com", "start_date": "2026-01-01", "end_date": ""}
        data.update(overrides)
        return data

    def test_an_admin_can_view_the_sponsor_list(self):
        Sponsor.objects.create(club=self.club, name="Acme Corp", start_date=datetime.date(2026, 1, 1))
        self.client.force_login(self.admin_user)

        response = self.club_get("sponsor_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Acme Corp")

    def test_a_management_position_cannot_view_the_sponsor_list(self):
        self.client.force_login(self.coach_manager)

        self.assertEqual(self.club_get("sponsor_list").status_code, 403)

    def test_plain_staff_cannot_view_the_sponsor_list(self):
        self.client.force_login(self.plain_staff)

        self.assertEqual(self.club_get("sponsor_list").status_code, 403)

    def test_an_admin_can_create_a_sponsor_with_a_logo(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("sponsor_create", self.sponsor_data(logo=make_image_file()))

        self.assertRedirects(response, reverse("management:sponsor_list"))
        sponsor = Sponsor.objects.get(club=self.club, name="Acme Corp")
        self.assertTrue(sponsor.logo)

    def test_a_management_position_cannot_create_a_sponsor(self):
        self.client.force_login(self.coach_manager)

        response = self.club_post("sponsor_create", self.sponsor_data())

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Sponsor.objects.filter(club=self.club, name="Acme Corp").exists())

    def test_an_end_date_before_the_start_date_is_a_form_error_not_a_500(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("sponsor_create", self.sponsor_data(start_date="2026-06-01", end_date="2026-01-01"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "before the start date")
        self.assertFalse(Sponsor.objects.filter(club=self.club).exists())

    def test_an_admin_can_edit_a_sponsor(self):
        sponsor = Sponsor.objects.create(club=self.club, name="Acme Corp", start_date=datetime.date(2026, 1, 1))
        self.client.force_login(self.admin_user)

        response = self.club_post("sponsor_update", self.sponsor_data(name="Acme Corp Renamed"), sponsor.pk)

        self.assertRedirects(response, reverse("management:sponsor_list"))
        sponsor.refresh_from_db()
        self.assertEqual(sponsor.name, "Acme Corp Renamed")

    def test_plain_staff_cannot_edit_a_sponsor(self):
        sponsor = Sponsor.objects.create(club=self.club, name="Acme Corp", start_date=datetime.date(2026, 1, 1))
        self.client.force_login(self.plain_staff)

        response = self.club_post("sponsor_update", self.sponsor_data(), sponsor.pk)

        self.assertEqual(response.status_code, 403)

    def test_an_admin_can_delete_a_sponsor(self):
        sponsor = Sponsor.objects.create(club=self.club, name="Doomed Corp", start_date=datetime.date(2026, 1, 1))
        self.client.force_login(self.admin_user)

        response = self.club_post("sponsor_delete", {}, sponsor.pk)

        self.assertRedirects(response, reverse("management:sponsor_list"))
        self.assertFalse(Sponsor.objects.filter(pk=sponsor.pk).exists())

    def test_plain_staff_cannot_delete_a_sponsor(self):
        sponsor = Sponsor.objects.create(club=self.club, name="Safe Corp", start_date=datetime.date(2026, 1, 1))
        self.client.force_login(self.plain_staff)

        response = self.club_post("sponsor_delete", {}, sponsor.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Sponsor.objects.filter(pk=sponsor.pk).exists())

    def test_nav_only_shows_sponsors_for_an_admin(self):
        # Sponsors lives under Finance's sub-nav now, which only expands on a
        # Finance-section page -- see ActiveNavHighlightTests's sibling comments.
        # Finance itself is admin-only, so a coach never reaches a page that would
        # expand it in the first place -- club_get raises PermissionDenied for them.
        self.client.force_login(self.admin_user)
        self.assertContains(self.club_get("membership_list"), "Sponsors")

        self.client.force_login(self.coach_manager)
        self.assertNotContains(self.club_get("home"), "Sponsors")


class ClubSettingsPreviewTests(ManagementTestBase):
    """The D9-alike live preview on the Club identity page -- see
    management/templates/management/club_settings.html. The Identity tab's
    own mock (phone/website, JS colour-bound) is separate from the Email/PDF
    tabs (real templates, server-rendered) -- see ClubSettingsDocumentTabTests
    for those."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_the_page_tabs_and_colour_swatches_render(self):
        response = self.club_get("club_settings")

        self.assertContains(response, "rounded-full bg-steel p-1")
        self.assertContains(response, 'id="primary_color_swatch"')
        self.assertContains(response, 'id="secondary_color_swatch"')

    def test_the_mobile_app_and_website_previews_render(self):
        response = self.club_get("club_settings")

        self.assertContains(response, "preview-phone-lg")
        self.assertContains(response, "preview-website")
        self.assertContains(response, "Where the brand shows up")

    def test_the_website_field_is_editable(self):
        response = self.club_get("club_settings")
        self.assertContains(response, 'name="website"')

    def test_identity_email_and_pdf_are_the_three_page_tabs(self):
        response = self.club_get("club_settings")

        self.assertContains(response, 'data-page-tab="identity"')
        self.assertContains(response, 'data-page-tab="email"')
        self.assertContains(response, 'data-page-tab="pdf"')

    def test_saving_still_updates_the_club(self):
        response = self.club_post(
            "club_settings",
            {"name": "Ajax United", "legal_name": "", "contact_email": "", "website": "https://ajax-united.example", "primary_color": "#123456", "secondary_color": "#654321"},
        )

        self.assertRedirects(response, reverse("management:club_settings"))
        self.club.refresh_from_db()
        self.assertEqual(self.club.primary_color, "#123456")
        self.assertEqual(self.club.secondary_color, "#654321")
        self.assertEqual(self.club.website, "https://ajax-united.example")

    def test_the_legal_address_can_be_set(self):
        response = self.club_post(
            "club_settings",
            {
                "name": "Ajax United",
                "legal_name": "",
                "legal_address": "Registered Office 5",
                "legal_zip_code": "9000",
                "legal_city": "Ghent",
                "contact_email": "",
                "website": "",
                "primary_color": "",
                "secondary_color": "",
            },
        )

        self.assertRedirects(response, reverse("management:club_settings"))
        self.club.refresh_from_db()
        self.assertEqual(self.club.legal_address, "Registered Office 5")
        self.assertEqual(self.club.legal_zip_code, "9000")
        self.assertEqual(self.club.legal_city, "Ghent")

    def test_the_payment_instructions_can_be_set(self):
        response = self.club_post(
            "club_settings",
            {
                "name": "Ajax United",
                "legal_name": "",
                "contact_email": "",
                "website": "",
                "primary_color": "",
                "secondary_color": "",
                "payment_instructions": "Bank transfer to BE00 0000 0000 0000",
            },
        )

        self.assertRedirects(response, reverse("management:club_settings"))
        self.club.refresh_from_db()
        self.assertEqual(self.club.payment_instructions, "Bank transfer to BE00 0000 0000 0000")

    def test_the_payment_instructions_field_is_editable(self):
        response = self.club_get("club_settings")

        self.assertContains(response, 'name="payment_instructions"')

    def test_the_event_background_can_be_uploaded(self):
        response = self.club_post(
            "club_settings",
            {
                "name": "Ajax United",
                "legal_name": "",
                "contact_email": "",
                "website": "",
                "primary_color": "",
                "secondary_color": "",
                "event_background": make_image_file("event-bg.png"),
            },
        )

        self.assertRedirects(response, reverse("management:club_settings"))
        self.club.refresh_from_db()
        self.assertTrue(self.club.event_background)

    def test_the_event_background_field_is_editable(self):
        response = self.club_get("club_settings")

        self.assertContains(response, 'name="event_background"')


class ClubSettingsDocumentTabTests(ManagementTestBase):
    """The Email and PDF tabs on the Club identity page -- every branded
    email (management.email_previews) and generated PDF (management.pdf_previews)
    this club can send, rendered against sample data. See
    EmailPreviewRenderViewTests/PDFPreviewRenderViewTests for the actual
    per-document render each card's iframe loads."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_renders_every_registered_email_and_pdf_preview(self):
        response = self.club_get("club_settings")

        self.assertEqual(response.status_code, 200)
        for preview in EMAIL_PREVIEWS:
            self.assertContains(response, preview.label)
        for preview in PDF_PREVIEWS:
            self.assertContains(response, preview.label)

    def test_each_card_links_to_its_own_render_view(self):
        response = self.club_get("club_settings")

        for preview in EMAIL_PREVIEWS:
            self.assertContains(response, reverse("management:email_preview_render", args=[preview.key]))
        for preview in PDF_PREVIEWS:
            self.assertContains(response, reverse("management:pdf_preview_render", args=[preview.key]))

    def test_the_subject_and_plain_text_body_render_for_emails(self):
        response = self.club_get("club_settings")

        self.assertContains(response, "Subject:")
        self.assertContains(response, "DUE-2026-00042")

    def test_pdf_cards_have_no_plain_text_toggle(self):
        # PDFs have no plain-text body to switch to -- unlike an email card,
        # a PDF card is just the one iframe. Exactly one toggle button per
        # email preview, none contributed by the PDF ones.
        response = self.club_get("club_settings")

        self.assertEqual(response.content.decode().count('data-view-btn="text"'), len(EMAIL_PREVIEWS))

    def test_claim_approved_was_dropped(self):
        # Not directly used right now -- see management/email_previews.py.
        response = self.club_get("club_settings")

        self.assertNotContains(response, "Parent claim approved")

    def test_no_real_data_is_needed(self):
        # The whole point: previews render with zero DuesInvoice/Event rows
        # in the database.
        self.assertFalse(DuesInvoice.objects.exists())
        self.assertFalse(Event.objects.filter(club=self.club).exists())

        response = self.club_get("club_settings")

        self.assertEqual(response.status_code, 200)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-doc-preview@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U18", short_name="U18")
        position = Position.objects.create(club=self.club, name="Coach14", short_name="C14", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("club_settings")

        self.assertEqual(response.status_code, 403)


class EmailPreviewRenderViewTests(ManagementTestBase):
    """The actual per-email HTML document each card's iframe loads via `src` --
    see EmailPreviewRenderView's own docstring for why this is a same-origin
    sub-request rather than a srcdoc="..." attribute."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_renders_the_full_html_document(self):
        response = self.client.get(reverse("management:email_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/html; charset=utf-8")
        self.assertContains(response, "<!DOCTYPE html>")
        self.assertContains(response, "DUE-2026-00042")

    def test_uses_the_clubs_own_branding(self):
        self.club.name = "Ajax United"
        self.club.secondary_color = "#123456"
        self.club.save(update_fields=["name", "secondary_color"])

        response = self.client.get(reverse("management:email_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "Ajax United")
        self.assertContains(response, "#123456")

    def test_the_password_reset_preview_renders(self):
        response = self.client.get(reverse("management:email_preview_render", args=["password_reset"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Reset your password")

    def test_allows_same_origin_framing(self):
        # The site-wide default (SecurityMiddleware, no X_FRAME_OPTIONS override
        # in settings.py) is DENY -- this response overrides that specifically,
        # since the Email tab's own iframe needs to load it.
        response = self.client.get(reverse("management:email_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response["X-Frame-Options"], "SAMEORIGIN")

    def test_an_unknown_key_404s(self):
        response = self.client.get(reverse("management:email_preview_render", args=["not-a-real-key"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 404)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-email-render@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U19", short_name="U19")
        position = Position.objects.create(club=self.club, name="Coach15", short_name="C15", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.client.get(reverse("management:email_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 403)


class PDFPreviewRenderViewTests(ManagementTestBase):
    """The actual per-PDF HTML document each PDF card's iframe loads via
    `src` -- the same underlying HTML WeasyPrint would turn into a PDF, shown
    directly rather than round-tripping through WeasyPrint for a preview."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_renders_the_dues_invoice_pdf_html(self):
        response = self.client.get(reverse("management:pdf_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<!DOCTYPE html>")
        self.assertContains(response, "DUE-2026-00042")

    def test_renders_the_referee_form_pdf_html(self):
        response = self.client.get(reverse("management:pdf_preview_render", args=["referee_form"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Referee payment form")
        self.assertContains(response, "Jamie Doe")

    def test_allows_same_origin_framing(self):
        response = self.client.get(reverse("management:pdf_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response["X-Frame-Options"], "SAMEORIGIN")

    def test_an_unknown_key_404s(self):
        response = self.client.get(reverse("management:pdf_preview_render", args=["not-a-real-key"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 404)

    def test_non_admin_gets_403(self):
        coach_user = User.objects.create_user(email="coach-pdf-render@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U20", short_name="U20")
        position = Position.objects.create(club=self.club, name="Coach16", short_name="C16", staff_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.client.get(reverse("management:pdf_preview_render", args=["dues_invoice"]), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.status_code, 403)


class BillingEndingBannerTests(ManagementTestBase):
    """The club dashboard's "billing is about to stop" warning -- see
    management.views.HomeView and management/templates/management/home.html.
    Admin-only, only within the plan's own renewal lead of the period ending, and only when
    nothing is owed -- an unpaid club gets the louder billing notice instead."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.plan = Plan.objects.create(name="Standard")
        PlanPrice.objects.create(plan=cls.plan, active_from=cls.season.start_date - datetime.timedelta(days=1200), amount=Decimal("500.00"))

    def settle(self):
        """The "period ends soon" notice only shows when nothing is owed."""
        record_payment(self.club.dues.first(), Decimal("500.00"))

    def test_admin_sees_the_banner_when_the_period_ends_soon(self):
        subscribe(self.club, self.plan, start=timezone.localdate() - datetime.timedelta(days=350), auto_renew=False)
        self.settle()
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertContains(response, "billing is about to stop")

    def test_admin_does_not_see_the_banner_when_the_period_is_not_ending_soon(self):
        subscribe(self.club, self.plan, start=timezone.localdate())
        self.settle()
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertNotContains(response, "billing is about to stop")

    def test_a_non_admin_manager_never_sees_the_banner(self):
        subscribe(self.club, self.plan, start=timezone.localdate() - datetime.timedelta(days=350), auto_renew=False)
        team = Team.objects.create(club=self.club, name="First Team", short_name="1st")
        position = Position.objects.create(club=self.club, name="Coach", short_name="C", staff_position=True, management_position=True)
        coach_user = User.objects.create_user(email="coach-banner@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.settle()
        self.client.force_login(coach_user)

        response = self.club_get("home")

        self.assertNotContains(response, "billing is about to stop")

    def test_a_club_with_no_subscription_shows_no_banner(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertNotContains(response, "billing is about to stop")
        self.assertNotContains(response, "will renew automatically")

    def test_an_auto_renewing_club_gets_a_reassuring_banner_instead(self):
        subscribe(self.club, self.plan, start=timezone.localdate() - datetime.timedelta(days=350), auto_renew=True)
        self.settle()
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertContains(response, "will renew automatically")
        self.assertNotContains(response, "billing is about to stop")


class RecurrenceUiTests(TestCase):
    """management.recurrence_ui's friendly builder <-> raw RRULE round-trip."""

    def test_weekly_round_trips(self):
        rrule = build_rrule("weekly", 2, ["WE", "MO"])

        self.assertEqual(rrule, "FREQ=WEEKLY;INTERVAL=2;BYDAY=MO,WE")
        self.assertEqual(parse_rrule(rrule), {"frequency": "weekly", "interval": 2, "weekdays": ["MO", "WE"]})
        self.assertEqual(str(describe_rrule(rrule)), "Every 2 weeks on Mon, Wed")

    def test_monthly_round_trips(self):
        rrule = build_rrule("monthly", 1)

        self.assertEqual(rrule, "FREQ=MONTHLY;INTERVAL=1")
        self.assertEqual(parse_rrule(rrule), {"frequency": "monthly", "interval": 1, "weekdays": []})
        self.assertEqual(str(describe_rrule(rrule)), "Every month")

    def test_an_unrecognised_rrule_falls_back_to_the_raw_string(self):
        self.assertIsNone(parse_rrule("FREQ=DAILY;COUNT=5"))
        self.assertEqual(describe_rrule("FREQ=DAILY;COUNT=5"), "FREQ=DAILY;COUNT=5")


class EventManagementTests(ManagementTestBase):
    """Event CRUD -- permissions are scoped per-team (like roster/staff), not
    club-wide, since an event's teams field is M2M: a manager of at least one
    of an event's current teams can edit it, see club.mixins.EventManagerRequiredMixin."""

    def setUp(self):
        # A couple of tests below flip the "officials" flag on for self.club;
        # waffle caches a flag's active-club ids with cache.add (add-if-absent,
        # see features/models.py's Flag._get_club_ids), which isn't part of the
        # per-test transaction -- without clearing it, that stays cached (and
        # officials_enabled_for(club) stays True) for every test that runs
        # after, even once the DB association itself is rolled back.
        cache.clear()
        self.addCleanup(cache.clear)

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.own_team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.home_ground = Location.objects.create(club=cls.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE", is_home=True)
        cls.coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)

        # Manager of own_team only -- "may for my team, may not for theirs" is the
        # line nearly every test here checks, so it's a standing fixture.
        cls.own_team_coach = User.objects.create_user(email="coach-events@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=cls.own_team_coach, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=cls.own_team, member=coach_member, season=cls.season, position=cls.coach_position)

        # Staff, but with no management position anywhere: reaches the app, manages nothing.
        cls.plain_staff = User.objects.create_user(email="physio-events@example.com", password="pw-secret-123")
        staff_member = Member.objects.create(user=cls.plain_staff, first_name="Pat", last_name="Physio")
        staff_position = Position.objects.create(club=cls.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=cls.own_team, member=staff_member, season=cls.season, position=staff_position)

    def make_group_member(self, group, email="committee-events@example.com"):
        # Staff (so they can reach the management app at all -- ClubStaffRequiredMixin
        # excludes a plain MEMBER-only club member entirely), but NOT a team manager of
        # anything: a non-management StaffAssignment on an unrelated team, so
        # teams_managed_by(this user) is empty and the group claim is what's really
        # being tested, isolated from any team-manager claim.
        member_user = User.objects.create_user(email=email, password="pw-secret-123")
        member = Member.objects.create(user=member_user, first_name="Gale", last_name="Group")
        physio_position = Position.objects.create(club=self.club, name="Committee Physio", short_name="CP", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=self.other_team, member=member, season=self.season, position=physio_position)
        GroupMembership.objects.create(group=group, member=member)
        return member_user

    def event_data(self, **overrides):
        data = {
            "title": "Training",
            "kind": "training",
            "teams": [str(self.own_team.pk)],
            "invited_members": [],
            "excluded_members": [],
            "location": "",
            "opponent": "",
            "start": "2026-09-01T18:00",
            "end": "",
            "gathering": "",
            "deadline": "",
            "max_referees": "2",
        }
        data.update(overrides)
        return data

    def test_a_team_manager_can_create_an_event_for_their_own_team(self):
        self.client.force_login(self.own_team_coach)

        response = self.club_post("event_create", self.event_data())

        event = Event.objects.get(title="Training")
        self.assertRedirects(response, reverse("management:event_detail", args=[event.pk]))
        self.assertIn(self.own_team, event.teams.all())

    def test_creating_an_event_notifies_the_invited_roster(self):
        position = Position.objects.create(club=self.club, name="Forward", short_name="F")
        player = Member.objects.create(first_name="Poul", last_name="Player")
        TeamMembership.objects.create(team=self.own_team, member=player, season=self.season, position=position)
        self.client.force_login(self.own_team_coach)

        # dispatch_notify_new_event fires on a real background thread now (see
        # events/services/notifications.py) -- run it inline here instead, via
        # side_effect, so the notification is guaranteed to exist by the time this
        # assertion runs rather than racing a thread that may not have finished yet.
        with mock.patch("management.views.dispatch_notify_new_event", side_effect=notify_new_event):
            self.club_post("event_create", self.event_data())

        event = Event.objects.get(title="Training")
        notification = Notification.objects.get(member=player, content_type__model="event", object_id=str(event.pk))
        self.assertEqual(notification.title, "Training")

    def test_creating_an_event_with_a_same_club_location_does_not_raise_a_cross_club_error(self):
        # Regression: Event.clean() rejects a location from another club by
        # comparing against self.club_id, which was still None on a brand-new
        # instance at validation time (club is only auto-assigned in save(),
        # which runs after full_clean()) -- so a same-club location falsely
        # failed as "must belong to the same club". See EventCreateView.get_form_kwargs.
        location = Location.objects.create(club=self.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE")
        self.client.force_login(self.own_team_coach)

        response = self.club_post("event_create", self.event_data(location=str(location.pk)))

        event = Event.objects.get(title="Training")
        self.assertRedirects(response, reverse("management:event_detail", args=[event.pk]))
        self.assertEqual(event.location, location)

    def test_the_location_dropdown_shows_the_city(self):
        Location.objects.create(club=self.club, name="Sportcentrum", address="Straat 1", city="Mechelen", zip_code="2800", country="BE")
        self.client.force_login(self.admin_user)

        response = self.club_get("event_create")

        self.assertContains(response, "Sportcentrum — Mechelen")

    def test_the_location_dropdown_adds_the_country_when_not_belgium(self):
        Location.objects.create(club=self.club, name="Rival Hall", address="Rue 2", city="Lille", zip_code="59000", country="FR")
        self.client.force_login(self.admin_user)

        response = self.club_get("event_create")

        self.assertContains(response, "Rival Hall — Lille, France")

    def test_the_location_field_is_searchable(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("event_create")

        self.assertContains(response, 'name="location"')
        self.assertContains(response, "data-searchable")

    def test_a_group_member_can_create_an_event_for_their_own_group(self):
        group = Group.objects.create(club=self.club, name="Committee")
        self.client.force_login(self.make_group_member(group))

        response = self.club_post("event_create", self.event_data(teams=[], groups=[str(group.pk)]))

        event = Event.objects.get(title="Training")
        self.assertRedirects(response, reverse("management:event_detail", args=[event.pk]))
        self.assertIn(group, event.groups.all())

    def test_a_group_member_cannot_pick_a_group_they_do_not_belong_to(self):
        own_group = Group.objects.create(club=self.club, name="Committee")
        other_group = Group.objects.create(club=self.club, name="Other Committee")
        self.client.force_login(self.make_group_member(own_group))

        response = self.club_post("event_create", self.event_data(teams=[], groups=[str(other_group.pk)]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Event.objects.filter(title="Training").exists())

    def test_a_non_admin_with_neither_team_nor_group_is_rejected(self):
        self.client.force_login(self.make_group_member(Group.objects.create(club=self.club, name="Committee")))

        response = self.club_post("event_create", self.event_data(teams=[], groups=[]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Event.objects.filter(title="Training").exists())

    def test_club_wide_is_not_offered_to_a_non_admin(self):
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_create")

        self.assertNotContains(response, 'name="club_wide"')

    def test_a_non_admin_cannot_force_a_club_wide_event(self):
        # club_wide isn't in the form for a non-admin, so even a forged POST
        # value must not slip through as a create-time claim.
        self.client.force_login(self.own_team_coach)

        response = self.club_post("event_create", self.event_data(teams=[], club_wide="on"))

        self.assertEqual(response.status_code, 200)
        event = Event.objects.filter(title="Training").first()
        self.assertIsNone(event)

    def test_an_admin_can_schedule_a_club_wide_event(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("event_create", self.event_data(teams=[], club_wide="on"))

        event = Event.objects.get(title="Training")
        self.assertRedirects(response, reverse("management:event_detail", args=[event.pk]))
        self.assertTrue(event.club_wide)

    def test_club_wide_cannot_be_combined_with_teams(self):
        self.client.force_login(self.admin_user)

        response = self.club_post("event_create", self.event_data(club_wide="on"))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Event.objects.filter(title="Training").exists())

    def test_a_group_member_who_is_not_a_team_manager_can_edit_their_group_event(self):
        group = Group.objects.create(club=self.club, name="Committee")
        member_user = self.make_group_member(group)
        self.client.force_login(member_user)
        self.club_post("event_create", self.event_data(teams=[], groups=[str(group.pk)]))
        event = Event.objects.get(title="Training")

        response = self.club_get("event_update", event.pk)

        self.assertEqual(response.status_code, 200)

    def test_the_new_event_forms_competition_dropdown_only_offers_flags_active_for_this_club(self):
        # Same filtering as the Django-admin form (events.admin.EventAdminForm) --
        # see management.forms.EventForm. A flagless competition, or one whose
        # flag isn't active for this club, never appears; fetch_game_info
        # (events.services.competitions) would silently do nothing for either
        # anyway, so offering them here would just be a dead end.
        active_flag = get_waffle_flag_model().objects.create(name="active-competition")
        active_flag.clubs.add(self.club)
        inactive_flag = get_waffle_flag_model().objects.create(name="inactive-competition")
        Competition.objects.create(name="Active Cup", module="events.competition.active", flag=active_flag)
        Competition.objects.create(name="Inactive Cup", module="events.competition.inactive", flag=inactive_flag)
        Competition.objects.create(name="Flagless Cup", module="events.competition.flagless")
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_create")

        self.assertContains(response, "Active Cup")
        self.assertNotContains(response, "Inactive Cup")
        self.assertNotContains(response, "Flagless Cup")

    def test_a_team_manager_cannot_create_an_event_for_a_team_they_dont_manage(self):
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(teams=[str(self.other_team.pk)]))

        self.assertFalse(Event.objects.filter(title="Training").exists())

    def test_a_plain_staff_member_gets_403_creating_an_event(self):
        self.client.force_login(self.plain_staff)

        response = self.club_post("event_create", self.event_data())

        self.assertEqual(response.status_code, 403)

    def test_a_plain_staff_member_can_still_view_the_event_list(self):
        self.client.force_login(self.plain_staff)

        self.assertEqual(self.club_get("event_list", params={"view": "list"}).status_code, 200)

    def test_the_list_shows_edit_and_delete_only_for_events_the_manager_manages(self):
        own_event = Event.objects.create(club=self.club, title="My event", start=timezone.now() + datetime.timedelta(days=1))
        own_event.teams.add(self.own_team)
        other_event = Event.objects.create(club=self.club, title="Other event", start=timezone.now() + datetime.timedelta(days=2))
        other_event.teams.add(self.other_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        # The Edit link goes to the detail page, not straight to the edit form (same
        # convention as Teams/News), so what actually distinguishes a manageable row
        # is the delete action being present.
        self.assertContains(response, reverse("management:event_delete", args=[own_event.pk]))
        self.assertNotContains(response, reverse("management:event_delete", args=[other_event.pk]))

    def test_the_list_shows_open_and_claimed_tasks(self):
        event = Event.objects.create(club=self.club, title="Training", start=timezone.now() + datetime.timedelta(days=1))
        event.teams.add(self.own_team)
        claimed = EventTask.objects.create(event=event, title="Bring fruit", needed_quantity=1)
        claimer = Member.objects.create(first_name="Fran", last_name="Family")
        EventTaskClaim.objects.create(task=claimed, member=claimer)
        EventTask.objects.create(event=event, title="Penalty bench", needed_quantity=2)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertContains(response, "Bring fruit: Fran Family")
        self.assertContains(response, "Penalty bench: open")

    def test_the_list_shows_referee_names_for_a_home_game(self):
        game = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, start=timezone.now() + datetime.timedelta(days=1), location=self.home_ground)
        game.teams.add(self.own_team)
        referee = Member.objects.create(first_name="Ref", last_name="Eree")
        EventReferee.objects.create(event=game, member=referee)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertContains(response, "Referees: Ref Eree")

    def test_the_list_does_not_show_referees_for_an_away_game(self):
        game = Event.objects.create(club=self.club, title="Away game", kind=Event.EventKind.GAME, start=timezone.now() + datetime.timedelta(days=1))
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertNotContains(response, "Referees:")

    def test_the_list_shows_a_friendly_badge_for_a_friendly_game(self):
        game = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, start=timezone.now() + datetime.timedelta(days=1), is_friendly=True)
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertContains(response, "Friendly")

    def test_the_list_does_not_show_a_friendly_badge_for_a_league_game(self):
        game = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, start=timezone.now() + datetime.timedelta(days=1), is_friendly=False)
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertNotContains(response, "Friendly")

    def test_the_week_calendar_marks_a_friendly_game(self):
        game = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, start=timezone.now(), is_friendly=True)
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list")  # default view=calendar, range=week

        self.assertContains(response, 'aria-label="Friendly"')

    def test_the_month_calendar_marks_a_friendly_game(self):
        game = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, start=timezone.now(), is_friendly=True)
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"range": "month"})

        self.assertContains(response, 'aria-label="Friendly"')

    def test_the_calendar_page_highlights_the_events_sidebar_item(self):
        # Regression: EventListView's own "calendar_nav" context (prev/next/today
        # for the Week/Month view) used to be keyed "nav", shadowing
        # management.context_processors.active_nav_section's identically-named
        # "nav" -- the value _nav_items.html reads to mark the Events sub-item
        # active. That silently broke the highlight on this page alone.
        self.client.force_login(self.admin_user)

        response = self.club_get("event_list")

        self.assertContains(response, '<a class="nav-item active" href="/manage/events/">')
        self.assertContains(response, '<a class="nav-subitem active" href="/manage/events/">')

    def test_the_week_calendar_marks_a_series_occurrence_with_the_repeat_icon(self):
        series = EventSeries.objects.create(club=self.club, title="Weekly Training", kind=Event.EventKind.TRAINING, dtstart=timezone.now(), rrule="FREQ=WEEKLY;COUNT=1")
        occurrence = Event.objects.create(club=self.club, title="Weekly Training", start=timezone.now(), series=series)
        occurrence.teams.add(self.own_team)
        standalone = Event.objects.create(club=self.club, title="One-off Training", start=timezone.now() + datetime.timedelta(hours=1))
        standalone.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list")  # default view=calendar, range=week

        self.assertContains(response, 'aria-label="Part of a series"', count=1)

    def test_the_week_calendar_carries_the_first_event_hour_for_the_initial_scroll(self):
        local_today = timezone.localdate()
        start = timezone.make_aware(datetime.datetime.combine(local_today, datetime.time(18, 0)))
        event = Event.objects.create(club=self.club, title="Evening practice", start=start)
        event.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list")  # default view=calendar, range=week

        self.assertContains(response, 'data-first-event-hour="18"')

    def test_the_week_calendar_first_event_hour_is_blank_with_no_events(self):
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"date": "2030-01-07"})  # a Monday with nothing scheduled

        self.assertContains(response, 'data-first-event-hour=""')

    def test_the_month_calendar_marks_a_series_occurrence_with_the_repeat_icon(self):
        series = EventSeries.objects.create(club=self.club, title="Weekly Training", kind=Event.EventKind.TRAINING, dtstart=timezone.now(), rrule="FREQ=WEEKLY;COUNT=1")
        occurrence = Event.objects.create(club=self.club, title="Weekly Training", start=timezone.now(), series=series)
        occurrence.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"range": "month"})

        self.assertContains(response, 'aria-label="Part of a series"', count=1)

    def test_a_manager_only_sees_events_for_teams_they_manage(self):
        own_event = Event.objects.create(club=self.club, title="My event", start=timezone.now() + datetime.timedelta(days=1))
        own_event.teams.add(self.own_team)
        other_event = Event.objects.create(club=self.club, title="Other event", start=timezone.now() + datetime.timedelta(days=2))
        other_event.teams.add(self.other_team)
        Event.objects.create(club=self.club, title="AGM", start=timezone.now() + datetime.timedelta(days=3))
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertContains(response, "My event")
        self.assertNotContains(response, "Other event")
        self.assertContains(response, "AGM")  # team-less events stay visible to everyone

    def test_a_group_member_sees_their_groups_event_without_managing_any_team(self):
        group = Group.objects.create(club=self.club, name="Committee")
        group_event = Event.objects.create(club=self.club, title="Committee meeting", start=timezone.now() + datetime.timedelta(days=1))
        group_event.groups.add(group)
        self.client.force_login(self.make_group_member(group))

        response = self.club_get("event_list", params={"view": "list"})

        self.assertContains(response, "Committee meeting")

    def test_a_non_member_does_not_see_another_groups_event(self):
        own_group = Group.objects.create(club=self.club, name="Committee")
        other_group = Group.objects.create(club=self.club, name="Other Committee")
        other_event = Event.objects.create(club=self.club, title="Other group event", start=timezone.now() + datetime.timedelta(days=1))
        other_event.groups.add(other_group)
        self.client.force_login(self.make_group_member(own_group))

        response = self.club_get("event_list", params={"view": "list"})

        self.assertNotContains(response, "Other group event")

    def test_an_admin_sees_every_event_regardless_of_team(self):
        own_event = Event.objects.create(club=self.club, title="My event", start=timezone.now() + datetime.timedelta(days=1))
        own_event.teams.add(self.own_team)
        other_event = Event.objects.create(club=self.club, title="Other event", start=timezone.now() + datetime.timedelta(days=2))
        other_event.teams.add(self.other_team)
        self.client.force_login(self.admin_user)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertContains(response, "My event")
        self.assertContains(response, "Other event")

    def test_a_manager_cannot_open_another_teams_event_by_url(self):
        other_event = Event.objects.create(club=self.club, title="Other event", start=timezone.now() + datetime.timedelta(days=2))
        other_event.teams.add(self.other_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_detail", other_event.pk)

        self.assertEqual(response.status_code, 404)

    def test_the_list_groups_into_this_week_next_week_and_month_dividers(self):
        # Mirrors mobile.views.CalendarView's own This week/Next week/by-month
        # agenda grouping -- see EventListView._list_groups.
        _this_week_start, this_week_end = week_bounds(timezone.localdate())
        this_week_event_start = timezone.make_aware(datetime.datetime.combine(this_week_end, datetime.time(18, 0)))
        if this_week_event_start <= timezone.now():
            # The fixed 18:00 anchor is only "this week" if it's still ahead of
            # now -- when this_week_end is today and it's already past 18:00,
            # fall back to a few minutes from now (still always this week)
            # instead of a start the view's own start__gte=now filter would
            # exclude as already past.
            this_week_event_start = timezone.now() + datetime.timedelta(minutes=5)
        this_week_event = Event.objects.create(club=self.club, title="This week event", start=this_week_event_start)
        this_week_event.teams.add(self.own_team)
        next_week_event = Event.objects.create(club=self.club, title="Next week event", start=timezone.make_aware(datetime.datetime.combine(this_week_end + datetime.timedelta(days=3), datetime.time(18, 0))))
        next_week_event.teams.add(self.own_team)
        later_event = Event.objects.create(club=self.club, title="Later event", start=timezone.make_aware(datetime.datetime.combine(this_week_end + datetime.timedelta(days=60), datetime.time(18, 0))))
        later_event.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list"})

        self.assertEqual([event.title for event in response.context["list_this_week"]], ["This week event"])
        self.assertEqual([event.title for event in response.context["list_next_week"]], ["Next week event"])
        later_months = response.context["list_months"]
        self.assertEqual(len(later_months), 1)
        self.assertEqual([event.title for event in later_months[0]["items"]], ["Later event"])
        self.assertContains(response, "This week")
        self.assertContains(response, "Next week")
        self.assertContains(response, later_event.start.strftime("%B %Y"))

    def test_show_past_groups_by_month_without_this_next_week_labels(self):
        past_event = Event.objects.create(club=self.club, title="Past event", start=timezone.now() - datetime.timedelta(days=10))
        past_event.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_list", params={"view": "list", "show_past": "1"})

        self.assertEqual(response.context["list_this_week"], [])
        self.assertEqual(response.context["list_next_week"], [])
        self.assertEqual(len(response.context["list_months"]), 1)
        self.assertEqual([event.title for event in response.context["list_months"][0]["items"]], ["Past event"])
        self.assertNotContains(response, "This week")

    def test_the_dashboard_only_shows_upcoming_events_for_managed_teams(self):
        own_event = Event.objects.create(club=self.club, title="My event", start=timezone.now() + datetime.timedelta(days=1))
        own_event.teams.add(self.own_team)
        other_event = Event.objects.create(club=self.club, title="Other event", start=timezone.now() + datetime.timedelta(days=2))
        other_event.teams.add(self.other_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("home")

        self.assertContains(response, "My event")
        self.assertNotContains(response, "Other event")

    def test_a_team_less_event_is_refused_for_a_non_admin(self):
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(teams=[]))

        self.assertFalse(Event.objects.filter(title="Training").exists())

    def test_an_admin_can_create_a_team_less_event(self):
        self.client.force_login(self.admin_user)

        self.club_post("event_create", self.event_data(teams=[]))

        self.assertTrue(Event.objects.filter(title="Training").exists())

    def test_a_manager_of_one_team_cannot_edit_an_event_for_another_team(self):
        event = Event.objects.create(club=self.club, title="Other's event", start=timezone.now() + datetime.timedelta(days=1))
        event.teams.add(self.other_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_update", event.pk)

        self.assertEqual(response.status_code, 403)

    def test_a_manager_can_edit_their_own_teams_event(self):
        event = Event.objects.create(club=self.club, title="My event", start=timezone.now() + datetime.timedelta(days=1))
        event.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        self.club_post("event_update", self.event_data(title="Renamed"), event.pk)

        event.refresh_from_db()
        self.assertEqual(event.title, "Renamed")

    def test_deleting_a_one_off_event_deletes_it(self):
        event = Event.objects.create(club=self.club, title="Gone", start=timezone.now() + datetime.timedelta(days=1))
        event.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        self.club_post("event_delete", {}, event.pk)

        self.assertFalse(Event.objects.filter(pk=event.pk).exists())

    def test_creating_a_game_records_competition_and_external_id_but_not_score(self):
        # Score/live status don't exist yet for a game that's only just being
        # scheduled -- the add form doesn't even offer those fields (see
        # test_the_add_form_has_no_score_or_live_fields below), so posting them
        # here has no effect.
        flag = get_waffle_flag_model().objects.create(name="regional-cup")
        flag.clubs.add(self.club)
        Competition.objects.create(name="Regional Cup", module="events.competition.regional", flag=flag)
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(kind="game", competition="Regional Cup", external_game_id="ext-42", score_for="3", score_against="1", is_live="on", is_friendly="on"))

        game = Event.objects.get(title="Training")
        self.assertEqual(game.kind, Event.EventKind.GAME)
        self.assertEqual(game.competition, "Regional Cup")
        self.assertEqual(game.external_game_id, "ext-42")
        self.assertTrue(game.is_friendly)
        self.assertIsNone(game.score_for)
        self.assertFalse(game.is_live)

    def test_creating_a_game_auto_titles_it_from_team_and_opponent(self):
        opponent = Opponent.objects.create(club=self.club, name="Rivals FC")
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(title="Whatever I typed", kind="game", opponent=str(opponent.pk)))

        game = Event.objects.get(opponent=opponent)
        self.assertEqual(game.title, f"{self.own_team.short_name} vs {opponent}")

    def test_editing_a_game_re_titles_it_when_the_opponent_changes(self):
        opponent = Opponent.objects.create(club=self.club, name="Rivals FC")
        new_opponent = Opponent.objects.create(club=self.club, name="Bears HC")
        game = Event.objects.create(club=self.club, title="Old title", kind=Event.EventKind.GAME, start=timezone.now() + datetime.timedelta(days=1), opponent=opponent)
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        self.club_post("event_update", self.event_data(title="Old title", kind="game", opponent=str(new_opponent.pk)), game.pk)

        game.refresh_from_db()
        self.assertEqual(game.title, f"{self.own_team.short_name} vs {new_opponent}")

    def test_a_game_with_no_opponent_yet_keeps_its_typed_title(self):
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(title="Cup semi-final", kind="game"))

        game = Event.objects.get(title="Cup semi-final")
        self.assertEqual(game.title, "Cup semi-final")

    def test_a_training_keeps_its_typed_title_even_with_an_opponent_posted(self):
        # Only Game auto-titles -- a tampered/leftover opponent value in the
        # POST for a non-game kind shouldn't retitle it.
        opponent = Opponent.objects.create(club=self.club, name="Rivals FC")
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(title="Practice", kind="training", opponent=str(opponent.pk)))

        event = Event.objects.get(title="Practice")
        self.assertEqual(event.title, "Practice")

    def test_the_add_form_has_no_score_or_live_fields(self):
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_create")

        self.assertNotContains(response, 'name="score_for"')
        self.assertNotContains(response, 'name="is_live"')
        self.assertContains(response, 'name="competition"')

    def test_the_add_form_has_no_max_officials_field_when_the_flag_is_off(self):
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_create")

        self.assertNotContains(response, 'name="max_officials"')

    def test_the_add_form_has_a_max_officials_field_when_the_flag_is_on(self):
        get_waffle_flag_model().objects.get_or_create(name="officials")[0].clubs.add(self.club)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_create")

        self.assertContains(response, 'name="max_officials"')

    def test_creating_a_game_records_max_officials_when_the_flag_is_on(self):
        get_waffle_flag_model().objects.get_or_create(name="officials")[0].clubs.add(self.club)
        self.client.force_login(self.own_team_coach)

        self.club_post("event_create", self.event_data(kind="game", max_officials="3"))

        game = Event.objects.get(title="Training")
        self.assertEqual(game.max_officials, 3)

    def test_editing_a_game_can_record_its_score_and_live_status(self):
        flag = get_waffle_flag_model().objects.create(name="regional-cup")
        flag.clubs.add(self.club)
        Competition.objects.create(name="Regional Cup", module="events.competition.regional", flag=flag)
        game = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, start=timezone.now() + datetime.timedelta(days=1))
        game.teams.add(self.own_team)
        self.client.force_login(self.own_team_coach)

        response = self.club_get("event_update", game.pk)
        self.assertContains(response, 'name="score_for"')
        self.assertContains(response, 'name="is_live"')

        self.club_post("event_update", self.event_data(kind="game", competition="Regional Cup", external_game_id="ext-42", score_for="3", score_against="1", is_live="on"), game.pk)

        game.refresh_from_db()
        self.assertEqual(game.score_for, 3)
        self.assertEqual(game.score_against, 1)
        self.assertTrue(game.is_live)

    def test_the_game_kind_choice_is_no_longer_called_match(self):
        self.assertNotIn("match", dict(Event.EventKind.choices))
        self.assertEqual(dict(Event.EventKind.choices)["game"], "Game")


class EventSeriesManagementTests(ManagementTestBase):
    """EventSeries CRUD + occurrence lifecycle actions (cancel/detach/stop)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.other_team = Team.objects.create(club=cls.club, name="Second Team", short_name="2nd")
        cls.coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)

        # A manager of each team: the series is created by the first, and the
        # second is the "different team's manager" the access checks refuse.
        cls.team_coach = cls.make_coach(cls.team, "coach-series@example.com")
        cls.other_team_coach = cls.make_coach(cls.other_team, "other-coach@example.com")

    @classmethod
    def make_coach(cls, team, email):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=team, member=coach_member, season=cls.season, position=cls.coach_position)
        return coach_user

    def series_data(self, **overrides):
        data = {
            "title": "Weekly training",
            "kind": "training",
            "dtstart": "2026-09-01T18:00",
            "until": "",
            "teams": [str(self.team.pk)],
            "invited_members": [],
            "excluded_members": [],
            "location": "",
            "opponent": "",
            "frequency": "weekly",
            "interval": "1",
            "weekdays": ["MO", "WE"],
            "duration_hours": "1",
            "duration_minutes": "0",
            "gathering_minutes_before": "",
            "deadline_minutes_before": "",
            "advanced_rrule": "",
        }
        data.update(overrides)
        return data

    def create_series(self):
        self.club_post("event_series_create", self.series_data())
        return EventSeries.objects.get(title="Weekly training")

    def test_creating_a_series_with_a_same_club_location_does_not_raise_a_cross_club_error(self):
        # Same regression as EventManagementTests' equivalent -- EventSeries.clean()
        # has the same self.club_id-is-still-None-at-validation-time problem.
        location = Location.objects.create(club=self.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE")
        self.client.force_login(self.team_coach)

        response = self.club_post("event_series_create", self.series_data(location=str(location.pk)))

        series = EventSeries.objects.get(title="Weekly training")
        self.assertRedirects(response, reverse("management:event_series_detail", args=[series.pk]))
        self.assertEqual(series.location, location)

    def test_creating_a_series_generates_occurrences_immediately(self):
        self.client.force_login(self.team_coach)

        series = self.create_series()

        self.assertTrue(series.occurrences.exists())

    def test_a_group_member_can_create_a_series_for_their_own_group(self):
        group = Group.objects.create(club=self.club, name="Committee")
        member_user = User.objects.create_user(email="committee-series@example.com", password="pw-secret-123")
        member = Member.objects.create(user=member_user, first_name="Gale", last_name="Group")
        physio_position = Position.objects.create(club=self.club, name="Series Physio", short_name="SP", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=self.other_team, member=member, season=self.season, position=physio_position)
        GroupMembership.objects.create(group=group, member=member)
        self.client.force_login(member_user)

        response = self.club_post("event_series_create", self.series_data(teams=[], groups=[str(group.pk)]))

        series = EventSeries.objects.get(title="Weekly training")
        self.assertRedirects(response, reverse("management:event_series_detail", args=[series.pk]))
        self.assertIn(group, series.groups.all())

    def test_generated_occurrences_copy_groups_and_club_wide(self):
        self.client.force_login(self.admin_user)
        series = EventSeries.objects.create(club=self.club, title="AGM", kind=Event.EventKind.MEETING, dtstart=timezone.now() + datetime.timedelta(days=1), rrule="FREQ=WEEKLY;COUNT=1", club_wide=True)

        generate_occurrences(series)

        occurrence = series.occurrences.get()
        self.assertTrue(occurrence.club_wide)

    def test_editing_a_series_propagates_to_future_occurrences_but_not_a_detached_one(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        detached = series.occurrences.filter(start__gte=timezone.now()).order_by("start").first()
        other = series.occurrences.exclude(pk=detached.pk).filter(start__gte=timezone.now()).first()
        detach_occurrence(detached)

        self.club_post("event_series_update", self.series_data(title="Renamed training"), series.pk)

        detached.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(detached.title, "Weekly training")
        self.assertEqual(other.title, "Renamed training")

    def test_cancelling_an_occurrence_records_an_excluded_date_not_a_raw_delete(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        occurrence = series.occurrences.order_by("start").first()
        start_iso = occurrence.start.isoformat()

        self.club_post("event_delete", {}, occurrence.pk)

        self.assertFalse(Event.objects.filter(pk=occurrence.pk).exists())
        series.refresh_from_db()
        self.assertIn(start_iso, series.excluded_dates)

    def test_cancelling_with_keep_record_marks_it_cancelled_instead_of_deleting(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        occurrence = series.occurrences.order_by("start").first()

        self.club_post("event_delete", {"keep_record": "on"}, occurrence.pk)

        occurrence.refresh_from_db()
        self.assertTrue(occurrence.cancelled)

    def test_stop_repeating_prevents_new_occurrences_without_touching_existing_ones(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        count_before = series.occurrences.count()

        self.club_post("event_series_stop", {}, series.pk)
        series.refresh_from_db()
        generate_occurrences(series)

        self.assertEqual(series.occurrences.count(), count_before)

    def test_a_manager_of_a_different_team_cannot_edit_the_series(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        self.client.force_login(self.other_team_coach)

        response = self.club_get("event_series_update", series.pk)

        self.assertEqual(response.status_code, 403)

    def test_a_manager_of_a_different_team_cannot_view_the_series_either(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        self.client.force_login(self.other_team_coach)

        response = self.club_get("event_series_detail", series.pk)

        self.assertEqual(response.status_code, 404)

    def test_deleting_a_series_deletes_its_occurrences(self):
        self.client.force_login(self.team_coach)
        series = self.create_series()
        occurrence_ids = list(series.occurrences.values_list("pk", flat=True))

        self.club_post("event_series_delete", {}, series.pk)

        self.assertFalse(EventSeries.objects.filter(pk=series.pk).exists())
        self.assertFalse(Event.objects.filter(pk__in=occurrence_ids).exists())


class EventDetailDisplayTests(ManagementTestBase):
    """The event detail page's RSVP breakdown/modal and the game "fetch info"
    stub -- see management.views.EventDetailView/EventFetchGameInfoView and
    events.services.competitions.fetch_game_info."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.position = Position.objects.create(club=cls.club, name="Forward", short_name="FW")
        cls.player = Member.objects.create(first_name="Peter", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        TeamMembership.objects.create(team=cls.team, season=cls.season, member=cls.player, position=cls.position)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_event(self, **kwargs):
        kwargs.setdefault("title", "Training")
        kwargs.setdefault("kind", Event.EventKind.TRAINING)
        kwargs.setdefault("start", timezone.now() + datetime.timedelta(days=1))
        event = Event.objects.create(club=self.club, **kwargs)
        event.teams.add(self.team)
        return event

    def test_the_rsvp_breakdown_shows_every_status_even_at_zero(self):
        event = self.make_event()

        response = self.club_get("event_detail", event.pk)

        for label in ["Present", "Absent", "Excused", "Selected", "Not selected", "Maybe", "No response"]:
            self.assertContains(response, label)

    def test_the_rsvp_modal_shows_who_responded_and_their_note(self):
        event = self.make_event()
        Attendance.objects.filter(event=event, member=self.player).update(status=Attendance.AttendanceStatus.PRESENT, note="Bringing the kit bag")

        response = self.club_get("event_detail", event.pk)

        self.assertContains(response, "Peter Player")
        self.assertContains(response, "Bringing the kit bag")

    def test_the_rsvp_modal_offers_a_filter_pill_per_status_with_responses(self):
        event = self.make_event()
        Attendance.objects.filter(event=event, member=self.player).update(status=Attendance.AttendanceStatus.PRESENT)

        response = self.club_get("event_detail", event.pk)

        # One pill per status that actually has a response ("Present"), none for
        # the others (Absent, Excused, ... all sit at 0 for this single-player event).
        self.assertContains(response, 'data-status="present"')
        self.assertNotContains(response, 'data-status="absent"')
        self.assertContains(response, 'data-status=""')  # the "All" pill

    def test_the_rsvp_modal_tags_each_row_with_its_status_for_client_side_filtering(self):
        event = self.make_event()
        Attendance.objects.filter(event=event, member=self.player).update(status=Attendance.AttendanceStatus.PRESENT)

        response = self.club_get("event_detail", event.pk)

        self.assertContains(response, '<tr data-status="present">')

    def test_the_details_card_shows_gathering_and_deadline_even_when_unset(self):
        event = self.make_event()

        response = self.club_get("event_detail", event.pk)

        self.assertContains(response, "Gathering")
        self.assertContains(response, "Registration deadline")

    def test_the_fetch_button_only_shows_for_a_game_with_a_competition_set(self):
        game_with_competition = self.make_event(title="Cup game", kind=Event.EventKind.GAME, competition="Regional Cup")
        game_without_competition = self.make_event(title="Friendly game", kind=Event.EventKind.GAME)
        training = self.make_event(title="Training session")

        self.assertContains(self.club_get("event_detail", game_with_competition.pk), "Fetch new game info")
        self.assertNotContains(self.club_get("event_detail", game_without_competition.pk), "Fetch new game info")
        self.assertNotContains(self.club_get("event_detail", training.pk), "Fetch new game info")

    def test_fetching_game_info_reports_that_nothing_is_configured_yet(self):
        # The competition's flag must be active for this club, or fetch_game_info
        # gates before ever getting as far as "no data source configured" -- see
        # test_fetching_game_info_is_a_silent_no_op_when_the_flag_is_not_active.
        # module points nowhere real -- a genuine configuration problem
        # (ModuleNotFoundError resolving it), not a fetch failure, so this is
        # the one case that still reports "isn't wired up".
        Flag = get_waffle_flag_model()
        flag = Flag.objects.create(name="regional-cup")
        flag.clubs.add(self.club)
        Competition.objects.create(name="Regional Cup", module="events.competition.regional", flag=flag)
        game = self.make_event(title="Cup game", kind=Event.EventKind.GAME, competition="Regional Cup")

        redirect = self.club_post("event_fetch_game_info", {}, game.pk)
        response = self.club_get("event_detail", game.pk)

        self.assertRedirects(redirect, reverse("management:event_detail", args=[game.pk]))
        self.assertContains(response, "wired up to a real data source yet")

    def test_fetching_game_info_reports_a_real_fetch_failure_distinctly(self):
        # A competition that resolves fine (real module/class) but fails for
        # some other reason (network, bad response, ...) must not be told
        # it's "not configured" -- that's exactly what confused the report
        # this test guards against: a genuinely-linked competition that
        # looked broken because the *real* error was masked.
        # RBIHF is seeded by a data migration (events/migrations/0017/0018)
        # into every database, own flag included -- creating a second
        # Competition row named "RBIHF" here would just be a duplicate
        # .filter(name=...).first() might not even pick, so this grants the
        # seeded one's own flag access instead.
        competition = Competition.objects.get(name="RBIHF")
        competition.flag.clubs.add(self.club)
        game = self.make_event(title="Cup game", kind=Event.EventKind.GAME, competition="RBIHF", external_game_id="1234")

        with mock.patch("events.competition.hockey.requests.get", side_effect=ConnectionError("boom")):
            redirect = self.club_post("event_fetch_game_info", {}, game.pk)
        response = self.club_get("event_detail", game.pk)

        self.assertRedirects(redirect, reverse("management:event_detail", args=[game.pk]))
        self.assertNotContains(response, "wired up to a real data source")
        self.assertContains(response, "Could not fetch game info from")

    def test_fetching_game_info_resolves_the_season_when_the_event_has_none_set(self):
        # Event.season is blank whenever nobody explicitly picked one --
        # update_game_information used to assume it was always set and
        # crashed with an AttributeError, which fetch_game_info's own old
        # blanket except then reported as "not configured" even though the
        # competition was linked correctly.
        competition = Competition.objects.get(name="RBIHF")
        competition.flag.clubs.add(self.club)
        game = self.make_event(title="Cup game", kind=Event.EventKind.GAME, competition="RBIHF", external_game_id="1234", season=None)
        self.assertIsNone(game.season)

        with mock.patch("events.competition.hockey.requests.get") as mock_get:
            mock_get.return_value.status_code = 200
            mock_get.return_value.json.return_value = {"live": True, "scoreA": 2, "scoreB": 1}
            redirect = self.club_post("event_fetch_game_info", {}, game.pk)

        self.assertRedirects(redirect, reverse("management:event_detail", args=[game.pk]))
        game.refresh_from_db()
        self.assertTrue(game.is_live)

    def test_fetching_game_info_is_a_silent_no_op_when_the_flag_is_not_active(self):
        # No matching Competition row at all -- same "nothing to gate on" outcome
        # as one that exists but whose flag isn't active for this club.
        game = self.make_event(title="Cup game", kind=Event.EventKind.GAME, competition="Regional Cup")

        redirect = self.club_post("event_fetch_game_info", {}, game.pk)
        response = self.club_get("event_detail", game.pk)

        self.assertRedirects(redirect, reverse("management:event_detail", args=[game.pk]))
        self.assertNotContains(response, "No competition data source is configured yet")
        self.assertContains(response, "is not enabled for this club")


class EventRefereeManagementTests(ManagementTestBase):
    """Assigning/removing referees from the event detail page's Referees
    panel -- home games only, see management.views.EventRefereeAssignView/
    EventRefereeRemoveView and events.services.referees."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.home_ground = Location.objects.create(club=cls.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE", is_home=True)
        cls.away_ground = Location.objects.create(club=cls.club, name="Away Ground", address="2 St", city="Town", zip_code="1000", country="BE")
        cls.level = RefereeLevel.objects.create(club=cls.club, name="Regional")
        cls.level.teams.add(cls.team)
        cls.referee = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=cls.club, member=cls.referee, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        RefereeProfile.objects.create(member=cls.referee, level=cls.level, valid_until=timezone.localdate() + datetime.timedelta(days=30))

        cls.coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        # Manages this team, which still isn't enough to touch referees -- that's
        # admin-only, so this actor recurs throughout the 403 checks below.
        cls.team_coach = cls.make_coach(cls.team, "coach-referees@example.com")

    @classmethod
    def make_coach(cls, team, email):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=team, member=coach_member, season=cls.season, position=cls.coach_position)
        return coach_user

    def make_game(self, **kwargs):
        kwargs.setdefault("title", "Cup game")
        kwargs.setdefault("kind", Event.EventKind.GAME)
        kwargs.setdefault("location", self.home_ground)
        kwargs.setdefault("start", timezone.now() + datetime.timedelta(days=1))
        event = Event.objects.create(club=self.club, **kwargs)
        event.teams.add(self.team)
        return event

    def test_the_referees_panel_only_shows_for_a_home_game(self):
        # "Referees" alone also matches the nav link on every page -- assert on
        # text unique to the panel itself.
        home_game = self.make_game()
        away_game = self.make_game(title="Away game", location=self.away_ground)
        self.client.force_login(self.admin_user)

        self.assertContains(self.club_get("event_detail", home_game.pk), "No referees assigned yet.")
        self.assertNotContains(self.club_get("event_detail", away_game.pk), "No referees assigned yet.")

    def test_the_referees_panel_is_replaced_by_a_note_for_a_federation_managed_team(self):
        self.team.referee_management = Team.RefereeManagement.FEDERATION
        self.team.save(update_fields=["referee_management"])
        home_game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", home_game.pk)

        self.assertNotContains(response, "No referees assigned yet.")
        self.assertContains(response, "managed by the federation")

    def test_cannot_assign_a_referee_to_a_federation_managed_teams_game(self):
        self.team.referee_management = Team.RefereeManagement.FEDERATION
        self.team.save(update_fields=["referee_management"])
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        self.assertEqual(response.status_code, 404)
        self.assertFalse(EventReferee.objects.filter(event=game).exists())

    def test_a_teams_own_coach_gets_403_assigning_a_referee(self):
        # Admin-only for now, even for the coach who manages this team --
        # see management.views.EventRefereeAssignView.
        game = self.make_game()
        self.client.force_login(self.team_coach)

        response = self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(EventReferee.objects.filter(event=game, member=self.referee).exists())

    def test_a_teams_own_coach_gets_403_removing_a_referee(self):
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.team_coach)

        response = self.club_post("event_referee_remove", {}, game.pk, assignment.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(EventReferee.objects.filter(event=game, member=self.referee).exists())

    def test_a_teams_coach_sees_the_referees_panel_but_not_the_assign_or_remove_controls(self):
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.team_coach)

        response = self.club_get("event_detail", game.pk)

        # The panel itself, and who's assigned, are still visible...
        self.assertContains(response, "Ref Eree")
        self.assertContains(response, "1 / 2")
        # ...but not the controls to change it.
        self.assertNotContains(response, reverse("management:event_referee_assign", args=[game.pk]))
        self.assertNotContains(response, reverse("management:event_referee_remove", args=[game.pk, assignment.pk]))

    def test_admin_can_assign_an_eligible_referee(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))
        self.assertTrue(EventReferee.objects.filter(event=game, member=self.referee).exists())

    def test_assigning_records_which_admin_assigned_them(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        assignment = EventReferee.objects.get(event=game, member=self.referee)
        self.assertEqual(assignment.assigned_by, self.admin_member)

    def test_cannot_assign_beyond_max_referees(self):
        game = self.make_game(max_referees=1)
        second_referee = Member.objects.create(first_name="Second", last_name="Ref")
        RefereeProfile.objects.create(member=second_referee, level=self.level, valid_until=timezone.localdate() + datetime.timedelta(days=30))
        self.client.force_login(self.admin_user)
        self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        response = self.club_post("event_referee_assign", {"member": str(second_referee.pk)}, game.pk)

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))
        self.assertEqual(EventReferee.objects.filter(event=game).count(), 1)

    def test_cannot_assign_a_referee_to_an_away_game(self):
        # eligible_referees() is already empty for a non-home game, so the
        # attempted member isn't found at all -- same 404 as any other
        # crafted POST naming someone who isn't a legitimate candidate.
        game = self.make_game(location=self.away_ground)
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        self.assertEqual(response.status_code, 404)
        self.assertFalse(EventReferee.objects.filter(event=game).exists())

    def test_cannot_assign_someone_not_eligible_via_a_crafted_post(self):
        game = self.make_game()
        ineligible = Member.objects.create(first_name="Not", last_name="Eligible")
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_assign", {"member": str(ineligible.pk)}, game.pk)

        self.assertEqual(response.status_code, 404)
        self.assertFalse(EventReferee.objects.filter(event=game).exists())

    def test_a_different_teams_coach_cannot_assign_a_referee(self):
        other_team = Team.objects.create(club=self.club, name="Second Team", short_name="2nd")
        game = self.make_game()
        self.client.force_login(self.make_coach(other_team, "coach-referees-other@example.com"))

        response = self.club_post("event_referee_assign", {"member": str(self.referee.pk)}, game.pk)

        self.assertEqual(response.status_code, 403)

    def test_can_remove_an_assigned_referee(self):
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_remove", {}, game.pk, assignment.pk)

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))
        self.assertFalse(EventReferee.objects.filter(event=game, member=self.referee).exists())

    def test_conflict_warning_shown_but_does_not_block_the_assign_control(self):
        # The referee is also on this team's roster and expected at an
        # overlapping training -- shown as a warning, still selectable.
        position = Position.objects.create(club=self.club, name="Forward", short_name="FW")
        TeamMembership.objects.create(team=self.team, member=self.referee, season=self.season, position=position)
        game = self.make_game(start=timezone.now() + datetime.timedelta(days=1))
        clashing_training = Event.objects.create(club=self.club, title="Clashing training", kind=Event.EventKind.TRAINING, start=game.start, end=game.start + datetime.timedelta(hours=1))
        clashing_training.teams.add(self.team)
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertContains(response, "⚠")
        self.assertContains(response, f'value="{self.referee.pk}"')

    def test_admin_can_add_an_external_referee(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_add_external", {"name": "Guest Referee"}, game.pk)

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))
        assignment = EventReferee.objects.get(event=game, external_name="Guest Referee")
        self.assertIsNone(assignment.member)

    def test_adding_an_external_referee_with_a_blank_name_is_rejected(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        self.club_post("event_referee_add_external", {"name": "  "}, game.pk)

        self.assertFalse(EventReferee.objects.filter(event=game).exists())

    def test_a_coach_gets_403_adding_an_external_referee(self):
        game = self.make_game()
        self.client.force_login(self.team_coach)

        response = self.club_post("event_referee_add_external", {"name": "Guest Referee"}, game.pk)

        self.assertEqual(response.status_code, 403)

    def test_event_detail_page_shows_an_external_referee(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, external_name="Guest Referee", assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertContains(response, "Guest Referee")
        self.assertContains(response, "External")

    def test_admin_can_set_a_referees_fee(self):
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_fee_update", {"fee": "25.00", "km": "40", "km_rate": "0.35"}, game.pk, assignment.pk)

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))
        assignment.refresh_from_db()
        self.assertEqual(assignment.fee, Decimal("25.00"))
        self.assertEqual(assignment.total_payable, Decimal("39.00"))

    def test_a_km_rate_with_more_than_two_decimals_is_accepted(self):
        # e.g. a per-km rate of €0.083 -- the km_rate input must not be pinned
        # to money-style step="0.01" the way the fee field is.
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_post("event_referee_fee_update", {"fee": "0", "km": "40", "km_rate": "0.083"}, game.pk, assignment.pk)

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))
        assignment.refresh_from_db()
        self.assertEqual(assignment.km_rate, Decimal("0.083"))

    def test_event_detail_page_shows_the_total_due_once_a_fee_is_set(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"))
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertContains(response, "25.00")

    def test_the_total_due_is_shown_with_at_most_two_decimals(self):
        # A per-km rate like 0.083 pushes the raw total to 3+ decimals -- the
        # "due" summary must still round to money-style 2.
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"), km=Decimal("40"), km_rate=Decimal("0.083"))
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertContains(response, "28.32 due")
        self.assertNotContains(response, "28.320")

    def test_the_fee_is_entered_inline_pre_filled_with_the_current_value(self):
        # Editing a fee is a single input right on the referee's row -- no
        # separate modal to click through, so it's less likely to be skipped.
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"))
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertContains(response, 'name="fee" value="25.00"')

    def test_an_unset_fee_is_flagged_on_the_referees_panel(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertContains(response, "Fee not set")

    def test_a_fee_already_set_is_not_flagged(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"))
        self.client.force_login(self.admin_user)

        response = self.club_get("event_detail", game.pk)

        self.assertNotContains(response, "Fee not set")

    def test_a_coach_gets_403_setting_a_fee(self):
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.team_coach)

        response = self.club_post("event_referee_fee_update", {"fee": "25.00"}, game.pk, assignment.pk)

        self.assertEqual(response.status_code, 403)


class EventRefereeFormPdfTests(ManagementTestBase):
    """Downloadable referee payment form -- see management.views.EventRefereeFormPdfView,
    modeled on the club's existing paper form."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.home_ground = Location.objects.create(club=cls.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE", is_home=True)
        cls.referee = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=cls.club, member=cls.referee, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

    def make_coach(self, team, email="coach-refpdf@example.com"):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        return coach_user

    def make_game(self, **kwargs):
        kwargs.setdefault("title", "Cup game")
        kwargs.setdefault("kind", Event.EventKind.GAME)
        kwargs.setdefault("location", self.home_ground)
        kwargs.setdefault("start", timezone.now() + datetime.timedelta(days=1))
        event = Event.objects.create(club=self.club, **kwargs)
        event.teams.add(self.team)
        return event

    def test_downloads_as_a_pdf(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        with mock.patch("management.views.event_referee_form_pdf", return_value=b"%PDF-fake") as renderer:
            response = self.club_get("event_referee_form_pdf", game.pk)

        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(".pdf", response["Content-Disposition"])
        self.assertEqual(response.content, b"%PDF-fake")
        renderer.assert_called_once()

    def test_uses_the_clubs_legal_name_and_home_location_when_set(self):
        self.club.legal_name = "Ajax United VZW"
        self.club.save(update_fields=["legal_name"])
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"), km=Decimal("40"), km_rate=Decimal("0.35"))
        self.client.force_login(self.admin_user)

        with mock.patch("management.views.event_referee_form_pdf", return_value=b"%PDF-fake") as renderer:
            self.club_get("event_referee_form_pdf", game.pk)

        context = renderer.call_args[0][0]
        self.assertEqual(context["club"].official_name, "Ajax United VZW")
        self.assertEqual(context["document_address"], self.home_ground)
        self.assertEqual(list(context["referees"]), [EventReferee.objects.get(event=game)])

    def test_the_grand_total_sums_every_referees_total_payable(self):
        game = self.make_game()
        other_referee = Member.objects.create(first_name="Other", last_name="Ref")
        ClubMembership.objects.create(club=self.club, member=other_referee, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"), km=Decimal("40"), km_rate=Decimal("0.083"))
        EventReferee.objects.create(event=game, member=other_referee, assigned_by=self.admin_member, fee=Decimal("20.00"))
        self.client.force_login(self.admin_user)

        with mock.patch("management.views.event_referee_form_pdf", return_value=b"%PDF-fake") as renderer:
            self.club_get("event_referee_form_pdf", game.pk)

        self.assertEqual(renderer.call_args[0][0]["grand_total"], Decimal("48.320"))

    def test_the_external_referee_pill_is_not_rendered(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, external_name="Guest Referee", assigned_by=self.admin_member)

        html = render_to_string("management/event_referee_form_pdf.html", {"club": self.club, "event": game, "referees": list(game.referees.all()), "document_address": self.home_ground, "grand_total": Decimal("0")})

        self.assertIn("Guest Referee", html)
        self.assertNotIn("External", html)

    def test_amounts_render_with_at_most_two_decimals(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"), km=Decimal("40"), km_rate=Decimal("0.083"))

        html = render_to_string("management/event_referee_form_pdf.html", {"club": self.club, "event": game, "referees": list(game.referees.all()), "document_address": self.home_ground, "grand_total": Decimal("28.320")})

        self.assertIn("€28.32<", html)
        self.assertNotIn("28.320", html)

    def test_the_pdf_uses_the_clubs_colours_when_set(self):
        self.club.primary_color = "#0f766e"
        self.club.secondary_color = "#f59e0b"
        self.club.save(update_fields=["primary_color", "secondary_color"])
        game = self.make_game()

        html = render_to_string("management/event_referee_form_pdf.html", {"club": self.club, "event": game, "referees": [], "document_address": self.home_ground, "grand_total": Decimal("0"), **referee_form_colors(self.club)})

        self.assertIn("--accent: #0f766e", html)

    def test_the_pdf_falls_back_to_default_colours_when_unset(self):
        game = self.make_game()

        html = render_to_string("management/event_referee_form_pdf.html", {"club": self.club, "event": game, "referees": [], "document_address": self.home_ground, "grand_total": Decimal("0"), **referee_form_colors(self.club)})

        self.assertIn("--accent: #3730a3", html)

    def test_the_info_card_background_is_not_left_to_unsupported_css(self):
        # WeasyPrint doesn't support color-mix() -- the background must be a
        # plain computed hex value baked into the template, or the card
        # silently renders with no background at all.
        game = self.make_game()

        html = render_to_string("management/event_referee_form_pdf.html", {"club": self.club, "event": game, "referees": [], "document_address": self.home_ground, "grand_total": Decimal("0"), **referee_form_colors(self.club)})

        self.assertNotIn("color-mix(", html)
        self.assertIn("--info-card-bg: #", html)

    def test_the_info_card_falls_back_to_secondary_when_primary_is_near_white(self):
        self.club.primary_color = "#ffffff"
        self.club.secondary_color = "#f59e0b"
        self.club.save(update_fields=["primary_color", "secondary_color"])

        colors = referee_form_colors(self.club)

        self.assertEqual(colors["accent_color"], "#ffffff")
        self.assertNotEqual(colors["info_card_color"], "#ffffff")

    def test_the_info_card_uses_primary_when_it_is_not_near_black_or_white(self):
        self.club.primary_color = "#0f766e"
        self.club.save(update_fields=["primary_color"])

        colors = referee_form_colors(self.club)

        self.assertEqual(colors["info_card_color"], _tint_with_white("#0f766e"))

    def test_a_missing_pdf_library_is_reported_rather_than_a_500(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        with mock.patch("management.views.event_referee_form_pdf", side_effect=PDFExportError("PDF rendering needs the native pango/cairo libraries.")):
            response = self.club_get("event_referee_form_pdf", game.pk)
        response = self.client.get(response.url, HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "pango")

    def test_a_coach_gets_403(self):
        game = self.make_game()
        self.client.force_login(self.make_coach(self.team))

        response = self.club_get("event_referee_form_pdf", game.pk)

        self.assertEqual(response.status_code, 403)


class RefereeManagementDashboardTests(ManagementTestBase):
    """The admin-only one-stop view of upcoming home games needing a
    club-arranged referee -- see management.views.RefereeManagementDashboardView."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")
        cls.federation_team = Team.objects.create(club=cls.club, name="Federation Team", short_name="Fed", referee_management=Team.RefereeManagement.FEDERATION)
        cls.home_ground = Location.objects.create(club=cls.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE", is_home=True)
        cls.away_ground = Location.objects.create(club=cls.club, name="Away Ground", address="2 St", city="Town", zip_code="1000", country="BE")
        cls.level = RefereeLevel.objects.create(club=cls.club, name="Regional")
        cls.level.teams.add(cls.team)
        cls.referee = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=cls.club, member=cls.referee, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        RefereeProfile.objects.create(member=cls.referee, level=cls.level, valid_until=timezone.localdate() + datetime.timedelta(days=30))

    def make_coach(self, team, email="coach-refdash@example.com"):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        return coach_user

    def make_game(self, team=None, **kwargs):
        kwargs.setdefault("title", "Cup game")
        kwargs.setdefault("kind", Event.EventKind.GAME)
        kwargs.setdefault("location", self.home_ground)
        kwargs.setdefault("start", timezone.now() + datetime.timedelta(days=1))
        event = Event.objects.create(club=self.club, **kwargs)
        event.teams.add(team or self.team)
        return event

    def test_is_admin_only(self):
        self.client.force_login(self.make_coach(self.team))
        self.assertEqual(self.club_get("referee_management").status_code, 403)

    def test_lists_an_upcoming_club_managed_home_game(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertContains(response, reverse("management:event_detail", args=[game.pk]))

    def test_lists_an_upcoming_club_managed_home_tournament(self):
        # A home tournament needs officiating arranged just as much as a home
        # game -- Event.is_home_fixture, not the Game-only is_home_game, is
        # what this dashboard's own queries key off.
        tournament = self.make_game(title="Regional Cup", kind=Event.EventKind.TOURNAMENT)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertContains(response, reverse("management:event_detail", args=[tournament.pk]))

    def test_each_game_tile_links_straight_to_the_referee_form_pdf(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertContains(response, reverse("management:event_referee_form_pdf", args=[game.pk]))

    def test_the_game_tile_shows_assigned_referee_names(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertContains(response, str(self.referee))

    def test_a_game_with_an_unpaid_referee_is_flagged_as_fees_pending(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_fees_pending"], 1)
        self.assertContains(response, "Fee not set")

    def test_a_game_with_a_fee_already_set_is_not_flagged(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member, fee=Decimal("25.00"))
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_fees_pending"], 0)
        self.assertNotContains(response, "Fee not set")

    def test_a_game_with_no_referee_at_all_is_not_counted_as_fees_pending(self):
        # Missing a referee entirely is a different, already-visible problem
        # (kpi_no_referee) -- it shouldn't also inflate the fees-pending count.
        self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_fees_pending"], 0)

    def test_excludes_a_federation_managed_teams_game(self):
        game = self.make_game(team=self.federation_team)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertNotContains(response, reverse("management:event_detail", args=[game.pk]))

    def test_excludes_an_away_game(self):
        game = self.make_game(location=self.away_ground)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertNotContains(response, reverse("management:event_detail", args=[game.pk]))

    def test_excludes_a_past_game(self):
        game = self.make_game(start=timezone.now() - datetime.timedelta(days=1))
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertNotContains(response, reverse("management:event_detail", args=[game.pk]))

    def test_excludes_a_cancelled_game(self):
        game = self.make_game(cancelled=True)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertNotContains(response, reverse("management:event_detail", args=[game.pk]))

    def test_an_out_of_range_value_falls_back_to_the_default(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")
        response_with_bad_range = self.client.get(f"{reverse('management:referee_management')}?range=bogus", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.context["range_choice"], "10")
        self.assertEqual(response_with_bad_range.context["range_choice"], "10")

    def test_a_valid_range_is_honoured(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(f"{reverse('management:referee_management')}?range=25", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertEqual(response.context["range_choice"], "25")

    def test_the_week_range_excludes_a_game_beyond_this_week(self):
        today = timezone.localdate()
        end_of_this_week = today + datetime.timedelta(days=6 - today.weekday())
        game_this_week = self.make_game(start=timezone.now() + datetime.timedelta(minutes=5))
        game_next_week = self.make_game(start=timezone.make_aware(datetime.datetime.combine(end_of_this_week + datetime.timedelta(days=1), datetime.time(10, 0))))
        self.client.force_login(self.admin_user)

        response = self.client.get(f"{reverse('management:referee_management')}?range=week", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, reverse("management:event_detail", args=[game_this_week.pk]))
        self.assertNotContains(response, reverse("management:event_detail", args=[game_next_week.pk]))

    def test_kpis_count_games_by_referee_staffing(self):
        self.make_game()
        partially_staffed = self.make_game()
        EventReferee.objects.create(event=partially_staffed, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_total"], 2)
        self.assertEqual(response.context["kpi_no_referee"], 1)
        self.assertEqual(response.context["kpi_understaffed"], 1)
        self.assertEqual(response.context["kpi_fully_staffed"], 0)

    def test_an_assigned_referee_gets_a_fee_form_for_the_dashboard_modal(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertContains(response, 'name="fee"')

    def test_assigning_from_the_dashboard_redirects_back_to_the_dashboard(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("management:event_referee_assign", args=[game.pk]),
            {"member": str(self.referee.pk), "next": reverse("management:referee_management")},
            HTTP_HOST="ajax-united.rosterchief.app",
        )

        self.assertRedirects(response, reverse("management:referee_management"))
        self.assertTrue(EventReferee.objects.filter(event=game, member=self.referee).exists())

    def test_removing_from_the_dashboard_redirects_back_to_the_dashboard(self):
        game = self.make_game()
        assignment = EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("management:event_referee_remove", args=[game.pk, assignment.pk]),
            {"next": reverse("management:referee_management")},
            HTTP_HOST="ajax-united.rosterchief.app",
        )

        self.assertRedirects(response, reverse("management:referee_management"))
        self.assertFalse(EventReferee.objects.filter(event=game, member=self.referee).exists())

    def test_an_unsafe_next_is_ignored(self):
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse("management:event_referee_assign", args=[game.pk]),
            {"member": str(self.referee.pk), "next": "https://evil.example.com/"},
            HTTP_HOST="ajax-united.rosterchief.app",
        )

        self.assertRedirects(response, reverse("management:event_detail", args=[game.pk]))

    def test_shows_a_pending_self_service_invite(self):
        # make_game()'s own event.teams.add(...) already triggers
        # events.services.referees.sync_referee_invites via events/signals.py --
        # self.referee (eligible for self.team) gets invited automatically.
        game = self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertTrue(RefereeSignup.objects.filter(event=game, member=self.referee, status=RefereeSignup.Status.INVITED).exists())
        self.assertContains(response, "Invited, no response yet")
        self.assertContains(response, str(self.referee))

    def test_no_pending_invite_shown_once_the_game_is_full(self):
        game = self.make_game(max_referees=1)
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        other_referee = Member.objects.create(first_name="Other", last_name="Ref")
        RefereeSignup.objects.create(event=game, member=other_referee, status=RefereeSignup.Status.INVITED)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertNotContains(response, "Invited, no response yet")

    def test_self_signed_up_referee_is_flagged_distinctly_from_an_admin_assignment(self):
        game = self.make_game()
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=None)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertContains(response, "Self sign-up")

    def test_workload_stats_count_games_refereed_this_season(self):
        game_one = self.make_game()
        game_two = self.make_game(title="Second game", start=timezone.now() + datetime.timedelta(days=2))
        EventReferee.objects.create(event=game_one, member=self.referee, assigned_by=self.admin_member, fee=Decimal("20.00"))
        EventReferee.objects.create(event=game_two, member=self.referee, assigned_by=self.admin_member, fee=Decimal("15.00"))
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_active_referees"], 1)
        self.assertEqual(response.context["kpi_total_assignments"], 2)
        self.assertEqual(response.context["kpi_avg_games_per_referee"], 2)

    def test_workload_stats_exclude_external_referees(self):
        game = self.make_game()
        add_external_referee(game, "Guest Ref", assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_active_referees"], 0)

    def test_workload_stats_empty_state(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["kpi_active_referees"], 0)
        self.assertEqual(response.context["kpi_avg_games_per_referee"], 0)

    # --- the "need" toggle (referees / officials / both) ---

    def setUp_officials(self):
        # get_or_create, not create -- events/migrations/0031_create_officials_flag
        # already seeded this row at migrate time, so a plain create() collides with
        # it. waffle also caches Flag lookups outside the DB transaction each test
        # rolls back -- see FeatureGatedSectionsTests' own identical setUp for why.
        cache.clear()
        self.addCleanup(cache.clear)
        get_waffle_flag_model().objects.get_or_create(name="officials")[0].clubs.add(self.club)

    def test_the_toggle_is_hidden_when_the_officials_flag_is_off(self):
        self.make_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertNotContains(response, "Officials")

    def test_a_fully_staffed_game_is_not_missing_a_referee(self):
        # Regression: one internal + one external referee is 2 rows against
        # Event.max_referees' own default of 2 -- fully staffed, so this game
        # must not appear in "Missing referees" (or its count) just because
        # it's still under this club's referee governance. Governance
        # (upcoming_games_needing_referee_management) and "missing" (referee_
        # count < max_referees) are different questions -- this toggle answers
        # the second one.
        game = self.make_game(title="Fully staffed")
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        add_external_referee(game, "Guest Ref", assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management", params={"need": "referees"})

        self.assertEqual(response.context["need_counts"]["referees"], 0)
        self.assertEqual(list(response.context["games"]), [])

    def test_a_fully_staffed_game_is_excluded_from_missing_either_too(self):
        game = self.make_game(title="Fully staffed")
        EventReferee.objects.create(event=game, member=self.referee, assigned_by=self.admin_member)
        add_external_referee(game, "Guest Ref", assigned_by=self.admin_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["need_choice"], "both")
        self.assertEqual(list(response.context["games"]), [])

    def test_each_buckets_count_matches_what_it_actually_shows(self):
        self.setUp_officials()
        # self.team defaults to referee_management=CLUB *and* official_management=CLUB,
        # so this game needs both. federation_team opted out of referee tools
        # (RefereeManagement.FEDERATION) but still defaults to official_management=CLUB,
        # so its game needs officials only -- see teams.models.Team's own docstring on
        # why the two are independent flags.
        self.make_game(title="Needs both")
        self.make_game(title="Needs officials only", team=self.federation_team)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["need_counts"], {"referees": 1, "officials": 2, "both": 2})

    def test_referees_filter_excludes_an_officials_only_game(self):
        self.setUp_officials()
        both_game = self.make_game(title="Needs both")
        self.make_game(title="Needs officials only", team=self.federation_team)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management", params={"need": "referees"})

        self.assertEqual(response.context["need_choice"], "referees")
        games = response.context["games"]
        self.assertEqual([game.pk for game in games], [both_game.pk])

    def test_officials_filter_excludes_a_referees_only_game(self):
        self.setUp_officials()
        # official_management=FEDERATION opts a team's games out of officials
        # tools entirely, same as referee_management=FEDERATION does for
        # referees -- see teams.models.Team's own docstring.
        referees_only_team = Team.objects.create(club=self.club, name="Referees Only", short_name="RO", official_management=Team.OfficialManagement.FEDERATION)
        officials_only_game = self.make_game(title="Needs officials only", team=self.federation_team)
        self.make_game(title="Needs referees only", team=referees_only_team)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management", params={"need": "officials"})

        games = response.context["games"]
        self.assertEqual([game.pk for game in games], [officials_only_game.pk])

    def test_both_is_the_default_and_shows_every_game(self):
        self.setUp_officials()
        both_game = self.make_game(title="Needs both")
        officials_only_game = self.make_game(title="Needs officials only", team=self.federation_team)
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management")

        self.assertEqual(response.context["need_choice"], "both")
        games = response.context["games"]
        self.assertEqual({game.pk for game in games}, {both_game.pk, officials_only_game.pk})

    def test_an_unknown_need_value_falls_back_to_both(self):
        self.setUp_officials()
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management", params={"need": "bogus"})

        self.assertEqual(response.context["need_choice"], "both")

    def test_need_officials_falls_back_to_both_when_the_flag_is_off(self):
        # No setUp_officials() here -- the flag stays off for this club.
        self.client.force_login(self.admin_user)

        response = self.club_get("referee_management", params={"need": "officials"})

        self.assertEqual(response.context["need_choice"], "both")


class FeatureGatedSectionsTests(ManagementTestBase):
    """The Shop and Forms sections are still stubs (StubListMixin) and, on top
    of being admin-only, only exist for a club at all once their own waffle
    Flag ("shop" / "formbuilder") is active for it -- see
    club.mixins.FeatureRequiredMixin and management.context_processors.feature_sections."""

    def setUp(self):
        super().setUp()
        # waffle caches Flag lookups outside the DB transaction each test rolls
        # back, so a flag created in one test can otherwise leak a stale/invalid
        # pk into the next -- see FeatureViewTests in controlpanel/tests.py.
        cache.clear()
        self.addCleanup(cache.clear)

    def activate(self, flag_name):
        flag = get_waffle_flag_model().objects.create(name=flag_name)
        flag.clubs.add(self.club)

    def make_plain_staff(self, email="physio-features@example.com"):
        staff_user = User.objects.create_user(email=email, password="pw-secret-123")
        staff_member = Member.objects.create(user=staff_user, first_name="Pat", last_name="Physio")
        ClubMembership.objects.create(club=self.club, member=staff_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        team = Team.objects.create(club=self.club, name="Physio Team", short_name="PHY")
        position = Position.objects.create(club=self.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=team, member=staff_member, season=self.season, position=position)
        return staff_user

    def test_shop_views_404_when_the_flag_is_not_active(self):
        self.client.force_login(self.admin_user)

        # invoice_list is deliberately absent here -- it also covers dues/
        # registration invoices, which have nothing to do with shop, so it
        # must stay reachable (200, not 404) regardless of this flag. See
        # InvoiceListViewTests.test_an_admin_can_view_the_invoice_list_with_shop_disabled.
        for name in ["product_list", "order_list", "discount_list"]:
            with self.subTest(name=name):
                self.assertEqual(self.club_get(name).status_code, 404)

    def test_forms_view_404_when_the_flag_is_not_active(self):
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("form_list").status_code, 404)

    def test_shop_views_are_reachable_once_the_flag_is_active(self):
        self.activate("shop")
        self.client.force_login(self.admin_user)

        for name in ["product_list", "order_list", "discount_list", "invoice_list"]:
            with self.subTest(name=name):
                self.assertEqual(self.club_get(name).status_code, 200)

    def test_forms_view_is_reachable_once_its_own_flag_is_active(self):
        self.activate("formbuilder")
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("form_list").status_code, 200)

    def test_the_shop_flag_does_not_also_enable_forms(self):
        # Different flags -- see the "(but different feature)" ask.
        self.activate("shop")
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("form_list").status_code, 404)

    def test_a_non_admin_still_gets_404_not_403_when_the_flag_is_off(self):
        # The section doesn't exist for this club at all -- not a permissions
        # question, so even someone who'd otherwise be refused (403) for lack
        # of admin rights sees the same 404 an admin would.
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("product_list").status_code, 404)

    def test_a_non_admin_gets_403_once_the_flag_is_active(self):
        self.activate("shop")
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("product_list").status_code, 403)

    def test_nav_hides_shop_and_forms_when_their_flags_are_off(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertNotContains(response, "Products")
        self.assertNotContains(response, "Forms")

    def test_nav_shows_shop_once_its_flag_is_active(self):
        # Products lives under Finance's sub-nav now, which only expands on a
        # Finance-section page -- see test_the_families_nav_link_is_reachable's comment.
        self.activate("shop")
        self.client.force_login(self.admin_user)

        response = self.club_get("membership_list")

        self.assertContains(response, "Products")
        self.assertNotContains(response, "Forms")


class EvaluationManagementTestBase(ManagementTestBase):
    """Shared fixture for the evaluations desktop UI: the "evaluations" waffle
    flag active for the club, plus a plain EvaluationManager grant (broader
    than ADMIN, see EvaluationManagerRequiredMixin) and a plain staff member
    with neither."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.player = Member.objects.create(first_name="Paul", last_name="Player")
        ClubMembership.objects.create(club=cls.club, member=cls.player, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)

        cls.eval_manager_user = User.objects.create_user(email="eval-manager@example.com", password="pw-secret-123")
        cls.eval_manager_member = Member.objects.create(user=cls.eval_manager_user, first_name="Erin", last_name="Evaluator")
        ClubMembership.objects.create(club=cls.club, member=cls.eval_manager_member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        EvaluationManager.objects.create(club=cls.club, member=cls.eval_manager_member)
        # Also a plain (non-management) StaffAssignment -- has_management_access
        # (club.mixins.ClubStaffRequiredMixin's own gate for the whole management
        # app, member_detail included) doesn't consider a bare EvaluationManager/
        # ShopManager grant "staff" by itself, same already-accepted layering gap
        # noted on FormManagerRequiredMixin's own docstring. A real evaluation
        # manager who needs to actually reach a member's page in practice holds
        # some other staff standing too -- mirrors physio-style fixtures elsewhere
        # in this file (e.g. FormManagementTests.plain_staff).
        eval_team = Team.objects.create(club=cls.club, name="Evaluators Team", short_name="EVL")
        eval_position = Position.objects.create(club=cls.club, name="Technical Director", short_name="TD", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=eval_team, member=cls.eval_manager_member, season=cls.season, position=eval_position)
        # And self.player on that same team's roster -- members_visible_to's
        # own non-admin branch (club/services/access.py, used generally by
        # MemberDetailView and every other member-scoped view, not just
        # evaluations) only surfaces players on a team the requester is
        # actually staffed on, regardless of an EvaluationManager grant's own
        # club-wide reach. Out of scope to change here; this fixture just
        # reflects the realistic case of an evaluation manager who's also
        # this player's own team's staff.
        TeamMembership.objects.create(team=eval_team, member=cls.player, season=cls.season)

    def setUp(self):
        super().setUp()
        cache.clear()
        self.addCleanup(cache.clear)
        # get_or_create, not create -- evaluations/migrations/0002_create_evaluations_flag.py
        # already seeds this flag (off by default), unlike "shop"/"formbuilder" above.
        flag, _created = get_waffle_flag_model().objects.get_or_create(name="evaluations")
        flag.clubs.add(self.club)

    def make_plain_staff(self, email="physio-eval@example.com"):
        staff_user = User.objects.create_user(email=email, password="pw-secret-123")
        staff_member = Member.objects.create(user=staff_user, first_name="Pat", last_name="Physio")
        ClubMembership.objects.create(club=self.club, member=staff_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        team = Team.objects.create(club=self.club, name="Physio Team", short_name="PHY")
        position = Position.objects.create(club=self.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=team, member=staff_member, season=self.season, position=position)
        # self.player on this same team's roster -- members_visible_to's
        # non-admin branch (club/services/access.py) only surfaces players on
        # a team the requester is staffed on, so without this the physio
        # couldn't even reach member_detail for self.player to see whether
        # the Evaluations section is hidden from them.
        TeamMembership.objects.create(team=team, member=self.player, season=self.season)
        return staff_user

    def make_rubric(self):
        form_obj = FormBuilderForm.objects.create(club=self.club, title="Evaluation form")
        FormBuilderField.objects.create(form=form_obj, label="Ball control", field_type=FormBuilderField.FieldType.NUMBER, order=1)
        EvaluationSettings.objects.create(club=self.club, form=form_obj)
        return form_obj


class EvaluationFeatureGatingTests(EvaluationManagementTestBase):
    """The whole surface is gated behind the "evaluations" waffle flag on top
    of its own permission tiers -- rubric editing is ADMIN-only
    (FeatureRequiredMixin), the rest is EvaluationManagerRequiredMixin
    (ADMIN or an EvaluationManager grant)."""

    def test_rubric_view_404s_when_the_flag_is_off(self):
        get_waffle_flag_model().objects.filter(name="evaluations").first().clubs.remove(self.club)
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("evaluation_rubric").status_code, 404)

    def test_admin_can_reach_the_rubric_editor(self):
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("evaluation_rubric").status_code, 200)

    def test_evaluation_manager_gets_403_on_the_rubric_editor(self):
        # Rubric editing is narrower than EvaluationManagerRequiredMixin --
        # only ADMIN may redefine what everyone scores players against.
        self.client.force_login(self.eval_manager_user)

        self.assertEqual(self.club_get("evaluation_rubric").status_code, 403)

    def test_plain_staff_gets_403_on_the_rubric_editor(self):
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("evaluation_rubric").status_code, 403)

    def test_evaluation_manager_can_open_the_new_evaluation_form(self):
        self.make_rubric()
        self.client.force_login(self.eval_manager_user)

        self.assertEqual(self.club_get("evaluation_create", self.player.pk).status_code, 200)

    def test_plain_staff_gets_403_on_the_new_evaluation_form(self):
        self.make_rubric()
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("evaluation_create", self.player.pk).status_code, 403)

    def test_nav_hides_evaluations_when_the_flag_is_off(self):
        # Evaluations sits under Settings' own sub-nav, which only expands on
        # a Settings-section page -- same reasoning as
        # test_nav_shows_shop_once_its_flag_is_active visiting membership_list
        # for Finance's own sub-nav.
        get_waffle_flag_model().objects.filter(name="evaluations").first().clubs.remove(self.club)
        self.client.force_login(self.admin_user)

        response = self.club_get("club_settings")

        self.assertNotContains(response, "Evaluations")

    def test_nav_shows_evaluations_once_the_flag_is_active(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("club_settings")

        self.assertContains(response, "Evaluations")


class EvaluationRubricViewTests(EvaluationManagementTestBase):
    """The rubric editor -- always saves through start_new_rubric_version,
    never mutates the current Form's Fields in place."""

    def rubric_post_data(self, **rows):
        data = {
            "criteria-TOTAL_FORMS": str(len(rows)),
            "criteria-INITIAL_FORMS": "0",
            "criteria-MIN_NUM_FORMS": "0",
            "criteria-MAX_NUM_FORMS": "1000",
        }
        for index, row in enumerate(rows.values()):
            for key, value in row.items():
                data[f"criteria-{index}-{key}"] = value
        return data

    def test_first_save_creates_a_rubric_from_nothing(self):
        self.client.force_login(self.admin_user)
        data = self.rubric_post_data(
            row0={"label": "Ball control", "field_type": "number", "required": "on", "help_text": "", "options": "", "order": "1"},
        )

        response = self.club_post("evaluation_rubric", data)

        self.assertRedirects(response, reverse("management:evaluation_rubric"))
        form_obj = current_rubric_form(self.club)
        self.assertIsNotNone(form_obj)
        field = form_obj.fields.get()
        self.assertEqual(field.label, "Ball control")
        self.assertEqual(field.key, "ball-control")

    def test_saving_an_edit_creates_a_new_version_and_keeps_the_old_one_intact(self):
        original = self.make_rubric()
        self.client.force_login(self.admin_user)
        data = self.rubric_post_data(
            row0={"label": "Ball control", "field_type": "number", "required": "on", "help_text": "", "options": "", "order": "1"},
            row1={"label": "Passing", "field_type": "number", "required": "on", "help_text": "", "options": "", "order": "2"},
        )

        self.club_post("evaluation_rubric", data)

        new_form = current_rubric_form(self.club)
        self.assertNotEqual(new_form.pk, original.pk)
        self.assertEqual(list(new_form.fields.order_by("order").values_list("label", flat=True)), ["Ball control", "Passing"])
        # The old version's own Field is untouched -- an already-scored
        # evaluation would still point at exactly this.
        self.assertTrue(original.fields.filter(label="Ball control").exists())

    def test_a_blank_row_is_dropped(self):
        self.client.force_login(self.admin_user)
        data = self.rubric_post_data(
            row0={"label": "Ball control", "field_type": "number", "required": "on", "help_text": "", "options": "", "order": "1"},
            row1={"label": "", "field_type": "number", "required": "", "help_text": "", "options": "", "order": "2"},
        )

        self.club_post("evaluation_rubric", data)

        self.assertEqual(current_rubric_form(self.club).fields.count(), 1)

    def test_clearing_every_row_is_rejected(self):
        self.make_rubric()
        self.client.force_login(self.admin_user)
        data = self.rubric_post_data(row0={"label": "", "field_type": "number", "required": "", "help_text": "", "options": "", "order": "1"})

        response = self.club_post("evaluation_rubric", data)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Add at least one criterion")

    def test_a_choice_criterion_without_options_is_rejected(self):
        self.client.force_login(self.admin_user)
        data = self.rubric_post_data(row0={"label": "Overall", "field_type": "choice", "required": "on", "help_text": "", "options": "", "order": "1"})

        response = self.club_post("evaluation_rubric", data)

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(current_rubric_form(self.club))

    def test_reordering_rows_persists_the_new_order(self):
        self.client.force_login(self.admin_user)
        data = self.rubric_post_data(
            row0={"label": "Passing", "field_type": "number", "required": "on", "help_text": "", "options": "", "order": "2"},
            row1={"label": "Ball control", "field_type": "number", "required": "on", "help_text": "", "options": "", "order": "1"},
        )

        self.club_post("evaluation_rubric", data)

        labels = list(current_rubric_form(self.club).fields.order_by("order").values_list("label", flat=True))
        self.assertEqual(labels, ["Ball control", "Passing"])


class EvaluationCreateViewTests(EvaluationManagementTestBase):
    """Filling in an evaluation for a player, from that player's member page."""

    def test_no_rubric_configured_shows_the_blocked_screen_not_a_500(self):
        self.client.force_login(self.eval_manager_user)

        response = self.club_get("evaluation_create", self.player.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "hasn't set up an evaluation rubric")

    def test_submitting_a_valid_evaluation_saves_it(self):
        self.make_rubric()
        self.client.force_login(self.eval_manager_user)

        response = self.club_post("evaluation_create", {"ball-control": "8"}, self.player.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.player.pk]))
        evaluation = PlayerEvaluation.objects.get(club=self.club, player=self.player)
        self.assertEqual(evaluation.submission.answers.get().value, "8")

    def test_a_missing_required_answer_is_rejected_with_the_field_flagged(self):
        self.make_rubric()
        self.client.force_login(self.eval_manager_user)

        response = self.club_post("evaluation_create", {"ball-control": ""}, self.player.pk)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(PlayerEvaluation.objects.filter(club=self.club, player=self.player).exists())
        self.assertTrue(response.context["form"].errors)

    def test_no_current_season_shows_the_blocked_screen_not_a_500(self):
        self.make_rubric()
        self.season.start_date = datetime.date.today() + datetime.timedelta(days=10)
        self.season.end_date = datetime.date.today() + datetime.timedelta(days=100)
        self.season.save()
        self.client.force_login(self.eval_manager_user)

        response = self.club_get("evaluation_create", self.player.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "no current season")


class EvaluationDetailViewTests(EvaluationManagementTestBase):
    def test_shows_the_answers_against_the_rubric_that_was_actually_scored(self):
        self.make_rubric()
        self.client.force_login(self.eval_manager_user)
        self.club_post("evaluation_create", {"ball-control": "8"}, self.player.pk)
        evaluation = PlayerEvaluation.objects.get(club=self.club, player=self.player)

        response = self.club_get("evaluation_detail", evaluation.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Ball control")
        self.assertContains(response, "8")

    def test_an_evaluation_from_another_club_404s(self):
        self.make_rubric()
        self.client.force_login(self.eval_manager_user)
        self.club_post("evaluation_create", {"ball-control": "8"}, self.player.pk)
        evaluation = PlayerEvaluation.objects.get(club=self.club, player=self.player)

        other_club = Club.objects.create(name="Rival FC", slug="rival-fc-evals")
        evaluation.club = other_club
        # Bypass PlayerEvaluation.clean()'s own club-scope guard -- this is
        # simulating cross-club data, not exercising that guard here.
        PlayerEvaluation.objects.filter(pk=evaluation.pk).update(club=other_club)

        response = self.club_get("evaluation_detail", evaluation.pk)

        self.assertEqual(response.status_code, 404)


class MemberDetailEvaluationsSectionTests(EvaluationManagementTestBase):
    """The Evaluations section on a member's own page -- entirely hidden, not
    just disabled, when the flag is off or the viewer lacks
    can_manage_evaluations."""

    def test_hidden_when_the_flag_is_off(self):
        get_waffle_flag_model().objects.filter(name="evaluations").first().clubs.remove(self.club)
        self.client.force_login(self.admin_user)

        response = self.club_get("member_detail", self.player.pk)

        self.assertNotContains(response, "Evaluations")

    def test_hidden_for_a_plain_staff_member(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_get("member_detail", self.player.pk)

        self.assertNotContains(response, "Evaluations")

    def test_shown_for_an_evaluation_manager(self):
        self.client.force_login(self.eval_manager_user)

        response = self.club_get("member_detail", self.player.pk)

        self.assertContains(response, "Evaluations")

    def test_lists_existing_evaluations(self):
        self.make_rubric()
        self.client.force_login(self.eval_manager_user)
        self.club_post("evaluation_create", {"ball-control": "8"}, self.player.pk)

        response = self.club_get("member_detail", self.player.pk)

        self.assertContains(response, reverse("management:evaluation_detail", args=[PlayerEvaluation.objects.get(club=self.club, player=self.player).pk]))


RBIHF_SAMPLE_HTML = """<html><body>
<div class="block"><div class="block-header"><h2>Sportoase Antwerp Phantoms</h2></div></div>
<div class="block"><div class="block-header"><h2 id="games-upcoming">Upcoming games</h2></div>
<div class="block-content"><table>
<tr><th class="game-nr">#</th><th class="date">Date</th><th class="hour">Hour</th><th>Location</th><th>Home</th><th>Visit</th></tr>
<tr>
<td class="game-nr"><a href="/game/5002" title="Game 5002">5002</a></td>
<td class="date">2026-09-12</td>
<td class="hour">12:15</td>
<td>Deurne</td>
<td><a href="/league/team/4460" title="Sportoase Antwerp Phantoms">Sportoase Antwerp Phantoms</a></td>
<td><a href="/league/team/4464" title="Amsterdam Tigers">Amsterdam Tigers</a></td>
</tr>
</table></div></div>
</body></html>"""


class RBIHFImportViewTests(ManagementTestBase):
    """The Events page's "Import from RBIHF" button/flow -- admin-only, and
    only when the "RBIHF" waffle Flag is active for the club, same gating
    machinery as FeatureGatedSectionsTests above. The scrape/diff/apply logic
    itself (events.services.rbihf_import) has its own offline tests in
    events/tests.py; these exercise the views end to end via a mocked
    fetch_html, never touching the network."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")

    def setUp(self):
        super().setUp()
        cache.clear()
        self.addCleanup(cache.clear)

    def activate_flag(self):
        # "RBIHF" is already seeded (migration 0018 links it to the Competition
        # row of the same name) -- get_or_create, not create.
        flag, _created = get_waffle_flag_model().objects.get_or_create(name="RBIHF")
        flag.clubs.add(self.club)

    def make_coach(self, email="coach-rbihf@example.com"):
        coach_user = User.objects.create_user(email=email, password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        ClubMembership.objects.create(club=self.club, member=coach_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=position)
        return coach_user

    def test_button_only_shows_for_admin_with_the_flag_active(self):
        self.activate_flag()
        self.client.force_login(self.admin_user)

        response = self.club_get("event_list")

        self.assertContains(response, "Import from RBIHF")

    def test_button_hidden_when_the_flag_is_not_active(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("event_list")

        self.assertNotContains(response, "Import from RBIHF")

    def test_button_hidden_from_a_non_admin_even_with_the_flag_active(self):
        self.activate_flag()
        self.client.force_login(self.make_coach())

        response = self.club_get("event_list")

        self.assertNotContains(response, "Import from RBIHF")

    def test_the_import_views_404_when_the_flag_is_not_active(self):
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("rbihf_import").status_code, 404)
        self.assertEqual(self.club_post("rbihf_import_confirm", {}).status_code, 404)

    def test_a_non_admin_gets_403_when_the_flag_is_active(self):
        self.activate_flag()
        self.client.force_login(self.make_coach())

        self.assertEqual(self.club_get("rbihf_import").status_code, 403)

    @mock.patch("management.views.fetch_html", return_value=RBIHF_SAMPLE_HTML)
    def test_submitting_the_form_shows_a_preview(self, mock_fetch):
        self.activate_flag()
        self.client.force_login(self.admin_user)

        response = self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sportoase Antwerp Phantoms")
        self.assertContains(response, "5002")
        self.assertContains(response, 'name="opponent_5002"')
        self.assertContains(response, 'name="location_5002"')
        mock_fetch.assert_called_once_with("https://www.rbihf.be/league/team/4460")

    @mock.patch("management.views.fetch_html", return_value=RBIHF_SAMPLE_HTML)
    def test_confirming_creates_the_event(self, mock_fetch):
        self.activate_flag()
        self.client.force_login(self.admin_user)
        self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        response = self.club_post("rbihf_import_confirm", {})

        self.assertRedirects(response, reverse("management:event_list"))
        event = Event.objects.get(club=self.club, external_game_id="5002")
        self.assertEqual(event.opponent.name, "Amsterdam Tigers")
        self.assertIn(self.team, event.teams.all())

    @mock.patch("management.views.fetch_html", return_value=RBIHF_SAMPLE_HTML)
    def test_confirming_respects_the_chosen_location(self, mock_fetch):
        self.activate_flag()
        location = Location.objects.create(club=self.club, name="Deurne Ice Hall", address="1 St", city="Deurne", zip_code="2100", country="BE")
        self.client.force_login(self.admin_user)
        self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        self.club_post("rbihf_import_confirm", {"location_5002": str(location.pk)})

        event = Event.objects.get(club=self.club, external_game_id="5002")
        self.assertEqual(event.location, location)

    @mock.patch("management.views.fetch_html", return_value=RBIHF_SAMPLE_HTML)
    def test_confirming_respects_the_chosen_opponent(self, mock_fetch):
        self.activate_flag()
        renamed = Opponent.objects.create(club=self.club, name="Amsterdam Tigers HC")
        self.client.force_login(self.admin_user)
        self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        self.club_post("rbihf_import_confirm", {"opponent_5002": str(renamed.pk)})

        event = Event.objects.get(club=self.club, external_game_id="5002")
        self.assertEqual(event.opponent, renamed)
        self.assertFalse(Opponent.objects.filter(club=self.club, name="Amsterdam Tigers").exists())

    @mock.patch("management.views.fetch_html", return_value=RBIHF_SAMPLE_HTML)
    def test_a_blank_opponent_choice_falls_back_to_the_scraped_name(self, mock_fetch):
        self.activate_flag()
        self.client.force_login(self.admin_user)
        self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        self.club_post("rbihf_import_confirm", {})

        event = Event.objects.get(club=self.club, external_game_id="5002")
        self.assertEqual(event.opponent.name, "Amsterdam Tigers")

    def test_confirming_with_nothing_stashed_redirects_with_a_notice(self):
        self.activate_flag()
        self.client.force_login(self.admin_user)

        response = self.club_post("rbihf_import_confirm", {})

        self.assertRedirects(response, reverse("management:rbihf_import"))

    def test_a_non_rbihf_url_is_a_form_error_not_a_500(self):
        self.activate_flag()
        self.client.force_login(self.admin_user)

        response = self.club_post("rbihf_import", {"url": "https://evil.example.com/x", "team": str(self.team.pk)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "RBIHF team page")

    @mock.patch("management.views.fetch_html", side_effect=RBIHFImportError("Could not reach the page."))
    def test_a_fetch_failure_is_a_form_error_not_a_500(self, mock_fetch):
        self.activate_flag()
        self.client.force_login(self.admin_user)

        response = self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Could not reach the page.")

    @mock.patch("management.views.fetch_html", return_value=RBIHF_SAMPLE_HTML)
    def test_re_running_the_same_import_shows_it_as_unchanged(self, mock_fetch):
        self.activate_flag()
        self.client.force_login(self.admin_user)
        self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})
        self.club_post("rbihf_import_confirm", {})

        response = self.club_post("rbihf_import", {"url": "https://www.rbihf.be/league/team/4460", "team": str(self.team.pk)})

        self.assertContains(response, "already up to date")
        self.assertEqual(Event.objects.filter(club=self.club, external_game_id="5002").count(), 1)


class MemberListKindFilterTests(ManagementTestBase):
    """?kind= on the member list -- see MemberListView.get_queryset. "member"
    (the default) matches the page's original guardian-excluding behaviour;
    "guardian" and "both" turn that exclusion into an explicit, reversible
    choice instead of a parent silently disappearing after being registered."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.family = Family.objects.create()
        cls.parent = Member.objects.create(first_name="Pat", last_name="Guardian")
        cls.child = Member.objects.create(first_name="Cody", last_name="Kid")
        FamilyMembership.objects.create(family=cls.family, member=cls.parent, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=cls.family, member=cls.child, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=cls.club, member=cls.parent, season=cls.season, kind=ClubMembership.Kind.GUARDIAN, status=ClubMembership.StatusChoices.ACTIVE)
        ClubMembership.objects.create(club=cls.club, member=cls.child, season=cls.season, kind=ClubMembership.Kind.MEMBER, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def get(self, kind=None):
        url = reverse("management:member_list")
        if kind is not None:
            url += f"?kind={kind}"
        return self.client.get(url, HTTP_HOST="ajax-united.rosterchief.app")

    def test_default_excludes_the_guardian(self):
        response = self.get()

        self.assertContains(response, "Cody Kid")
        self.assertNotContains(response, "Pat Guardian")
        self.assertEqual(response.context["selected_kind"], "member")

    def test_kind_guardian_shows_only_the_guardian(self):
        response = self.get("guardian")

        self.assertContains(response, "Pat Guardian")
        self.assertNotContains(response, "Cody Kid")

    def test_kind_both_shows_everyone(self):
        response = self.get("both")

        self.assertContains(response, "Pat Guardian")
        self.assertContains(response, "Cody Kid")

    def test_kind_both_flags_the_guardian_row_but_not_the_member_row(self):
        # Both rows show up mixed together in "both" -- the badge is the only thing
        # telling them apart, so it must land on the guardian and nowhere else.
        response = self.get("both")

        self.assertContains(response, ">Guardian<", count=1)

    def test_the_guardian_badge_only_shows_in_the_both_view(self):
        response = self.get("guardian")

        self.assertContains(response, "Pat Guardian")
        self.assertNotContains(response, ">Guardian<")

    def test_an_unrecognised_kind_falls_back_to_the_default(self):
        response = self.get("bogus")

        self.assertNotContains(response, "Pat Guardian")
        self.assertEqual(response.context["selected_kind"], "member")

    def test_a_non_admins_guardian_view_stays_within_their_own_visibility(self):
        # _guardians_only(club) is club-wide -- it's intersected with
        # members_visible_to() so a non-admin only ever sees guardians of
        # someone already visible to them, never every guardian in the club.
        coach_user = User.objects.create_user(email="coach-kind@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-kind", short_name="CK", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.get("both")

        self.assertNotContains(response, "Pat Guardian")


class MemberListFilterAndColumnsTests(ManagementTestBase):
    """Status/fee status/team filters, the Licence and (admin-only) Dues columns,
    and the search-as-you-type input on the member list."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="U16", short_name="U16")
        cls.position = Position.objects.create(club=cls.club, name="Forward", staff_position=False)

        cls.active_paid = Member.objects.create(first_name="Ada", last_name="Active")
        cls.active_membership = ClubMembership.objects.create(club=cls.club, member=cls.active_paid, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.PAID, license="BE-123")
        TeamMembership.objects.create(team=cls.team, member=cls.active_paid, season=cls.season, position=cls.position)

        cls.pending_unpaid = Member.objects.create(first_name="Pip", last_name="Pending")
        ClubMembership.objects.create(club=cls.club, member=cls.pending_unpaid, season=cls.season, status=ClubMembership.StatusChoices.PENDING, fee_status=ClubMembership.FeeStatus.UNPAID)

    def setUp(self):
        self.client.force_login(self.admin_user)

    def get(self, **params):
        url = reverse("management:member_list")
        if params:
            url += "?" + "&".join(f"{key}={value}" for key, value in params.items())
        return self.client.get(url, HTTP_HOST="ajax-united.rosterchief.app")

    def test_status_filter_narrows_to_the_matching_membership(self):
        response = self.get(status="pending")

        self.assertContains(response, "Pip Pending")
        self.assertNotContains(response, "Ada Active")
        self.assertEqual(response.context["selected_status"], "pending")

    def test_fee_status_filter_narrows_to_the_matching_membership(self):
        response = self.get(fee_status="unpaid")

        self.assertContains(response, "Pip Pending")
        self.assertNotContains(response, "Ada Active")

    def test_team_filter_narrows_to_that_teams_roster(self):
        response = self.get(team=str(self.team.pk))

        self.assertContains(response, "Ada Active")
        self.assertNotContains(response, "Pip Pending")

    def test_search_input_has_no_submit_button_and_autosubmits(self):
        response = self.get()

        self.assertNotContains(response, ">Search<")
        self.assertContains(response, "data-autosubmit")

    def test_licence_column_shows_the_current_membership_license(self):
        response = self.get()

        self.assertContains(response, "BE-123")

    def test_dues_column_only_shows_for_an_admin(self):
        admin_response = self.get()
        self.assertContains(admin_response, "Paid")

        coach_user = User.objects.create_user(email="coach-dues@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        position = Position.objects.create(club=self.club, name="Head Coach", short_name="HCD", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        coach_response = self.get()
        self.assertNotContains(coach_response, ">Dues<")

    def test_clear_filters_link_only_shows_once_a_filter_is_active(self):
        clean_response = self.get()
        self.assertNotContains(clean_response, "Clear filters")

        filtered_response = self.get(status="pending")
        self.assertContains(filtered_response, "Clear filters")

    def test_unrostered_filter_matches_the_dashboards_own_needs_attention_row(self):
        # Same condition as attention.unrostered on the dashboard (club_attention) --
        # active, no TeamMembership this season. cls.active_paid is rostered (U16);
        # cls.pending_unpaid isn't ACTIVE at all, so neither counts here.
        unrostered_member = Member.objects.create(first_name="Uma", last_name="Unrostered")
        ClubMembership.objects.create(club=self.club, member=unrostered_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.PAID)

        response = self.get(unrostered="1")

        self.assertContains(response, "Uma Unrostered")
        self.assertNotContains(response, "Ada Active")
        self.assertNotContains(response, "Pip Pending")

    def test_docs_open_filter_matches_the_dashboards_own_missing_documentation_row(self):
        OnboardingRequirement.objects.create(club=self.club, name="Medical certificate")
        # Neither fixture has a MemberRequirementStatus row at all -- "required, no
        # row yet" is itself open, per OnboardingRequirement's own docstring -- so
        # both match regardless of ClubMembership.status (docs tracking doesn't
        # care whether they're PENDING or ACTIVE, only fee-driven status does).

        response = self.get(docs="open")

        self.assertContains(response, "Ada Active")
        self.assertContains(response, "Pip Pending")

    def test_docs_open_filter_excludes_a_member_who_resolved_everything(self):
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Medical certificate")
        mark_complete(self.active_membership, requirement, user=self.admin_user)

        response = self.get(docs="open")

        self.assertNotContains(response, "Ada Active")

    def test_documents_column_only_shows_once_the_club_has_requirements(self):
        no_requirements_response = self.get()
        self.assertNotContains(no_requirements_response, ">Documents<")

        OnboardingRequirement.objects.create(club=self.club, name="Medical certificate")
        OnboardingRequirement.objects.create(club=self.club, name="Photo")
        mark_complete(self.active_membership, OnboardingRequirement.objects.get(name="Photo"), user=self.admin_user)

        response = self.get()

        self.assertContains(response, ">Documents<")
        self.assertContains(response, "1/2")


class FamilyListViewTests(ManagementTestBase):
    """The dedicated Families page -- one row per family, parents/guardians and
    children in separate columns, with the family name and a per-row "Edit"
    action both landing on family_detail (the same overview-page convention
    the Groups list uses for its own Edit button, not a standalone form)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.family = Family.objects.create(name="The Smiths")
        cls.parent = Member.objects.create(first_name="Pat", last_name="Smith")
        cls.child = Member.objects.create(first_name="Cody", last_name="Smith")
        FamilyMembership.objects.create(family=cls.family, member=cls.parent, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=cls.family, member=cls.child, role=FamilyMembership.FamilyRole.CHILD)
        ClubMembership.objects.create(club=cls.club, member=cls.parent, season=cls.season, kind=ClubMembership.Kind.GUARDIAN, status=ClubMembership.StatusChoices.ACTIVE)
        ClubMembership.objects.create(club=cls.club, member=cls.child, season=cls.season, kind=ClubMembership.Kind.MEMBER, status=ClubMembership.StatusChoices.ACTIVE)

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_lists_the_family_with_parents_and_children_columns(self):
        response = self.club_get("family_list")

        self.assertContains(response, "The Smiths")
        self.assertContains(response, "Pat Smith")
        self.assertContains(response, "Cody Smith")

    def test_the_family_name_and_edit_action_both_link_to_family_detail(self):
        response = self.club_get("family_list")

        detail_url = reverse("management:family_detail", args=[self.family.pk])
        self.assertContains(response, f'href="{detail_url}"', count=2)

    def test_a_family_from_another_club_is_excluded(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc")
        other_season = make_season(other_club)
        other_family = Family.objects.create()
        other_member = Member.objects.create(first_name="Other", last_name="Kid")
        ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.ACTIVE)
        FamilyMembership.objects.create(family=other_family, member=other_member, role=FamilyMembership.FamilyRole.CHILD)

        response = self.club_get("family_list")

        self.assertNotContains(response, "Other Kid")

    def test_the_families_nav_link_is_reachable(self):
        # The sidebar is two levels now (design_handoff_rosterchief_platform/README.md,
        # D1/D2): Households only expands under Members while a Members-section page is
        # current, not from every page the way the old flat nav showed it everywhere.
        response = self.club_get("member_list")

        self.assertContains(response, reverse("management:family_list"))

    def test_a_non_admin_only_sees_families_they_have_a_reason_to(self):
        # Same visibility rule as FamilyDetailView: a coach with no tie to this
        # family shouldn't learn it exists just by opening the Families list.
        coach_user = User.objects.create_user(email="coach-fam@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U11", short_name="U11")
        position = Position.objects.create(club=self.club, name="Coach-fam", short_name="CF", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("family_list")

        self.assertNotContains(response, "The Smiths")

    def test_search_matches_a_parents_name(self):
        response = self.client.get(reverse("management:family_list") + "?q=Pat", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "The Smiths")

    def test_search_matches_a_childs_name(self):
        response = self.client.get(reverse("management:family_list") + "?q=Cody", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "The Smiths")

    def test_search_matches_a_last_name(self):
        response = self.client.get(reverse("management:family_list") + "?q=Smith", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "The Smiths")

    def test_search_with_no_match_excludes_the_family(self):
        response = self.client.get(reverse("management:family_list") + "?q=Nobody", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertNotContains(response, "The Smiths")

    def test_search_term_is_kept_in_the_input(self):
        response = self.client.get(reverse("management:family_list") + "?q=Pat", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, 'value="Pat"')

    def test_search_still_respects_visibility(self):
        # A search term matching someone real must not surface a family the
        # requester otherwise has no reason to see.
        coach_user = User.objects.create_user(email="coach-fam2@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U12", short_name="U12")
        position = Position.objects.create(club=self.club, name="Coach-fam2", short_name="CF2", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.client.get(reverse("management:family_list") + "?q=Pat", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertNotContains(response, "The Smiths")


class SidebarCounterTests(ManagementTestBase):
    """The nav's admin-only badges -- pending parent claims, upcoming
    club-managed games nobody's down to referee yet, and the Sign-up queue.
    See management.context_processors.sidebar_counters."""

    def setUp(self):
        # The fixture admin's own ClubMembership (ManagementTestBase.setUpTestData)
        # is ACTIVE but fee_status defaults to UNPAID -- genuinely "not clean" by
        # signup_queue_count's own rule, which would otherwise put a baseline 1 on
        # every count below regardless of what each test itself sets up.
        ClubMembership.objects.filter(club=self.club, member=self.admin_member, season=self.season).update(fee_status=ClubMembership.FeeStatus.PAID)

    def make_pending_claim(self):
        return ParentClaim.objects.create(
            club=self.club,
            parent_first_name="Pat",
            parent_last_name="Parent",
            parent_email="pat-claim@example.com",
            child_first_name="Cody",
            child_last_name="Child",
            child_date_of_birth=datetime.date(2015, 1, 1),
        )

    def make_home_game(self, referee=None):
        team = Team.objects.create(club=self.club, name="First Team", short_name="1st")
        home_ground = Location.objects.create(club=self.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE", is_home=True)
        event = Event.objects.create(club=self.club, title="Cup game", kind=Event.EventKind.GAME, location=home_ground, start=timezone.now() + datetime.timedelta(days=1))
        event.teams.add(team)
        if referee is not None:
            EventReferee.objects.create(event=event, member=referee, assigned_by=self.admin_member)
        return event

    def test_zero_is_shown_explicitly_when_nothing_is_pending(self):
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["pending_parent_claims_count"], 0)
        self.assertEqual(response.context["games_missing_referees_count"], 0)
        self.assertEqual(response.context["signup_pending_count"], 0)

    def test_counts_reflect_a_pending_claim_and_an_unrefereed_game(self):
        self.make_pending_claim()
        self.make_home_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["pending_parent_claims_count"], 1)
        self.assertEqual(response.context["games_missing_referees_count"], 1)

    def test_signup_count_reflects_a_pending_membership(self):
        member = Member.objects.create(first_name="Sig", last_name="Nup")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.PENDING)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")
        self.assertEqual(response.context["signup_pending_count"], 1)

        # The rolled-up badge on the always-visible "Members" nav item --
        # the Sign-up sub-item's own badge only renders once that section is
        # expanded (nav_section == "members"), so its own link isn't on the
        # Home page's nav at all regardless of the count.
        self.assertContains(response, ">1<")
        member_list_response = self.club_get("member_list")
        self.assertContains(member_list_response, reverse("management:signup_list"))

    def test_an_already_active_and_clean_membership_does_not_count(self):
        member = Member.objects.create(first_name="Done", last_name="Already")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.PAID)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["signup_pending_count"], 0)

    def test_an_unconfirmed_registration_does_not_count_yet(self):
        # Billing review comes first now -- see registration.services.
        # invoicing's own module docstring. A registration nobody's confirmed
        # yet hasn't even reached the Sign-up queue, so it shouldn't inflate
        # this badge either.
        member = Member.objects.create(first_name="Awaiting", last_name="Confirm")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.PENDING)
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="sidebar-pat@example.com")
        RegistrationDetails.objects.create(membership=membership, batch=batch)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["signup_pending_count"], 0)

    def test_a_refereed_game_does_not_count(self):
        referee = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=self.club, member=referee, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.make_home_game(referee=referee)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["games_missing_referees_count"], 0)

    def test_a_game_missing_both_referee_and_official_counts_once_not_twice(self):
        # get_or_create, not create -- events/migrations/0031_create_officials_flag
        # already seeds this row at migrate time, so a plain create() collides.
        cache.clear()
        self.addCleanup(cache.clear)
        get_waffle_flag_model().objects.get_or_create(name="officials")[0].clubs.add(self.club)
        # make_home_game's team defaults to CLUB for both referee_management and
        # official_management, so one unstaffed game needs both -- the naive
        # games_missing_referees_count(...) + games_missing_officials_count(...)
        # would count it twice (1 + 1 = 2) for what's actually one game.
        self.make_home_game()
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["games_missing_referees_count"], 1)

    def test_a_referee_assigned_game_still_counts_when_missing_an_official(self):
        cache.clear()
        self.addCleanup(cache.clear)
        get_waffle_flag_model().objects.get_or_create(name="officials")[0].clubs.add(self.club)
        referee = Member.objects.create(first_name="Ref", last_name="Eree")
        ClubMembership.objects.create(club=self.club, member=referee, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        # A referee assigned takes this game out of the referee side (see
        # test_a_refereed_game_does_not_count above), but it has no official
        # assigned at all -- the combined count must still surface it as 1,
        # not silently drop it because the referee side alone is settled.
        self.make_home_game(referee=referee)
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["games_missing_referees_count"], 1)

    def test_an_already_reviewed_claim_does_not_count(self):
        claim = self.make_pending_claim()
        claim.status = ParentClaim.Status.APPROVED
        claim.save(update_fields=["status"])
        self.client.force_login(self.admin_user)

        response = self.club_get("home")

        self.assertEqual(response.context["pending_parent_claims_count"], 0)

    def test_a_fully_staffed_game_does_not_crowd_out_the_next_limit_window(self):
        # Regression: the badge used to slice to the next `limit` games BEFORE
        # dropping fully-staffed ones, while RefereeManagementDashboardView's own
        # kpi_no_referee (for the same default range) filters to understaffed
        # games first and only then slices. A fully-staffed game chronologically
        # ahead of the unstaffed ones used to eat a slot in the badge's window
        # without eating one in the dashboard's, so the two disagreed -- the
        # inconsistency reported from production.
        team = Team.objects.create(club=self.club, name="First Team", short_name="1st")
        home_ground = Location.objects.create(club=self.club, name="Home Ground", address="1 St", city="Town", zip_code="1000", country="BE", is_home=True)
        referees = [Member.objects.create(first_name=f"Ref{i}", last_name="Eree") for i in range(2)]
        for referee in referees:
            ClubMembership.objects.create(club=self.club, member=referee, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        staffed = Event.objects.create(club=self.club, title="Staffed game", kind=Event.EventKind.GAME, location=home_ground, start=timezone.now() + datetime.timedelta(days=1))
        staffed.teams.add(team)
        for referee in referees:
            EventReferee.objects.create(event=staffed, member=referee, assigned_by=self.admin_member)

        for day in range(2, 12):
            event = Event.objects.create(club=self.club, title=f"Unstaffed game {day}", kind=Event.EventKind.GAME, location=home_ground, start=timezone.now() + datetime.timedelta(days=day))
            event.teams.add(team)

        self.client.force_login(self.admin_user)

        response = self.club_get("home")
        dashboard_response = self.club_get("referee_management")

        self.assertEqual(response.context["games_missing_referees_count"], 10)
        self.assertEqual(dashboard_response.context["kpi_no_referee"], 10)

    def test_a_non_admin_gets_no_counters_and_no_badge_links(self):
        coach_user = User.objects.create_user(email="coach-badge@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U10", short_name="U10")
        position = Position.objects.create(club=self.club, name="Coach-badge", short_name="CB", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        response = self.club_get("home")

        self.assertIsNone(response.context["pending_parent_claims_count"])
        self.assertIsNone(response.context["games_missing_referees_count"])
        self.assertIsNone(response.context["signup_pending_count"])
        self.assertNotContains(response, reverse("management:parent_claim_list"))
        self.assertNotContains(response, reverse("management:referee_management"))
        self.assertNotContains(response, reverse("management:signup_list"))


class NotificationBellTests(ManagementTestBase):
    """The topbar bell (every page) and the dashboard's fuller list -- see
    management.context_processors.notification_bell. Unlike SidebarCounterTests'
    admin-only badges above, any signed-in staff member with a Member row can
    have notifications."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_none_when_nothing_is_pending_but_zero_is_shown_once_therere_notifications(self):
        response = self.club_get("home")

        self.assertEqual(response.context["unread_notification_count"], 0)
        self.assertEqual(list(response.context["recent_notifications"]), [])

    def test_unread_count_and_recent_list_reflect_real_notifications(self):
        Notification.objects.create(club=self.club, member=self.admin_member, title="Big win", body="We won 3-0.")

        response = self.club_get("home")

        self.assertEqual(response.context["unread_notification_count"], 1)
        self.assertContains(response, "Big win")

    def test_a_read_notification_does_not_count_as_unread(self):
        Notification.objects.create(club=self.club, member=self.admin_member, title="Old news", body="Body.", read_at=timezone.now())

        response = self.club_get("home")

        self.assertEqual(response.context["unread_notification_count"], 0)

    def test_only_this_members_own_notifications_show(self):
        other_member = Member.objects.create(first_name="Other", last_name="Staff")
        Notification.objects.create(club=self.club, member=other_member, title="Not for you", body="Body.")

        response = self.club_get("home")

        self.assertEqual(response.context["unread_notification_count"], 0)
        self.assertNotContains(response, "Not for you")

    def test_the_bell_badge_renders_on_every_page_not_just_the_dashboard(self):
        Notification.objects.create(club=self.club, member=self.admin_member, title="Big win", body="We won 3-0.")

        response = self.club_get("member_list")

        self.assertContains(response, 'id="notification-menu"')

    def test_anonymous_gets_no_bell(self):
        self.client.logout()

        response = self.client.get(reverse("account_login"), HTTP_HOST="ajax-united.rosterchief.app")

        self.assertNotContains(response, 'id="notification-menu"')


class NotificationMarkAllReadViewTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_marks_every_unread_notification_read(self):
        first = Notification.objects.create(club=self.club, member=self.admin_member, title="One", body="Body.")
        second = Notification.objects.create(club=self.club, member=self.admin_member, title="Two", body="Body.")

        self.club_post("notification_mark_all_read", {})

        first.refresh_from_db()
        second.refresh_from_db()
        self.assertIsNotNone(first.read_at)
        self.assertIsNotNone(second.read_at)

    def test_does_not_touch_another_members_notification(self):
        other_member = Member.objects.create(first_name="Other", last_name="Staff")
        other_notification = Notification.objects.create(club=self.club, member=other_member, title="Not yours", body="Body.")

        self.club_post("notification_mark_all_read", {})

        other_notification.refresh_from_db()
        self.assertIsNone(other_notification.read_at)

    def test_redirects_to_next_when_given(self):
        response = self.club_post("notification_mark_all_read", {"next": reverse("management:member_list")})

        self.assertRedirects(response, reverse("management:member_list"))


class NotificationClearAllViewTests(ManagementTestBase):
    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def test_deletes_every_notification_read_or_not(self):
        Notification.objects.create(club=self.club, member=self.admin_member, title="One", body="Body.", read_at=timezone.now())
        Notification.objects.create(club=self.club, member=self.admin_member, title="Two", body="Body.")

        self.club_post("notification_clear_all", {})

        self.assertFalse(Notification.objects.filter(member=self.admin_member).exists())

    def test_does_not_touch_another_members_notification(self):
        other_member = Member.objects.create(first_name="Other", last_name="Staff")
        other_notification = Notification.objects.create(club=self.club, member=other_member, title="Not yours", body="Body.")

        self.club_post("notification_clear_all", {})

        self.assertTrue(Notification.objects.filter(pk=other_notification.pk).exists())

    def test_redirects_to_next_when_given(self):
        response = self.club_post("notification_clear_all", {"next": reverse("management:member_list")})

        self.assertRedirects(response, reverse("management:member_list"))


class ManagementListPaginationTests(ManagementTestBase):
    """paginate_by on the six paginated lists (MemberListView/EventListView/
    TeamListView/GroupListView/NewsListView/FamilyListView) and the shared
    management/_pagination.html pager. Exercised with a patched-down page size
    rather than dozens of fixture rows."""

    def setUp(self):
        super().setUp()
        self.client.force_login(self.admin_user)

    def make_members(self, count, last_name="Match"):
        members = []
        for i in range(count):
            member = Member.objects.create(first_name=f"Search{i}", last_name=last_name)
            ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
            members.append(member)
        return members

    def test_page_2_shows_different_rows_than_page_1(self):
        self.make_members(5)

        with mock.patch("management.views.MemberListView.paginate_by", 2):
            page1 = self.client.get(reverse("management:member_list") + "?q=Match", HTTP_HOST="ajax-united.rosterchief.app")
            page2 = self.client.get(reverse("management:member_list") + "?q=Match&page=2", HTTP_HOST="ajax-united.rosterchief.app")

        page1_ids = {member.pk for member in page1.context["members"]}
        page2_ids = {member.pk for member in page2.context["members"]}
        self.assertEqual(len(page1_ids), 2)
        self.assertEqual(len(page2_ids), 2)
        self.assertEqual(page1_ids & page2_ids, set())

    def test_a_pagination_link_preserves_the_search_query_string(self):
        self.make_members(5)

        with mock.patch("management.views.MemberListView.paginate_by", 2):
            response = self.client.get(reverse("management:member_list") + "?q=Match", HTTP_HOST="ajax-united.rosterchief.app")

        self.assertContains(response, "q=Match&amp;page=2")

    def test_family_list_pagination_is_wired_and_shows_a_page_count(self):
        for i in range(3):
            family = Family.objects.create()
            member = Member.objects.create(first_name=f"Fam{i}", last_name="Ily")
            FamilyMembership.objects.create(family=family, member=member, role=FamilyMembership.FamilyRole.CHILD)
            ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        with mock.patch("management.views.FamilyListView.paginate_by", 2):
            response = self.club_get("family_list")

        self.assertTrue(response.context["is_paginated"])
        self.assertContains(response, "Page 1 of 2")

    def test_event_team_group_news_lists_are_all_wired_for_pagination(self):
        # A lighter "is it wired" check for the remaining four -- MemberListView
        # and FamilyListView above already cover the actual pager mechanics
        # (page split + query-string preservation), which is shared markup
        # (_pagination.html) and shared ListView machinery (paginate_by), not
        # something that differs meaningfully per view.
        for name in ["event_list", "team_list", "group_list", "news_list"]:
            # event_list defaults to the unpaginated D5-style calendar now --
            # ?view=list is what opts back into the paginated table this loop means to check.
            params = {"view": "list"} if name == "event_list" else None
            response = self.club_get(name, params=params)
            self.assertIsNotNone(response.context["paginator"], name)


class OnboardingRequirementManagementTests(ManagementTestBase):
    """Club-admin CRUD for OnboardingRequirement -- management:onboarding_requirement_*."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def test_an_admin_can_create_a_requirement(self):
        response = self.club_post("onboarding_requirement_create", {"name": "Photo", "requires_document": "", "is_active": "on"})

        self.assertRedirects(response, reverse("management:onboarding_requirement_list"))
        self.assertTrue(OnboardingRequirement.objects.filter(club=self.club, name="Photo").exists())

    def test_a_coach_cannot_create_one(self):
        coach = User.objects.create_user(email="coach-req@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Coach", last_name="Person")
        ClubMembership.objects.create(club=self.club, member=coach_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(coach)

        response = self.club_post("onboarding_requirement_create", {"name": "Photo"})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(OnboardingRequirement.objects.filter(club=self.club).exists())

    def test_deleting_a_requirement_keeps_existing_statuses_for_the_record(self):
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Waiver")
        member = Member.objects.create(first_name="Sam", last_name="Roe")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season)
        status = MemberRequirementStatus.objects.create(membership=membership, requirement=requirement, is_complete=True)

        self.club_post("onboarding_requirement_delete", {}, requirement.pk)

        self.assertFalse(OnboardingRequirement.objects.filter(pk=requirement.pk).exists())
        # The cascade is deliberate here (unlike the note above) -- see
        # club.models.MemberRequirementStatus.requirement's on_delete=CASCADE:
        # a deleted requirement has nothing left to be a status *of*.
        self.assertFalse(MemberRequirementStatus.objects.filter(pk=status.pk).exists())


class MemberRequirementChecklistTests(ManagementTestBase):
    """Marking a member's checklist items done/undone from their own page, and
    downloading whatever document was attached -- management:member_requirement_*.

    Uploads land on rosterchief.storage.private_storage's real, configured location
    (PRIVATE_MEDIA_ROOT) rather than a tempdir substituted via override_settings: that
    storage is a module-level singleton built once at process start (see
    rosterchief/storage.py), and Django's override_settings machinery only reloads
    django.core.files.storage.storages/default_storage on a STORAGES change -- it has
    no way to know about a hand-built FileSystemStorage instance a model field
    happens to hold a reference to, so overriding the setting here would silently do
    nothing and this class would look green while actually writing outside a tempdir
    it thinks it's using. Each uploading test cleans up after itself instead, the
    same way FeePayment/Club-logo-style tests elsewhere in this suite do.
    """

    def setUp(self):
        self.client.force_login(self.admin_user)
        self.requirement = OnboardingRequirement.objects.create(club=self.club, name="Medical certificate", requires_document=True)
        self.member = Member.objects.create(first_name="Noor", last_name="Somers")
        self.membership = ClubMembership.objects.create(club=self.club, member=self.member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.PAID)

    def upload_certificate(self):
        upload = SimpleUploadedFile("certificate.pdf", b"%PDF-1.4 fake", content_type="application/pdf")
        self.club_post("member_requirement_complete", {"document": upload, "note": ""}, self.member.pk, self.requirement.pk)
        status = MemberRequirementStatus.objects.get(membership=self.membership, requirement=self.requirement)
        self.addCleanup(status.document.delete, save=False)
        return status

    def test_staff_can_mark_a_requirement_complete_with_a_document(self):
        upload = SimpleUploadedFile("certificate.pdf", b"%PDF-1.4 fake", content_type="application/pdf")
        response = self.club_post("member_requirement_complete", {"document": upload, "note": "handed in at practice"}, self.member.pk, self.requirement.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.member.pk]))
        status = MemberRequirementStatus.objects.get(membership=self.membership, requirement=self.requirement)
        self.addCleanup(status.document.delete, save=False)
        self.assertTrue(status.is_complete)
        self.assertEqual(status.completed_by, self.admin_user)
        self.assertTrue(status.document.name)

        # The whole point: paid AND active, but still had an open requirement a moment
        # ago -- marking it complete must never touch either field.
        self.membership.refresh_from_db()
        self.assertEqual(self.membership.status, ClubMembership.StatusChoices.ACTIVE)
        self.assertEqual(self.membership.fee_status, ClubMembership.FeeStatus.PAID)

    def test_a_plain_coach_cannot_mark_one_complete(self):
        # Only an admin/MEMBER_ADMIN may touch a member's checklist
        # (MemberAdminRequiredMixin) -- a coach can still *see* it on a member's
        # profile (ClubStaffRequiredMixin's read access there), but not mark
        # anything done or reopen it. A *real* staff assignment on the member's
        # own team is what would have made this coach able to reach the page at
        # all pre-checklist-permission, so the fixture still sets that up to
        # isolate this from a simpler "can't reach the page" 403.
        team = Team.objects.create(club=self.club, name="U16")
        coach_position = Position.objects.create(club=self.club, name="Head Coach", staff_position=True, management_position=True)
        player_position = Position.objects.create(club=self.club, name="Forward", staff_position=False)

        coach = User.objects.create_user(email="coach-checklist@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Coach", last_name="Person")
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=coach_position)
        enrol_mfa(coach)

        TeamMembership.objects.create(team=team, member=self.member, season=self.season, position=player_position, jersey_number=9)
        self.client.force_login(coach)

        response = self.club_post("member_requirement_complete", {"note": ""}, self.member.pk, self.requirement.pk)

        self.assertEqual(response.status_code, 403)

    def test_a_member_admin_can_mark_one_complete(self):
        member_admin_user = User.objects.create_user(email="memberadmin-checklist@example.com", password="pw-secret-123")
        member_admin_member = Member.objects.create(user=member_admin_user, first_name="Mia", last_name="Admin")
        ClubMembership.objects.create(club=self.club, member=member_admin_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        # The active membership above already granted a plain MEMBER role via
        # club/signals.py's post_save sync -- update that row rather than
        # creating a second one (ClubRole is unique per club/member).
        ClubRole.objects.filter(club=self.club, member=member_admin_member).update(role=ClubRole.Roles.MEMBER_ADMIN)
        enrol_mfa(member_admin_user)
        self.client.force_login(member_admin_user)

        response = self.club_post("member_requirement_complete", {"note": ""}, self.member.pk, self.requirement.pk)

        # fetch_redirect_response=False: this member_admin has no team/family tie
        # to self.member, so members_visible_to would 404 the redirect target --
        # a separate, pre-existing visibility question this test isn't about.
        self.assertRedirects(response, reverse("management:member_detail", args=[self.member.pk]), fetch_redirect_response=False)
        self.assertTrue(MemberRequirementStatus.objects.get(membership=self.membership, requirement=self.requirement).is_complete)

    def test_marking_incomplete_deletes_the_document_on_file(self):
        # Reopening something that already has a document on file *is* how
        # staff say "this was wrong, redo it" (e.g. the wrong photo) --
        # club.services.onboarding.mark_incomplete's own docstring.
        self.upload_certificate()

        self.club_post("member_requirement_incomplete", {}, self.member.pk, self.requirement.pk)

        status = MemberRequirementStatus.objects.get(membership=self.membership, requirement=self.requirement)
        self.assertFalse(status.is_complete)
        self.assertFalse(status.document)

    def test_downloading_the_document_requires_management_access(self):
        self.upload_certificate()

        response = self.club_get("member_requirement_document", self.member.pk, self.requirement.pk)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(b"".join(response.streaming_content), b"%PDF-1.4 fake")

        self.client.logout()
        anonymous_response = self.club_get("member_requirement_document", self.member.pk, self.requirement.pk)
        self.assertEqual(anonymous_response.status_code, 302)

    def test_downloading_with_no_document_uploaded_is_a_404(self):
        response = self.club_get("member_requirement_document", self.member.pk, self.requirement.pk)

        self.assertEqual(response.status_code, 404)

    def test_the_document_is_never_reachable_through_the_public_media_url(self):
        # The whole point of a separate, private storage (rosterchief/storage.py):
        # nothing here ever gets a /media/ URL at all.
        status = self.upload_certificate()

        with self.assertRaises(ValueError):
            _ = status.document.url

    def test_staff_can_bypass_a_requirement_with_a_note(self):
        response = self.club_post("member_requirement_bypass", {"note": "already has a recent photo on file"}, self.member.pk, self.requirement.pk)

        self.assertRedirects(response, reverse("management:member_detail", args=[self.member.pk]))
        status = MemberRequirementStatus.objects.get(membership=self.membership, requirement=self.requirement)
        self.assertTrue(status.is_bypassed)
        self.assertFalse(status.is_complete)
        self.assertEqual(status.completed_by, self.admin_user)

    def test_bypassing_without_a_note_is_rejected(self):
        self.club_post("member_requirement_bypass", {"note": ""}, self.member.pk, self.requirement.pk)

        self.assertFalse(MemberRequirementStatus.objects.filter(membership=self.membership, requirement=self.requirement, is_bypassed=True).exists())

    def test_bypass_honours_next_for_the_signup_page(self):
        response = self.club_post("member_requirement_bypass", {"note": "waived", "next": reverse("management:signup_list")}, self.member.pk, self.requirement.pk)

        self.assertRedirects(response, reverse("management:signup_list"))

    def test_complete_honours_next_for_the_signup_page(self):
        response = self.club_post("member_requirement_complete", {"note": "", "next": reverse("management:signup_list")}, self.member.pk, self.requirement.pk)

        self.assertRedirects(response, reverse("management:signup_list"))


class SignupDashboardTests(ManagementTestBase):
    """The admin-only Sign-up page -- management:signup_*."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def make_pending_member(self, fee_status=ClubMembership.FeeStatus.UNPAID):
        member = Member.objects.create(first_name="Noor", last_name="Somers")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.PENDING, fee_status=fee_status)
        return member, membership

    def test_is_admin_only(self):
        coach = User.objects.create_user(email="coach-signup@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-signup", short_name="CS", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach)

        self.assertEqual(self.club_get("signup_list").status_code, 403)

    def test_a_manually_created_membership_shows_immediately(self):
        # No registration behind it at all -- nothing to wait on, unaffected
        # by the new billing-first gate below.
        member, _membership = self.make_pending_member()

        response = self.club_get("signup_list")

        self.assertContains(response, member.first_name)

    def test_an_unconfirmed_registration_is_hidden(self):
        _member, membership = self.make_pending_member()
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat-unconfirmed@example.com")
        RegistrationDetails.objects.create(membership=membership, batch=batch)

        response = self.club_get("signup_list")

        self.assertNotIn(membership, response.context["memberships"])

    def test_a_confirmed_registration_appears(self):
        _member, membership = self.make_pending_member()
        batch = RegistrationBatch.objects.create(
            club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat-confirmed@example.com", invoice_sent_at=timezone.now()
        )
        RegistrationDetails.objects.create(membership=membership, batch=batch)

        response = self.club_get("signup_list")

        self.assertIn(membership, response.context["memberships"])

    def test_lists_every_current_season_member_kind_membership(self):
        member, _membership = self.make_pending_member()

        response = self.club_get("signup_list")

        self.assertContains(response, "Somers")
        self.assertContains(response, member.first_name)

    def test_links_straight_to_the_members_own_page(self):
        # Sign-up's own checklist is download-only -- uploading a document
        # only happens on the member's own page, so both the row and its
        # drawer need a direct jump there.
        member, _membership = self.make_pending_member()

        response = self.club_get("signup_list")

        self.assertContains(response, reverse("management:member_detail", args=[member.pk]), count=2)

    def test_loads_searchable_select_js_for_the_link_to_member_picker(self):
        # Without this, the "Existing member" dropdown falls back to a plain
        # native <select> -- fine for a handful of members, unusable for a
        # club with hundreds.
        response = self.club_get("signup_list")

        self.assertContains(response, "searchable-select.js")

    def test_the_table_shows_each_members_fee_status(self):
        # FeeStatus labels are stored lowercase ("paid") -- .badge's own CSS
        # (text-transform: uppercase) is what makes it read "PAID" on screen.
        self.make_pending_member(fee_status=ClubMembership.FeeStatus.PAID)

        response = self.club_get("signup_list")

        self.assertContains(response, ">paid<")

    def test_a_guardian_kind_membership_is_excluded(self):
        guardian = Member.objects.create(first_name="Pat", last_name="Guardian")
        ClubMembership.objects.create(club=self.club, member=guardian, season=self.season, kind=ClubMembership.Kind.GUARDIAN, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_get("signup_list")

        self.assertNotContains(response, "Guardian")

    def test_shows_the_teams_a_member_has_been_placed_on(self):
        member, _membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Forward", staff_position=False)
        TeamMembership.objects.create(team=team, member=member, season=self.season, position=position)

        response = self.club_get("signup_list")

        self.assertContains(response, "U9")

    def test_shows_a_link_to_view_an_uploaded_document(self):
        member, membership = self.make_pending_member()
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Medical certificate", requires_document=True)
        status = MemberRequirementStatus.objects.create(
            membership=membership, requirement=requirement, document=SimpleUploadedFile("certificate.pdf", b"%PDF-1.4 fake", content_type="application/pdf")
        )
        self.addCleanup(status.document.delete, save=False)

        response = self.club_get("signup_list")

        self.assertContains(response, reverse("management:member_requirement_document", args=[member.pk, requirement.pk]))

    def test_no_document_link_when_nothing_was_uploaded(self):
        member, membership = self.make_pending_member()
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Medical certificate", requires_document=True)
        MemberRequirementStatus.objects.create(membership=membership, requirement=requirement, is_bypassed=True, note="already on file elsewhere")

        response = self.club_get("signup_list")

        self.assertNotContains(response, reverse("management:member_requirement_document", args=[member.pk, requirement.pk]))

    def test_shows_a_volunteers_confirmed_staff_placement_too(self):
        # teams_registered used to only ever look at TeamMembership (players)
        # -- a volunteer auto-placed as staff at confirm time (RegistrationInvoiceReviewView.
        # place_confirmed_entries) read as "Not placed yet" here forever, even
        # with a real StaffAssignment already on record.
        member, membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Team Manager", staff_position=True)
        StaffAssignment.objects.create(team=team, member=member, season=self.season, position=position)

        response = self.club_get("signup_list")

        found = next(row for row in response.context["memberships"] if row.pk == membership.pk)
        self.assertIn(team, found.teams_registered)

    def test_shows_the_requested_jersey_number(self):
        _member, membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=team, requested_jersey_number=7)

        response = self.club_get("signup_list")

        self.assertContains(response, "Requested #7.")

    def test_the_requested_team_chip_is_highlighted(self):
        _member, membership = self.make_pending_member()
        requested_team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        Team.objects.create(club=self.club, name="U11", short_name="U11")
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=requested_team)

        response = self.club_get("signup_list")

        rendered_membership = next(m for m in response.context["memberships"] if m.pk == membership.pk)
        self.assertEqual(rendered_membership.requested_team_ids, {requested_team.pk})
        self.assertContains(response, f'data-team-cell="team_cell_{membership.pk}"')

    def test_approve_all_clean_activates_a_paid_and_fully_checked_member(self):
        _member, membership = self.make_pending_member(fee_status=ClubMembership.FeeStatus.PAID)
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Photo")
        mark_complete(membership, requirement, user=self.admin_user)

        response = self.club_post("signup_approve_all_clean", {})

        self.assertRedirects(response, reverse("management:signup_list"))
        membership.refresh_from_db()
        self.assertEqual(membership.status, ClubMembership.StatusChoices.ACTIVE)

    def test_approve_all_clean_is_admin_only(self):
        coach = User.objects.create_user(email="coach-approve@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-approve", short_name="CA", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach)

        self.assertEqual(self.club_post("signup_approve_all_clean", {}).status_code, 403)

    def test_approve_one_activates_a_clean_pending_membership(self):
        _member, membership = self.make_pending_member(fee_status=ClubMembership.FeeStatus.PAID)

        response = self.club_post("signup_approve_one", {}, membership.member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        membership.refresh_from_db()
        self.assertEqual(membership.status, ClubMembership.StatusChoices.ACTIVE)

    def test_approve_one_leaves_an_unpaid_membership_pending(self):
        _member, membership = self.make_pending_member(fee_status=ClubMembership.FeeStatus.UNPAID)

        self.club_post("signup_approve_one", {}, membership.member.pk)

        membership.refresh_from_db()
        self.assertEqual(membership.status, ClubMembership.StatusChoices.PENDING)

    def test_approve_one_is_admin_only(self):
        _member, membership = self.make_pending_member(fee_status=ClubMembership.FeeStatus.PAID)
        coach = User.objects.create_user(email="coach-approveone@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-approveone", short_name="CO", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach)

        self.assertEqual(self.club_post("signup_approve_one", {}, membership.member.pk).status_code, 403)

    def test_the_longest_waiting_pending_member_is_highlighted(self):
        _older_member, older = self.make_pending_member()
        older.created = timezone.now() - datetime.timedelta(days=5)
        older.save(update_fields=["created"])
        self.make_pending_member()

        response = self.club_get("signup_list")

        by_pk = {membership.pk: membership for membership in response.context["memberships"]}
        self.assertTrue(by_pk[older.pk].is_oldest_pending)
        self.assertEqual(sum(1 for m in response.context["memberships"] if m.is_oldest_pending), 1)

    def test_an_already_active_membership_is_never_flagged_as_oldest_pending(self):
        member = Member.objects.create(first_name="Ada", last_name="Active")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_status=ClubMembership.FeeStatus.PAID)

        response = self.club_get("signup_list")

        self.assertFalse(any(m.is_oldest_pending for m in response.context["memberships"]))

    def test_an_active_and_fully_clean_membership_is_dropped_from_the_queue(self):
        # Regression: an already-ACTIVE member whose checklist is fully resolved
        # and fees are paid has nothing left for this page to do -- it used to
        # trail at the bottom of the table forever, reading as if it still
        # needed processing.
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Photo")
        _member, membership = self.make_pending_member(fee_status=ClubMembership.FeeStatus.PAID)
        membership.status = ClubMembership.StatusChoices.ACTIVE
        membership.save(update_fields=["status"])
        mark_complete(membership, requirement, user=self.admin_user)

        response = self.club_get("signup_list")

        self.assertNotIn(membership.pk, {m.pk for m in response.context["memberships"]})

    def test_an_active_but_not_yet_clean_membership_still_shows(self):
        _member, membership = self.make_pending_member(fee_status=ClubMembership.FeeStatus.UNPAID)
        membership.status = ClubMembership.StatusChoices.ACTIVE
        membership.save(update_fields=["status"])

        response = self.club_get("signup_list")

        self.assertIn(membership.pk, {m.pk for m in response.context["memberships"]})

    def test_a_requested_team_from_registration_shows_as_a_hint(self):
        _member, membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=team)

        response = self.club_get("signup_list")

        self.assertContains(response, "Requested U9 at registration.")

    def test_two_members_sharing_an_unstarted_requirement_get_distinct_bypass_dialog_ids(self):
        # Regression: the bypass dialog id used to fall back to the requirement's
        # own pk whenever a member hadn't started it yet, so every member sharing
        # an untouched requirement rendered an *identical* dialog id -- clicking
        # "Bypass" for anyone but the first member in the list opened the first
        # member's dialog instead, since getElementById only ever returns the
        # first DOM match.
        requirement = OnboardingRequirement.objects.create(club=self.club, name="Photo")
        _first_member, first_membership = self.make_pending_member()
        _second_member, second_membership = self.make_pending_member()

        response = self.club_get("signup_list")

        self.assertContains(response, f'id="bypass_{first_membership.pk}_{requirement.pk}"')
        self.assertContains(response, f'id="bypass_{second_membership.pk}_{requirement.pk}"')

    def test_place_in_team_rosters_a_still_pending_member(self):
        member, _membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        Position.objects.create(club=self.club, name="Forward", staff_position=False)

        response = self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        self.assertTrue(TeamMembership.objects.filter(team=team, member=member, season=self.season).exists())

    def test_place_in_team_over_ajax_returns_json_instead_of_redirecting(self):
        # signup_list.html's own fetch() handler -- X-Requested-With is what tells
        # the view to answer with JSON rather than a normal redirect+Django message.
        member, _membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        Position.objects.create(club=self.club, name="Forward", staff_position=False)

        response = self.client.post(
            reverse("management:signup_place_in_team", args=[member.pk]),
            {"team": team.pk},
            HTTP_HOST="ajax-united.rosterchief.app",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], True)
        self.assertTrue(TeamMembership.objects.filter(team=team, member=member, season=self.season).exists())

    def test_place_in_team_over_ajax_reports_failure_without_placing_twice(self):
        member, _membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Forward", staff_position=False)
        TeamMembership.objects.create(team=team, member=member, season=self.season, position=position)

        response = self.client.post(
            reverse("management:signup_place_in_team", args=[member.pk]),
            {"team": team.pk},
            HTTP_HOST="ajax-united.rosterchief.app",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["ok"], False)
        self.assertEqual(TeamMembership.objects.filter(team=team, member=member, season=self.season).count(), 1)

    def test_place_in_team_does_not_require_eligible_roster_members(self):
        # eligible_roster_members(club) -- what TeamMembershipForm's own dropdown
        # uses -- would exclude a still-PENDING member entirely; this path must not.
        member, membership = self.make_pending_member()
        self.assertNotIn(member, eligible_roster_members(self.club))
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        Position.objects.create(club=self.club, name="Forward", staff_position=False)

        self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        self.assertTrue(TeamMembership.objects.filter(team=team, member=member, season=self.season).exists())
        membership.refresh_from_db()
        self.assertEqual(membership.status, ClubMembership.StatusChoices.PENDING)

    def test_place_in_team_is_position_free_and_leaves_it_blank(self):
        # No dropdown -- just the team. Position stays blank (TeamMembership.
        # position is nullable) rather than guessed at: it's the team
        # manager's own call to get right later (team_roster_update), and a
        # blank position is also what groups every just-placed member
        # together (e.g. via the API) until it's set. Jersey number is
        # different -- see test_place_in_team_honours_a_requested_jersey_number
        # below -- since it's never guessed, only ever a pool-checked replay
        # of what was already explicitly requested at registration.
        member, _membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        Position.objects.create(club=self.club, name="Forward", staff_position=False)

        self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        roster_spot = TeamMembership.objects.get(team=team, member=member, season=self.season)
        self.assertIsNone(roster_spot.position)
        self.assertIsNone(roster_spot.jersey_number)

    def test_place_in_team_honours_a_requested_jersey_number(self):
        member, membership = self.make_pending_member()
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=20)
        team = Team.objects.create(club=self.club, name="U9", short_name="U9", pool=pool)
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=team, requested_jersey_number=9)

        self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        roster_spot = TeamMembership.objects.get(team=team, member=member, season=self.season)
        self.assertEqual(roster_spot.jersey_number, 9)

    def test_place_in_team_stamps_the_result_back_onto_registration_details(self):
        # Without this, cancelling this membership later would have no way
        # to know this placement came from a registration at all, and
        # wouldn't clean it up (club.services.cancellation.cancel_membership).
        member, membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        details = RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=team)

        self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        details.refresh_from_db()
        roster_spot = TeamMembership.objects.get(team=team, member=member, season=self.season)
        self.assertEqual(details.resulting_team_membership, roster_spot)

    def test_place_in_team_drops_a_requested_jersey_number_someone_else_already_holds(self):
        member, membership = self.make_pending_member()
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=20)
        team = Team.objects.create(club=self.club, name="U9", short_name="U9", pool=pool)
        other_member = Member.objects.create(first_name="Otto", last_name="Other", date_of_birth=datetime.date(2015, 1, 1))
        TeamMembership.objects.create(team=team, member=other_member, season=self.season, jersey_number=9)
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=team, requested_jersey_number=9)

        self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        roster_spot = TeamMembership.objects.get(team=team, member=member, season=self.season)
        self.assertIsNone(roster_spot.jersey_number)

    def test_place_in_team_with_no_pool_leaves_a_requested_jersey_number_unset(self):
        member, membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Pat", contact_last_name="Parent", contact_email="pat@example.com", invoice_sent_at=timezone.now())
        RegistrationDetails.objects.create(membership=membership, batch=batch, requested_team=team, requested_jersey_number=9)

        self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        roster_spot = TeamMembership.objects.get(team=team, member=member, season=self.season)
        self.assertIsNone(roster_spot.jersey_number)

    def test_place_in_team_disabled_button_for_an_already_placed_team_is_a_no_op(self):
        # The button for a team the member is already on renders disabled (see
        # signup_list.html) -- but the endpoint itself has to handle a resend
        # gracefully too (unique_member_per_team_per_season), not 500.
        member, _membership = self.make_pending_member()
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Forward", staff_position=False)
        TeamMembership.objects.create(team=team, member=member, season=self.season, position=position)

        response = self.club_post("signup_place_in_team", {"team": team.pk}, member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        self.assertEqual(TeamMembership.objects.filter(team=team, member=member, season=self.season).count(), 1)

    def test_cancel_soft_cancels_a_returning_member(self):
        member, membership = self.make_pending_member()
        previous_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 1, 1), end_date=datetime.date(2020, 12, 31))
        ClubMembership.objects.create(club=self.club, member=member, season=previous_season, status=ClubMembership.StatusChoices.LAPSED)

        response = self.club_post("signup_cancel", {}, member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        membership.refresh_from_db()
        self.assertEqual(membership.status, ClubMembership.StatusChoices.CANCELLED)

    def test_cancel_deletes_a_brand_new_member(self):
        member, _membership = self.make_pending_member()

        response = self.club_post("signup_cancel", {}, member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        self.assertFalse(Member.objects.filter(pk=member.pk).exists())

    def test_cancel_is_admin_only(self):
        member, _membership = self.make_pending_member()
        coach = User.objects.create_user(email="coach-cancel@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-cancel", short_name="CC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach)

        response = self.club_post("signup_cancel", {}, member.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Member.objects.filter(pk=member.pk).exists())

    def test_link_to_member_repoints_the_membership(self):
        member, membership = self.make_pending_member()
        real = Member.objects.create(first_name="Real", last_name="Person")
        previous_season = Season.objects.create(club=self.club, start_date=datetime.date(2020, 1, 1), end_date=datetime.date(2020, 12, 31))
        ClubMembership.objects.create(club=self.club, member=real, season=previous_season, status=ClubMembership.StatusChoices.LAPSED)

        response = self.club_post("signup_link_to_member", {"member": str(real.pk)}, member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        membership.refresh_from_db()
        self.assertEqual(membership.member, real)
        self.assertFalse(Member.objects.filter(pk=member.pk).exists())

    def test_link_to_member_shows_an_error_when_the_target_already_has_one_this_season(self):
        member, _membership = self.make_pending_member()
        real = Member.objects.create(first_name="Real", last_name="Person")
        ClubMembership.objects.create(club=self.club, member=real, season=self.season, status=ClubMembership.StatusChoices.PENDING)

        response = self.club_post("signup_link_to_member", {"member": str(real.pk)}, member.pk)

        self.assertRedirects(response, reverse("management:signup_list"))
        self.assertTrue(Member.objects.filter(pk=member.pk).exists())

    def test_link_to_member_is_admin_only(self):
        member, _membership = self.make_pending_member()
        real = Member.objects.create(first_name="Real", last_name="Person")
        coach = User.objects.create_user(email="coach-link@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-link", short_name="CL", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach)

        response = self.club_post("signup_link_to_member", {"member": str(real.pk)}, member.pk)

        self.assertEqual(response.status_code, 403)


class MemberAdminAccessTests(ManagementTestBase):
    """MEMBER_ADMIN: full read/write on people, teams, referee setup, and onboarding
    requirements -- but not Finance/Shop, Club identity, Sponsors, Roles, or the
    Sign-up page (all of those stay ClubAdminRequiredMixin, real admin only)."""

    def setUp(self):
        self.member_admin_user = User.objects.create_user(email="memberadmin@example.com", password="pw-secret-123")
        member_admin_member = Member.objects.create(user=self.member_admin_user, first_name="Mo", last_name="Admin")
        ClubMembership.objects.create(club=self.club, member=member_admin_member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=self.club, member=member_admin_member).update(role=ClubRole.Roles.MEMBER_ADMIN)
        enrol_mfa(self.member_admin_user)
        self.client.force_login(self.member_admin_user)

    def test_can_create_a_member(self):
        response = self.club_post("member_create", {"first_name": "New", "last_name": "Player", "email": "newplayer@example.com"})

        self.assertTrue(Member.objects.filter(first_name="New", last_name="Player").exists())
        self.assertNotEqual(response.status_code, 403)

    def test_can_create_a_team(self):
        response = self.club_post("team_create", {"name": "U12", "short_name": "U12", "referee_management": Team.RefereeManagement.CLUB})

        self.assertTrue(Team.objects.filter(club=self.club, name="U12").exists())
        self.assertNotEqual(response.status_code, 403)

    def test_can_reach_referee_management(self):
        self.assertEqual(self.club_get("referee_management").status_code, 200)

    def test_can_reach_onboarding_requirements(self):
        self.assertEqual(self.club_get("onboarding_requirement_list").status_code, 200)

    def test_can_reach_group_list(self):
        self.assertEqual(self.club_get("group_list").status_code, 200)

    def test_can_reach_volunteer_list(self):
        self.assertEqual(self.club_get("volunteer_list").status_code, 200)

    def test_cannot_reach_finance(self):
        self.assertEqual(self.club_get("membership_list").status_code, 403)

    def test_cannot_reach_club_settings(self):
        self.assertEqual(self.club_get("club_settings").status_code, 403)

    def test_cannot_reach_roles(self):
        self.assertEqual(self.club_get("role_list").status_code, 403)

    def test_cannot_grant_a_role(self):
        target = Member.objects.create(first_name="Target", last_name="Person")
        ClubMembership.objects.create(club=self.club, member=target, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_post("role_create", {"member": target.pk, "role": ClubRole.Roles.ADMIN})

        self.assertEqual(response.status_code, 403)

    def test_cannot_reach_sponsors(self):
        self.assertEqual(self.club_get("sponsor_list").status_code, 403)

    def test_cannot_reach_signup(self):
        self.assertEqual(self.club_get("signup_list").status_code, 403)

    def test_cannot_reach_positions(self):
        self.assertEqual(self.club_get("position_create").status_code, 403)


class VolunteerListTests(ManagementTestBase):
    """management:volunteer_list/volunteer_place -- the club-wide staff
    roster plus pending registration.models.RegistrationDetails volunteer
    entries awaiting placement into a real StaffAssignment."""

    def setUp(self):
        self.client.force_login(self.admin_user)
        self.team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        self.position = Position.objects.create(club=self.club, name="Coach", short_name="CO", staff_position=True)

    def make_pending_volunteer(self, requested_team=None, requested_position=None):
        member = Member.objects.create(first_name="Val", last_name="Volunteer")
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.PENDING)
        batch = RegistrationBatch.objects.create(club=self.club, season=self.season, contact_first_name="Val", contact_last_name="Volunteer", contact_email="val@example.com")
        details = RegistrationDetails.objects.create(membership=membership, batch=batch, entry_kind=RegistrationDetails.EntryKind.VOLUNTEER, requested_team=requested_team, requested_position=requested_position)
        return member, details

    def test_is_member_admin_gated(self):
        plain = User.objects.create_user(email="plain-volunteer@example.com", password="pw-secret-123")
        Member.objects.create(user=plain, first_name="Plain", last_name="Staffer")
        self.client.force_login(plain)

        self.assertEqual(self.club_get("volunteer_list").status_code, 403)

    def test_shows_existing_staff_assignments(self):
        coach_member = Member.objects.create(first_name="Cara", last_name="Coach")
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=self.position)

        response = self.club_get("volunteer_list")

        self.assertContains(response, "Cara Coach")
        self.assertContains(response, "U9")

    def test_groups_multiple_assignments_for_the_same_volunteer_into_one_row(self):
        coach_member = Member.objects.create(first_name="Cara", last_name="Coach")
        other_team = Team.objects.create(club=self.club, name="U11", short_name="U11")
        StaffAssignment.objects.create(team=self.team, member=coach_member, season=self.season, position=self.position)
        StaffAssignment.objects.create(team=other_team, member=coach_member, season=self.season, position=self.position)

        response = self.club_get("volunteer_list")

        self.assertEqual(len(response.context["volunteers"]), 1)
        self.assertEqual(len(response.context["volunteers"][0]["assignments"]), 2)

    def test_shows_a_pending_volunteer_awaiting_placement(self):
        member, _details = self.make_pending_volunteer(requested_team=self.team, requested_position=self.position)

        response = self.club_get("volunteer_list")

        self.assertContains(response, member.get_full_name())
        self.assertEqual(len(response.context["pending"]), 1)

    def test_placing_creates_a_staff_assignment_and_clears_the_pending_row(self):
        member, details = self.make_pending_volunteer(requested_team=self.team, requested_position=self.position)

        response = self.club_post("volunteer_place", {"team": self.team.pk, "position": self.position.pk}, details.pk)

        self.assertRedirects(response, reverse("management:volunteer_list"))
        assignment = StaffAssignment.objects.get(member=member, team=self.team, season=self.season)
        self.assertEqual(assignment.position, self.position)
        details.refresh_from_db()
        self.assertEqual(details.resulting_staff_assignment, assignment)
        self.assertEqual(len(self.club_get("volunteer_list").context["pending"]), 0)

    def test_placing_into_an_already_held_team_is_rejected(self):
        member, details = self.make_pending_volunteer()
        StaffAssignment.objects.create(team=self.team, member=member, season=self.season, position=self.position)
        other_position = Position.objects.create(club=self.club, name="Manager", short_name="MG", staff_position=True)

        self.club_post("volunteer_place", {"team": self.team.pk, "position": other_position.pk}, details.pk)

        details.refresh_from_db()
        self.assertIsNone(details.resulting_staff_assignment)
        self.assertEqual(StaffAssignment.objects.filter(member=member, team=self.team).count(), 1)

    def test_placement_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-volunteer")
        other_season = make_season(other_club)
        other_member = Member.objects.create(first_name="Out", last_name="Sider")
        other_membership = ClubMembership.objects.create(club=other_club, member=other_member, season=other_season, status=ClubMembership.StatusChoices.PENDING)
        other_batch = RegistrationBatch.objects.create(club=other_club, season=other_season, contact_first_name="Out", contact_last_name="Sider", contact_email="out@example.com")
        other_details = RegistrationDetails.objects.create(membership=other_membership, batch=other_batch, entry_kind=RegistrationDetails.EntryKind.VOLUNTEER)

        response = self.club_post("volunteer_place", {"team": self.team.pk, "position": self.position.pk}, other_details.pk)

        self.assertEqual(response.status_code, 404)


class ShopTestBase(ManagementTestBase):
    """Shared fixture for shop management tests: the "shop" waffle flag active
    for this club (see FeatureGatedSectionsTests's own comment on why cache.clear()
    matters here), plus one product to hang orders/discounts off of."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.product = Product.objects.create(club=cls.club, name="Home Jersey", price=Decimal("25.00"))

    def setUp(self):
        super().setUp()
        cache.clear()
        self.addCleanup(cache.clear)
        flag = get_waffle_flag_model().objects.create(name="shop")
        flag.clubs.add(self.club)

    def make_shop_manager(self, email="shopadmin@example.com"):
        user = User.objects.create_user(email=email, password="pw-secret-123")
        member = Member.objects.create(user=user, first_name="Sam", last_name="ShopAdmin")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        ShopManager.objects.create(club=self.club, member=member)
        return user

    def make_plain_staff(self, email="physio-shop@example.com"):
        user = User.objects.create_user(email=email, password="pw-secret-123")
        member = Member.objects.create(user=user, first_name="Pat", last_name="Physio")
        team = Team.objects.create(club=self.club, name="Shop Physio Team", short_name="SPT")
        position = Position.objects.create(club=self.club, name="Physio Shop", short_name="PHS", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=team, member=member, season=self.season, position=position)
        return user


class ShopManagerRBACBoundaryTests(ShopTestBase):
    """The three-way boundary ShopManagerRequiredMixin is built for -- see
    club.mixins.ShopManagerRequiredMixin/club.services.access.can_manage_shop."""

    def test_a_plain_editor_without_a_shop_grant_gets_403(self):
        user = User.objects.create_user(email="editor-only@example.com", password="pw-secret-123")
        member = Member.objects.create(user=user, first_name="Edie", last_name="Editor")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=self.club, member=member).update(role=ClubRole.Roles.EDITOR)
        enrol_mfa(user)  # ClubRole EDITOR requires a second factor.
        self.client.force_login(user)

        self.assertEqual(self.club_get("product_list").status_code, 403)

    def test_a_shop_manager_with_no_club_role_at_all_can_reach_shop(self):
        # No ClubMembership, no ClubRole row of any kind -- the ShopManager grant
        # alone is enough, same as StaffAssignment granting coach access on top of
        # nothing at all.
        user = User.objects.create_user(email="shop-only@example.com", password="pw-secret-123")
        member = Member.objects.create(user=user, first_name="Sol", last_name="ShopOnly")
        ShopManager.objects.create(club=self.club, member=member)
        self.client.force_login(user)

        self.assertEqual(self.club_get("product_list").status_code, 200)

    def test_an_admin_reaches_shop_regardless_of_a_shop_manager_grant(self):
        self.client.force_login(self.admin_user)

        self.assertFalse(ShopManager.objects.filter(club=self.club, member=self.admin_member).exists())
        self.assertEqual(self.club_get("product_list").status_code, 200)


class ProductManagementTests(ShopTestBase):
    def product_data(self, **overrides):
        data = {
            "name": "Away Jersey",
            "description": "",
            "price": "30.00",
            "product_type": Product.ProductType.MERCHANDISE,
            "is_active": "on",
            # ProductCreateView's own variant formset (management/forms.py's
            # ProductVariantFormSet) -- zero submitted rows is a valid "no
            # variants yet" formset, same as its extra=3 blank rows unfilled.
            "variants-TOTAL_FORMS": "0",
            "variants-INITIAL_FORMS": "0",
            "variants-MIN_NUM_FORMS": "0",
            "variants-MAX_NUM_FORMS": "1000",
            # ProductRegistrantDiscountTierFormSet -- both ProductCreateView
            # and ProductUpdateView now save this alongside ProductForm; zero
            # submitted rows is a valid "no tiers" formset, same reasoning as
            # the variant formset above.
            "tiers-TOTAL_FORMS": "0",
            "tiers-INITIAL_FORMS": "0",
            "tiers-MIN_NUM_FORMS": "0",
            "tiers-MAX_NUM_FORMS": "1000",
        }
        data.update(overrides)
        return data

    def test_the_season_field_excludes_a_completed_season(self):
        today = timezone.localdate()
        past_season = Season.objects.create(club=self.club, start_date=today - datetime.timedelta(days=400), end_date=today - datetime.timedelta(days=31))
        future_season = Season.objects.create(club=self.club, start_date=today + datetime.timedelta(days=301), end_date=today + datetime.timedelta(days=600))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_create")

        available_seasons = response.context["form"].fields["season"].queryset
        self.assertIn(self.season, available_seasons)
        self.assertIn(future_season, available_seasons)
        self.assertNotIn(past_season, available_seasons)

    def test_the_season_dropdown_shows_the_full_year_label(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_create")

        # "Season 2026-2027", not Season.name's short "26-27" form.
        self.assertContains(response, f"Season {self.season.start_date.year}-{self.season.end_date.year}")

    def test_a_shop_manager_can_view_the_product_list(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Home Jersey")

    def test_a_membership_products_season_shows_on_the_list(self):
        self.product.product_type = Product.ProductType.MEMBERSHIP
        self.product.season = self.season
        self.product.save(update_fields=["product_type", "season"])
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_list")

        self.assertContains(response, self.season.name)

    def test_a_non_membership_products_season_is_not_shown(self):
        # Product.season can technically still be set on a non-membership
        # row (e.g. leftover from before the type was changed) -- the list
        # should still read as "not applicable" for it, not show stale data.
        self.product.season = self.season
        self.product.save(update_fields=["season"])
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_list")

        self.assertNotContains(response, self.season.name)

    def test_plain_staff_cannot_view_the_product_list(self):
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("product_list").status_code, 403)

    def test_the_sold_column_reflects_paid_orders_only(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-sold@example.com")
        paid = Order.objects.create(club=self.club, purchaser=purchaser, payment_status=Order.PaymentStatus.PAID, total=Decimal("50.00"))
        OrderLine.objects.create(order=paid, product=self.product, quantity=2, unit_price=Decimal("25.00"), line_total=Decimal("50.00"))
        pending = Order.objects.create(club=self.club, purchaser=purchaser, payment_status=Order.PaymentStatus.PENDING, total=Decimal("999.00"))
        OrderLine.objects.create(order=pending, product=self.product, quantity=99, unit_price=Decimal("999.00"), line_total=Decimal("999.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_list")

        self.assertEqual(response.context["products"][0].sold_count, 2)

    def test_a_shop_manager_can_create_a_product(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_create", self.product_data())

        self.assertRedirects(response, reverse("management:product_list"))
        self.assertTrue(Product.objects.filter(club=self.club, name="Away Jersey").exists())

    def test_an_admin_can_create_a_product_with_a_photo(self):
        self.client.force_login(self.admin_user)

        self.club_post("product_create", self.product_data(image=make_image_file()))

        product = Product.objects.get(club=self.club, name="Away Jersey")
        self.assertTrue(product.image)

    def test_plain_staff_cannot_create_a_product(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("product_create", self.product_data())

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Product.objects.filter(club=self.club, name="Away Jersey").exists())

    def test_a_shop_manager_can_edit_a_product(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_update", self.product_data(name="Home Jersey (Renamed)"), self.product.pk)

        self.assertRedirects(response, reverse("management:product_list"))
        self.product.refresh_from_db()
        self.assertEqual(self.product.name, "Home Jersey (Renamed)")

    def test_plain_staff_cannot_edit_a_product(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("product_update", self.product_data(), self.product.pk)

        self.assertEqual(response.status_code, 403)

    def test_toggling_a_product_flips_is_active(self):
        self.client.force_login(self.make_shop_manager())
        self.assertTrue(self.product.is_active)

        self.club_post("product_toggle_active", {}, self.product.pk)
        self.product.refresh_from_db()
        self.assertFalse(self.product.is_active)

        self.club_post("product_toggle_active", {}, self.product.pk)
        self.product.refresh_from_db()
        self.assertTrue(self.product.is_active)

    def test_plain_staff_cannot_toggle_a_product(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("product_toggle_active", {}, self.product.pk)

        self.assertEqual(response.status_code, 403)
        self.product.refresh_from_db()
        self.assertTrue(self.product.is_active)

    def test_there_is_no_product_delete_view(self):
        # Product is PROTECTed by OrderLine -- deactivating is the only removal
        # path, see ProductToggleActiveView's own docstring.
        with self.assertRaises(NoReverseMatch):
            reverse("management:product_delete")

    def test_a_product_can_be_created_with_a_category(self):
        category = ProductCategory.objects.create(club=self.club, name="Merchandise")
        self.client.force_login(self.make_shop_manager())

        self.club_post("product_create", self.product_data(category=category.pk))

        product = Product.objects.get(club=self.club, name="Away Jersey")
        self.assertEqual(product.category, category)

    def test_a_category_from_another_club_is_rejected(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-cat")
        other_category = ProductCategory.objects.create(club=other_club, name="Foreign")
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_create", self.product_data(category=other_category.pk))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(club=self.club, name="Away Jersey").exists())


class ProductCreateWithVariantsTests(ShopTestBase):
    """The create-product page's own variant formset (ProductVariantFormSet) --
    lets a club fill in "Small"/"Medium"/"Large" rows right on the create
    form instead of saving the product first. See ProductManagementTests'
    own product_data() for the formset's management-form fields."""

    def formset_data(self, *rows):
        data = {"variants-TOTAL_FORMS": str(len(rows)), "variants-INITIAL_FORMS": "0", "variants-MIN_NUM_FORMS": "0", "variants-MAX_NUM_FORMS": "1000"}
        for index, row in enumerate(rows):
            for field, value in row.items():
                data[f"variants-{index}-{field}"] = value
        return data

    def tier_formset_data(self):
        # Zero submitted rows is a valid "no tiers" formset -- ProductRegistrantDiscountTierFormSetTests
        # below covers actually submitting tier rows.
        return {"tiers-TOTAL_FORMS": "0", "tiers-INITIAL_FORMS": "0", "tiers-MIN_NUM_FORMS": "0", "tiers-MAX_NUM_FORMS": "1000"}

    def test_variant_rows_are_created_alongside_the_product(self):
        self.client.force_login(self.make_shop_manager())
        data = {"name": "Away Jersey", "description": "", "price": "30.00", "product_type": Product.ProductType.MERCHANDISE, "is_active": "on"}
        data.update(self.formset_data({"name": "Small"}, {"name": "Large", "price": "35.00"}))
        data.update(self.tier_formset_data())

        response = self.club_post("product_create", data)

        self.assertRedirects(response, reverse("management:product_list"))
        product = Product.objects.get(club=self.club, name="Away Jersey")
        self.assertEqual(product.variants.count(), 2)
        self.assertEqual(product.variants.get(name="Large").price, Decimal("35.00"))

    def test_a_blank_row_is_simply_skipped(self):
        self.client.force_login(self.make_shop_manager())
        data = {"name": "Away Jersey", "description": "", "price": "30.00", "product_type": Product.ProductType.MERCHANDISE, "is_active": "on"}
        data.update(self.formset_data({"name": "Small"}, {"name": ""}))
        data.update(self.tier_formset_data())

        self.club_post("product_create", data)

        product = Product.objects.get(club=self.club, name="Away Jersey")
        self.assertEqual(product.variants.count(), 1)

    def test_two_rows_with_the_same_name_are_rejected(self):
        self.client.force_login(self.make_shop_manager())
        data = {"name": "Away Jersey", "description": "", "price": "30.00", "is_active": "on"}
        data.update(self.formset_data({"name": "Small"}, {"name": "small"}))
        data.update(self.tier_formset_data())

        response = self.club_post("product_create", data)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(club=self.club, name="Away Jersey").exists())


class ProductRegistrantDiscountTierManagementTests(ShopTestBase):
    """The product form's own registrant-discount-tier formset
    (ProductRegistrantDiscountTierFormSet) -- "x people = x% off, y people =
    y% off" staircase pricing, available on both create and edit (unlike
    variants, which only get their inline formset on create)."""

    def base_data(self, **overrides):
        data = {
            "name": "Player Registration",
            "description": "",
            "price": "100.00",
            "product_type": Product.ProductType.MEMBERSHIP,
            "is_active": "on",
            "variants-TOTAL_FORMS": "0",
            "variants-INITIAL_FORMS": "0",
            "variants-MIN_NUM_FORMS": "0",
            "variants-MAX_NUM_FORMS": "1000",
        }
        data.update(overrides)
        return data

    def tier_formset_data(self, *rows, total=None):
        data = {"tiers-TOTAL_FORMS": str(total if total is not None else len(rows)), "tiers-INITIAL_FORMS": "0", "tiers-MIN_NUM_FORMS": "0", "tiers-MAX_NUM_FORMS": "1000"}
        for index, row in enumerate(rows):
            for field, value in row.items():
                data[f"tiers-{index}-{field}"] = value
        return data

    def test_tier_rows_are_created_alongside_the_product(self):
        self.client.force_login(self.make_shop_manager())
        data = self.base_data()
        data.update(self.tier_formset_data({"min_registrants": "2", "discount_type": "percentage", "discount_amount": "5"}, {"min_registrants": "4", "discount_type": "percentage", "discount_amount": "10"}))

        response = self.club_post("product_create", data)

        self.assertRedirects(response, reverse("management:product_list"))
        product = Product.objects.get(club=self.club, name="Player Registration")
        self.assertEqual(product.registrant_discount_tiers.count(), 2)
        self.assertEqual(product.registrant_discount_tiers.get(min_registrants=4).discount_amount, Decimal("10.00"))

    def test_the_registrant_discount_scope_dropdown_has_no_blank_option(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_create")

        self.assertEqual(response.context["form"].fields["registrant_discount_scope"].choices, Product.RegistrantDiscountScope.choices)

    def test_registrant_discount_scope_defaults_to_per_person(self):
        self.client.force_login(self.make_shop_manager())
        data = self.base_data()
        data.update(self.tier_formset_data({"min_registrants": "2", "discount_type": "percentage", "discount_amount": "5"}))

        self.club_post("product_create", data)

        product = Product.objects.get(club=self.club, name="Player Registration")
        self.assertEqual(product.registrant_discount_scope, Product.RegistrantDiscountScope.PER_PERSON)

    def test_registrant_discount_scope_can_be_set_to_per_order(self):
        self.client.force_login(self.make_shop_manager())
        data = self.base_data(registrant_discount_scope=Product.RegistrantDiscountScope.PER_ORDER)
        data.update(self.tier_formset_data({"min_registrants": "2", "discount_type": "percentage", "discount_amount": "5"}))

        self.club_post("product_create", data)

        product = Product.objects.get(club=self.club, name="Player Registration")
        self.assertEqual(product.registrant_discount_scope, Product.RegistrantDiscountScope.PER_ORDER)

    def test_a_blank_row_is_simply_skipped(self):
        self.client.force_login(self.make_shop_manager())
        data = self.base_data()
        data.update(self.tier_formset_data({"min_registrants": "2", "discount_type": "percentage", "discount_amount": "5"}, {}))

        self.club_post("product_create", data)

        product = Product.objects.get(club=self.club, name="Player Registration")
        self.assertEqual(product.registrant_discount_tiers.count(), 1)

    def test_two_tiers_with_the_same_threshold_are_rejected(self):
        self.client.force_login(self.make_shop_manager())
        data = self.base_data()
        data.update(self.tier_formset_data({"min_registrants": "3", "discount_type": "percentage", "discount_amount": "5"}, {"min_registrants": "3", "discount_type": "percentage", "discount_amount": "10"}))

        response = self.club_post("product_create", data)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(club=self.club, name="Player Registration").exists())

    def test_editing_a_product_replaces_its_tiers(self):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, price=Decimal("100.00"))
        ProductRegistrantDiscountTier.objects.create(product=product, min_registrants=2, discount_type="percentage", discount_amount=Decimal("5"))
        self.client.force_login(self.make_shop_manager())
        data = self.base_data()
        data.update(self.tier_formset_data({"min_registrants": "3", "discount_type": "fixed_amount", "discount_amount": "15"}))

        response = self.club_post("product_update", data, product.pk)

        self.assertRedirects(response, reverse("management:product_list"))
        tiers = list(product.registrant_discount_tiers.all())
        self.assertEqual(len(tiers), 1)
        self.assertEqual(tiers[0].min_registrants, 3)
        self.assertEqual(tiers[0].discount_amount, Decimal("15.00"))

    def test_the_edit_form_is_prefilled_with_existing_tiers(self):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, price=Decimal("100.00"))
        ProductRegistrantDiscountTier.objects.create(product=product, min_registrants=2, discount_type="percentage", discount_amount=Decimal("5"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_update", product.pk)

        self.assertContains(response, 'value="2"')

    def test_a_tier_below_2_people_is_rejected(self):
        self.client.force_login(self.make_shop_manager())
        data = self.base_data()
        data.update(self.tier_formset_data({"min_registrants": "1", "discount_type": "percentage", "discount_amount": "5"}))

        response = self.club_post("product_create", data)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Product.objects.filter(club=self.club, name="Player Registration").exists())


class ProductCategoryManagementTests(ShopTestBase):
    def test_a_shop_manager_can_add_a_category(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_category_create", {"name": "Merchandise"})

        self.assertRedirects(response, reverse("management:product_list"))
        self.assertTrue(ProductCategory.objects.filter(club=self.club, name="Merchandise").exists())

    def test_plain_staff_cannot_add_a_category(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("product_category_create", {"name": "Merchandise"})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(ProductCategory.objects.filter(club=self.club, name="Merchandise").exists())

    def test_a_shop_manager_can_edit_a_category(self):
        category = ProductCategory.objects.create(club=self.club, name="Merch")
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_category_update", {"name": "Merchandise"}, category.pk)

        self.assertRedirects(response, reverse("management:product_list"))
        category.refresh_from_db()
        self.assertEqual(category.name, "Merchandise")

    def test_categories_are_ordered_alphabetically(self):
        ProductCategory.objects.create(club=self.club, name="Merchandise")
        ProductCategory.objects.create(club=self.club, name="Fees")

        # registration_kind="" -- the two system categories every club gets
        # (shop.signals.create_registration_categories) aren't what this is
        # testing.
        self.assertEqual(list(ProductCategory.objects.filter(club=self.club, registration_kind="").values_list("name", flat=True)), ["Fees", "Merchandise"])

    def test_deleting_a_category_uncategorises_its_products(self):
        category = ProductCategory.objects.create(club=self.club, name="Merch")
        self.product.category = category
        self.product.save(update_fields=["category"])
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_category_delete", {}, category.pk)

        self.assertRedirects(response, reverse("management:product_list"))
        self.assertFalse(ProductCategory.objects.filter(pk=category.pk).exists())

    def test_a_system_category_cannot_be_deleted_via_the_view(self):
        category = ProductCategory.objects.get(club=self.club, registration_kind=ProductCategory.RegistrationKind.PLAYER)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_category_delete", {}, category.pk)

        self.assertRedirects(response, reverse("management:product_list"))
        self.assertTrue(ProductCategory.objects.filter(pk=category.pk).exists())

    def test_the_product_list_page_has_no_delete_action_for_a_system_category(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_list")

        self.assertContains(response, "Player")
        self.assertContains(response, "System")
        self.assertNotContains(response, "product_category_delete")
        self.product.refresh_from_db()
        self.assertIsNone(self.product.category)

    def test_a_category_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-cat2")
        other_category = ProductCategory.objects.create(club=other_club, name="Foreign")
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_category_update", {"name": "Hijacked"}, other_category.pk)

        self.assertEqual(response.status_code, 404)


class ProductVariantManagementTests(ShopTestBase):
    """"Small", "Medium — Red", whatever label the club actually sells by --
    see ProductVariant's own docstring for why it's one free-text field, not
    separate size/colour axes."""

    def test_the_edit_page_breaks_sold_count_down_per_variant(self):
        small = ProductVariant.objects.create(product=self.product, name="Small")
        ProductVariant.objects.create(product=self.product, name="Large")
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-variant-sold@example.com")
        order = Order.objects.create(club=self.club, purchaser=purchaser, payment_status=Order.PaymentStatus.PAID, fulfillment_status=Order.FulfillmentStatus.DELIVERED, total=Decimal("50.00"))
        OrderLine.objects.create(order=order, product=self.product, variant=small, quantity=3, unit_price=Decimal("25.00"), line_total=Decimal("75.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_update", self.product.pk)

        variants_by_name = {variant.name: variant for variant in response.context["variants"]}
        self.assertEqual(variants_by_name["Small"].sold_count, 3)
        self.assertEqual(variants_by_name["Large"].sold_count, 0)
        self.assertEqual(response.context["product_sold_count"], 3)

    def test_a_shop_manager_can_add_a_variant(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_variant_create", {"name": "Small", "is_active": "on"}, self.product.pk)

        self.assertRedirects(response, reverse("management:product_update", kwargs={"pk": self.product.pk}))
        self.assertTrue(self.product.variants.filter(name="Small").exists())

    def test_a_variant_can_override_the_products_price(self):
        self.client.force_login(self.make_shop_manager())

        self.club_post("product_variant_create", {"name": "XXL", "price": "35.00", "is_active": "on"}, self.product.pk)

        variant = self.product.variants.get(name="XXL")
        self.assertEqual(variant.price, Decimal("35.00"))
        self.assertEqual(variant.effective_price, Decimal("35.00"))

    def test_plain_staff_cannot_add_a_variant(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("product_variant_create", {"name": "Small", "is_active": "on"}, self.product.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(self.product.variants.filter(name="Small").exists())

    def test_a_shop_manager_can_edit_a_variant(self):
        variant = ProductVariant.objects.create(product=self.product, name="Small")
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_variant_update", {"name": "S", "is_active": "on"}, self.product.pk, variant.pk)

        self.assertRedirects(response, reverse("management:product_update", kwargs={"pk": self.product.pk}))
        variant.refresh_from_db()
        self.assertEqual(variant.name, "S")

    def test_a_shop_manager_can_toggle_a_variant_active(self):
        variant = ProductVariant.objects.create(product=self.product, name="Small")
        self.client.force_login(self.make_shop_manager())

        self.club_post("product_variant_toggle_active", {}, self.product.pk, variant.pk)

        variant.refresh_from_db()
        self.assertFalse(variant.is_active)

    def test_plain_staff_cannot_toggle_a_variant(self):
        variant = ProductVariant.objects.create(product=self.product, name="Small")
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("product_variant_toggle_active", {}, self.product.pk, variant.pk)

        self.assertEqual(response.status_code, 403)
        variant.refresh_from_db()
        self.assertTrue(variant.is_active)

    def test_a_variant_from_another_products_edit_page_404s(self):
        other_product = Product.objects.create(club=self.club, name="Away Jersey")
        variant = ProductVariant.objects.create(product=other_product, name="Small")
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("product_variant_toggle_active", {}, self.product.pk, variant.pk)

        self.assertEqual(response.status_code, 404)

    def test_there_is_no_variant_delete_view(self):
        with self.assertRaises(NoReverseMatch):
            reverse("management:product_variant_delete")

    def test_the_product_edit_page_lists_its_variants(self):
        ProductVariant.objects.create(product=self.product, name="Small")
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("product_update", self.product.pk)

        self.assertContains(response, "Small")


class DiscountManagementTests(ShopTestBase):
    def discount_data(self, **overrides):
        data = {"name": "Summer sale", "description": "", "code": "summer10", "discount_type": DiscountType.PERCENTAGE, "discount_amount": "10", "is_active": "on"}
        data.update(overrides)
        return data

    def test_a_shop_manager_can_view_the_discount_list(self):
        Discount.objects.create(club=self.club, name="Existing", code="EXIST", discount_amount=Decimal("5"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("discount_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Existing")

    def test_a_shop_manager_can_create_a_discount_and_the_code_gets_uppercased(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("discount_create", self.discount_data())

        self.assertRedirects(response, reverse("management:discount_list"))
        discount = Discount.objects.get(club=self.club, name="Summer sale")
        self.assertEqual(discount.code, "SUMMER10")

    def test_plain_staff_cannot_create_a_discount(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("discount_create", self.discount_data())

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Discount.objects.filter(club=self.club, name="Summer sale").exists())

    def test_a_shop_manager_can_edit_a_discount(self):
        discount = Discount.objects.create(club=self.club, name="Existing", code="EXIST", discount_amount=Decimal("5"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("discount_update", self.discount_data(name="Renamed sale"), discount.pk)

        self.assertRedirects(response, reverse("management:discount_list"))
        discount.refresh_from_db()
        self.assertEqual(discount.name, "Renamed sale")

    def test_toggling_a_discount_flips_is_active(self):
        discount = Discount.objects.create(club=self.club, name="Existing", code="EXIST", discount_amount=Decimal("5"))
        self.client.force_login(self.make_shop_manager())
        self.assertTrue(discount.is_active)

        self.club_post("discount_toggle_active", {}, discount.pk)

        discount.refresh_from_db()
        self.assertFalse(discount.is_active)

    def test_plain_staff_cannot_toggle_a_discount(self):
        discount = Discount.objects.create(club=self.club, name="Existing", code="EXIST", discount_amount=Decimal("5"))
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("discount_toggle_active", {}, discount.pk)

        self.assertEqual(response.status_code, 403)


class VoucherManagementTests(ShopTestBase):
    def voucher_data(self, **overrides):
        data = {"issued_to": "", "amount": "50.00", "expiry_date": "2030-01-01", "is_active": "on", "notes": ""}
        data.update(overrides)
        return data

    def test_a_shop_manager_can_view_the_voucher_list(self):
        Voucher.objects.create(club=self.club, amount=Decimal("20.00"), expiry_date=datetime.date(2030, 1, 1))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("voucher_list")

        self.assertEqual(response.status_code, 200)

    def test_plain_staff_cannot_view_the_voucher_list(self):
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("voucher_list").status_code, 403)

    def test_a_shop_manager_can_create_a_voucher_and_gets_a_number(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("voucher_create", self.voucher_data())

        voucher = Voucher.objects.get(club=self.club, amount=Decimal("50.00"))
        self.assertRedirects(response, reverse("management:voucher_detail", args=[voucher.pk]))
        self.assertTrue(voucher.number.startswith("VCH-"))
        self.assertEqual(voucher.consumed_amount, Decimal("0"))

    def test_a_voucher_can_be_issued_to_a_member(self):
        member = Member.objects.create(first_name="Gigi", last_name="Given")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.make_shop_manager())

        self.club_post("voucher_create", self.voucher_data(issued_to=member.pk))

        voucher = Voucher.objects.get(club=self.club, amount=Decimal("50.00"))
        self.assertEqual(voucher.issued_to, member)

    def test_plain_staff_cannot_create_a_voucher(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("voucher_create", self.voucher_data())

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Voucher.objects.filter(club=self.club).exists())

    def test_a_shop_manager_can_edit_a_voucher(self):
        voucher = Voucher.objects.create(club=self.club, amount=Decimal("20.00"), expiry_date=datetime.date(2030, 1, 1))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("voucher_update", self.voucher_data(amount="30.00"), voucher.pk)

        self.assertRedirects(response, reverse("management:voucher_detail", args=[voucher.pk]))
        voucher.refresh_from_db()
        self.assertEqual(voucher.amount, Decimal("30.00"))

    def test_a_voucher_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-voucher")
        voucher = Voucher.objects.create(club=other_club, amount=Decimal("20.00"), expiry_date=datetime.date(2030, 1, 1))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("voucher_update", self.voucher_data(), voucher.pk)

        self.assertEqual(response.status_code, 404)


class VoucherConsumptionManagementTests(ShopTestBase):
    def make_voucher(self, amount=Decimal("50.00")):
        return Voucher.objects.create(club=self.club, amount=amount, expiry_date=datetime.date(2030, 1, 1))

    def test_a_shop_manager_can_record_a_manual_consumption(self):
        voucher = self.make_voucher()
        manager = self.make_shop_manager()
        self.client.force_login(manager)

        response = self.club_post("voucher_consumption_create", {"amount": "20.00", "note": "Cash payment at the clubhouse."}, voucher.pk)

        self.assertRedirects(response, reverse("management:voucher_detail", args=[voucher.pk]))
        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("20.00"))
        consumption = VoucherConsumption.objects.get(voucher=voucher)
        self.assertEqual(consumption.amount, Decimal("20.00"))
        self.assertEqual(consumption.note, "Cash payment at the clubhouse.")
        self.assertEqual(consumption.recorded_by, manager)

    def test_a_consumption_exceeding_the_available_amount_is_rejected(self):
        voucher = self.make_voucher(amount=Decimal("20.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("voucher_consumption_create", {"amount": "50.00", "note": "Too much."}, voucher.pk)

        self.assertRedirects(response, reverse("management:voucher_detail", args=[voucher.pk]))
        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("0"))
        self.assertFalse(VoucherConsumption.objects.filter(voucher=voucher).exists())

    def test_a_consumption_without_a_note_is_rejected(self):
        voucher = self.make_voucher()
        self.client.force_login(self.make_shop_manager())

        self.club_post("voucher_consumption_create", {"amount": "10.00", "note": ""}, voucher.pk)

        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("0"))
        self.assertFalse(VoucherConsumption.objects.filter(voucher=voucher).exists())

    def test_plain_staff_cannot_record_a_consumption(self):
        voucher = self.make_voucher()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("voucher_consumption_create", {"amount": "10.00", "note": "Nope."}, voucher.pk)

        self.assertEqual(response.status_code, 403)
        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("0"))

    def test_deleting_a_consumption_restores_the_available_amount(self):
        voucher = self.make_voucher()
        consumption = VoucherConsumption.objects.create(voucher=voucher, amount=Decimal("15.00"), note="Refundable test entry")
        voucher.consumed_amount = Decimal("15.00")
        voucher.save(update_fields=["consumed_amount"])
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("voucher_consumption_delete", {}, voucher.pk, consumption.pk)

        self.assertRedirects(response, reverse("management:voucher_detail", args=[voucher.pk]))
        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("0"))
        self.assertFalse(VoucherConsumption.objects.filter(pk=consumption.pk).exists())

    def test_plain_staff_cannot_delete_a_consumption(self):
        voucher = self.make_voucher()
        consumption = VoucherConsumption.objects.create(voucher=voucher, amount=Decimal("15.00"), note="Entry")
        voucher.consumed_amount = Decimal("15.00")
        voucher.save(update_fields=["consumed_amount"])
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("voucher_consumption_delete", {}, voucher.pk, consumption.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(VoucherConsumption.objects.filter(pk=consumption.pk).exists())

    def test_a_consumption_from_another_clubs_voucher_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-consumption")
        voucher = Voucher.objects.create(club=other_club, amount=Decimal("20.00"), expiry_date=datetime.date(2030, 1, 1))
        consumption = VoucherConsumption.objects.create(voucher=voucher, amount=Decimal("5.00"), note="Entry")
        self.client.force_login(self.make_shop_manager())

        create_response = self.club_post("voucher_consumption_create", {"amount": "5.00", "note": "Entry"}, voucher.pk)
        delete_response = self.club_post("voucher_consumption_delete", {}, voucher.pk, consumption.pk)

        self.assertEqual(create_response.status_code, 404)
        self.assertEqual(delete_response.status_code, 404)

    def test_the_voucher_detail_page_shows_manual_and_order_history_together(self):
        voucher = self.make_voucher()
        VoucherConsumption.objects.create(voucher=voucher, amount=Decimal("10.00"), note="Cash payment for the tournament fee.")
        voucher.consumed_amount = Decimal("10.00")
        voucher.save(update_fields=["consumed_amount"])
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        order = Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("10.00"))
        Payment.objects.create(order=order, voucher=voucher, amount=Decimal("10.00"), method=Payment.PaymentMethod.VOUCHER, status=Payment.PaymentStatus.CONFIRMED)
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("voucher_detail", voucher.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cash payment for the tournament fee.")
        self.assertContains(response, order.number)
        history = response.context["history"]
        self.assertEqual(len(history), 2)
        self.assertEqual({row["kind"] for row in history}, {"manual", "order"})


class OrderManagementTests(ShopTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly@example.com")

    def make_order(self, payment_status=Order.PaymentStatus.PENDING, fulfillment_status=Order.FulfillmentStatus.NOT_READY, total=Decimal("50.00")):
        order = Order.objects.create(club=self.club, purchaser=self.purchaser, payment_status=payment_status, fulfillment_status=fulfillment_status, total=total)
        OrderLine.objects.create(order=order, product=self.product, quantity=2, unit_price=Decimal("25.00"), line_total=total)
        create_invoice_for_order(order)
        return order

    def test_a_shop_manager_can_view_the_order_list(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("order_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, order.number)

    def test_plain_staff_cannot_view_the_order_list(self):
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("order_list").status_code, 403)

    def test_the_kpi_strip_reflects_paid_orders_regardless_of_the_status_filter(self):
        self.make_order(payment_status=Order.PaymentStatus.PAID, total=Decimal("30.00"))
        self.make_order(payment_status=Order.PaymentStatus.PENDING, total=Decimal("999.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("order_list", params={"payment_status": Order.PaymentStatus.PAID})

        self.assertEqual(response.context["kpis"]["total_orders"], 2)
        self.assertEqual(response.context["kpis"]["open_orders"], 2)
        self.assertEqual(response.context["kpis"]["total_sold"], Decimal("30.00"))

    def test_the_payment_status_filter_narrows_the_list(self):
        pending = self.make_order()
        paid = self.make_order(payment_status=Order.PaymentStatus.PAID)
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list", params={"payment_status": Order.PaymentStatus.PAID})

        self.assertContains(response, paid.number)
        self.assertNotContains(response, pending.number)

    def test_the_default_view_hides_delivered_and_paid_orders(self):
        pending = self.make_order()
        closed = self.make_order(payment_status=Order.PaymentStatus.PAID, fulfillment_status=Order.FulfillmentStatus.DELIVERED)
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list")

        self.assertContains(response, pending.number)
        self.assertNotContains(response, closed.number)

    def test_the_default_view_still_shows_delivered_but_unpaid_orders(self):
        # Payment still needs collecting -- delivered alone isn't "closed"
        # (Order.is_closed's own docstring; this is the exact case that used
        # to be wrongly hidden before payment/fulfillment status were split).
        delivered_unpaid = self.make_order(payment_status=Order.PaymentStatus.PENDING, fulfillment_status=Order.FulfillmentStatus.DELIVERED)
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list")

        self.assertContains(response, delivered_unpaid.number)

    def test_show_all_shows_closed_orders_too(self):
        closed = self.make_order(payment_status=Order.PaymentStatus.PAID, fulfillment_status=Order.FulfillmentStatus.DELIVERED)
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list", params={"show": "all"})

        self.assertContains(response, closed.number)

    def test_explicitly_filtering_on_delivered_still_shows_it(self):
        delivered = self.make_order(fulfillment_status=Order.FulfillmentStatus.DELIVERED)
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list", params={"fulfillment_status": Order.FulfillmentStatus.DELIVERED})

        self.assertContains(response, delivered.number)

    def test_search_matches_the_purchasers_name(self):
        match = Member.objects.create(first_name="Zelda", last_name="Zephyr")
        order = Order.objects.create(club=self.club, purchaser=match, payment_status=Order.PaymentStatus.PENDING, total=Decimal("10.00"))
        other = self.make_order()
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list", params={"q": "Zelda"})

        self.assertContains(response, order.number)
        self.assertNotContains(response, other.number)

    def test_search_matches_by_family_name(self):
        parent = Member.objects.create(first_name="Priya", last_name="Family")
        child = Member.objects.create(first_name="Cy", last_name="Family")
        family = Family.objects.create(name="Family")
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=family, member=child, role=FamilyMembership.FamilyRole.CHILD)
        order = Order.objects.create(club=self.club, purchaser=child, payment_status=Order.PaymentStatus.PENDING, total=Decimal("10.00"))
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list", params={"q": "Family"})

        self.assertContains(response, order.number)

    def test_search_matches_the_order_number(self):
        order = self.make_order()
        other = self.make_order()
        self.client.force_login(self.admin_user)

        response = self.club_get("order_list", params={"q": order.number})

        self.assertContains(response, order.number)
        self.assertNotContains(response, other.number)

    def test_a_shop_manager_can_view_order_detail(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("order_detail", order.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Home Jersey")
        self.assertContains(response, self.purchaser.first_name)

    def test_plain_staff_cannot_view_order_detail(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("order_detail", order.pk).status_code, 403)

    def test_marking_an_order_paid_creates_a_confirmed_cash_payment_by_default(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_mark_paid", {"method": Payment.PaymentMethod.CASH, "reference": "receipt-1"}, order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PAID)

        payment = Payment.objects.get(order=order)
        self.assertEqual(payment.amount, order.total)
        self.assertEqual(payment.method, Payment.PaymentMethod.CASH)
        self.assertEqual(payment.status, Payment.PaymentStatus.CONFIRMED)
        self.assertEqual(payment.reference, "receipt-1")
        self.assertIsNotNone(payment.paid_at)

    def test_marking_an_order_paid_by_bank_transfer(self):
        order = self.make_order()
        self.client.force_login(self.admin_user)

        self.club_post("order_mark_paid", {"method": Payment.PaymentMethod.BANK_TRANSFER, "reference": ""}, order.pk)

        payment = Payment.objects.get(order=order)
        self.assertEqual(payment.method, Payment.PaymentMethod.BANK_TRANSFER)

    def test_plain_staff_cannot_mark_an_order_paid(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_mark_paid", {"method": Payment.PaymentMethod.CASH, "reference": ""}, order.pk)

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Payment.objects.filter(order=order).exists())

    def test_marking_an_order_delivered_does_not_require_paid_first(self):
        order = self.make_order()
        self.client.force_login(self.admin_user)

        response = self.club_post("order_mark_delivered", {}, order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.DELIVERED)

    def test_marking_an_order_ready_for_pickup(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_mark_ready_for_pickup", {"pickup_instructions": "Ask at reception."}, order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.READY_FOR_PICKUP)
        self.assertEqual(order.pickup_instructions, "Ask at reception.")

    def test_pickup_instructions_are_optional(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_mark_ready_for_pickup", {"pickup_instructions": ""}, order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.READY_FOR_PICKUP)
        self.assertEqual(order.pickup_instructions, "")

    def test_marking_ready_for_pickup_notifies_the_purchaser(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_mark_ready_for_pickup", {"pickup_instructions": ""}, order.pk)

        notification = Notification.objects.get(member=self.purchaser)
        self.assertIn(order.number, notification.title)
        self.assertEqual(notification.source, order)

    def test_plain_staff_cannot_mark_ready_for_pickup(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_mark_ready_for_pickup", {"pickup_instructions": ""}, order.pk)

        self.assertEqual(response.status_code, 403)
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.NOT_READY)
        self.assertFalse(Notification.objects.filter(member=self.purchaser).exists())

    def test_an_order_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-order")
        other_purchaser = Member.objects.create(first_name="Otto", last_name="Other", email="otto@example.com")
        order = Order.objects.create(club=other_club, purchaser=other_purchaser, payment_status=Order.PaymentStatus.PENDING, total=Decimal("10.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_mark_ready_for_pickup", {"pickup_instructions": ""}, order.pk)

        self.assertEqual(response.status_code, 404)

    def test_bulk_mark_paid_marks_every_selected_order(self):
        first = self.make_order()
        second = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_bulk_mark_paid", {"order_ids": [str(first.pk), str(second.pk)], "method": Payment.PaymentMethod.CASH})

        self.assertRedirects(response, reverse("management:order_list"))
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.payment_status, Order.PaymentStatus.PAID)
        self.assertEqual(second.payment_status, Order.PaymentStatus.PAID)
        self.assertEqual(Payment.objects.filter(order__in=[first, second], status=Payment.PaymentStatus.CONFIRMED).count(), 2)

    def test_bulk_mark_paid_uses_the_chosen_method(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_bulk_mark_paid", {"order_ids": [str(order.pk)], "method": Payment.PaymentMethod.BANK_TRANSFER})

        payment = Payment.objects.get(order=order)
        self.assertEqual(payment.method, Payment.PaymentMethod.BANK_TRANSFER)
        self.assertEqual(payment.reference, "")

    def test_bulk_mark_paid_skips_cancelled_orders(self):
        cancelled = self.make_order(fulfillment_status=Order.FulfillmentStatus.CANCELLED)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_bulk_mark_paid", {"order_ids": [str(cancelled.pk)], "method": Payment.PaymentMethod.CASH})

        cancelled.refresh_from_db()
        self.assertEqual(cancelled.fulfillment_status, Order.FulfillmentStatus.CANCELLED)
        self.assertFalse(Payment.objects.filter(order=cancelled).exists())

    def test_bulk_mark_paid_with_nothing_selected_shows_an_error(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_bulk_mark_paid", {})

        self.assertRedirects(response, reverse("management:order_list"))
        self.assertFalse(Payment.objects.exists())

    def test_bulk_mark_paid_never_touches_another_clubs_order(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-bulk-paid")
        other_purchaser = Member.objects.create(first_name="Otto", last_name="Other", email="otto-bulk-paid@example.com")
        foreign_order = Order.objects.create(club=other_club, purchaser=other_purchaser, payment_status=Order.PaymentStatus.PENDING, total=Decimal("10.00"))
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_bulk_mark_paid", {"order_ids": [str(foreign_order.pk)], "method": Payment.PaymentMethod.CASH})

        foreign_order.refresh_from_db()
        self.assertEqual(foreign_order.payment_status, Order.PaymentStatus.PENDING)

    def test_plain_staff_cannot_bulk_mark_paid(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_bulk_mark_paid", {"order_ids": [str(order.pk)], "method": Payment.PaymentMethod.CASH})

        self.assertEqual(response.status_code, 403)
        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PENDING)

    def test_bulk_mark_ready_for_pickup_applies_the_shared_instructions_to_every_order(self):
        first = self.make_order()
        second = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_bulk_mark_ready_for_pickup", {"order_ids": [str(first.pk), str(second.pk)], "pickup_instructions": "Ring the bell."})

        self.assertRedirects(response, reverse("management:order_list"))
        first.refresh_from_db()
        second.refresh_from_db()
        self.assertEqual(first.fulfillment_status, Order.FulfillmentStatus.READY_FOR_PICKUP)
        self.assertEqual(second.fulfillment_status, Order.FulfillmentStatus.READY_FOR_PICKUP)
        self.assertEqual(first.pickup_instructions, "Ring the bell.")
        self.assertEqual(second.pickup_instructions, "Ring the bell.")

    def test_bulk_mark_ready_for_pickup_notifies_each_purchaser(self):
        first = self.make_order()
        second = self.make_order()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_bulk_mark_ready_for_pickup", {"order_ids": [str(first.pk), str(second.pk)], "pickup_instructions": ""})

        self.assertEqual(Notification.objects.filter(member=self.purchaser, title__icontains=first.number).count(), 1)
        self.assertEqual(Notification.objects.filter(member=self.purchaser, title__icontains=second.number).count(), 1)

    def test_bulk_mark_ready_for_pickup_skips_cancelled_orders(self):
        cancelled = self.make_order(fulfillment_status=Order.FulfillmentStatus.CANCELLED)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_bulk_mark_ready_for_pickup", {"order_ids": [str(cancelled.pk)], "pickup_instructions": ""})

        cancelled.refresh_from_db()
        self.assertEqual(cancelled.fulfillment_status, Order.FulfillmentStatus.CANCELLED)

    def test_bulk_mark_ready_for_pickup_with_nothing_selected_shows_an_error(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_bulk_mark_ready_for_pickup", {"pickup_instructions": ""})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_plain_staff_cannot_bulk_mark_ready_for_pickup(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_bulk_mark_ready_for_pickup", {"order_ids": [str(order.pk)], "pickup_instructions": ""})

        self.assertEqual(response.status_code, 403)
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.NOT_READY)

    def test_cancelling_an_order(self):
        order = self.make_order()
        self.client.force_login(self.admin_user)

        response = self.club_post("order_cancel", {}, order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.CANCELLED)

    def test_plain_staff_cannot_cancel_an_order(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_cancel", {}, order.pk)

        self.assertEqual(response.status_code, 403)
        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.NOT_READY)


class PaymentManagementTests(ShopTestBase):
    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-payment@example.com")

    def make_order(self, payment_status=Order.PaymentStatus.PAID, fulfillment_status=Order.FulfillmentStatus.NOT_READY, total=Decimal("50.00")):
        return Order.objects.create(club=self.club, purchaser=self.purchaser, payment_status=payment_status, fulfillment_status=fulfillment_status, total=total)

    def make_payment(self, order, **overrides):
        kwargs = {"order": order, "amount": order.total, "method": Payment.PaymentMethod.CASH, "status": Payment.PaymentStatus.CONFIRMED, "paid_at": timezone.now()}
        kwargs.update(overrides)
        return Payment.objects.create(**kwargs)

    def test_a_shop_manager_can_edit_a_payments_method_and_reference(self):
        order = self.make_order()
        payment = self.make_payment(order)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("payment_update", {"method": Payment.PaymentMethod.BANK_TRANSFER, "reference": "TX-42"}, order.pk, payment.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        payment.refresh_from_db()
        self.assertEqual(payment.method, Payment.PaymentMethod.BANK_TRANSFER)
        self.assertEqual(payment.reference, "TX-42")

    def test_editing_a_payment_never_touches_its_amount_or_status(self):
        order = self.make_order()
        payment = self.make_payment(order)
        self.client.force_login(self.make_shop_manager())

        self.club_post("payment_update", {"method": Payment.PaymentMethod.CREDIT_CARD, "reference": ""}, order.pk, payment.pk)

        payment.refresh_from_db()
        self.assertEqual(payment.amount, order.total)
        self.assertEqual(payment.status, Payment.PaymentStatus.CONFIRMED)

    def test_plain_staff_cannot_edit_a_payment(self):
        order = self.make_order()
        payment = self.make_payment(order)
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("payment_update", {"method": Payment.PaymentMethod.CASH, "reference": "nope"}, order.pk, payment.pk)

        self.assertEqual(response.status_code, 403)
        payment.refresh_from_db()
        self.assertNotEqual(payment.reference, "nope")

    def test_a_payment_from_another_order_404s(self):
        order = self.make_order()
        other_order = self.make_order()
        payment = self.make_payment(other_order)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("payment_update", {"method": Payment.PaymentMethod.CASH, "reference": ""}, order.pk, payment.pk)

        self.assertEqual(response.status_code, 404)

    def test_deleting_the_only_payment_drops_the_order_back_to_pending(self):
        order = self.make_order()
        payment = self.make_payment(order)
        self.client.force_login(self.make_shop_manager())

        url = reverse("management:payment_delete", args=[order.pk, payment.pk])
        response = self.client.post(url, {}, HTTP_HOST="ajax-united.rosterchief.app", follow=True)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        self.assertFalse(Payment.objects.filter(pk=payment.pk).exists())
        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PENDING)
        self.assertContains(response, "back to pending")

    def test_deleting_one_of_several_payments_recomputes_the_remaining_balance(self):
        # order.total is 50.00 (make_order's own default); two payments of 20 +
        # 30 sum to exactly that, hence PAID initially. Deleting the 20 leaves
        # only 30 collected -- genuinely PARTIALLY_PAID now, not still PAID.
        order = self.make_order()
        first = self.make_payment(order, amount=Decimal("20.00"))
        self.make_payment(order, amount=Decimal("30.00"))
        self.client.force_login(self.make_shop_manager())

        self.club_post("payment_delete", {}, order.pk, first.pk)

        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PARTIALLY_PAID)
        self.assertEqual(order.payments.count(), 1)

    def test_deleting_the_only_payment_on_a_cancelled_order_still_drops_payment_to_pending(self):
        # fulfillment_status=CANCELLED is untouched -- only payment_status ==
        # REFUNDED blocks the pending-revert (PaymentDeleteView's own docstring).
        order = self.make_order(fulfillment_status=Order.FulfillmentStatus.CANCELLED)
        payment = self.make_payment(order)
        self.client.force_login(self.make_shop_manager())

        self.club_post("payment_delete", {}, order.pk, payment.pk)

        order.refresh_from_db()
        self.assertEqual(order.fulfillment_status, Order.FulfillmentStatus.CANCELLED)
        self.assertEqual(order.payment_status, Order.PaymentStatus.PENDING)

    def test_deleting_the_only_payment_on_a_refunded_order_leaves_payment_status_alone(self):
        order = self.make_order(payment_status=Order.PaymentStatus.REFUNDED)
        payment = self.make_payment(order)
        self.client.force_login(self.make_shop_manager())

        self.club_post("payment_delete", {}, order.pk, payment.pk)

        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.REFUNDED)

    def test_plain_staff_cannot_delete_a_payment(self):
        order = self.make_order()
        payment = self.make_payment(order)
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("payment_delete", {}, order.pk, payment.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(Payment.objects.filter(pk=payment.pk).exists())

    def test_a_payment_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-payment")
        other_purchaser = Member.objects.create(first_name="Otto", last_name="Other", email="otto-payment@example.com")
        other_order = Order.objects.create(club=other_club, purchaser=other_purchaser, payment_status=Order.PaymentStatus.PAID, total=Decimal("10.00"))
        payment = self.make_payment(other_order)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("payment_delete", {}, other_order.pk, payment.pk)

        self.assertEqual(response.status_code, 404)


class OrderAddPaymentTests(ShopTestBase):
    """OrderAddPaymentView -- the flexible counterpart to OrderMarkPaidView's
    quick "pay everything now" shortcut. Covers what the voucher feature is
    actually for: a voucher covering part of an order, cash/OrderMarkPaidView
    covering the rest, in two separate submits."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-addpay@example.com")

    def make_order(self, total=Decimal("50.00")):
        return Order.objects.create(club=self.club, purchaser=self.purchaser, total=total)

    def make_voucher(self, amount=Decimal("30.00"), expiry_date=datetime.date(2030, 1, 1), club=None, **overrides):
        return Voucher.objects.create(club=club or self.club, amount=amount, expiry_date=expiry_date, **overrides)

    def add_payment_data(self, **overrides):
        data = {"add_payment-amount": "20.00", "add_payment-method": Payment.PaymentMethod.CASH, "add_payment-voucher": "", "add_payment-reference": ""}
        data.update(overrides)
        return data

    def test_a_shop_manager_can_record_a_partial_cash_payment(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_add_payment", self.add_payment_data(), order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PARTIALLY_PAID)
        payment = Payment.objects.get(order=order)
        self.assertEqual(payment.amount, Decimal("20.00"))

    def test_a_voucher_payment_deducts_from_the_vouchers_balance(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("30.00"))
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("30.00"))
        payment = Payment.objects.get(order=order)
        self.assertEqual(payment.voucher, voucher)
        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PARTIALLY_PAID)

    def test_a_voucher_issued_to_the_purchaser_is_selectable(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("30.00"), issued_to=self.purchaser)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertTrue(Payment.objects.filter(order=order, voucher=voucher).exists())

    def test_a_voucher_issued_to_a_family_member_is_selectable(self):
        order = self.make_order()
        parent = Member.objects.create(first_name="Pat", last_name="Orderer")
        family = Family.objects.create(name="Orderer family")
        FamilyMembership.objects.create(family=family, member=self.purchaser, role=FamilyMembership.FamilyRole.CHILD)
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)
        voucher = self.make_voucher(amount=Decimal("30.00"), issued_to=parent)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertTrue(Payment.objects.filter(order=order, voucher=voucher).exists())

    def test_an_unassigned_voucher_is_selectable(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("30.00"))  # no issued_to
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertTrue(Payment.objects.filter(order=order, voucher=voucher).exists())

    def test_a_voucher_issued_to_someone_outside_the_family_is_not_selectable(self):
        order = self.make_order()
        outsider = Member.objects.create(first_name="Otto", last_name="Outsider")
        voucher = self.make_voucher(amount=Decimal("30.00"), issued_to=outsider)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertFalse(Payment.objects.filter(order=order).exists())
        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("0"))

    def test_a_voucher_and_then_cash_together_fully_settle_the_order(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("30.00"))
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)
        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "20.00"}), order.pk)

        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PAID)
        self.assertEqual(order.payments.count(), 2)

    def test_a_voucher_payment_over_the_available_balance_is_rejected(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("10.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "20.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        self.assertFalse(Payment.objects.filter(order=order).exists())
        voucher.refresh_from_db()
        self.assertEqual(voucher.consumed_amount, Decimal("0"))

    def test_an_expired_voucher_is_not_selectable(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("50.00"), expiry_date=datetime.date(2020, 1, 1))
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "20.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertFalse(Payment.objects.filter(order=order).exists())

    def test_no_voucher_selected_is_rejected(self):
        order = self.make_order()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "20.00", "add_payment-method": Payment.PaymentMethod.VOUCHER}), order.pk)

        self.assertFalse(Payment.objects.filter(order=order).exists())

    def test_a_voucher_from_another_club_is_not_selectable(self):
        order = self.make_order()
        other_club = Club.objects.create(name="Other Club", slug="other-club-voucher-pay")
        voucher = self.make_voucher(amount=Decimal("50.00"), club=other_club)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "20.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.assertFalse(Payment.objects.filter(order=order).exists())

    def test_an_amount_over_the_orders_remaining_balance_is_rejected(self):
        order = self.make_order(total=Decimal("50.00"))
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "100.00"}), order.pk)

        self.assertFalse(Payment.objects.filter(order=order).exists())

    def test_mark_paid_only_tops_up_the_remaining_balance_after_a_voucher_payment(self):
        order = self.make_order()
        voucher = self.make_voucher(amount=Decimal("30.00"))
        self.client.force_login(self.make_shop_manager())
        self.club_post("order_add_payment", self.add_payment_data(**{"add_payment-amount": "30.00", "add_payment-method": Payment.PaymentMethod.VOUCHER, "add_payment-voucher": voucher.pk}), order.pk)

        self.club_post("order_mark_paid", {"method": Payment.PaymentMethod.CASH, "reference": ""}, order.pk)

        order.refresh_from_db()
        self.assertEqual(order.payment_status, Order.PaymentStatus.PAID)
        cash_payment = Payment.objects.get(order=order, method=Payment.PaymentMethod.CASH)
        self.assertEqual(cash_payment.amount, Decimal("20.00"))

    def test_plain_staff_cannot_add_a_payment(self):
        order = self.make_order()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_add_payment", self.add_payment_data(), order.pk)

        self.assertEqual(response.status_code, 403)


class OrderLineManagementTests(ShopTestBase):
    """OrderLineUpdateView -- fixing a mistake on an existing line item after
    the fact. product/unit_price/line_total stay derived, not directly
    editable (OrderLineEditForm/OrderLineUpdateView's own docstrings)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-lines@example.com")

    def make_order_with_line(self, quantity=2, unit_price=Decimal("25.00")):
        order = Order.objects.create(club=self.club, purchaser=self.purchaser, total=unit_price * quantity)
        line = OrderLine.objects.create(order=order, product=self.product, quantity=quantity, unit_price=unit_price, line_total=unit_price * quantity)
        return order, line

    def test_a_shop_manager_can_change_the_quantity_and_the_order_total_recomputes(self):
        order, line = self.make_order_with_line(quantity=2, unit_price=Decimal("25.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_line_update", {"quantity": "3", "beneficiary": "", "variant": "", "production_status": ProductionStatus.PENDING}, order.pk, line.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        line.refresh_from_db()
        self.assertEqual(line.quantity, 3)
        self.assertEqual(line.line_total, Decimal("75.00"))
        order.refresh_from_db()
        self.assertEqual(order.total, Decimal("75.00"))

    def test_changing_the_variant_recomputes_the_unit_price(self):
        order, line = self.make_order_with_line(quantity=1, unit_price=Decimal("25.00"))
        variant = ProductVariant.objects.create(product=self.product, name="XL", price=Decimal("30.00"))
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_line_update", {"quantity": "1", "beneficiary": "", "variant": variant.pk, "production_status": ProductionStatus.PENDING}, order.pk, line.pk)

        line.refresh_from_db()
        self.assertEqual(line.unit_price, Decimal("30.00"))
        self.assertEqual(line.line_total, Decimal("30.00"))

    def test_growing_the_total_drops_a_paid_order_back_to_partially_paid(self):
        order, line = self.make_order_with_line(quantity=1, unit_price=Decimal("25.00"))
        Payment.objects.create(order=order, amount=Decimal("25.00"), method=Payment.PaymentMethod.CASH, status=Payment.PaymentStatus.CONFIRMED)
        order.payment_status = Order.PaymentStatus.PAID
        order.save(update_fields=["payment_status"])
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_line_update", {"quantity": "2", "beneficiary": "", "variant": "", "production_status": ProductionStatus.PENDING}, order.pk, line.pk)

        order.refresh_from_db()
        self.assertEqual(order.total, Decimal("50.00"))
        self.assertEqual(order.payment_status, Order.PaymentStatus.PARTIALLY_PAID)

    def test_personalization_fields_are_only_editable_when_the_product_allows_it(self):
        self.product.personalization_enabled = True
        self.product.save(update_fields=["personalization_enabled"])
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post(
            "order_line_update", {"quantity": "2", "beneficiary": "", "variant": "", "personalization_number": "7", "personalization_name": "SMITH", "production_status": ProductionStatus.PENDING}, order.pk, line.pk
        )

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        line.refresh_from_db()
        self.assertEqual(line.personalization_number, "7")
        self.assertEqual(line.personalization_name, "SMITH")

    def test_personalization_is_ignored_when_the_product_doesnt_allow_it(self):
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_line_update", {"quantity": "2", "beneficiary": "", "variant": "", "personalization_number": "7", "personalization_name": "SMITH"}, order.pk, line.pk)

        line.refresh_from_db()
        self.assertEqual(line.personalization_number, "")
        self.assertEqual(line.personalization_name, "")

    def test_a_beneficiary_can_be_set(self):
        order, line = self.make_order_with_line()
        beneficiary = Member.objects.create(first_name="Cy", last_name="Child")
        ClubMembership.objects.create(club=self.club, member=beneficiary, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_line_update", {"quantity": "2", "beneficiary": beneficiary.pk, "variant": "", "production_status": ProductionStatus.PENDING}, order.pk, line.pk)

        line.refresh_from_db()
        self.assertEqual(line.beneficiary, beneficiary)

    def test_production_status_can_be_set_by_hand_for_a_merchandise_product(self):
        self.product.product_type = Product.ProductType.MERCHANDISE
        self.product.save(update_fields=["product_type"])
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_line_update", {"quantity": "2", "beneficiary": "", "variant": "", "production_status": ProductionStatus.RECEIVED}, order.pk, line.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        line.refresh_from_db()
        self.assertEqual(line.production_status, ProductionStatus.RECEIVED)
        order.refresh_from_db()
        self.assertEqual(order.production_status, ProductionStatus.RECEIVED)

    def test_production_status_is_ignored_for_a_non_merchandise_product(self):
        self.product.product_type = Product.ProductType.MEMBERSHIP
        self.product.save(update_fields=["product_type"])
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_line_update", {"quantity": "2", "beneficiary": "", "variant": "", "production_status": ProductionStatus.RECEIVED}, order.pk, line.pk)

        line.refresh_from_db()
        self.assertEqual(line.production_status, ProductionStatus.PENDING)

    def test_plain_staff_cannot_edit_a_line_item(self):
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_line_update", {"quantity": "3", "beneficiary": "", "variant": ""}, order.pk, line.pk)

        self.assertEqual(response.status_code, 403)
        line.refresh_from_db()
        self.assertEqual(line.quantity, 2)

    def test_a_line_item_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-line")
        other_purchaser = Member.objects.create(first_name="Otto", last_name="Other", email="otto-line@example.com")
        other_product = Product.objects.create(club=other_club, name="Other Jersey", price=Decimal("25.00"))
        other_order = Order.objects.create(club=other_club, purchaser=other_purchaser, total=Decimal("25.00"))
        other_line = OrderLine.objects.create(order=other_order, product=other_product, quantity=1, unit_price=Decimal("25.00"), line_total=Decimal("25.00"))
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_line_update", {"quantity": "2", "beneficiary": "", "variant": ""}, other_order.pk, other_line.pk)

        self.assertEqual(response.status_code, 404)


class OrderLineMarkReceivedTests(ShopTestBase):
    """OrderLineMarkReceivedView -- the one-click "Mark received" button next
    to a line already IN_PRODUCTION (order_detail.html), an alternative to
    OrderLineEditForm's production_status dropdown for the common case."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.product.product_type = Product.ProductType.MERCHANDISE
        cls.product.save(update_fields=["product_type"])
        cls.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-received@example.com")

    def make_order_with_line(self, production_status=ProductionStatus.IN_PRODUCTION, quantity=1, unit_price=Decimal("25.00")):
        order = Order.objects.create(club=self.club, purchaser=self.purchaser, total=unit_price * quantity)
        line = OrderLine.objects.create(order=order, product=self.product, quantity=quantity, unit_price=unit_price, line_total=unit_price * quantity, production_status=production_status)
        return order, line

    def test_a_shop_manager_can_mark_a_line_received(self):
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_line_mark_received", {}, order.pk, line.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[order.pk]))
        line.refresh_from_db()
        self.assertEqual(line.production_status, ProductionStatus.RECEIVED)
        order.refresh_from_db()
        self.assertEqual(order.production_status, ProductionStatus.RECEIVED)

    def test_plain_staff_cannot_mark_a_line_received(self):
        order, line = self.make_order_with_line()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_line_mark_received", {}, order.pk, line.pk)

        self.assertEqual(response.status_code, 403)
        line.refresh_from_db()
        self.assertEqual(line.production_status, ProductionStatus.IN_PRODUCTION)

    def test_a_line_item_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-received")
        other_purchaser = Member.objects.create(first_name="Otto", last_name="Other", email="otto-received@example.com")
        other_product = Product.objects.create(club=other_club, name="Other Jersey", price=Decimal("25.00"))
        other_order = Order.objects.create(club=other_club, purchaser=other_purchaser, total=Decimal("25.00"))
        other_line = OrderLine.objects.create(order=other_order, product=other_product, quantity=1, unit_price=Decimal("25.00"), line_total=Decimal("25.00"), production_status=ProductionStatus.IN_PRODUCTION)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_line_mark_received", {}, other_order.pk, other_line.pk)

        self.assertEqual(response.status_code, 404)


class OrderProductionExportTests(ShopTestBase):
    """OrderProductionExportView -- the Orders list's own "Download order
    list" modal. Exports pending (not-yet-in-production) lines for the
    selected merchandise products and marks them in production, so a repeat
    download only ever includes what's new since.

    Redirects to order_list?production_download=<token> rather than
    returning the file directly (see the view's own docstring for why); the
    tests below follow that redirect and hit the download URL themselves,
    the same two requests order_list.html's own auto-download script fires
    for real in a browser."""

    def setUp(self):
        super().setUp()
        self.product.product_type = Product.ProductType.MERCHANDISE
        self.product.personalization_enabled = True
        self.product.save(update_fields=["product_type", "personalization_enabled"])
        self.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-export@example.com")

    def make_line(self, product=None, order=None, **overrides):
        order = order or Order.objects.create(club=self.club, purchaser=self.purchaser, total=Decimal("25.00"))
        kwargs = {"order": order, "product": product or self.product, "quantity": 1, "unit_price": Decimal("25.00"), "line_total": Decimal("25.00")}
        kwargs.update(overrides)
        return OrderLine.objects.create(**kwargs)

    def download(self, response):
        """Follows an export POST's own redirect to the file it stashed."""
        self.assertEqual(response.status_code, 302)
        token = parse_qs(urlparse(response.url).query)["production_download"][0]
        return self.club_get("order_production_download", token)

    def test_the_export_lists_beneficiary_and_personalization(self):
        beneficiary = Member.objects.create(first_name="Cy", last_name="Child")
        line = self.make_line(beneficiary=beneficiary, personalization_number="7", personalization_name="SMITH")
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})
        download = self.download(response)

        self.assertEqual(download.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(download.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertIn((line.order.number, "Child", "Cy", None, "7", "SMITH", 1), rows)

    def test_redirects_to_order_list_with_a_download_token(self):
        self.make_line()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})

        self.assertTrue(response.url.startswith(reverse("management:order_list")))
        self.assertIn("production_download=", response.url)

    def test_exported_lines_are_marked_in_production_before_the_file_is_even_downloaded(self):
        # The marking is a side effect of the POST itself, not of following
        # the redirect or downloading the file -- this is exactly the "does
        # it actually mark, regardless of what the browser does with the
        # download" regression check.
        line = self.make_line()
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})

        line.refresh_from_db()
        self.assertEqual(line.production_status, ProductionStatus.IN_PRODUCTION)
        line.order.refresh_from_db()
        self.assertEqual(line.order.production_status, ProductionStatus.IN_PRODUCTION)

    def test_the_download_token_is_one_time_use(self):
        self.make_line()
        self.client.force_login(self.make_shop_manager())
        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})
        token = parse_qs(urlparse(response.url).query)["production_download"][0]

        first = self.club_get("order_production_download", token)
        second = self.club_get("order_production_download", token)

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 404)

    def test_a_repeat_download_has_nothing_pending_left_to_export(self):
        self.make_line()
        self.client.force_login(self.make_shop_manager())
        self.download(self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]}))

        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})

        # Nothing pending -> notify + redirect straight to order_list, no token.
        self.assertRedirects(response, reverse("management:order_list"))

    def test_cancelled_orders_are_excluded(self):
        order = Order.objects.create(club=self.club, purchaser=self.purchaser, total=Decimal("25.00"), fulfillment_status=Order.FulfillmentStatus.CANCELLED)
        self.make_line(order=order)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_only_selected_products_are_included(self):
        other_product = Product.objects.create(club=self.club, name="Away Kit", price=Decimal("20.00"), product_type=Product.ProductType.MERCHANDISE)
        self.make_line()
        other_line = self.make_line(product=other_product)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})
        download = self.download(response)

        workbook = openpyxl.load_workbook(BytesIO(download.content))
        self.assertEqual(workbook.sheetnames, [self.product.name])
        other_line.refresh_from_db()
        self.assertEqual(other_line.production_status, ProductionStatus.PENDING)

    def test_no_products_selected_shows_an_error(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_export", {})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_plain_staff_cannot_download_the_export(self):
        self.make_line()
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_production_export", {"product_ids": [str(self.product.pk)]})

        self.assertEqual(response.status_code, 403)

    def test_a_download_token_from_another_club_404s(self):
        other_club = Club.objects.create(name="Other Club", slug="other-club-production-dl")
        other_purchaser = Member.objects.create(first_name="Otto", last_name="Other")
        other_product = Product.objects.create(club=other_club, name="Other Jersey", price=Decimal("25.00"), product_type=Product.ProductType.MERCHANDISE)
        other_order = Order.objects.create(club=other_club, purchaser=other_purchaser, total=Decimal("25.00"))
        other_line = OrderLine.objects.create(order=other_order, product=other_product, quantity=1, unit_price=Decimal("25.00"), line_total=Decimal("25.00"))
        token = stash_production_export(other_club, build_production_export([other_line]), "other-club.xlsx")
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("order_production_download", token)

        self.assertEqual(response.status_code, 404)


class OrderProductionReprintTests(ShopTestBase):
    """OrderProductionReprintView -- the Orders list's own "Redownload in
    production" action. A fresh, unchanged copy of whatever's already been
    sent to a manufacturer for the selected products; unlike
    OrderProductionExportView, this never marks anything."""

    def setUp(self):
        super().setUp()
        self.product.product_type = Product.ProductType.MERCHANDISE
        self.product.save(update_fields=["product_type"])
        self.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly-reprint@example.com")

    def make_line(self, product=None, order=None, **overrides):
        order = order or Order.objects.create(club=self.club, purchaser=self.purchaser, total=Decimal("25.00"))
        kwargs = {"order": order, "product": product or self.product, "quantity": 1, "unit_price": Decimal("25.00"), "line_total": Decimal("25.00")}
        kwargs.update(overrides)
        return OrderLine.objects.create(**kwargs)

    def test_reprints_lines_currently_in_production(self):
        line = self.make_line(production_status=ProductionStatus.IN_PRODUCTION)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_reprint", {"product_ids": [str(self.product.pk)]})

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertIn((line.order.number, "Orderer", "Olly", None, None, None, 1), rows)

    def test_does_not_mark_anything(self):
        line = self.make_line(production_status=ProductionStatus.IN_PRODUCTION)
        self.client.force_login(self.make_shop_manager())

        self.club_post("order_production_reprint", {"product_ids": [str(self.product.pk)]})

        line.refresh_from_db()
        self.assertEqual(line.production_status, ProductionStatus.IN_PRODUCTION)

    def test_pending_lines_are_not_included(self):
        self.make_line()  # still PENDING, never exported
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_reprint", {"product_ids": [str(self.product.pk)]})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_received_lines_are_not_included(self):
        self.make_line(production_status=ProductionStatus.RECEIVED)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_reprint", {"product_ids": [str(self.product.pk)]})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_cancelled_orders_are_excluded(self):
        order = Order.objects.create(club=self.club, purchaser=self.purchaser, total=Decimal("25.00"), fulfillment_status=Order.FulfillmentStatus.CANCELLED)
        self.make_line(order=order, production_status=ProductionStatus.IN_PRODUCTION)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_reprint", {"product_ids": [str(self.product.pk)]})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_no_products_selected_shows_an_error(self):
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("order_production_reprint", {})

        self.assertRedirects(response, reverse("management:order_list"))

    def test_plain_staff_cannot_reprint(self):
        self.make_line(production_status=ProductionStatus.IN_PRODUCTION)
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("order_production_reprint", {"product_ids": [str(self.product.pk)]})

        self.assertEqual(response.status_code, 403)


class InvoicePdfViewTests(ShopTestBase):
    """Mirrors club.tests.InvoicePdfTests/shop.tests.InvoicePdfTests's own mocking
    technique -- patched where management.views imported it, not at its source."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.purchaser = Member.objects.create(first_name="Olly", last_name="Orderer", email="olly@example.com")
        cls.order = Order.objects.create(club=cls.club, purchaser=cls.purchaser, total=Decimal("25.00"))
        OrderLine.objects.create(order=cls.order, product=cls.product, quantity=1, unit_price=Decimal("25.00"), line_total=Decimal("25.00"))
        cls.invoice = create_invoice_for_order(cls.order)

    def test_a_shop_manager_can_download_the_pdf(self):
        self.client.force_login(self.make_shop_manager())

        with mock.patch("management.views.render_invoice_pdf", return_value=b"%PDF-1.4 fake") as renderer:
            response = self.club_get("order_invoice_pdf", self.order.pk)

        renderer.assert_called_once_with(self.invoice)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(self.invoice.number, response["Content-Disposition"])

    def test_a_pdf_error_notifies_and_redirects_back_to_the_order(self):
        self.client.force_login(self.admin_user)

        with mock.patch("management.views.render_invoice_pdf", side_effect=ShopInvoicePDFError("boom")):
            response = self.club_get("order_invoice_pdf", self.order.pk)

        self.assertRedirects(response, reverse("management:order_detail", args=[self.order.pk]))

    def test_plain_staff_cannot_download_the_pdf(self):
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("order_invoice_pdf", self.order.pk).status_code, 403)


class InvoiceListViewTests(ShopTestBase):
    def test_a_shop_manager_can_view_the_invoice_list(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        order = Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("10.00"))
        invoice = create_invoice_for_order(order)
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("invoice_list")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, invoice.number)
        self.assertContains(response, reverse("management:order_detail", args=[order.pk]))

    def test_plain_staff_cannot_view_the_invoice_list(self):
        self.client.force_login(self.make_plain_staff())

        self.assertEqual(self.club_get("invoice_list").status_code, 403)

    def test_an_admin_can_view_the_invoice_list_with_shop_disabled(self):
        # Dues/registration invoices have nothing to do with the shop --
        # an admin must still reach this page at a club that's never turned
        # shop on at all.
        get_waffle_flag_model().objects.filter(name="shop").first().clubs.remove(self.club)
        self.client.force_login(self.admin_user)

        self.assertEqual(self.club_get("invoice_list").status_code, 200)

    def test_a_non_admin_shop_manager_cannot_view_the_invoice_list_with_shop_disabled(self):
        shop_manager = self.make_shop_manager()
        get_waffle_flag_model().objects.filter(name="shop").first().clubs.remove(self.club)
        self.client.force_login(shop_manager)

        self.assertEqual(self.club_get("invoice_list").status_code, 403)

    def test_every_invoice_created_gets_its_own_number(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        first = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("1")))
        second = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("2")))

        self.assertNotEqual(first.number, second.number)
        self.assertEqual(Invoice.objects.filter(club=self.club).count(), 2)

    def make_dues_invoice(self, member, **overrides):
        membership = ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE, fee_amount=Decimal("100.00"))
        kwargs = {"due_in_days": 14, "recipient_email": member.email or "parent@example.com", "sent_to_guardian": False}
        kwargs.update(overrides)
        return create_or_resend_invoice(membership, **kwargs)

    def test_dues_invoices_appear_for_an_admin(self):
        member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana@example.com")
        invoice = self.make_dues_invoice(member)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertContains(response, invoice.number)
        self.assertContains(response, "Dana")

    def test_a_dues_invoice_for_a_cancelled_membership_is_excluded(self):
        member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana-cancelled@example.com")
        invoice = self.make_dues_invoice(member)
        invoice.membership.status = ClubMembership.StatusChoices.CANCELLED
        invoice.membership.save()
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertNotContains(response, invoice.number)

    def test_dues_invoices_are_hidden_from_a_non_admin_shop_manager(self):
        member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana@example.com")
        invoice = self.make_dues_invoice(member)
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("invoice_list")

        self.assertNotContains(response, invoice.number)
        self.assertNotContains(response, '<option value="dues"')

    def test_a_non_admin_requesting_type_dues_still_sees_no_dues_rows(self):
        member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana@example.com")
        invoice = self.make_dues_invoice(member)
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("invoice_list", params={"type": "dues"})

        self.assertNotContains(response, invoice.number)

    def test_type_filter_order_excludes_dues(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        order_invoice = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("10.00")))
        dues_member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana2@example.com")
        dues_invoice = self.make_dues_invoice(dues_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list", params={"type": "order"})

        self.assertContains(response, order_invoice.number)
        self.assertNotContains(response, dues_invoice.number)

    def test_type_filter_dues_excludes_orders(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        order_invoice = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("10.00")))
        dues_member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana3@example.com")
        dues_invoice = self.make_dues_invoice(dues_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list", params={"type": "dues"})

        self.assertContains(response, dues_invoice.number)
        self.assertNotContains(response, order_invoice.number)

    def test_search_matches_the_purchasers_name(self):
        match = Member.objects.create(first_name="Zelda", last_name="Zephyr")
        other = Member.objects.create(first_name="Olly", last_name="Orderer")
        matching_invoice = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=match, total=Decimal("10.00")))
        other_invoice = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=other, total=Decimal("10.00")))
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list", params={"q": "Zelda"})

        self.assertContains(response, matching_invoice.number)
        self.assertNotContains(response, other_invoice.number)

    def test_search_matches_by_family_name(self):
        parent = Member.objects.create(first_name="Priya", last_name="Family", email="priya@example.com")
        child = Member.objects.create(first_name="Cy", last_name="Family")
        family = Family.objects.create(name="Family")
        FamilyMembership.objects.create(family=family, member=parent, role=FamilyMembership.FamilyRole.PARENT)
        FamilyMembership.objects.create(family=family, member=child, role=FamilyMembership.FamilyRole.CHILD)
        invoice = self.make_dues_invoice(child, recipient_email=parent.email, sent_to_guardian=True)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list", params={"q": "Family"})

        self.assertContains(response, invoice.number)

    def test_download_links_point_at_the_right_pdf_view(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        order = Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("10.00"))
        create_invoice_for_order(order)
        dues_member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana4@example.com")
        dues_invoice = self.make_dues_invoice(dues_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertContains(response, reverse("management:order_invoice_pdf", args=[order.pk]))
        self.assertContains(response, reverse("management:membership_invoice_pdf", args=[dues_invoice.membership.pk]))

    def test_rows_are_sorted_newest_first_across_both_kinds(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        older_order = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("1")))
        older_order.issued_at = timezone.now() - datetime.timedelta(days=5)
        older_order.save(update_fields=["issued_at"])
        dues_member = Member.objects.create(first_name="Dana", last_name="Duesworth", email="dana5@example.com")
        newer_dues = self.make_dues_invoice(dues_member)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertLess(response.content.find(newer_dues.number.encode()), response.content.find(older_order.number.encode()))

    def test_pagination_splits_the_merged_list(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        for _i in range(26):
            create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("1")))
        self.client.force_login(self.admin_user)

        first_page = self.club_get("invoice_list")
        second_page = self.club_get("invoice_list", params={"page": 2})

        self.assertEqual(len(first_page.context["invoices"]), 25)
        self.assertEqual(len(second_page.context["invoices"]), 1)

    def make_registration_batch(self, contact_email="reg-parent@example.com", price=Decimal("80.00")):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=price)
        variant = ProductVariant.objects.create(product=product, name="Standard", price=price)
        entries = [EntryInput(first_name="Kid", last_name="Testerson", date_of_birth=datetime.date(2016, 5, 1), product_variant=variant)]
        return submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email=contact_email, entries=entries)

    def test_an_unconfirmed_registration_does_not_appear(self):
        # Nothing financial reaches the Invoices list until staff has
        # reviewed and confirmed it on the Registrations queue -- see
        # RegistrationInvoiceQueueViewTests for where an unconfirmed one
        # actually shows up instead.
        batch = self.make_registration_batch()
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertNotContains(response, batch.contact_email)
        self.assertEqual(len(response.context["invoices"]), 0)

    def test_a_confirmed_registration_appears(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertContains(response, batch.invoice_number)
        self.assertContains(response, "Reg Parent")
        self.assertContains(response, "80.00")

    def test_registrations_are_grouped_one_row_per_batch_not_per_person(self):
        product = Product.objects.create(club=self.club, name="Sibling Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("50.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("50.00"))
        entries = [
            EntryInput(first_name="Kid", last_name="One", date_of_birth=datetime.date(2015, 5, 1), product_variant=variant),
            EntryInput(first_name="Kid", last_name="Two", date_of_birth=datetime.date(2017, 5, 1), product_variant=variant),
        ]
        batch = submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email="reg-siblings@example.com", entries=entries)
        confirm_and_send_invoice(batch, due_in_days=14)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertContains(response, "100.00")
        self.assertEqual(len(response.context["invoices"]), 1)

    def test_registration_rows_are_hidden_from_a_non_admin_shop_manager(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("invoice_list")

        self.assertNotContains(response, batch.invoice_number)
        self.assertNotContains(response, '<option value="registration"')

    def test_type_filter_registration_excludes_order_and_dues(self):
        purchaser = Member.objects.create(first_name="Olly", last_name="Orderer")
        order_invoice = create_invoice_for_order(Order.objects.create(club=self.club, purchaser=purchaser, total=Decimal("10.00")))
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list", params={"type": "registration"})

        self.assertContains(response, batch.invoice_number)
        self.assertNotContains(response, order_invoice.number)

    def test_a_registration_still_appears_once_a_dues_invoice_is_also_sent_for_it(self):
        # The old "already covered by a DuesInvoice" exclusion is gone --
        # a registration invoice and a DuesInvoice are two independent ways
        # a membership can end up billed, and this list only ever shows
        # confirmed registration invoices now, so there's no draft-vs-
        # DuesInvoice ambiguity left to resolve.
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        membership = batch.entries.get().membership
        create_or_resend_invoice(membership, due_in_days=14, recipient_email="reg-parent@example.com", sent_to_guardian=False)
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertContains(response, batch.invoice_number)

    def test_a_fully_cancelled_registration_is_excluded(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        membership = batch.entries.get().membership
        membership.status = ClubMembership.StatusChoices.CANCELLED
        membership.save()
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertNotContains(response, batch.invoice_number)

    def test_registration_row_links_to_its_own_checkout_screen_and_pdf(self):
        # Not the contact's member page any more -- a registration covers a
        # whole group, not any one person (management.views.
        # RegistrationInvoiceDetailView).
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.admin_user)

        response = self.club_get("invoice_list")

        self.assertContains(response, reverse("management:registration_invoice_detail", args=[batch.pk]))
        self.assertContains(response, reverse("management:registration_invoice_pdf", args=[batch.pk]))

    def test_downloading_a_registration_invoice_pdf(self):
        batch = self.make_registration_batch()
        self.client.force_login(self.admin_user)

        with mock.patch("management.views.batch_invoice_pdf", return_value=b"%PDF-fake"):
            response = self.club_get("registration_invoice_pdf", batch.pk)

        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertEqual(response.content, b"%PDF-fake")

    def test_downloading_a_registration_invoice_pdf_is_admin_only(self):
        batch = self.make_registration_batch()
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("registration_invoice_pdf", batch.pk)

        self.assertEqual(response.status_code, 403)


class RegistrationInvoiceDetailViewTests(ShopTestBase):
    """management:registration_invoice_detail -- the read-only "checkout
    screen" InvoiceListView's own registration rows link to."""

    def make_registration_batch(self, contact_email="reg-parent@example.com", price=Decimal("80.00")):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=price)
        variant = ProductVariant.objects.create(product=product, name="Standard", price=price)
        entries = [EntryInput(first_name="Kid", last_name="Testerson", date_of_birth=datetime.date(2016, 5, 1), product_variant=variant)]
        return submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email=contact_email, entries=entries)

    def test_shows_every_line_and_the_total(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.admin_user)

        response = self.club_get("registration_invoice_detail", batch.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Testerson")
        self.assertContains(response, batch.invoice_number)

    def test_404s_before_the_invoice_is_confirmed(self):
        # Still being worked from the Registrations review screen itself --
        # nothing to view here yet.
        batch = self.make_registration_batch()
        self.client.force_login(self.admin_user)

        response = self.club_get("registration_invoice_detail", batch.pk)

        self.assertEqual(response.status_code, 404)

    def test_is_admin_only(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.make_shop_manager())

        response = self.club_get("registration_invoice_detail", batch.pk)

        self.assertEqual(response.status_code, 403)

    def test_shows_a_mark_as_paid_button_while_unpaid(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.admin_user)

        response = self.club_get("registration_invoice_detail", batch.pk)

        self.assertContains(response, reverse("management:registration_invoice_mark_paid", args=[batch.pk]))

    def test_hides_the_mark_as_paid_button_once_settled(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        membership = batch.entries.get().membership
        membership.fee_status = ClubMembership.FeeStatus.PAID
        membership.save()
        self.client.force_login(self.admin_user)

        response = self.club_get("registration_invoice_detail", batch.pk)

        self.assertNotContains(response, reverse("management:registration_invoice_mark_paid", args=[batch.pk]))


class RegistrationInvoiceMarkPaidViewTests(ShopTestBase):
    """management:registration_invoice_mark_paid -- the checkout screen's own
    secondary path to what Dues & billing's per-row/bulk "Mark ... paid"
    buttons already do."""

    def make_registration_batch(self, contact_email="reg-mark-paid@example.com", price=Decimal("80.00")):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=price)
        variant = ProductVariant.objects.create(product=product, name="Standard", price=price)
        entries = [EntryInput(first_name="Kid", last_name="Payer", date_of_birth=datetime.date(2016, 5, 1), product_variant=variant)]
        return submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email=contact_email, entries=entries)

    def test_marks_every_membership_in_the_batch_as_paid(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        membership = batch.entries.get().membership
        self.assertNotEqual(membership.fee_status, ClubMembership.FeeStatus.PAID)
        self.client.force_login(self.admin_user)

        response = self.club_post("registration_invoice_mark_paid", {}, batch.pk)

        self.assertRedirects(response, reverse("management:registration_invoice_detail", args=[batch.pk]))
        membership.refresh_from_db()
        self.assertEqual(membership.fee_status, ClubMembership.FeeStatus.PAID)

    def test_404s_before_the_invoice_is_confirmed(self):
        batch = self.make_registration_batch()
        self.client.force_login(self.admin_user)

        response = self.club_post("registration_invoice_mark_paid", {}, batch.pk)

        self.assertEqual(response.status_code, 404)

    def test_is_admin_only(self):
        batch = self.make_registration_batch()
        confirm_and_send_invoice(batch, due_in_days=14)
        batch.refresh_from_db()
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("registration_invoice_mark_paid", {}, batch.pk)

        self.assertEqual(response.status_code, 403)


class RegistrationInvoiceQueueViewTests(ManagementTestBase):
    """management:registration_invoice_queue -- the review queue of
    registrations whose invoice hasn't been confirmed/sent yet."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def make_batch(self, contact_email="reg-parent@example.com", price=Decimal("80.00")):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=price)
        variant = ProductVariant.objects.create(product=product, name="Standard", price=price)
        entries = [EntryInput(first_name="Kid", last_name="Testerson", date_of_birth=datetime.date(2016, 5, 1), product_variant=variant)]
        return submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email=contact_email, entries=entries)

    def test_is_admin_only(self):
        self.make_batch()
        coach_user = User.objects.create_user(email="coach-reginv@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=coach_user, first_name="Cara", last_name="Coach")
        team = Team.objects.create(club=self.club, name="U9", short_name="U9")
        position = Position.objects.create(club=self.club, name="Coach-reginv", short_name="CR", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=team, member=coach_member, season=self.season, position=position)
        self.client.force_login(coach_user)

        self.assertEqual(self.club_get("registration_invoice_queue").status_code, 403)

    def test_lists_an_unconfirmed_registration(self):
        self.make_batch()

        response = self.club_get("registration_invoice_queue")

        self.assertContains(response, "Reg Parent")
        self.assertContains(response, "Kid Testerson")
        self.assertContains(response, "80.00")

    def test_a_confirmed_registration_is_not_listed(self):
        batch = self.make_batch()
        confirm_and_send_invoice(batch, due_in_days=14)

        response = self.club_get("registration_invoice_queue")

        self.assertNotContains(response, "Reg Parent")

    def test_a_fully_cancelled_registration_is_not_listed(self):
        batch = self.make_batch()
        membership = batch.entries.get().membership
        membership.status = ClubMembership.StatusChoices.CANCELLED
        membership.save()

        response = self.club_get("registration_invoice_queue")

        self.assertNotContains(response, "Reg Parent")

    def test_search_matches_the_entrys_own_name(self):
        self.make_batch()

        hit = self.club_get("registration_invoice_queue", params={"q": "Testerson"})
        miss = self.club_get("registration_invoice_queue", params={"q": "Zzzznotfound"})

        self.assertContains(hit, "Reg Parent")
        self.assertNotContains(miss, "Reg Parent")

    def test_review_link_points_at_the_review_view(self):
        batch = self.make_batch()

        response = self.club_get("registration_invoice_queue")

        self.assertContains(response, reverse("management:registration_invoice_review", args=[batch.pk]))


class RegistrationInvoiceReviewViewTests(ManagementTestBase):
    """management:registration_invoice_review -- edit/exclude lines, apply
    an overall discount, then confirm & send in one action."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def make_batch(self, contact_email="reg-parent@example.com", price=Decimal("80.00")):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=price)
        variant = ProductVariant.objects.create(product=product, name="Standard", price=price)
        entries = [EntryInput(first_name="Kid", last_name="Testerson", date_of_birth=datetime.date(2016, 5, 1), product_variant=variant)]
        return submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email=contact_email, entries=entries)

    def _post_data(self, entries, price_overrides=None, excluded_ids=None, team_by_entry=None, jersey_number_by_entry=None, position_by_entry=None, **confirm_overrides):
        price_overrides = price_overrides or {}
        excluded_ids = excluded_ids or set()
        team_by_entry = team_by_entry or {}
        jersey_number_by_entry = jersey_number_by_entry or {}
        position_by_entry = position_by_entry or {}
        data = {
            "lines-TOTAL_FORMS": str(len(entries)),
            "lines-INITIAL_FORMS": str(len(entries)),
            "lines-MIN_NUM_FORMS": "0",
            "lines-MAX_NUM_FORMS": "1000",
            "manual_discount_amount": str(confirm_overrides.get("manual_discount_amount", "0")),
            "manual_discount_note": confirm_overrides.get("manual_discount_note", ""),
            "due_in_days": str(confirm_overrides.get("due_in_days", 14)),
        }
        for i, entry in enumerate(entries):
            data[f"lines-{i}-id"] = str(entry.pk)
            data[f"lines-{i}-price"] = str(price_overrides.get(entry.pk, entry.price))
            if entry.pk in excluded_ids:
                data[f"lines-{i}-excluded_from_invoice"] = "on"
            if entry.pk in team_by_entry:
                data[f"lines-{i}-requested_team"] = str(team_by_entry[entry.pk].pk)
            if entry.pk in jersey_number_by_entry:
                data[f"lines-{i}-requested_jersey_number"] = str(jersey_number_by_entry[entry.pk])
            if entry.pk in position_by_entry:
                data[f"lines-{i}-requested_position"] = str(position_by_entry[entry.pk].pk)
        return data

    def test_is_admin_only(self):
        batch = self.make_batch()
        plain_user = User.objects.create_user(email="plain-reginv@example.com", password="pw-secret-123")
        ClubMembership.objects.create(club=self.club, member=Member.objects.create(user=plain_user, first_name="Pat", last_name="Plain"), season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(plain_user)

        self.assertEqual(self.club_get("registration_invoice_review", batch.pk).status_code, 403)

    def test_get_renders_the_line_formset(self):
        batch = self.make_batch()

        response = self.club_get("registration_invoice_review", batch.pk)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Kid Testerson")

    def test_an_already_confirmed_batch_404s(self):
        batch = self.make_batch()
        confirm_and_send_invoice(batch, due_in_days=14)

        self.assertEqual(self.club_get("registration_invoice_review", batch.pk).status_code, 404)

    def test_editing_a_price_persists_it(self):
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries, price_overrides={entries[0].pk: Decimal("60.00")})

        self.club_post("registration_invoice_review", data, batch.pk)

        entries[0].refresh_from_db()
        self.assertEqual(entries[0].price, Decimal("60.00"))

    def test_editing_a_price_updates_the_memberships_fee_amount(self):
        # Without this, editing the price here only ever changed what this
        # batch's own PDF/queue total showed -- Dues & billing/mobile/
        # reminders all read ClubMembership.fee_amount directly and would
        # keep citing the original, unedited 80.00.
        batch = self.make_batch(price=Decimal("80.00"))
        entries = list(batch.entries.all())
        membership = entries[0].membership
        data = self._post_data(entries, price_overrides={entries[0].pk: Decimal("60.00")})

        self.club_post("registration_invoice_review", data, batch.pk)

        membership.refresh_from_db()
        self.assertEqual(membership.fee_amount, Decimal("60.00"))

    def test_excluding_a_line_drops_it_from_the_confirmed_total(self):
        product = Product.objects.create(club=self.club, name="Sibling Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("50.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("50.00"))
        entries_input = [
            EntryInput(first_name="Kid", last_name="One", date_of_birth=datetime.date(2015, 5, 1), product_variant=variant),
            EntryInput(first_name="Kid", last_name="Two", date_of_birth=datetime.date(2017, 5, 1), product_variant=variant),
        ]
        batch = submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email="reg-sib@example.com", entries=entries_input)
        entries = list(batch.entries.all())
        data = self._post_data(entries, excluded_ids={entries[0].pk})

        self.club_post("registration_invoice_review", data, batch.pk)

        entries[0].refresh_from_db()
        self.assertTrue(entries[0].excluded_from_invoice)
        batch.refresh_from_db()
        _subtotal, _discount_amount, total = batch_totals(active_batch_entries(batch), batch.manual_discount_amount)
        self.assertEqual(total, Decimal("50.00"))

    def test_excluding_a_line_zeroes_out_that_memberships_fee_amount(self):
        product = Product.objects.create(club=self.club, name="Sibling Registration 2", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("50.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("50.00"))
        entries_input = [
            EntryInput(first_name="Kid", last_name="Three", date_of_birth=datetime.date(2015, 5, 1), product_variant=variant),
            EntryInput(first_name="Kid", last_name="Four", date_of_birth=datetime.date(2017, 5, 1), product_variant=variant),
        ]
        batch = submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email="reg-sib2@example.com", entries=entries_input)
        entries = list(batch.entries.all())
        excluded_membership = entries[0].membership
        kept_membership = entries[1].membership
        data = self._post_data(entries, excluded_ids={entries[0].pk})

        self.club_post("registration_invoice_review", data, batch.pk)

        excluded_membership.refresh_from_db()
        kept_membership.refresh_from_db()
        self.assertEqual(excluded_membership.fee_amount, Decimal("0.00"))
        self.assertEqual(kept_membership.fee_amount, Decimal("50.00"))

    def test_a_manual_discount_reduces_the_total_and_keeps_its_note(self):
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries, manual_discount_amount="10.00", manual_discount_note="Loyalty discount")

        self.club_post("registration_invoice_review", data, batch.pk)

        batch.refresh_from_db()
        self.assertEqual(batch.manual_discount_amount, Decimal("10.00"))
        self.assertEqual(batch.manual_discount_note, "Loyalty discount")
        _subtotal, _discount_amount, total = batch_totals(active_batch_entries(batch), batch.manual_discount_amount)
        self.assertEqual(total, Decimal("70.00"))

    def test_a_manual_discount_reduces_the_memberships_fee_amount(self):
        batch = self.make_batch(price=Decimal("80.00"))
        entries = list(batch.entries.all())
        membership = entries[0].membership
        data = self._post_data(entries, manual_discount_amount="10.00")

        self.club_post("registration_invoice_review", data, batch.pk)

        membership.refresh_from_db()
        self.assertEqual(membership.fee_amount, Decimal("70.00"))

    def test_a_manual_discount_splits_proportionally_across_siblings(self):
        product = Product.objects.create(club=self.club, name="Sibling Registration 3", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("30.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("30.00"))
        other_product = Product.objects.create(club=self.club, name="Sibling Registration 4", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("70.00"))
        other_variant = ProductVariant.objects.create(product=other_product, name="Standard", price=Decimal("70.00"))
        entries_input = [
            EntryInput(first_name="Kid", last_name="Five", date_of_birth=datetime.date(2015, 5, 1), product_variant=variant),
            EntryInput(first_name="Kid", last_name="Six", date_of_birth=datetime.date(2017, 5, 1), product_variant=other_variant),
        ]
        batch = submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email="reg-sib3@example.com", entries=entries_input)
        entries = list(batch.entries.all())
        membership_a = entries[0].membership
        membership_b = entries[1].membership
        # 30/100 and 70/100 of the batch -- a 10.00 discount splits 3.00/7.00.
        data = self._post_data(entries, manual_discount_amount="10.00")

        self.club_post("registration_invoice_review", data, batch.pk)

        membership_a.refresh_from_db()
        membership_b.refresh_from_db()
        self.assertEqual(membership_a.fee_amount, Decimal("27.00"))
        self.assertEqual(membership_b.fee_amount, Decimal("63.00"))
        # The split is always exact -- never lost (or gained) to rounding.
        self.assertEqual((Decimal("30.00") - membership_a.fee_amount) + (Decimal("70.00") - membership_b.fee_amount), Decimal("10.00"))

    def test_confirming_allocates_a_number_and_sends_the_email(self):
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries)

        response = self.club_post("registration_invoice_review", data, batch.pk)

        self.assertRedirects(response, reverse("management:registration_invoice_queue"))
        batch.refresh_from_db()
        self.assertTrue(batch.invoice_number.startswith("REG-"))
        self.assertIsNotNone(batch.invoice_sent_at)
        self.assertEqual(batch.invoice_due_date, timezone.localdate() + datetime.timedelta(days=14))
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["reg-parent@example.com"])

    def test_confirming_emails_every_parent_guardian_on_record_not_just_the_submitter(self):
        batch = self.make_batch(contact_email="bernard@example.com")
        entries = list(batch.entries.all())
        child = entries[0].membership.member
        family = child.family_memberships.first().family
        second_parent = Member.objects.create(first_name="Charlotte", last_name="Baes", email="charlotte@example.com")
        FamilyMembership.objects.create(family=family, member=second_parent, role=FamilyMembership.FamilyRole.PARENT)
        data = self._post_data(entries)

        self.club_post("registration_invoice_review", data, batch.pk)

        self.assertEqual(len(mail.outbox), 1)
        self.assertCountEqual(mail.outbox[0].to, ["bernard@example.com", "charlotte@example.com"])

    def test_excluding_every_line_does_not_confirm(self):
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries, excluded_ids={entries[0].pk})

        self.club_post("registration_invoice_review", data, batch.pk)

        batch.refresh_from_db()
        self.assertIsNone(batch.invoice_sent_at)

    def test_confirming_a_line_with_a_team_and_number_places_the_player(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth Confirm", min_number=1, max_number=99)
        team = Team.objects.create(club=self.club, name="U9", short_name="U9", pool=pool)
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries, team_by_entry={entries[0].pk: team}, jersey_number_by_entry={entries[0].pk: 7})

        self.club_post("registration_invoice_review", data, batch.pk)

        entries[0].refresh_from_db()
        member = entries[0].membership.member
        team_membership = TeamMembership.objects.get(team=team, season=self.season, member=member)
        self.assertEqual(team_membership.jersey_number, 7)
        self.assertEqual(entries[0].resulting_team_membership, team_membership)

    def test_confirming_a_line_with_no_team_chosen_does_not_block_confirmation(self):
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries)

        response = self.club_post("registration_invoice_review", data, batch.pk)

        self.assertRedirects(response, reverse("management:registration_invoice_queue"))
        entries[0].refresh_from_db()
        self.assertIsNone(entries[0].resulting_team_membership)
        self.assertFalse(TeamMembership.objects.filter(member=entries[0].membership.member).exists())

    def test_confirming_a_volunteer_line_with_a_team_creates_a_staff_assignment(self):
        team = Team.objects.create(club=self.club, name="U11", short_name="U11")
        position = Position.objects.create(club=self.club, name="Assistant Coach", short_name="AC", staff_position=True)
        product = Product.objects.create(club=self.club, name="Volunteer Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("0.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("0.00"))
        entries_input = [EntryInput(first_name="Val", last_name="Volunteer", date_of_birth=datetime.date(1985, 5, 1), entry_kind=RegistrationDetails.EntryKind.VOLUNTEER, product_variant=variant)]
        batch = submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email="reg-vol@example.com", entries=entries_input)
        entries = list(batch.entries.all())
        data = self._post_data(entries, team_by_entry={entries[0].pk: team}, position_by_entry={entries[0].pk: position})

        self.club_post("registration_invoice_review", data, batch.pk)

        entries[0].refresh_from_db()
        member = entries[0].membership.member
        staff_assignment = StaffAssignment.objects.get(team=team, member=member, season=self.season)
        self.assertEqual(staff_assignment.position, position)
        self.assertEqual(entries[0].resulting_staff_assignment, staff_assignment)

    def test_confirming_rejects_a_jersey_number_taken_since_it_was_requested(self):
        pool = NumberPool.objects.create(club=self.club, name="Youth", min_number=1, max_number=99)
        team = Team.objects.create(club=self.club, name="U13", short_name="U13", pool=pool)
        other_member = Member.objects.create(first_name="Other", last_name="Player")
        TeamMembership.objects.create(team=team, season=self.season, member=other_member, jersey_number=7)
        batch = self.make_batch()
        entries = list(batch.entries.all())
        data = self._post_data(entries, team_by_entry={entries[0].pk: team}, jersey_number_by_entry={entries[0].pk: 7})

        response = self.club_post("registration_invoice_review", data, batch.pk)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["formset"].is_valid())
        batch.refresh_from_db()
        self.assertIsNone(batch.invoice_sent_at)


class RegistrationBatchCancelViewTests(ManagementTestBase):
    """management:registration_invoice_cancel -- withdraws a whole
    not-yet-confirmed registration in one action."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def make_batch(self, contact_email="reg-cancel@example.com"):
        product = Product.objects.create(club=self.club, name="Player Registration Cancel", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("80.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("80.00"))
        entries = [
            EntryInput(first_name="Kid", last_name="One", date_of_birth=datetime.date(2015, 5, 1), product_variant=variant),
            EntryInput(first_name="Kid", last_name="Two", date_of_birth=datetime.date(2017, 5, 1), product_variant=variant),
        ]
        return submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email=contact_email, entries=entries)

    def test_is_admin_only(self):
        batch = self.make_batch()
        plain_user = User.objects.create_user(email="plain-regcancel@example.com", password="pw-secret-123")
        ClubMembership.objects.create(club=self.club, member=Member.objects.create(user=plain_user, first_name="Pat", last_name="Plain"), season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(plain_user)

        self.assertEqual(self.club_post("registration_invoice_cancel", {}, batch.pk).status_code, 403)

    def test_cancels_every_membership_in_the_batch(self):
        # Both kids here only exist because of this one registration -- per
        # club.services.cancellation.cancel_membership's own docstring, that
        # means each cancel deletes the Member outright rather than leaving
        # a cancelled shell, so what's checked is that neither membership
        # (nor the ClubMembership row backing it) survives, not a CANCELLED
        # status on a row that no longer exists.
        batch = self.make_batch()
        membership_ids = [entry.membership_id for entry in batch.entries.all()]

        response = self.club_post("registration_invoice_cancel", {}, batch.pk)

        self.assertRedirects(response, reverse("management:registration_invoice_queue"))
        self.assertFalse(ClubMembership.objects.filter(pk__in=membership_ids).exists())

    def test_cancels_a_returning_members_membership_without_deleting_them(self):
        batch = self.make_batch()
        entry = batch.entries.first()
        member = entry.membership.member
        # A second, unrelated season is enough to make this member "returning"
        # rather than "born this season" -- cancel_membership's own is_new_member
        # check is club-wide across seasons, not scoped to this one registration.
        other_season = Season.objects.create(club=self.club, start_date=self.season.start_date.replace(year=self.season.start_date.year - 1), end_date=self.season.end_date.replace(year=self.season.end_date.year - 1))
        ClubMembership.objects.create(club=self.club, member=member, season=other_season, status=ClubMembership.StatusChoices.ACTIVE)

        response = self.club_post("registration_invoice_cancel", {}, batch.pk)

        self.assertRedirects(response, reverse("management:registration_invoice_queue"))
        entry.membership.refresh_from_db()
        self.assertEqual(entry.membership.status, ClubMembership.StatusChoices.CANCELLED)
        self.assertTrue(Member.objects.filter(pk=member.pk).exists())

    def test_an_already_confirmed_batch_404s(self):
        batch = self.make_batch()
        confirm_and_send_invoice(batch, due_in_days=14)

        self.assertEqual(self.club_post("registration_invoice_cancel", {}, batch.pk).status_code, 404)


class RegistrationInvoiceReminderTests(ManagementTestBase):
    """The Dues & billing "Send reminders" button and its overdue KPI now
    cover confirmed registration invoices too, not just DuesInvoice."""

    def setUp(self):
        self.client.force_login(self.admin_user)

    def make_confirmed_overdue_batch(self):
        product = Product.objects.create(club=self.club, name="Player Registration", product_type=Product.ProductType.MEMBERSHIP, season=self.season, price=Decimal("80.00"))
        variant = ProductVariant.objects.create(product=product, name="Standard", price=Decimal("80.00"))
        entries = [EntryInput(first_name="Kid", last_name="Testerson", date_of_birth=datetime.date(2016, 5, 1), product_variant=variant)]
        batch = submit_registration(self.club, contact_first_name="Reg", contact_last_name="Parent", contact_email="reg-overdue@example.com", entries=entries)
        confirm_and_send_invoice(batch, due_in_days=1)
        batch.invoice_due_date = timezone.localdate() - datetime.timedelta(days=5)
        batch.save(update_fields=["invoice_due_date"])
        return batch

    def test_overdue_registration_counts_toward_the_kpi(self):
        self.make_confirmed_overdue_batch()

        response = self.club_get("membership_list")

        self.assertEqual(response.context["kpi_overdue_invoices"], 1)

    def test_send_reminders_emails_an_overdue_registration(self):
        batch = self.make_confirmed_overdue_batch()
        mail.outbox.clear()

        self.club_post("membership_send_invoice_reminders", {})

        self.assertEqual(len(mail.outbox), 1)
        batch.refresh_from_db()
        self.assertEqual(batch.invoice_reminder_count, 1)


class ShopToggleViewTests(ShopTestBase):
    def test_an_admin_can_open_the_shop(self):
        self.assertFalse(self.club.shop_open)
        self.client.force_login(self.admin_user)

        response = self.club_post("shop_toggle", {})

        self.assertRedirects(response, reverse("management:product_list"))
        self.club.refresh_from_db()
        self.assertTrue(self.club.shop_open)

    def test_a_shop_manager_can_close_the_shop(self):
        self.club.shop_open = True
        self.club.save(update_fields=["shop_open"])
        self.client.force_login(self.make_shop_manager())

        self.club_post("shop_toggle", {})

        self.club.refresh_from_db()
        self.assertFalse(self.club.shop_open)

    def test_plain_staff_cannot_toggle_the_shop(self):
        self.client.force_login(self.make_plain_staff())

        response = self.club_post("shop_toggle", {})

        self.assertEqual(response.status_code, 403)
        self.club.refresh_from_db()
        self.assertFalse(self.club.shop_open)


class ShopManagerGrantRevokeTests(ShopTestBase):
    """Granting/revoking the ShopManager flag itself -- admin-only ground (see
    ClubRoleCreateView/ShopManagerRevokeView's own docstrings), unlike the shop
    views it unlocks. Granting goes through the roles page's single "Grant
    role" flow ("Shop admin" is one of its dropdown values, not a separate
    modal/endpoint) -- revoking still has its own view, since ShopManager
    isn't a ClubRole row."""

    def test_an_admin_can_grant_shop_manager(self):
        target = Member.objects.create(first_name="Target", last_name="Person")
        ClubMembership.objects.create(club=self.club, member=target, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)

        response = self.club_post("role_create", {"member": str(target.pk), "role": "shop_admin"})

        self.assertRedirects(response, reverse("management:role_list"))
        self.assertTrue(ShopManager.objects.filter(club=self.club, member=target).exists())

    def test_a_shop_manager_cannot_grant_shop_manager_to_someone_else(self):
        target = Member.objects.create(first_name="Target", last_name="Person")
        ClubMembership.objects.create(club=self.club, member=target, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.make_shop_manager())

        response = self.club_post("role_create", {"member": str(target.pk), "role": "shop_admin"})

        self.assertEqual(response.status_code, 403)
        self.assertFalse(ShopManager.objects.filter(club=self.club, member=target).exists())

    def test_an_admin_can_revoke_shop_manager(self):
        shop_manager_user = self.make_shop_manager()
        grant = ShopManager.objects.get(club=self.club, member__user=shop_manager_user)
        self.client.force_login(self.admin_user)

        response = self.club_post("shop_admin_revoke", {}, grant.pk)

        self.assertRedirects(response, reverse("management:role_list"))
        self.assertFalse(ShopManager.objects.filter(pk=grant.pk).exists())

    def test_a_shop_manager_cannot_revoke_their_own_grant(self):
        shop_manager_user = self.make_shop_manager()
        grant = ShopManager.objects.get(club=self.club, member__user=shop_manager_user)
        self.client.force_login(shop_manager_user)

        response = self.club_post("shop_admin_revoke", {}, grant.pk)

        self.assertEqual(response.status_code, 403)
        self.assertTrue(ShopManager.objects.filter(pk=grant.pk).exists())

    def test_granting_shop_manager_is_additive_to_an_existing_editor_role(self):
        target_user = User.objects.create_user(email="editor-shop@example.com", password="pw-secret-123")
        target = Member.objects.create(user=target_user, first_name="Edie", last_name="Editor")
        ClubMembership.objects.create(club=self.club, member=target, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=self.club, member=target).update(role=ClubRole.Roles.EDITOR)
        enrol_mfa(target_user)  # ClubRole EDITOR requires a second factor.
        self.client.force_login(self.admin_user)

        self.club_post("role_create", {"member": str(target.pk), "role": "shop_admin"})

        self.assertTrue(ShopManager.objects.filter(club=self.club, member=target).exists())
        self.assertEqual(ClubRole.objects.get(club=self.club, member=target).role, ClubRole.Roles.EDITOR)

        self.client.force_login(target_user)
        self.assertEqual(self.club_get("product_list").status_code, 200)

    def test_the_shop_admins_section_appears_on_the_roles_page(self):
        self.make_shop_manager(email="listed-shop-admin@example.com")
        self.client.force_login(self.admin_user)

        response = self.club_get("role_list")

        self.assertContains(response, "Shop admins")
        self.assertContains(response, "Sam ShopAdmin")

    def test_shop_admin_is_an_option_in_the_single_grant_role_modal(self):
        # Not a second modal/endpoint -- see this class's own docstring.
        self.client.force_login(self.admin_user)

        response = self.club_get("role_list")

        self.assertContains(response, 'id="grant_role_modal"')
        self.assertNotContains(response, 'id="grant_shop_admin_modal"')
        self.assertContains(response, "Shop admin")




class FormManagementTests(ManagementTestBase):
    """Form/Field/FormSend CRUD -- FormManagerRequiredMixin (ADMIN/EDITOR/
    coach-manager, gated behind the "formbuilder" waffle flag, see
    can_manage_forms/club.mixins.FormManagerRequiredMixin)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.team = Team.objects.create(club=cls.club, name="First Team", short_name="1st")

        cls.coach_manager = User.objects.create_user(email="coach-forms@example.com", password="pw-secret-123")
        coach_member = Member.objects.create(user=cls.coach_manager, first_name="Cara", last_name="Coach")
        coach_position = Position.objects.create(club=cls.club, name="Head Coach", short_name="HC", staff_position=True, management_position=True)
        StaffAssignment.objects.create(team=cls.team, member=coach_member, season=cls.season, position=coach_position)

        cls.plain_staff = User.objects.create_user(email="physio-forms@example.com", password="pw-secret-123")
        staff_member = Member.objects.create(user=cls.plain_staff, first_name="Pat", last_name="Physio")
        staff_position = Position.objects.create(club=cls.club, name="Physio", short_name="PH", staff_position=True, management_position=False)
        StaffAssignment.objects.create(team=cls.team, member=staff_member, season=cls.season, position=staff_position)

        cls.editor = User.objects.create_user(email="editor-forms@example.com", password="pw-secret-123")
        editor_member = Member.objects.create(user=cls.editor, first_name="Eve", last_name="Editor")
        ClubMembership.objects.create(club=cls.club, member=editor_member, season=cls.season, status=ClubMembership.StatusChoices.ACTIVE)
        ClubRole.objects.filter(club=cls.club, member=editor_member).update(role=ClubRole.Roles.EDITOR)
        enrol_mfa(cls.editor)

    def setUp(self):
        super().setUp()
        cache.clear()
        self.addCleanup(cache.clear)
        flag = get_waffle_flag_model().objects.create(name="formbuilder")
        flag.clubs.add(self.club)

    def make_form(self, **kwargs):
        kwargs.setdefault("club", self.club)
        kwargs.setdefault("title", "Sign-up")
        return FormBuilderForm.objects.create(**kwargs)

    # --- permissions -- broader than plain ClubAdminRequiredMixin, unlike
    # the flag-gating itself (FeatureGatedSectionsTests) which only proves
    # an admin can reach it. ---

    def test_coach_manager_can_reach_form_list(self):
        self.client.force_login(self.coach_manager)

        self.assertEqual(self.club_get("form_list").status_code, 200)

    def test_editor_can_reach_form_list(self):
        self.client.force_login(self.editor)

        self.assertEqual(self.club_get("form_list").status_code, 200)

    def test_plain_staff_gets_403(self):
        self.client.force_login(self.plain_staff)

        self.assertEqual(self.club_get("form_list").status_code, 403)

    # --- create ---

    def test_creates_a_form_with_its_questions(self):
        self.client.force_login(self.admin_user)
        data = {
            "title": "Sign-up",
            "description": "",
            "login_required": "",
            "is_active": "on",
            "fields-TOTAL_FORMS": "2",
            "fields-INITIAL_FORMS": "0",
            "fields-MIN_NUM_FORMS": "0",
            "fields-MAX_NUM_FORMS": "1000",
            "fields-0-label": "Shirt size",
            "fields-0-field_type": "choice",
            "fields-0-required": "on",
            "fields-0-options": "S\nM\nL",
            "fields-1-label": "",
            "fields-1-field_type": "text",
            "fields-1-required": "",
            "fields-1-options": "",
        }

        response = self.club_post("form_create", data)

        form_obj = FormBuilderForm.objects.get(club=self.club, title="Sign-up")
        self.assertRedirects(response, reverse("management:form_detail", args=[form_obj.pk]))
        field = form_obj.fields.get()
        self.assertEqual(field.label, "Shirt size")
        self.assertEqual(field.options, ["S", "M", "L"])
        # A blank key here means the row was bulk_create()'d straight to the
        # DB, bypassing Field.save()'s own auto-slugify -- and every question
        # created that way collapses onto the same `name=""` dynamic form
        # field, which a browser's own form serialization drops, making a
        # member's answer look permanently unanswered/required no matter
        # what they actually picked.
        self.assertEqual(field.key, "shirt-size")

    def test_each_question_gets_its_own_distinct_key(self):
        self.client.force_login(self.admin_user)
        data = {
            "title": "Sign-up",
            "description": "",
            "login_required": "",
            "is_active": "on",
            "fields-TOTAL_FORMS": "2",
            "fields-INITIAL_FORMS": "0",
            "fields-MIN_NUM_FORMS": "0",
            "fields-MAX_NUM_FORMS": "1000",
            "fields-0-label": "First question",
            "fields-0-field_type": "text",
            "fields-0-required": "on",
            "fields-0-options": "",
            "fields-1-label": "Second question",
            "fields-1-field_type": "text",
            "fields-1-required": "on",
            "fields-1-options": "",
        }

        self.club_post("form_create", data)

        form_obj = FormBuilderForm.objects.get(club=self.club, title="Sign-up")
        keys = list(form_obj.fields.order_by("order").values_list("key", flat=True))
        self.assertEqual(keys, ["first-question", "second-question"])
        self.assertNotIn("", keys)

    def test_a_blank_question_row_is_skipped(self):
        self.client.force_login(self.admin_user)
        data = {
            "title": "Sign-up",
            "description": "",
            "login_required": "",
            "is_active": "on",
            "fields-TOTAL_FORMS": "1",
            "fields-INITIAL_FORMS": "0",
            "fields-MIN_NUM_FORMS": "0",
            "fields-MAX_NUM_FORMS": "1000",
            "fields-0-label": "",
            "fields-0-field_type": "text",
            "fields-0-required": "",
            "fields-0-options": "",
        }

        self.club_post("form_create", data)

        form_obj = FormBuilderForm.objects.get(club=self.club, title="Sign-up")
        self.assertFalse(form_obj.fields.exists())

    def test_duplicate_question_text_is_rejected(self):
        self.client.force_login(self.admin_user)
        data = {
            "title": "Sign-up",
            "description": "",
            "login_required": "",
            "is_active": "on",
            "fields-TOTAL_FORMS": "2",
            "fields-INITIAL_FORMS": "0",
            "fields-MIN_NUM_FORMS": "0",
            "fields-MAX_NUM_FORMS": "1000",
            "fields-0-label": "Name",
            "fields-0-field_type": "text",
            "fields-0-required": "",
            "fields-0-options": "",
            "fields-1-label": "Name",
            "fields-1-field_type": "text",
            "fields-1-required": "",
            "fields-1-options": "",
        }

        self.club_post("form_create", data)

        self.assertFalse(FormBuilderForm.objects.filter(club=self.club, title="Sign-up").exists())

    # --- fields ---

    def test_can_add_a_question_from_the_detail_page(self):
        form_obj = self.make_form()
        self.client.force_login(self.admin_user)

        self.club_post("form_field_create", {"label": "Age", "field_type": "number", "required": "on", "help_text": "", "order": "1", "options": ""}, form_obj.pk)

        self.assertTrue(form_obj.fields.filter(label="Age").exists())

    def test_toggling_a_question_flips_is_active(self):
        form_obj = self.make_form()
        field = FormBuilderField.objects.create(form=form_obj, label="Name", order=1)
        self.client.force_login(self.admin_user)

        self.club_post("form_field_toggle_active", {}, form_obj.pk, field.pk)

        field.refresh_from_db()
        self.assertFalse(field.is_active)

    def test_a_question_from_another_club_404s(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc-forms")
        other_form = FormBuilderForm.objects.create(club=other_club, title="Other")
        other_field = FormBuilderField.objects.create(form=other_form, label="X", order=1)
        self.client.force_login(self.admin_user)

        response = self.club_post("form_field_toggle_active", {}, other_form.pk, other_field.pk)

        self.assertEqual(response.status_code, 404)

    # --- sends ---

    def test_creating_a_send_notifies_the_resolved_audience(self):
        form_obj = self.make_form()
        member = Member.objects.create(first_name="Jane", last_name="Doe")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        self.client.force_login(self.admin_user)
        data = {
            "club_wide": "",
            "teams": [str(self.team.pk)],
            "groups": [],
            "invited_members": [str(member.pk)],
            "excluded_members": [],
            "opens_at": "",
            "closes_at": "",
            "max_submissions_per_user": "",
            "is_active": "on",
        }

        with mock.patch("management.views.dispatch_notify_form_send", side_effect=notify_form_send):
            response = self.club_post("formsend_create", data, form_obj.pk)

        send = FormSend.objects.get(form=form_obj)
        self.assertRedirects(response, reverse("management:formsend_responses", args=[form_obj.pk, send.pk]))
        self.assertTrue(Notification.objects.filter(member=member).exists())

    def test_a_send_from_another_club_404s(self):
        other_club = Club.objects.create(name="Rival FC", slug="rival-fc-forms-2")
        other_form = FormBuilderForm.objects.create(club=other_club, title="Other")
        other_send = FormSend.objects.create(club=other_club, form=other_form)
        self.client.force_login(self.admin_user)

        response = self.club_get("formsend_responses", other_form.pk, other_send.pk)

        self.assertEqual(response.status_code, 404)

    # --- responses ---

    def test_responses_page_shows_the_audience_summary_and_the_table_together(self):
        form_obj = self.make_form()
        send = FormSend.objects.create(club=self.club, form=form_obj, club_wide=True)
        member = Member.objects.create(first_name="Jane", last_name="Doe")
        ClubMembership.objects.create(club=self.club, member=member, season=self.season, status=ClubMembership.StatusChoices.ACTIVE)
        Submission.objects.create(send=send, member=member)
        self.client.force_login(self.admin_user)

        response = self.club_get("formsend_responses", form_obj.pk, send.pk)

        self.assertEqual(response.status_code, 200)
        # club_wide -- also picks up every other active member setUpTestData
        # created (admin, editor), not just Jane -- hence >=1, not ==1.
        self.assertGreaterEqual(response.context["audience_count"], 1)
        self.assertEqual(response.context["submission_count"], 1)
        self.assertEqual(response.context["audience_count"] - response.context["not_yet_count"], 1)
        self.assertEqual(response.context["report"].count, 1)

    def test_responses_export_is_a_csv_of_the_sends_own_submissions(self):
        form_obj = self.make_form()
        field = FormBuilderField.objects.create(form=form_obj, key="name", label="Name", order=1)
        send = FormSend.objects.create(club=self.club, form=form_obj, club_wide=True)
        member = Member.objects.create(first_name="Jane", last_name="Doe")
        submission = Submission.objects.create(send=send, member=member)
        Answer.objects.create(submission=submission, field=field, value="Jane")
        self.client.force_login(self.admin_user)

        response = self.club_get("formsend_responses_export", form_obj.pk, send.pk)

        self.assertEqual(response["Content-Type"], "text/csv")
        content = response.content.decode()
        self.assertIn("Jane", content)
